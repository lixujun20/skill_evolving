"""SpreadsheetBench-Verified adapter.

The first implementation targets a credible baseline scaffold: the model writes
Python/openpyxl code against a copied workbook, the runner executes it, and the
verifier compares the declared answer range with the golden workbook.
"""
from __future__ import annotations

import asyncio
import ast
import copy
import hashlib
import json
import os
import re
import shutil
import subprocess
import tarfile
import tempfile
import textwrap
import time
import urllib.request
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl

from academic.benchmarks.core.artifacts import ArtifactStore
from academic.benchmarks.core.cost_accounting import make_cost_event
from academic.benchmarks.core.llm_text import ask_text_llm
from academic.benchmarks.core.maintenance_adapter import MaintenanceRunConfig, NoOpMaintenanceAdapter
from academic.benchmarks.core.skill_injector import BudgetSkillInjector
from academic.benchmarks.core.types import (
    BenchmarkResult,
    BenchmarkTask,
    SkillArtifact,
    SkillBundleCase,
    SkillEvidence,
    SkillInterface,
    SkillLineage,
    SkillTestCaseRun,
    SkillTestResult,
)
from academic.config import CODE_EXEC_TIMEOUT
from academic.skill_repository.llm_maintenance import (
    _ask_json,
    apply_refine_payload,
    normalize_skill_name,
    refine_skill_artifact_llm,
    summarize_dependency_context,
)

DATASET_URL = (
    "https://huggingface.co/datasets/KAKA22/SpreadsheetBench/resolve/main/"
    "spreadsheetbench_verified_400.tar.gz"
)

SPREADSHEET_SYSTEM = """You are a spreadsheet manipulation agent.

Write Python code using openpyxl to modify the workbook at INPUT_XLSX and save
the final answer workbook to OUTPUT_XLSX. Do not explain instead of editing the
file. Preserve sheets, formats, and unrelated cells when possible.

Retrieved skill package guidance:
{skills}

Callable function skills:
{callable_skills}

Return exactly one fenced python code block.
"""

SPREADSHEET_EXTRACT_SYSTEM = """\
You are the SpreadsheetBench skill extractor in a benchmark-agnostic skill
evolution system.

You receive one compact SpreadsheetBench task trace. Extract reusable,
testable spreadsheet skills only when the trace evidence supports them.

Field semantics:
- `artifacts`: [] when the trace is failed, speculative, too local, or already
  covered by an existing artifact.
- `name`: narrow snake_case name.
- `kind`: use `executable_tool` when the body contains a concrete reusable
  openpyxl code idiom; use `workflow_guardrail_card` for ordering/inspection
  workflows; use `interface_contract_card` for exact workbook/range contracts.
- `body`: actionable guidance shown to the model. For function/code skills,
  include a concise executable openpyxl snippet, required variables
  INPUT_XLSX/OUTPUT_XLSX, and non-applicability.
- `metadata.domains`: must include `SpreadsheetBench` and exact instruction
  type(s), not broad "all".
- `metadata.allowed_tools`: use ["openpyxl"].
- `metadata.intent_keywords`: terms that should retrieve this skill.
- `metadata.source_task_ids`: current task id.
- `metadata.evidence_span`: the compact trace/code/verifier evidence.
- `metadata.non_applicability`: when not to use the skill.
- `dependencies`: named skills this artifact explicitly relies on; [] when
  none.

Rules:
1. Preserve workbook sheets, formulas, styles, and unrelated cells unless the
   task explicitly requires replacement.
2. Prefer reusable openpyxl idioms over task transcripts.
3. Do not copy full code. Keep snippets short and parameterized by active sheet,
   answer range, or detected headers.
4. Do not invent benchmark answers or hidden workbook structure.
5. If the trace failed or score is below 0.9, extract only when verifier
   mismatches or stderr prove a concrete corrective contract. For example,
   predicted-vs-expected formula mismatches can become a narrow formula-pattern
   skill; do not extract from opaque failures.
6. Return strict JSON only. End every object and array explicitly.

Return schema:
{
  "artifacts": [
    {
      "name": "snake_case_name",
      "kind": "executable_tool | workflow_guardrail_card | interface_contract_card",
      "description": "short summary",
      "body": "actionable content",
      "interface": {
        "summary": "one-line contract",
        "usage": "when/how to use",
        "input_contract": {},
        "output_contract": {},
        "invocation_contract": {"injection_type": "informational"},
        "compatibility_notes": "non-applicability"
      },
      "metadata": {
        "domains": ["SpreadsheetBench"],
        "allowed_tools": ["openpyxl"],
        "intent_keywords": [],
        "source_task_ids": [],
        "source": "spreadsheet_llm_trace_extraction",
        "evidence_span": "",
        "scope": "",
        "non_applicability": "",
        "maintenance_action": "new_skill"
      },
      "dependencies": []
    }
  ]
}
"""

SPREADSHEET_CREDIT_SYSTEM = """\
You are the SpreadsheetBench credit assigner and maintenance-attribution judge.

You receive one compact task trace and only the retrieved candidate skills.
Judge whether each skill was helpful, harmful, neutral, or uncertain for this
specific task.

Field semantics:
- `judgment`: helpful if the skill likely improved correctness or efficiency;
  harmful if it likely caused wrong cells, wrong sheet/range, execution errors,
  formula/value mistakes, or irrelevant prompt pollution; neutral when present
  but irrelevant; uncertain when evidence is insufficient.
- `effect_type`: use one of correctness_gain, correctness_harm, token_saving,
  token_overhead, workflow_alignment, workflow_pollution, schema_help,
  schema_harm, domain_match, domain_mismatch, no_material_effect, unknown.
- `maintenance_actions`: skill-local actions. Use [] for neutral/uncertain.
- `refine_required`: true only when the skill should be edited before bundle
  testing due to concrete scope/schema/workflow/code evidence.
- `filter_candidate`: true only if the skill should be disabled or removed from
  retrieval, not merely because this task failed.
- `evidence_strength`: strong for direct retrieved-skill/code/verifier evidence,
  medium for close trace match, weak for circumstantial prompt noise.
- `attribution_scope`: prompt_influence, direct_use, retrieval_noise,
  integration_context, or none.
- `bundle_case_suggestions`: focused SpreadsheetBench task cases. The case
  should use only the official task snapshot and answer range. Do not invent a
  new golden workbook, answer cells, or expected values.
- `focus_turn_indices`: SpreadsheetBench is single-turn; use [0] only when a
  replayable official task fragment exists, otherwise [].
- `task_fragment_policy`: reuse_official_fragment when the official workbook,
  instruction, and answer range are enough to replay; no_replayable_fragment
  otherwise.

Rules:
1. Retrieved does not mean helpful. Judge causality from the code, stderr, cell
   mismatches, score, and skill scope.
2. For successful tasks with a relevant retrieved skill, positive suggestions
   require explicit correctness_gain, workflow_alignment, schema_help, or
   token_saving.
3. For failed tasks, mark harmful only when the failure matches the skill's
   instruction or code pattern, or the skill's scope clearly conflicts with the
   task.
4. If no candidate skill is causally implicated, return neutral/uncertain.
5. Return strict JSON only. End every object and array explicitly.

Return schema:
{
  "task_summary": {
    "task_id": "",
    "score": 0.0,
    "success": false,
    "total_tokens": 0
  },
  "skill_judgments": [
    {
      "skill_name": "",
      "judgment": "helpful | harmful | neutral | uncertain",
      "effect_type": "correctness_gain | correctness_harm | token_saving | token_overhead | workflow_alignment | workflow_pollution | schema_help | schema_harm | domain_match | domain_mismatch | no_material_effect | unknown",
      "confidence": 0.0,
      "reason": "",
      "maintenance_actions": [
        {
          "action": "keep | narrow_scope | fix_schema_contract | refine_workflow | disable_candidate | add_bundle_case | record_evidence",
          "reason": "",
          "target_scope": ""
        }
      ],
      "refine_required": false,
      "filter_candidate": false,
      "evidence_strength": "strong | medium | weak",
      "attribution_scope": "direct_use | prompt_influence | retrieval_noise | integration_context | none",
      "bundle_case_suggestions": [
        {
          "polarity": "positive | negative | integration",
          "reason": "",
          "source_task_id": "",
          "focus_turn_indices": [0],
          "required_context_turn_indices": [],
          "state_requirements": {},
          "expected_contract": "",
          "task_fragment_policy": "reuse_official_fragment | no_replayable_fragment"
        }
      ],
      "evidence": {
        "retrieved": true,
        "used": false,
        "relevant_turn_indices": [0],
        "related_tool_names": ["openpyxl"],
        "error_refs": [],
        "trace_signals": []
      }
    }
  ]
}
"""


@dataclass
class SpreadsheetTrace:
    task_id: str
    prompt: str = ""
    code: str = ""
    stdout: str = ""
    stderr: str = ""
    elapsed_s: float = 0.0
    retrieved_skills: List[str] = field(default_factory=list)
    prompt_injected_skills: List[str] = field(default_factory=list)
    callable_skills: List[Dict[str, Any]] = field(default_factory=list)
    filtered_skills: List[Dict[str, Any]] = field(default_factory=list)
    injector_events: List[Dict[str, Any]] = field(default_factory=list)
    cost_events: List[Dict[str, Any]] = field(default_factory=list)
    total_tokens: int = 0
    input_tokens: int = 0
    cache_input_tokens: int = 0
    completion_tokens: int = 0

    def as_dict(self) -> Dict[str, Any]:
        return {
            "task_id": self.task_id,
            "prompt": self.prompt,
            "code": self.code,
            "stdout": self.stdout,
            "stderr": self.stderr,
            "elapsed_s": self.elapsed_s,
            "retrieved_skills": self.retrieved_skills,
            "prompt_injected_skills": self.prompt_injected_skills,
            "callable_skills": self.callable_skills,
            "filtered_skills": self.filtered_skills,
            "injector_events": self.injector_events,
            "cost_events": self.cost_events,
            "total_tokens": self.total_tokens,
            "input_tokens": self.input_tokens,
            "cache_input_tokens": self.cache_input_tokens,
            "completion_tokens": self.completion_tokens,
        }


def ensure_spreadsheetbench(cache_dir: Path, refresh: bool = False) -> Path:
    cache_dir.mkdir(parents=True, exist_ok=True)
    archive = cache_dir / "spreadsheetbench_verified_400.tar.gz"
    extracted = cache_dir / "spreadsheetbench_verified_400"
    if refresh or not archive.exists():
        with urllib.request.urlopen(DATASET_URL, timeout=180) as response:
            archive.write_bytes(response.read())
    if refresh or not extracted.exists():
        if extracted.exists():
            shutil.rmtree(extracted)
        with tarfile.open(archive) as tf:
            tf.extractall(cache_dir)
    return extracted


def load_spreadsheet_tasks(
    *,
    cache_dir: Path,
    n_train: int = 200,
    n_test: int = 200,
    split_seed: int = 42,
    refresh: bool = False,
) -> Tuple[List[BenchmarkTask], List[BenchmarkTask]]:
    root = ensure_spreadsheetbench(cache_dir, refresh=refresh)
    dataset_path = root / "dataset.json"
    raw = json.loads(dataset_path.read_text())
    tasks: List[BenchmarkTask] = []
    for item in raw:
        folder = root / item["spreadsheet_path"]
        init_files = sorted(folder.glob("*_init.xlsx"))
        golden_files = sorted(folder.glob("*_golden.xlsx"))
        prompt_path = folder / "prompt.txt"
        if not init_files or not golden_files:
            continue
        tasks.append(
            BenchmarkTask(
                benchmark="spreadsheet",
                task_id=str(item["id"]),
                question=item["instruction"],
                expected={
                    "golden_xlsx": str(golden_files[0]),
                    "answer_sheet": item.get("answer_sheet"),
                    "answer_position": item.get("answer_position"),
                },
                input_artifacts={
                    "input_xlsx": str(init_files[0]),
                    "prompt_txt": prompt_path.read_text(errors="replace") if prompt_path.exists() else "",
                },
                metadata={
                    "instruction_type": item.get("instruction_type"),
                    "data_position": item.get("data_position"),
                    "spreadsheet_path": item.get("spreadsheet_path"),
                },
            )
        )

    import random

    shuffled = list(tasks)
    random.Random(split_seed).shuffle(shuffled)
    return shuffled[:n_train], shuffled[n_train : n_train + n_test]


async def run_spreadsheet_task(
    task: BenchmarkTask,
    *,
    llm_config: str,
    model_name: Optional[str] = None,
    artifact_store: Optional[ArtifactStore] = None,
    top_k_skills: int = 5,
    skill_injector_mode: str | None = None,
    skill_context_budget_chars: int | None = None,
    work_dir: Optional[Path] = None,
) -> BenchmarkResult:
    t0 = time.monotonic()
    trace = SpreadsheetTrace(task_id=task.task_id)
    query = str(task.question)
    retrieved = artifact_store.retrieve(query, top_k=top_k_skills) if artifact_store else []
    trace.retrieved_skills = [skill.name for skill in retrieved]
    injector_mode = (
        skill_injector_mode
        or os.environ.get("SPREADSHEET_SKILL_INJECTOR_MODE")
        or os.environ.get("SKILL_INJECTOR_MODE")
        or "full"
    ).strip().lower()
    injected = list(retrieved)
    skill_prompt = artifact_store.build_prompt(injected) if artifact_store else "(none)"
    if artifact_store and injector_mode not in {"", "full"}:
        injector = BudgetSkillInjector(
            mode=injector_mode,
            max_full_skills=int(os.environ.get("SPREADSHEET_SKILL_INJECTOR_MAX_FULL_SKILLS", os.environ.get("SKILL_INJECTOR_MAX_FULL_SKILLS", "0")) or "0"),
            max_summary_skills=int(os.environ.get("SPREADSHEET_SKILL_INJECTOR_MAX_SUMMARY_SKILLS", os.environ.get("SKILL_INJECTOR_MAX_SUMMARY_SKILLS", "2")) or "2"),
            budget_chars=int(skill_context_budget_chars or os.environ.get("SPREADSHEET_SKILL_CONTEXT_BUDGET_CHARS", os.environ.get("SKILL_INJECTOR_BUDGET_CHARS", "2200")) or "2200"),
            compact_chars_per_skill=int(os.environ.get("SPREADSHEET_SKILL_COMPACT_CHARS_PER_SKILL", os.environ.get("SKILL_INJECTOR_COMPACT_CHARS_PER_SKILL", "900")) or "900"),
        )
        injection = injector.select(
            retrieved,
            query=query,
            allowed_injection_types={"informational", "workflow", "functional"},
        )
        injected = injection.artifacts
        skill_prompt = injection.prompt()
        trace.filtered_skills = list(injection.filtered)
        trace.injector_events = [injection.as_event()]
        trace.cost_events.append(
            make_cost_event(
                role="injector",
                phase="executor",
                benchmark="spreadsheet",
                task_id=task.task_id,
                model=model_name or "",
                llm_config=llm_config,
                skill_prompt_chars=int(trace.injector_events[0].get("prompt_chars") or 0),
                metadata={
                    "mode": trace.injector_events[0].get("mode"),
                    "injected_count": trace.injector_events[0].get("injected_count"),
                    "filtered_count": trace.injector_events[0].get("filtered_count"),
                    "deterministic": True,
                },
            )
        )
    base_work_dir = work_dir or Path(tempfile.mkdtemp(prefix="spreadsheetbench_"))
    base_work_dir.mkdir(parents=True, exist_ok=True)
    callable_prompt = _write_spreadsheet_skill_library(injected, base_work_dir)
    trace.callable_skills = callable_prompt["skills"]
    skill_prompt_for_system = (
        skill_prompt
        + ("\n\nCallable skill import map:\n" + callable_prompt["prompt"] if callable_prompt["prompt"] else "")
    )
    trace.prompt_injected_skills = [skill.name for skill in injected]
    system = SPREADSHEET_SYSTEM.format(
        skills=skill_prompt,
        callable_skills=callable_prompt["prompt"] or "(no callable function skills available)",
    )
    input_copy = base_work_dir / f"{task.task_id}_input.xlsx"
    output_path = base_work_dir / f"{task.task_id}_output.xlsx"
    shutil.copyfile(task.input_artifacts["input_xlsx"], input_copy)
    shutil.copyfile(input_copy, output_path)

    prompt = _build_spreadsheet_prompt(task, input_copy, output_path)
    trace.prompt = prompt
    try:
        response = await ask_text_llm(
            llm_config=llm_config,
            model_name=model_name,
            system=system,
            prompt=prompt,
        )
        trace.input_tokens = response.prompt_tokens
        trace.cache_input_tokens = response.cache_input_tokens
        trace.total_tokens = response.prompt_tokens + response.cache_input_tokens + response.completion_tokens
        trace.completion_tokens = response.completion_tokens
        trace.code = _extract_code(response.content)
        trace.cost_events.append(
            make_cost_event(
                role="executor",
                phase="task_rollout",
                benchmark="spreadsheet",
                task_id=task.task_id,
                model=response.model_name or model_name or "",
                llm_config=llm_config,
                input_tokens=response.prompt_tokens,
                cache_input_tokens=response.cache_input_tokens,
                output_tokens=response.completion_tokens,
                prompt_chars=len(prompt),
                skill_prompt_chars=len(skill_prompt_for_system),
                system_prompt_chars=len(system),
                final_conversation_chars=len(system) + len(prompt) + len(response.content or ""),
                metadata={
                    "api_style": response.api_style,
                    "skill_injector_mode": injector_mode,
                    "prompt_injected_skills": list(trace.prompt_injected_skills),
                    "callable_skills": list(trace.callable_skills),
                },
            )
        )
        if not trace.code:
            trace.elapsed_s = round(time.monotonic() - t0, 3)
            return BenchmarkResult(
                benchmark="spreadsheet",
                task_id=task.task_id,
                success=False,
                score=0.0,
                metrics={"reason": "no_python_code"},
                trace=trace.as_dict(),
            )
        stdout, stderr, returncode = await asyncio.to_thread(
            _run_code, trace.code, input_copy, output_path, base_work_dir
        )
        trace.stdout = stdout
        trace.stderr = stderr
        verify = verify_spreadsheet_output(
            predicted_xlsx=output_path,
            golden_xlsx=Path(task.expected["golden_xlsx"]),
            sheet_name=task.expected.get("answer_sheet"),
            answer_range=task.expected.get("answer_position"),
        )
        verify["returncode"] = returncode
        verify["execution_ok"] = returncode == 0
        success = bool(returncode == 0 and verify["pass"])
    except Exception as exc:
        trace.elapsed_s = round(time.monotonic() - t0, 3)
        return BenchmarkResult(
            benchmark="spreadsheet",
            task_id=task.task_id,
            success=False,
            score=0.0,
            metrics={"exception": type(exc).__name__},
            trace=trace.as_dict(),
            error=str(exc),
        )

    trace.elapsed_s = round(time.monotonic() - t0, 3)
    verify["total_tokens"] = trace.total_tokens
    verify["input_tokens"] = trace.input_tokens
    verify["cache_input_tokens"] = trace.cache_input_tokens
    verify["completion_tokens"] = trace.completion_tokens
    verify["cost_events"] = trace.cost_events
    verify["elapsed_s"] = trace.elapsed_s
    verify["retrieved_skills"] = trace.retrieved_skills
    verify["prompt_injected_skills"] = trace.prompt_injected_skills
    verify["filtered_skills"] = trace.filtered_skills
    verify["injector_events"] = trace.injector_events
    verify["skill_injector_mode"] = injector_mode
    verify["model_name"] = response.model_name
    verify["llm_api_style"] = response.api_style
    return BenchmarkResult(
        benchmark="spreadsheet",
        task_id=task.task_id,
        success=success,
        score=float(verify["cell_accuracy"]),
        metrics=verify,
        trace=trace.as_dict(),
    )


def verify_spreadsheet_output(
    *,
    predicted_xlsx: Path,
    golden_xlsx: Path,
    sheet_name: Optional[str],
    answer_range: Optional[str],
) -> Dict[str, Any]:
    pred_wb = openpyxl.load_workbook(predicted_xlsx, data_only=False)
    gold_wb = openpyxl.load_workbook(golden_xlsx, data_only=False)
    requested_sheet = _first_sheet_name(sheet_name)
    refs = _answer_range_refs(answer_range, default_sheet=requested_sheet)
    if not refs:
        refs = [(requested_sheet, None)]
    mismatches = []
    checked = 0
    for ref_sheet, cell_range in refs:
        sheet = ref_sheet if ref_sheet in gold_wb.sheetnames else gold_wb.sheetnames[0]
        if sheet not in pred_wb.sheetnames:
            mismatches.append({"cell": "__sheet__", "predicted": None, "expected": sheet})
            continue
        pred_ws = pred_wb[sheet]
        gold_ws = gold_wb[sheet]
        cells = _cells_in_range(cell_range, gold_ws)
        checked += len(cells)
        for cell in cells:
            pv = _normalize_cell_value(pred_ws[cell].value)
            gv = _normalize_cell_value(gold_ws[cell].value)
            if pv != gv:
                mismatches.append({"cell": f"{sheet}!{cell}", "predicted": pv, "expected": gv})
    correct = checked - len(mismatches)
    return {
        "pass": len(mismatches) == 0 and checked > 0,
        "cell_accuracy": round(correct / max(checked, 1), 4),
        "checked_cells": checked,
        "mismatched_cells": mismatches[:20],
        "answer_sheet": requested_sheet,
        "answer_position": answer_range,
    }


def _build_spreadsheet_prompt(task: BenchmarkTask, input_copy: Path, output_path: Path) -> str:
    preview = _workbook_preview(input_copy)
    return (
        f"Instruction:\n{task.question}\n\n"
        f"Input workbook path: {input_copy}\n"
        f"Output workbook path: {output_path}\n"
        f"Answer sheet: {task.expected.get('answer_sheet')}\n"
        f"Answer range: {task.expected.get('answer_position')}\n"
        f"Data position: {task.metadata.get('data_position')}\n\n"
        f"Workbook preview:\n{preview}\n\n"
        "Write robust openpyxl code. It may inspect workbook sheets/cells, then save to OUTPUT_XLSX."
    )


def _workbook_preview(path: Path, max_rows: int = 8, max_cols: int = 8) -> str:
    wb = openpyxl.load_workbook(path, data_only=False)
    parts = []
    for sheet in wb.sheetnames[:5]:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, max_rows),
            min_col=1,
            max_col=min(ws.max_column, max_cols),
            values_only=True,
        ):
            rows.append([_jsonable_cell_value(value) for value in row])
        parts.append(f"## {sheet} ({ws.max_row}x{ws.max_column})\n{json.dumps(rows, ensure_ascii=False)}")
    return "\n\n".join(parts)


def _extract_code(text: str) -> str:
    match = re.search(r"```(?:python)?\s*(.*?)```", text or "", re.S | re.I)
    return match.group(1).strip() if match else ""


def _write_spreadsheet_skill_library(skills: Sequence[SkillArtifact], work_dir: Path) -> Dict[str, Any]:
    callable_rows: List[Dict[str, Any]] = []
    chunks = [
        "from pathlib import Path\n",
        "import openpyxl\n",
        "from openpyxl import load_workbook\n",
        "from openpyxl.utils import column_index_from_string\n",
        "from datetime import datetime, date, time, timedelta\n",
        "import math\n",
        "import re\n\n",
        "def _spreadsheet_column_value(value):\n",
        "    if isinstance(value, str) and re.fullmatch(r\"[A-Za-z]+\", value.strip()):\n",
        "        return column_index_from_string(value.strip().upper())\n",
        "    return value\n\n",
        "def _spreadsheet_callable_kwargs(kwargs, ws):\n",
        "    env = dict(kwargs or {})\n",
        "    for key, value in list(env.items()):\n",
        "        if key.endswith(\"_column\"):\n",
        "            base = key[: -len(\"_column\")]\n",
        "            env.setdefault(f\"{base}_col\", _spreadsheet_column_value(value))\n",
        "            if base.endswith(\"ies\"):\n",
        "                env.setdefault(f\"{base[:-3]}y_col\", _spreadsheet_column_value(value))\n",
        "        elif key.endswith(\"_col\"):\n",
        "            env[key] = _spreadsheet_column_value(value)\n",
        "            env.setdefault(f\"{key[: -len('_col')]}_column\", value)\n",
        "    env.setdefault(\"value_start_row\", 1)\n",
        "    env.setdefault(\"value_end_row\", ws.max_row)\n",
        "    env.setdefault(\"result_start_row\", 1)\n",
        "    env.setdefault(\"result_end_row\", ws.max_row)\n",
        "    return env\n\n",
    ]
    for skill in skills:
        if not _is_spreadsheet_callable_skill(skill):
            continue
        func_name = _spreadsheet_skill_function_name(skill.name)
        code = _spreadsheet_skill_code(skill)
        if not code:
            continue
        rendered = _render_spreadsheet_skill_function(func_name, code)
        if not rendered:
            continue
        try:
            ast.parse(rendered)
        except SyntaxError:
            continue
        chunks.append(f"\n# Skill: {skill.name}\n")
        chunks.append(rendered)
        chunks.append("\n")
        callable_rows.append(
            {
                "skill_name": skill.name,
                "function_name": func_name,
                "description": _spreadsheet_callable_description(skill),
            }
        )
    if callable_rows:
        (work_dir / "skill_library.py").write_text("".join(chunks))
    prompt = "\n".join(
        f"- `{row['function_name']}(INPUT_XLSX, OUTPUT_XLSX, **kwargs)` from skill `{row['skill_name']}`: {row['description'][:220]}"
        for row in callable_rows
    )
    if prompt:
        prompt = (
            "You may directly reuse executable spreadsheet skills by importing them in your returned code, e.g. "
            "`from skill_library import skill_name` then `skill_name(INPUT_XLSX, OUTPUT_XLSX)`. "
            "Prefer importing a callable over rewriting its implementation when its scope matches the workbook and requested answer range. "
            "Pass explicit keyword arguments for sheet names, ranges, columns, and row bounds when the description requires them.\n"
            + prompt
        )
    return {"prompt": prompt, "skills": callable_rows}


def _spreadsheet_callable_description(skill: SkillArtifact, *, limit: int = 520) -> str:
    parts: List[str] = []
    if skill.description:
        parts.append(str(skill.description).strip())
    body = str(skill.body or "")
    applicability = re.search(
        r"(?is)(Applicability\s*:\s*.*?)(?:\n\n|Reusable openpyxl idiom|```|Non-applicability\s*:)",
        body,
    )
    if applicability:
        parts.append(applicability.group(1).strip())
    required = []
    for line in body.splitlines()[:12]:
        stripped = line.strip().lstrip("#").strip()
        if re.search(r"(?i)\b(required variables|assumes|parameters|kwargs)\b", stripped):
            required.append(stripped)
    if required:
        parts.append(" ".join(required[:3]))
    text = " ".join(part for part in parts if part).strip()
    text = re.sub(r"\s+", " ", text)
    if len(text) > limit:
        text = text[: limit - 3].rstrip() + "..."
    return text or str(skill.description or skill.name)


def _is_spreadsheet_callable_skill(skill: SkillArtifact) -> bool:
    """Return whether a Spreadsheet skill should be exposed as importable code.

    Older evolved stores sometimes set ``metadata.injection_type`` to
    ``informational`` even for ``executable_tool`` artifacts.  Callable exposure
    should follow the artifact's semantic kind and code availability, not only
    the prompt injection label.  Workflow and knowledge/interface cards remain
    prompt-only.
    """
    kind = str(skill.kind or "").strip().lower()
    if kind not in {"executable_tool", "function_tool", "script_tool"}:
        return False
    return bool(_spreadsheet_skill_code(skill))


def _spreadsheet_skill_function_name(name: str) -> str:
    value = re.sub(r"\W+", "_", str(name or "").strip()).strip("_").lower()
    if not value:
        value = "spreadsheet_skill"
    if value[0].isdigit():
        value = f"skill_{value}"
    return value


def _spreadsheet_skill_code(skill: SkillArtifact) -> str:
    body = str(skill.body or "")
    match = re.search(r"```(?:python)?\s*(.*?)```", body, re.S | re.I)
    if match:
        code = match.group(1).strip()
    else:
        code = body.strip()
        if not _looks_like_spreadsheet_python(code):
            return ""
    code = re.sub(r"load_workbook\((['\"])[^'\"]*?_input\\.xlsx\1\)", "load_workbook(INPUT_XLSX)", code)
    code = re.sub(r"\\.save\((['\"])[^'\"]*?_output\\.xlsx\1\)", ".save(OUTPUT_XLSX)", code)
    return code


def _looks_like_spreadsheet_python(code: str) -> bool:
    if not code.strip():
        return False
    python_start = re.search(
        r"(?m)^\s*(?:from|import|def|class|for|while|if|try|with)\s+\w+|^\s*(?:wb|ws|worksheet|workbook)\b",
        code,
    )
    if python_start:
        return True
    spreadsheet_markers = [
        "INPUT_XLSX",
        "OUTPUT_XLSX",
        "load_workbook(",
        "openpyxl.",
        "wb.save(",
        ".cell(",
        "ws[",
    ]
    return any(marker in code for marker in spreadsheet_markers)


def _render_spreadsheet_skill_function(func_name: str, code: str) -> str:
    try:
        tree = ast.parse(code)
    except SyntaxError:
        return _wrap_spreadsheet_skill_snippet(func_name, code)
    functions = [node.name for node in tree.body if isinstance(node, ast.FunctionDef)]
    if functions:
        first = functions[0]
        return (
            code
            + "\n\n"
            + f"def {func_name}(INPUT_XLSX, OUTPUT_XLSX, **kwargs):\n"
            + "    try:\n"
            + f"        return {first}(INPUT_XLSX, OUTPUT_XLSX, **kwargs)\n"
            + "    except TypeError:\n"
            + f"        return {first}(INPUT_XLSX, OUTPUT_XLSX)\n"
        )
    return _wrap_spreadsheet_skill_snippet(func_name, code)


def _wrap_spreadsheet_skill_snippet(func_name: str, code: str) -> str:
    code_literal = repr(code.strip() or "pass")
    return (
        f"def {func_name}(INPUT_XLSX, OUTPUT_XLSX, **kwargs):\n"
        "    INPUT_XLSX = Path(INPUT_XLSX)\n"
        "    OUTPUT_XLSX = Path(OUTPUT_XLSX)\n"
        "    wb = load_workbook(INPUT_XLSX)\n"
        "    ws = wb.active\n"
        "    env = dict(globals())\n"
        "    env.update({\n"
        "        'INPUT_XLSX': INPUT_XLSX,\n"
        "        'OUTPUT_XLSX': OUTPUT_XLSX,\n"
        "        'wb': wb,\n"
        "        'ws': ws,\n"
        "        'worksheet': ws,\n"
        "        'workbook': wb,\n"
        "    })\n"
        "    env.update(_spreadsheet_callable_kwargs(kwargs, ws))\n"
        f"    exec({code_literal}, env, env)\n"
        "    out_wb = env.get('wb', wb)\n"
        "    if hasattr(out_wb, 'save'):\n"
        "        out_wb.save(OUTPUT_XLSX)\n"
        "    return OUTPUT_XLSX\n"
    )


def _spreadsheet_callable_kwargs(kwargs, ws):
    env = dict(kwargs or {})
    for key, value in list(env.items()):
        if key.endswith("_column"):
            base = key[: -len("_column")]
            env.setdefault(f"{base}_col", _spreadsheet_column_value(value))
            if base.endswith("ies"):
                env.setdefault(f"{base[:-3]}y_col", _spreadsheet_column_value(value))
        elif key.endswith("_col"):
            env[key] = _spreadsheet_column_value(value)
            env.setdefault(f"{key[: -len('_col')]}_column", value)
    env.setdefault("value_start_row", 1)
    env.setdefault("value_end_row", ws.max_row)
    env.setdefault("result_start_row", 1)
    env.setdefault("result_end_row", ws.max_row)
    return env


def _spreadsheet_column_value(value):
    if isinstance(value, str) and re.fullmatch(r"[A-Za-z]+", value.strip()):
        return column_index_from_string(value.strip().upper())
    return value


def _run_code(code: str, input_xlsx: Path, output_xlsx: Path, work_dir: Path) -> Tuple[str, str, int]:
    script = work_dir / "run_spreadsheet_solution.py"
    script.write_text(
        "from pathlib import Path\n"
        f"INPUT_XLSX = Path({str(input_xlsx)!r})\n"
        f"OUTPUT_XLSX = Path({str(output_xlsx)!r})\n\n"
        + code
        + "\n"
    )
    proc = subprocess.run(
        ["python", str(script)],
        cwd=str(work_dir),
        text=True,
        capture_output=True,
        timeout=CODE_EXEC_TIMEOUT,
    )
    return proc.stdout[-4000:], proc.stderr[-4000:], proc.returncode


def _cells_in_range(answer_range: Optional[str], ws: Any) -> List[str]:
    if not answer_range:
        return [
            cell.coordinate
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        ]
    target = ws[answer_range]
    if hasattr(target, "coordinate"):
        return [target.coordinate]
    cells = []
    for row in target:
        if isinstance(row, tuple):
            cells.extend(cell.coordinate for cell in row)
        else:
            cells.append(row.coordinate)
    return cells


class SpreadsheetMaintenanceAdapter(NoOpMaintenanceAdapter):
    """SpreadsheetBench adapter for the benchmark-agnostic evolution runner.

    Spreadsheet evolution is benchmark-native: successful traces can extract
    openpyxl/workflow skills, retrieved skills receive per-task credit, credit
    produces focused replayable SpreadsheetBench bundle cases, micro
    maintenance refines or filters implicated skills, and macro maintenance
    promotes/deduplicates/filters repository candidates at the window level.
    """

    benchmark = "spreadsheet"

    async def run_task(
        self,
        task: BenchmarkTask,
        *,
        store: ArtifactStore,
        config: MaintenanceRunConfig,
        phase: str,
        task_index: int,
        run_idx: int,
    ) -> BenchmarkResult:
        del phase, task_index, run_idx
        return await run_spreadsheet_task(
            task,
            llm_config=config.llm_config,
            model_name=config.model_name,
            artifact_store=store,
            top_k_skills=config.top_k_skills,
            skill_injector_mode=config.extra.get("skill_injector_mode"),
            skill_context_budget_chars=config.extra.get("skill_context_budget_chars"),
        )

    async def assign_credit(
        self,
        *,
        detail: Dict[str, Any],
        store: ArtifactStore,
        config: MaintenanceRunConfig,
        round_index: int,
        task_index: int,
    ) -> List[Dict[str, Any]]:
        del round_index
        projection = _spreadsheet_trace_projection(detail)
        retrieved_names = projection.get("retrieved_skills") or []
        candidate_artifacts = [
            artifact for name in retrieved_names for artifact in [store.get(str(name))] if artifact
        ]
        if not candidate_artifacts:
            return []
        try:
            payload = await _ask_json(
                system=SPREADSHEET_CREDIT_SYSTEM,
                user=_json_block(
                    {
                        "task": _spreadsheet_task_fragment(detail),
                        "trace_projection": projection,
                        "candidate_skills": [
                            _spreadsheet_skill_projection(artifact, projection=projection)
                            for artifact in candidate_artifacts
                        ],
                    }
                ),
                llm_config=config.llm_config,
                model_name=config.model_name,
                role="spreadsheet_credit_assigner",
                metadata={"task_id": detail.get("task_id"), "task_index": task_index},
            )
            events = _normalize_spreadsheet_credit_events(
                payload,
                detail=detail,
                candidate_artifacts=candidate_artifacts,
                projection=projection,
            )
        except Exception as exc:
            events = _heuristic_spreadsheet_credit_events(
                detail=detail,
                candidate_artifacts=candidate_artifacts,
                projection=projection,
                reason=f"credit_llm_failed:{type(exc).__name__}",
            )
        for event in events:
            artifact = store.get(event["skill_name"])
            if artifact is None:
                continue
            evidence_row = {
                "task_id": detail.get("task_id"),
                "judgment": event.get("judgment"),
                "effect_type": event.get("effect_type"),
                "confidence": event.get("confidence"),
                "reason": event.get("reason"),
                "score": projection.get("score"),
                "success": projection.get("success"),
            }
            if event.get("judgment") == "helpful":
                artifact.evidence.helpful_cases.append(evidence_row)
                artifact.success_count += 1
            elif event.get("judgment") == "harmful":
                artifact.evidence.harmful_cases.append(evidence_row)
            artifact.usage_count += 1
        return events

    async def apply_credit_bundle_cases(
        self,
        *,
        detail: Dict[str, Any],
        credit_events: Sequence[Dict[str, Any]],
        store: ArtifactStore,
        config: MaintenanceRunConfig,
        round_index: int,
        task_index: int,
    ) -> List[Dict[str, Any]]:
        del config, round_index, task_index
        created: List[Dict[str, Any]] = []
        for event in credit_events:
            artifact = store.get(str(event.get("skill_name") or ""))
            if artifact is None:
                continue
            for suggestion in event.get("bundle_case_suggestions") or []:
                case = _spreadsheet_case_from_credit_suggestion(
                    detail=detail,
                    artifact=artifact,
                    event=event,
                    suggestion=dict(suggestion or {}),
                )
                if case is None:
                    continue
                bucket = _bundle_bucket(case.polarity)
                existing = {item.case_id for item in getattr(artifact.bundle, bucket)}
                if case.case_id not in existing:
                    getattr(artifact.bundle, bucket).append(case)
                    artifact.bundle.bundle_version += 1
                    created.append(
                        {
                            "skill_name": artifact.name,
                            "case_id": case.case_id,
                            "polarity": case.polarity,
                            "source_task_id": detail.get("task_id"),
                            "reason": suggestion.get("reason") or event.get("reason"),
                        }
                    )
            _trim_spreadsheet_bundle_cases(artifact)
        return created

    async def run_micro_maintenance(self, **kwargs: Any) -> Dict[str, Any]:
        detail = kwargs.get("detail") or {}
        credit_events = list(kwargs.get("credit_events") or [])
        credit_bundle_cases = list(kwargs.get("credit_bundle_cases") or [])
        store: ArtifactStore = kwargs["store"]
        config: MaintenanceRunConfig = kwargs["config"]
        task_index = int(kwargs.get("task_index") or 0)
        projection = _spreadsheet_trace_projection(detail)
        extracted = await _extract_spreadsheet_skills_from_detail(
            detail,
            store=store,
            config=config,
            task_index=task_index,
        )
        extraction_reports: List[Dict[str, Any]] = []
        for artifact in extracted:
            store.add_pending(artifact)
            extraction_reports.append(
                {
                    "skill_name": artifact.name,
                    "status": artifact.status,
                    "description": artifact.description,
                    "source_task_ids": artifact.metadata.get("source_task_ids") or [],
                }
            )
        maintenance_targets = _spreadsheet_micro_targets(
            credit_events=credit_events,
            credit_bundle_cases=credit_bundle_cases,
            extracted=extracted,
        )
        refine_decisions: List[Dict[str, Any]] = []
        bundle_results: List[Dict[str, Any]] = []
        for skill_name in maintenance_targets:
            artifact = store.get(skill_name)
            if artifact is None:
                continue
            skill_credit = [event for event in credit_events if event.get("skill_name") == skill_name]
            pre_refine = await _refine_spreadsheet_skill_from_credit(
                artifact=artifact,
                credit_context=skill_credit,
                detail=detail,
                store=store,
                config=config,
            )
            refine_decisions.append(pre_refine)
            if pre_refine.get("updated_artifact"):
                store.add(pre_refine["updated_artifact"])
                artifact = store.get(skill_name) or pre_refine["updated_artifact"]
            if artifact.bundle.all_cases():
                result = await _execute_spreadsheet_bundle_tests(
                    artifact=artifact,
                    config=config,
                )
                store.add_test_result(result)
                bundle_results.append(result.as_dict())
                if not result.aggregate.get("passed"):
                    post_refine = await _refine_spreadsheet_skill_from_bundle(
                        artifact=store.get(skill_name) or artifact,
                        test_result=result,
                        credit_context=skill_credit,
                        store=store,
                        config=config,
                    )
                    refine_decisions.append(post_refine)
                    if post_refine.get("updated_artifact"):
                        store.add(post_refine["updated_artifact"])
        report = {
            "phase": "micro",
            "task_id": detail.get("task_id"),
            "task_index": task_index,
            "maintenance_targets": maintenance_targets,
            "maintenance_test_results": bundle_results,
            "refine_decisions": [
                {k: v for k, v in item.items() if k != "updated_artifact"}
                for item in refine_decisions
            ],
            "credit_bundle_cases": copy.deepcopy(credit_bundle_cases),
            "extraction_reports": extraction_reports,
            "reason": "spreadsheet_micro_maintenance",
        }
        report["trace_projection"] = _spreadsheet_trace_projection(detail)
        return report

    async def run_macro_maintenance(
        self,
        *,
        window_details: Sequence[Dict[str, Any]],
        all_train_details: Sequence[Dict[str, Any]],
        credit_events: Sequence[Dict[str, Any]],
        store: ArtifactStore,
        config: MaintenanceRunConfig,
        round_index: int,
        window_index: int,
        final_window: bool = False,
    ) -> Dict[str, Any]:
        del config, round_index
        promotions = _promote_spreadsheet_pending_from_window(store, window_details)
        dedupe = _dedupe_spreadsheet_skills(store)
        filtered = _filter_spreadsheet_harmful_skills(store, credit_events)
        store.refresh_all_dependencies()
        return {
            "phase": "macro_final" if final_window else "macro",
            "window_index": window_index,
            "task_ids": [str(item.get("task_id") or "") for item in window_details],
            "maintenance_targets": sorted(set(promotions + dedupe + filtered)),
            "maintenance_test_results": [],
            "refine_decisions": [],
            "overlap_refactor": {
                "benchmark": "spreadsheet",
                "attempts": [],
                "promoted_pending_skills": promotions,
                "deduplicated_skills": dedupe,
                "filtered_skills": filtered,
                "window_trace_segments": [
                    _spreadsheet_trace_projection(item) for item in window_details
                ],
                "all_train_count": len(all_train_details),
            },
            "run_overlap_refactor": True,
            "reason": "spreadsheet_macro_promote_dedupe_filter",
        }

    def store_snapshot(self, store: ArtifactStore) -> Dict[str, Any]:
        artifacts = store.all()
        return {
            "n_skills": len(artifacts),
            "n_active": sum(1 for item in artifacts if item.status == "active"),
            "n_pending": sum(1 for item in artifacts if item.status == "pending"),
            "n_disabled": sum(1 for item in artifacts if item.is_disabled()),
            "n_archived": sum(1 for item in artifacts if item.status == "archived"),
            "skill_names": [artifact.name for artifact in artifacts],
            "skill_versions": {artifact.name: artifact.version for artifact in artifacts},
            "skill_status": {artifact.name: artifact.status for artifact in artifacts},
            "bundle_case_counts": {
                artifact.name: len(artifact.bundle.all_cases()) for artifact in artifacts
            },
        }


def _spreadsheet_trace_projection(detail: Dict[str, Any]) -> Dict[str, Any]:
    runs = detail.get("runs") or []
    first = runs[0] if runs else {}
    metrics = first.get("metrics") or {}
    trace = first.get("trace") or {}
    return {
        "task_id": detail.get("task_id"),
        "success": first.get("success"),
        "score": first.get("score"),
        "answer_sheet": metrics.get("answer_sheet"),
        "answer_position": metrics.get("answer_position"),
        "checked_cells": metrics.get("checked_cells"),
        "mismatched_cells": metrics.get("mismatched_cells", [])[:5],
        "execution_ok": metrics.get("execution_ok"),
        "llm_api_style": metrics.get("llm_api_style"),
        "retrieved_skills": metrics.get("retrieved_skills") or trace.get("retrieved_skills") or [],
        "stderr_tail": str(trace.get("stderr") or "")[-800:],
    }


def _json_block(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, indent=2)


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _stable_id(*parts: Any, length: int = 10) -> str:
    raw = "\n".join(str(part) for part in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:length]


def _spreadsheet_task_fragment(detail: Dict[str, Any]) -> Dict[str, Any]:
    task = dict(detail.get("task") or {})
    expected = dict(task.get("expected") or {})
    metadata = dict(task.get("metadata") or {})
    return {
        "benchmark": task.get("benchmark") or "spreadsheet",
        "task_id": task.get("task_id") or detail.get("task_id"),
        "question": task.get("question"),
        "expected": {
            "answer_sheet": expected.get("answer_sheet"),
            "answer_position": expected.get("answer_position"),
            "golden_xlsx": expected.get("golden_xlsx"),
        },
        "input_artifacts": {
            "input_xlsx": (task.get("input_artifacts") or {}).get("input_xlsx"),
            "prompt_txt_preview": str((task.get("input_artifacts") or {}).get("prompt_txt") or "")[:1200],
        },
        "metadata": {
            "instruction_type": metadata.get("instruction_type"),
            "data_position": metadata.get("data_position"),
            "spreadsheet_path": metadata.get("spreadsheet_path"),
        },
    }


def _spreadsheet_code_snippet(code: str, limit: int = 1600) -> str:
    text = str(code or "").strip()
    if len(text) <= limit:
        return text
    lines = text.splitlines()
    kept: List[str] = []
    total = 0
    for line in lines:
        if total + len(line) + 1 > limit:
            break
        kept.append(line)
        total += len(line) + 1
    return "\n".join(kept).rstrip() + "\n# ... truncated ..."


def _spreadsheet_result_projection(detail: Dict[str, Any]) -> Dict[str, Any]:
    run = dict((detail.get("runs") or [{}])[0] or {})
    trace = dict(run.get("trace") or {})
    metrics = dict(run.get("metrics") or {})
    return {
        "task": _spreadsheet_task_fragment(detail),
        "success": run.get("success"),
        "score": run.get("score"),
        "metrics": {
            "answer_sheet": metrics.get("answer_sheet"),
            "answer_position": metrics.get("answer_position"),
            "cell_accuracy": metrics.get("cell_accuracy"),
            "checked_cells": metrics.get("checked_cells"),
            "mismatched_cells": (metrics.get("mismatched_cells") or [])[:8],
            "execution_ok": metrics.get("execution_ok"),
            "returncode": metrics.get("returncode"),
            "total_tokens": metrics.get("total_tokens"),
        },
        "trace": {
            "retrieved_skills": trace.get("retrieved_skills") or metrics.get("retrieved_skills") or [],
            "code_snippet": _spreadsheet_code_snippet(trace.get("code") or ""),
            "stderr_tail": str(trace.get("stderr") or "")[-1200:],
            "stdout_tail": str(trace.get("stdout") or "")[-800:],
        },
    }


def _spreadsheet_skill_projection(
    artifact: SkillArtifact,
    *,
    projection: Dict[str, Any],
) -> Dict[str, Any]:
    return {
        "skill_name": artifact.name,
        "version": artifact.version,
        "kind": artifact.kind,
        "status": artifact.status,
        "description": artifact.description,
        "body": artifact.body[:1800],
        "interface": artifact.interface.as_dict(),
        "metadata": {
            "domains": artifact.metadata.get("domains") or [],
            "allowed_tools": artifact.metadata.get("allowed_tools") or [],
            "intent_keywords": artifact.metadata.get("intent_keywords") or [],
            "source_task_ids": artifact.metadata.get("source_task_ids") or [],
            "non_applicability": artifact.metadata.get("non_applicability"),
        },
        "retrieved": artifact.name in set(projection.get("retrieved_skills") or []),
        "usage_count": artifact.usage_count,
        "success_count": artifact.success_count,
        "recent_helpful": artifact.evidence.helpful_cases[-3:],
        "recent_harmful": artifact.evidence.harmful_cases[-3:],
    }


def _coerce_spreadsheet_artifact(raw: Dict[str, Any], *, detail: Dict[str, Any]) -> SkillArtifact | None:
    name = normalize_skill_name(str(raw.get("name") or ""))
    description = str(raw.get("description") or "").strip()
    body = str(raw.get("body") or "").strip()
    if not name or not description or not body:
        return None
    if str(raw.get("kind") or "executable_tool") == "executable_tool" and "openpyxl" not in body.lower():
        body = "Use openpyxl with INPUT_XLSX and OUTPUT_XLSX.\n\n" + body
    task = _spreadsheet_task_fragment(detail)
    metadata = dict(raw.get("metadata") or {})
    domains = [str(item).strip() for item in (metadata.get("domains") or []) if str(item).strip()]
    if "SpreadsheetBench" not in domains:
        domains.insert(0, "SpreadsheetBench")
    instruction_type = str((task.get("metadata") or {}).get("instruction_type") or "").strip()
    if instruction_type and instruction_type not in domains:
        domains.append(instruction_type)
    metadata.update(
        {
            "domains": domains,
            "allowed_tools": ["openpyxl"],
            "source": metadata.get("source") or "spreadsheet_llm_trace_extraction",
            "source_task_ids": list(dict.fromkeys([*(metadata.get("source_task_ids") or []), str(task.get("task_id") or "")])),
            "benchmark": "spreadsheet",
            "instruction_type": instruction_type,
            "version_kind": "seed",
            "injection_type": metadata.get("injection_type") or "informational",
        }
    )
    tags = [
        "domain:SpreadsheetBench",
        "tool:openpyxl",
        *(f"intent:{item}" for item in metadata.get("intent_keywords") or []),
    ]
    artifact = SkillArtifact(
        name=name,
        kind=str(raw.get("kind") or "executable_tool"),
        description=description[:240],
        body=body[:1400],
        metadata=metadata,
        tags=list(dict.fromkeys(str(tag) for tag in tags if str(tag).strip())),
        interface=SkillInterface(
            summary=str(((raw.get("interface") or {}).get("summary")) or description),
            usage=str(((raw.get("interface") or {}).get("usage")) or "Use when the SpreadsheetBench instruction and workbook shape match this skill scope."),
            input_contract=dict(((raw.get("interface") or {}).get("input_contract")) or {"benchmark": "SpreadsheetBench"}),
            output_contract=dict(((raw.get("interface") or {}).get("output_contract")) or {"tool": "openpyxl"}),
            invocation_contract=dict(((raw.get("interface") or {}).get("invocation_contract")) or {"injection_type": "informational"}),
            compatibility_notes=str(((raw.get("interface") or {}).get("compatibility_notes")) or metadata.get("non_applicability") or ""),
        ),
        evidence=SkillEvidence(
            source_traces=[_spreadsheet_result_projection(detail)],
        ),
        lineage=SkillLineage(version_kind="seed"),
        dependencies=[str(item).strip() for item in (raw.get("dependencies") or []) if str(item).strip()],
        status="pending",
    )
    artifact.bundle.positive_cases.append(
        _spreadsheet_case_from_task(
            detail=detail,
            skill_name=artifact.name,
            polarity="positive",
            reason="successful source trace used to bootstrap the skill contract",
            source="distilled_success",
            confidence=0.7,
        )
    )
    return artifact


async def _extract_spreadsheet_skills_from_detail(
    detail: Dict[str, Any],
    *,
    store: ArtifactStore,
    config: MaintenanceRunConfig,
    task_index: int,
) -> List[SkillArtifact]:
    projection = _spreadsheet_trace_projection(detail)
    has_success_evidence = bool(projection.get("success")) and float(projection.get("score") or 0.0) >= 0.9
    has_repair_evidence = _spreadsheet_has_repair_evidence(projection)
    if not has_success_evidence and not has_repair_evidence:
        return []
    existing = [
        {
            "name": artifact.name,
            "description": artifact.description,
            "status": artifact.status,
            "domains": artifact.metadata.get("domains") or [],
            "intent_keywords": artifact.metadata.get("intent_keywords") or [],
        }
        for artifact in store.all()
        if artifact.status in {"active", "pending", "stale"}
    ]
    try:
        payload = await _ask_json(
            system=SPREADSHEET_EXTRACT_SYSTEM,
            user=_json_block(
                {
                    "existing_artifacts": existing[-40:],
                    "result": _spreadsheet_result_projection(detail),
                }
            ),
            llm_config=config.llm_config,
            model_name=config.model_name,
            role="spreadsheet_extractor",
            metadata={"task_id": detail.get("task_id"), "task_index": task_index},
        )
        raw_artifacts = list(payload.get("artifacts") or [])
    except Exception:
        raw_artifacts = [
            _heuristic_spreadsheet_artifact_payload(detail)
            if has_success_evidence
            else _heuristic_spreadsheet_repair_artifact_payload(detail)
        ]
    if not raw_artifacts and has_repair_evidence:
        raw_artifacts = [_heuristic_spreadsheet_repair_artifact_payload(detail)]
    artifacts: List[SkillArtifact] = []
    for raw in raw_artifacts[:3]:
        artifact = _coerce_spreadsheet_artifact(dict(raw or {}), detail=detail)
        if artifact is None:
            continue
        artifacts.append(artifact)
    return artifacts


def _heuristic_spreadsheet_artifact_payload(detail: Dict[str, Any]) -> Dict[str, Any]:
    task = _spreadsheet_task_fragment(detail)
    result = _spreadsheet_result_projection(detail)
    instruction_type = str((task.get("metadata") or {}).get("instruction_type") or "spreadsheet_task").strip()
    answer_position = str((task.get("expected") or {}).get("answer_position") or "").strip()
    question = str(task.get("question") or "")
    keywords = _spreadsheet_keywords(question, instruction_type)
    name = normalize_skill_name(f"spreadsheet_{instruction_type}_{'_'.join(keywords[:4])}")[:80]
    snippet = result["trace"]["code_snippet"]
    body = (
        f"Applicability: SpreadsheetBench {instruction_type} tasks with similar instruction terms "
        f"{keywords[:8]} and answer range `{answer_position}`. Use openpyxl, inspect workbook sheets, "
        "write only the requested answer cells, and preserve unrelated sheets/styles.\n\n"
        "Reusable openpyxl idiom from successful evidence:\n"
        "```python\n"
        f"{snippet}\n"
        "```\n\n"
        "Non-applicability: do not copy hard-coded cell values, sheet names, or row counts unless the "
        "current workbook inspection confirms them."
    )
    return {
        "name": name,
        "kind": "executable_tool",
        "description": f"Openpyxl pattern for {instruction_type} spreadsheet tasks.",
        "body": body,
        "interface": {
            "summary": f"SpreadsheetBench {instruction_type} openpyxl pattern.",
            "usage": "Retrieve for similar SpreadsheetBench instructions before writing Python code.",
            "input_contract": {"benchmark": "SpreadsheetBench", "instruction_type": instruction_type},
            "output_contract": {"tool": "openpyxl", "save_to": "OUTPUT_XLSX"},
            "invocation_contract": {"injection_type": "informational"},
            "compatibility_notes": "Validate workbook layout dynamically; avoid hard-coded values from the source task.",
        },
        "metadata": {
            "domains": ["SpreadsheetBench", instruction_type],
            "allowed_tools": ["openpyxl"],
            "intent_keywords": keywords,
            "source_task_ids": [str(task.get("task_id") or "")],
            "evidence_span": f"successful task score={result.get('score')} answer_position={answer_position}",
            "scope": f"SpreadsheetBench {instruction_type}",
            "non_applicability": "Current workbook layout or instruction type differs materially from the source evidence.",
            "maintenance_action": "new_skill",
        },
        "dependencies": [],
    }


def _spreadsheet_has_repair_evidence(projection: Dict[str, Any]) -> bool:
    if projection.get("execution_ok") is False and projection.get("stderr_tail"):
        return True
    mismatches = list(projection.get("mismatched_cells") or [])
    if not mismatches:
        return False
    expected_text = " ".join(str(item.get("expected") or "") for item in mismatches if isinstance(item, dict))
    predicted_text = " ".join(str(item.get("predicted") or "") for item in mismatches if isinstance(item, dict))
    return bool(expected_text.strip() or predicted_text.strip())


def _heuristic_spreadsheet_repair_artifact_payload(detail: Dict[str, Any]) -> Dict[str, Any]:
    task = _spreadsheet_task_fragment(detail)
    result = _spreadsheet_result_projection(detail)
    projection = _spreadsheet_trace_projection(detail)
    instruction_type = str((task.get("metadata") or {}).get("instruction_type") or "spreadsheet_repair").strip()
    mismatches = list(projection.get("mismatched_cells") or [])
    expected_examples = [
        {"cell": item.get("cell"), "expected": item.get("expected")}
        for item in mismatches
        if isinstance(item, dict)
    ][:4]
    predicted_examples = [
        {"cell": item.get("cell"), "predicted": item.get("predicted")}
        for item in mismatches
        if isinstance(item, dict)
    ][:4]
    question = str(task.get("question") or "")
    keywords = _spreadsheet_keywords(question, instruction_type)
    formula_tokens = _spreadsheet_formula_tokens(expected_examples)
    name = normalize_skill_name(
        f"spreadsheet_repair_{instruction_type}_{'_'.join((formula_tokens or keywords)[:4])}"
    )[:90]
    body = (
        f"Applicability: SpreadsheetBench {instruction_type} tasks with similar instruction terms "
        f"{keywords[:8]} and verifier evidence matching these expected formula/value patterns: "
        f"{json.dumps(expected_examples, ensure_ascii=False)}.\n\n"
        "Corrective rule: when generating formulas, preserve the official spreadsheet semantics "
        "visible in the task and workbook preview, including blank-cell guards, lookup ranges, "
        "absolute references, and formula strings. Prefer writing the formula pattern itself when "
        "the answer range expects formulas, not computed display values.\n\n"
        "Failure evidence to avoid:\n"
        f"predicted={json.dumps(predicted_examples, ensure_ascii=False)}\n"
        f"expected={json.dumps(expected_examples, ensure_ascii=False)}\n\n"
        "Openpyxl idiom:\n"
        "```python\n"
        "import openpyxl\n"
        "wb = openpyxl.load_workbook(INPUT_XLSX)\n"
        "# inspect workbook layout, then write formulas/values to the requested answer range\n"
        "wb.save(OUTPUT_XLSX)\n"
        "```\n\n"
        "Non-applicability: do not hard-code this exact formula unless the current task has the "
        "same workbook layout, answer range, and instruction semantics."
    )
    return {
        "name": name,
        "kind": "workflow_guardrail_card",
        "description": f"Repair pattern for {instruction_type} SpreadsheetBench formula/value mismatches.",
        "body": body,
        "interface": {
            "summary": f"SpreadsheetBench {instruction_type} mismatch repair pattern.",
            "usage": "Retrieve for similar formula/value mismatch-prone spreadsheet tasks.",
            "input_contract": {"benchmark": "SpreadsheetBench", "instruction_type": instruction_type},
            "output_contract": {"tool": "openpyxl", "save_to": "OUTPUT_XLSX", "preserve_formula_semantics": True},
            "invocation_contract": {"injection_type": "workflow"},
            "compatibility_notes": "Validate workbook layout and answer range before applying any source formula pattern.",
        },
        "metadata": {
            "domains": ["SpreadsheetBench", instruction_type],
            "allowed_tools": ["openpyxl"],
            "intent_keywords": list(dict.fromkeys([*keywords, *formula_tokens])),
            "source_task_ids": [str(task.get("task_id") or "")],
            "evidence_span": f"failed trace score={result.get('score')} mismatches={expected_examples}",
            "scope": f"SpreadsheetBench {instruction_type} mismatch repair",
            "non_applicability": "Do not apply when formula/value semantics, sheet layout, or answer range differ.",
            "maintenance_action": "new_skill",
            "extracted_from_failure": True,
        },
        "dependencies": [],
    }


def _spreadsheet_formula_tokens(examples: Sequence[Dict[str, Any]]) -> List[str]:
    text = " ".join(str(item.get("expected") or "") for item in examples)
    tokens = []
    for word in re.findall(r"[A-Za-z_][A-Za-z0-9_]+", text):
        lowered = word.lower()
        if lowered in {"if", "or", "and"} or len(lowered) >= 4:
            if lowered not in tokens:
                tokens.append(lowered)
        if len(tokens) >= 8:
            break
    return tokens


def _spreadsheet_keywords(question: str, instruction_type: str = "") -> List[str]:
    raw = f"{instruction_type} {question}"
    words = re.findall(r"[A-Za-z][A-Za-z0-9_]+", raw.lower())
    stop = {
        "the", "and", "for", "with", "into", "from", "this", "that", "workbook",
        "spreadsheet", "sheet", "cell", "cells", "please", "using", "make",
    }
    out: List[str] = []
    for word in words:
        if len(word) < 3 or word in stop:
            continue
        if word not in out:
            out.append(word)
        if len(out) >= 12:
            break
    return out or ["spreadsheet", "openpyxl"]


def _normalize_spreadsheet_credit_events(
    payload: Dict[str, Any],
    *,
    detail: Dict[str, Any],
    candidate_artifacts: Sequence[SkillArtifact],
    projection: Dict[str, Any],
) -> List[Dict[str, Any]]:
    known = {artifact.name for artifact in candidate_artifacts}
    events: List[Dict[str, Any]] = []
    for raw in payload.get("skill_judgments") or []:
        row = dict(raw or {})
        skill_name = str(row.get("skill_name") or "").strip()
        if skill_name not in known:
            continue
        judgment = str(row.get("judgment") or "uncertain").strip().lower()
        if judgment not in {"helpful", "harmful", "neutral", "uncertain"}:
            judgment = "uncertain"
        suggestions = []
        for item in row.get("bundle_case_suggestions") or []:
            suggestion = dict(item or {})
            suggestion.setdefault("skill_name", skill_name)
            suggestions.append(_normalize_spreadsheet_bundle_suggestion(suggestion, detail=detail))
        events.append(
            {
                "benchmark": "spreadsheet",
                "task_id": detail.get("task_id"),
                "skill_name": skill_name,
                "judgment": judgment,
                "effect_type": str(row.get("effect_type") or "unknown").strip().lower() or "unknown",
                "confidence": max(0.0, min(1.0, float(row.get("confidence") or 0.0))),
                "reason": str(row.get("reason") or ""),
                "maintenance_actions": [
                    dict(item or {}) for item in (row.get("maintenance_actions") or []) if isinstance(item, dict)
                ],
                "refine_required": bool(row.get("refine_required")),
                "filter_candidate": bool(row.get("filter_candidate")),
                "evidence_strength": str(row.get("evidence_strength") or "weak").strip().lower() or "weak",
                "attribution_scope": str(row.get("attribution_scope") or "none").strip().lower() or "none",
                "bundle_case_suggestions": suggestions,
                "evidence": copy.deepcopy(dict(row.get("evidence") or {})),
                "projection": copy.deepcopy(projection),
            }
        )
    return events


def _normalize_spreadsheet_bundle_suggestion(
    suggestion: Dict[str, Any],
    *,
    detail: Dict[str, Any],
) -> Dict[str, Any]:
    polarity = str(suggestion.get("polarity") or "").strip().lower()
    if polarity not in {"positive", "negative", "integration"}:
        polarity = "integration"
    policy = str(suggestion.get("task_fragment_policy") or "reuse_official_fragment").strip()
    return {
        "skill_name": str(suggestion.get("skill_name") or ""),
        "polarity": polarity,
        "reason": str(suggestion.get("reason") or ""),
        "source_task_id": str(suggestion.get("source_task_id") or detail.get("task_id") or ""),
        "focus_turn_indices": [0] if policy != "no_replayable_fragment" else [],
        "required_context_turn_indices": [],
        "state_requirements": copy.deepcopy(dict(suggestion.get("state_requirements") or {})),
        "expected_contract": str(suggestion.get("expected_contract") or "SpreadsheetBench verifier should match the official answer range in the golden workbook."),
        "task_fragment_policy": policy if policy in {"reuse_official_fragment", "no_replayable_fragment"} else "reuse_official_fragment",
    }


def _heuristic_spreadsheet_credit_events(
    *,
    detail: Dict[str, Any],
    candidate_artifacts: Sequence[SkillArtifact],
    projection: Dict[str, Any],
    reason: str,
) -> List[Dict[str, Any]]:
    events: List[Dict[str, Any]] = []
    success = bool(projection.get("success"))
    score = float(projection.get("score") or 0.0)
    stderr = str(projection.get("stderr_tail") or "")
    mismatches = projection.get("mismatched_cells") or []
    for artifact in candidate_artifacts:
        overlap = _spreadsheet_scope_overlap(artifact, detail)
        if success and score >= 0.9 and overlap:
            judgment = "helpful"
            effect = "correctness_gain"
            confidence = 0.55
            suggestions = [
                {
                    "skill_name": artifact.name,
                    "polarity": "positive",
                    "reason": "successful task with retrieved in-scope spreadsheet skill",
                    "source_task_id": detail.get("task_id"),
                    "focus_turn_indices": [0],
                    "required_context_turn_indices": [],
                    "state_requirements": {},
                    "expected_contract": "official SpreadsheetBench answer range should match the golden workbook",
                    "task_fragment_policy": "reuse_official_fragment",
                }
            ]
        elif (stderr or mismatches) and overlap:
            judgment = "uncertain"
            effect = "unknown"
            confidence = 0.35
            suggestions = []
        else:
            judgment = "neutral"
            effect = "no_material_effect"
            confidence = 0.5
            suggestions = []
        events.append(
            {
                "benchmark": "spreadsheet",
                "task_id": detail.get("task_id"),
                "skill_name": artifact.name,
                "judgment": judgment,
                "effect_type": effect,
                "confidence": confidence,
                "reason": reason,
                "maintenance_actions": [],
                "refine_required": False,
                "filter_candidate": False,
                "evidence_strength": "weak",
                "attribution_scope": "prompt_influence" if judgment == "helpful" else "none",
                "bundle_case_suggestions": suggestions,
                "evidence": {"retrieved": True, "trace_signals": [reason]},
                "projection": copy.deepcopy(projection),
            }
        )
    return events


def _spreadsheet_scope_overlap(artifact: SkillArtifact, detail: Dict[str, Any]) -> bool:
    task = _spreadsheet_task_fragment(detail)
    instruction_type = str((task.get("metadata") or {}).get("instruction_type") or "").lower()
    question = str(task.get("question") or "").lower()
    keywords = [str(item).lower() for item in (artifact.metadata.get("intent_keywords") or [])]
    domains = [str(item).lower() for item in (artifact.metadata.get("domains") or [])]
    if instruction_type and instruction_type in domains:
        return True
    return any(keyword and keyword in question for keyword in keywords)


def _spreadsheet_case_from_credit_suggestion(
    *,
    detail: Dict[str, Any],
    artifact: SkillArtifact,
    event: Dict[str, Any],
    suggestion: Dict[str, Any],
) -> SkillBundleCase | None:
    policy = str(suggestion.get("task_fragment_policy") or "").strip()
    if policy == "no_replayable_fragment":
        artifact.evidence.repeated_evidence.append(
            {
                "task_id": detail.get("task_id"),
                "skill_name": artifact.name,
                "reason": suggestion.get("reason") or event.get("reason"),
                "not_replayable": True,
            }
        )
        return None
    polarity = str(suggestion.get("polarity") or "").strip().lower()
    confidence = float(event.get("confidence") or 0.0)
    judgment = str(event.get("judgment") or "")
    used = bool((event.get("evidence") or {}).get("used"))
    if polarity == "negative" and not (judgment == "harmful" and (confidence >= 0.65 or used)):
        return None
    if polarity == "positive" and str(event.get("effect_type") or "") not in {
        "token_saving",
        "schema_help",
        "workflow_alignment",
        "correctness_gain",
    }:
        return None
    return _spreadsheet_case_from_task(
        detail=detail,
        skill_name=artifact.name,
        polarity=polarity,
        reason=str(suggestion.get("reason") or event.get("reason") or ""),
        source=f"credit_assigner_{polarity}",
        confidence=confidence,
        credit_event=event,
    )


def _spreadsheet_case_from_task(
    *,
    detail: Dict[str, Any],
    skill_name: str,
    polarity: str,
    reason: str,
    source: str,
    confidence: float,
    credit_event: Dict[str, Any] | None = None,
) -> SkillBundleCase:
    task = _spreadsheet_task_fragment(detail)
    case_id = f"{skill_name}:{polarity}:{_stable_id(task.get('task_id'), reason, source)}"
    return SkillBundleCase(
        case_id=case_id,
        source=source,
        prompt=str(task.get("question") or "")[:240],
        expected={
            "answer_sheet": (task.get("expected") or {}).get("answer_sheet"),
            "answer_position": (task.get("expected") or {}).get("answer_position"),
            "verifier": "spreadsheet_golden_range",
        },
        context={
            "task_fragment": copy.deepcopy(task),
            "source_task_id": task.get("task_id"),
            "focus_turns": [0],
            "focus_tools": ["openpyxl"],
            "reason": reason,
            "credit_event": copy.deepcopy(credit_event or {}),
            "confidence": confidence,
        },
        tags=["spreadsheet", polarity],
        polarity=polarity,
        contrast_protocol={"with_skill": True, "without_skill": polarity != "negative"},
    )


def _bundle_bucket(polarity: str) -> str:
    return {
        "positive": "positive_cases",
        "negative": "negative_cases",
        "integration": "integration_cases",
    }.get(str(polarity), "integration_cases")


def _bundle_case_limit_per_polarity() -> int:
    try:
        return max(1, int(os.environ.get("BFCL_BUNDLE_CASE_LIMIT_PER_POLARITY", "2") or "2"))
    except Exception:
        return 2


def _bundle_max_total_cases() -> int:
    try:
        return max(1, int(os.environ.get("BFCL_BUNDLE_MAX_TOTAL_CASES", "6") or "6"))
    except Exception:
        return 6


def _trim_spreadsheet_bundle_cases(artifact: SkillArtifact) -> None:
    per = _bundle_case_limit_per_polarity()
    total = _bundle_max_total_cases()
    for attr in ("positive_cases", "negative_cases", "integration_cases"):
        cases = list(getattr(artifact.bundle, attr) or [])
        if len(cases) > per:
            cases.sort(key=_spreadsheet_case_priority, reverse=True)
            setattr(artifact.bundle, attr, cases[:per])
    all_cases = [
        ("positive_cases", case) for case in artifact.bundle.positive_cases
    ] + [
        ("negative_cases", case) for case in artifact.bundle.negative_cases
    ] + [
        ("integration_cases", case) for case in artifact.bundle.integration_cases
    ]
    if len(all_cases) <= total:
        return
    all_cases.sort(key=lambda item: _spreadsheet_case_priority(item[1]), reverse=True)
    kept = {id(case) for _bucket, case in all_cases[:total]}
    for attr in ("positive_cases", "negative_cases", "integration_cases"):
        setattr(artifact.bundle, attr, [case for case in getattr(artifact.bundle, attr) if id(case) in kept])
    artifact.bundle.fixtures = {
        **dict(artifact.bundle.fixtures or {}),
        "bundle_trimmed": True,
        "bundle_split_count": len(all_cases) - total,
    }


def _spreadsheet_case_priority(case: SkillBundleCase) -> tuple[int, float, str]:
    ctx = dict(case.context or {})
    source = str(case.source or "")
    confidence = float(ctx.get("confidence") or (ctx.get("credit_event") or {}).get("confidence") or 0.0)
    credit = 1 if source.startswith("credit_assigner") else 0
    regression = 1 if case.polarity in {"negative", "integration"} else 0
    return (credit + regression, confidence, case.case_id)


def _spreadsheet_micro_targets(
    *,
    credit_events: Sequence[Dict[str, Any]],
    credit_bundle_cases: Sequence[Dict[str, Any]],
    extracted: Sequence[SkillArtifact],
) -> List[str]:
    targets: List[str] = []
    for event in credit_events:
        judgment = str(event.get("judgment") or "")
        confidence = float(event.get("confidence") or 0.0)
        if event.get("refine_required") or event.get("filter_candidate"):
            targets.append(str(event.get("skill_name") or ""))
        elif judgment == "harmful" and confidence >= 0.65:
            targets.append(str(event.get("skill_name") or ""))
        elif judgment == "helpful" and confidence >= 0.55:
            targets.append(str(event.get("skill_name") or ""))
    targets.extend(str(item.get("skill_name") or "") for item in credit_bundle_cases)
    return sorted({name for name in targets if name})


async def _refine_spreadsheet_skill_from_credit(
    *,
    artifact: SkillArtifact,
    credit_context: Sequence[Dict[str, Any]],
    detail: Dict[str, Any],
    store: ArtifactStore,
    config: MaintenanceRunConfig,
) -> Dict[str, Any]:
    strong = [
        event for event in credit_context
        if event.get("refine_required")
        or event.get("filter_candidate")
        or (event.get("judgment") == "harmful" and float(event.get("confidence") or 0.0) >= 0.65)
    ]
    if not strong:
        return {"skill_name": artifact.name, "action": "keep", "reason": "no_strong_credit_signal"}
    test_result = {
        "result_id": f"{artifact.name}:credit:{_stable_id(detail.get('task_id'), _now_iso())}",
        "aggregate": {"passed": False, "reason": "pre_bundle_credit_refine"},
        "failed_cases": [_spreadsheet_result_projection(detail)],
    }
    return await _run_spreadsheet_refiner(
        artifact=artifact,
        test_result=test_result,
        credit_context=list(strong),
        store=store,
        config=config,
        phase="spreadsheet_credit_pre_refine",
    )


async def _refine_spreadsheet_skill_from_bundle(
    *,
    artifact: SkillArtifact,
    test_result: SkillTestResult,
    credit_context: Sequence[Dict[str, Any]],
    store: ArtifactStore,
    config: MaintenanceRunConfig,
) -> Dict[str, Any]:
    return await _run_spreadsheet_refiner(
        artifact=artifact,
        test_result=test_result.as_dict(),
        credit_context=list(credit_context),
        store=store,
        config=config,
        phase="spreadsheet_bundle_refine",
    )


async def _run_spreadsheet_refiner(
    *,
    artifact: SkillArtifact,
    test_result: Dict[str, Any],
    credit_context: Sequence[Dict[str, Any]],
    store: ArtifactStore,
    config: MaintenanceRunConfig,
    phase: str,
) -> Dict[str, Any]:
    try:
        payload = await refine_skill_artifact_llm(
            artifact,
            test_result=test_result,
            credit_context=list(credit_context),
            refinement_history=artifact.history[-3:],
            dependency_summaries=summarize_dependency_context(store.all()),
            llm_config=config.llm_config,
            model_name=config.model_name,
            audit_context={"phase": phase, "benchmark": "spreadsheet"},
        )
    except Exception as exc:
        if any(event.get("filter_candidate") for event in credit_context):
            updated = copy.deepcopy(artifact)
            updated.status = "disabled"
            updated.metadata["disabled"] = True
            updated.metadata["disabled_reason"] = f"spreadsheet_refiner_unavailable_after_filter_credit:{type(exc).__name__}"
            return {
                "skill_name": artifact.name,
                "action": "disable",
                "reason": updated.metadata["disabled_reason"],
                "updated_artifact": updated,
            }
        return {
            "skill_name": artifact.name,
            "action": "keep",
            "reason": f"refiner_failed:{type(exc).__name__}",
        }
    decision = dict(payload.get("decision") or {})
    action = str(decision.get("action") or "keep").strip()
    if action == "keep":
        return {"skill_name": artifact.name, "action": "keep", "reason": decision.get("reason") or ""}
    updated = apply_refine_payload(artifact, payload)
    if _artifact_semantic_signature(updated) == _artifact_semantic_signature(artifact):
        return {"skill_name": artifact.name, "action": "keep", "reason": "refiner_returned_no_semantic_change"}
    return {
        "skill_name": artifact.name,
        "action": action,
        "reason": decision.get("reason") or "",
        "updated_artifact": updated,
    }


def _artifact_semantic_signature(artifact: SkillArtifact) -> str:
    payload = {
        "kind": artifact.kind,
        "description": artifact.description,
        "body": artifact.body,
        "metadata": artifact.metadata,
        "interface": artifact.interface.as_dict(),
        "status": artifact.status,
        "dependencies": artifact.dependencies,
    }
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


async def _execute_spreadsheet_bundle_tests(
    *,
    artifact: SkillArtifact,
    config: MaintenanceRunConfig,
) -> SkillTestResult:
    case_runs: List[SkillTestCaseRun] = []
    for case in artifact.bundle.all_cases():
        task_fragment = dict((case.context or {}).get("task_fragment") or {})
        task = BenchmarkTask(
            benchmark="spreadsheet",
            task_id=str(task_fragment.get("task_id") or case.case_id),
            question=task_fragment.get("question") or case.prompt,
            expected=copy.deepcopy(dict(task_fragment.get("expected") or {})),
            input_artifacts=copy.deepcopy(dict(task_fragment.get("input_artifacts") or {})),
            metadata=copy.deepcopy(dict(task_fragment.get("metadata") or {})),
        )
        if not task.input_artifacts.get("input_xlsx") or not task.expected.get("golden_xlsx"):
            case_runs.append(
                SkillTestCaseRun(
                    case_id=case.case_id,
                    variant="with_skill",
                    passed=False,
                    accuracy=0.0,
                    failure_summary="missing official workbook paths",
                    trace_ref=task.task_id,
                    bundle_case_snapshot=case.as_dict(),
                    metadata={"polarity": case.polarity, "contract_failure": "missing_workbook_paths"},
                )
            )
            continue
        result = await run_spreadsheet_task(
            task,
            llm_config=config.llm_config,
            model_name=config.model_name,
            artifact_store=ArtifactStore([copy.deepcopy(artifact)]),
            top_k_skills=1,
        )
        metrics = dict(result.metrics or {})
        case_runs.append(
            SkillTestCaseRun(
                case_id=case.case_id,
                variant="with_skill",
                passed=bool(result.success),
                accuracy=float(result.score or 0.0),
                tokens=int(metrics.get("total_tokens") or 0),
                failure_summary="" if result.success else str(result.error or metrics.get("reason") or metrics.get("mismatched_cells") or "failed"),
                trace_ref=result.task_id,
                trace=copy.deepcopy(result.trace),
                input_payload={"task": task.as_dict(), "skill_name": artifact.name},
                expected_behavior=copy.deepcopy(case.expected),
                actual_output={"metrics": metrics, "score": result.score, "success": result.success},
                trace_summary=_spreadsheet_trace_projection(
                    {"task_id": result.task_id, "task": task.as_dict(), "runs": [result.as_dict()]}
                ),
                skill_snapshot={"name": artifact.name, "version": artifact.version},
                bundle_case_snapshot=case.as_dict(),
                metadata={"polarity": case.polarity},
            )
        )
    passed = bool(case_runs) and all(run.passed for run in case_runs)
    avg_accuracy = round(
        sum(float(run.accuracy or 0.0) for run in case_runs) / max(len(case_runs), 1),
        4,
    )
    result = SkillTestResult(
        result_id=f"{artifact.name}:spreadsheet_bundle:{_stable_id(artifact.version, _now_iso())}",
        skill_name=artifact.name,
        skill_version=artifact.version,
        bundle_id=artifact.bundle.bundle_id or f"{artifact.name}.bundle",
        bundle_version=artifact.bundle.bundle_version,
        run_label="spreadsheet_bundle_unit",
        unit_case_runs=case_runs,
        aggregate={
            "passed": passed,
            "n_cases": len(case_runs),
            "n_passed": sum(1 for run in case_runs if run.passed),
            "avg_accuracy": avg_accuracy,
        },
        created_at=_now_iso(),
    )
    return result


def _promote_spreadsheet_pending_from_window(
    store: ArtifactStore,
    window_details: Sequence[Dict[str, Any]],
) -> List[str]:
    window_task_ids = {str(item.get("task_id") or "") for item in window_details}
    promoted: List[str] = []
    for artifact in store.pending_artifacts():
        source_ids = {str(item) for item in (artifact.metadata.get("source_task_ids") or [])}
        if not source_ids & window_task_ids:
            continue
        if not artifact.bundle.positive_cases:
            continue
        if store.promote_pending(
            artifact.name,
            reason="spreadsheet_successful_source_trace_in_macro_window",
            refactor_group_id=f"spreadsheet_macro:{_stable_id(*sorted(window_task_ids))}",
        ):
            promoted.append(artifact.name)
    return promoted


def _dedupe_spreadsheet_skills(store: ArtifactStore) -> List[str]:
    active = [artifact for artifact in store.all() if artifact.status == "active" and not artifact.is_disabled()]
    groups: Dict[str, List[SkillArtifact]] = {}
    for artifact in active:
        key = _spreadsheet_dedupe_key(artifact)
        groups.setdefault(key, []).append(artifact)
    archived: List[str] = []
    for group in groups.values():
        if len(group) <= 1:
            continue
        group.sort(
            key=lambda item: (
                item.success_count,
                item.usage_count,
                len(item.evidence.helpful_cases),
                -len(item.evidence.harmful_cases),
            ),
            reverse=True,
        )
        keeper = group[0]
        for duplicate in group[1:]:
            duplicate.status = "archived"
            duplicate.metadata["archived_reason"] = f"spreadsheet_duplicate_of:{keeper.name}"
            duplicate.metadata["duplicate_keeper"] = keeper.name
            keeper.evidence.repeated_evidence.append(
                {
                    "merged_duplicate": duplicate.name,
                    "source_task_ids": duplicate.metadata.get("source_task_ids") or [],
                }
            )
            archived.append(duplicate.name)
    return archived


def _spreadsheet_dedupe_key(artifact: SkillArtifact) -> str:
    instruction_type = str(artifact.metadata.get("instruction_type") or "").lower()
    keywords = sorted(str(item).lower() for item in (artifact.metadata.get("intent_keywords") or [])[:6])
    text = " ".join([instruction_type, artifact.kind, *keywords])
    return _stable_id(text, length=12)


def _filter_spreadsheet_harmful_skills(
    store: ArtifactStore,
    credit_events: Sequence[Dict[str, Any]],
) -> List[str]:
    harmful_by_skill: Dict[str, List[Dict[str, Any]]] = {}
    helpful_by_skill: Dict[str, int] = {}
    for event in credit_events:
        skill_name = str(event.get("skill_name") or "")
        if not skill_name:
            continue
        if event.get("judgment") == "harmful" and float(event.get("confidence") or 0.0) >= 0.65:
            harmful_by_skill.setdefault(skill_name, []).append(dict(event))
        if event.get("judgment") == "helpful":
            helpful_by_skill[skill_name] = helpful_by_skill.get(skill_name, 0) + 1
    filtered: List[str] = []
    for skill_name, rows in harmful_by_skill.items():
        if len(rows) < 2 or helpful_by_skill.get(skill_name, 0) > 0:
            continue
        artifact = store.get(skill_name)
        if artifact is None or artifact.is_disabled():
            continue
        artifact.status = "disabled"
        artifact.metadata["disabled"] = True
        artifact.metadata["disabled_reason"] = "spreadsheet_macro_repeated_high_confidence_harmful_credit"
        filtered.append(skill_name)
    return filtered


def _split_sheet_range(answer_range: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    if not answer_range:
        return None, answer_range
    text = str(answer_range).strip()
    if "!" not in text:
        return None, text.strip().strip("'").strip('"')
    sheet, cell_range = text.rsplit("!", 1)
    sheet = sheet.strip().strip("'").strip('"')
    cell_range = cell_range.strip().strip("'").strip('"')
    return sheet or None, cell_range


def _answer_range_refs(
    answer_range: Optional[str],
    *,
    default_sheet: Optional[str],
) -> List[Tuple[Optional[str], Optional[str]]]:
    parts = _split_answer_range_list(answer_range)
    refs: List[Tuple[Optional[str], Optional[str]]] = []
    for part in parts:
        sheet, cell_range = _split_sheet_range(_normalize_answer_range_text(part))
        refs.append((sheet or default_sheet, cell_range))
    if not refs and default_sheet:
        refs.append((default_sheet, None))
    return refs


def _split_answer_range_list(answer_range: Optional[str]) -> List[str]:
    text = str(answer_range or "").strip()
    if not text:
        return []
    parts: List[str] = []
    current: List[str] = []
    in_quote = False
    for char in text:
        if char == "'":
            in_quote = not in_quote
        if char == "," and not in_quote:
            part = "".join(current).strip()
            if part:
                parts.append(part)
            current = []
            continue
        current.append(char)
    part = "".join(current).strip()
    if part:
        parts.append(part)
    return parts


def _normalize_answer_range_text(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return text
    # Some SpreadsheetBench rows encode ranges as "'Sheet1!'A1:B2".
    text = re.sub(r"^'([^']+!)'([A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?)$", r"'\1\2", text)
    text = text.strip().strip('"')
    if text.count("'") == 1 and text.startswith("'") and "!" in text:
        text = text[1:]
    return text


def _first_sheet_name(sheet_name: Optional[str]) -> Optional[str]:
    if not sheet_name:
        return None
    return str(sheet_name).split(",", 1)[0].strip().strip("'").strip('"') or None


def _normalize_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 8)
    if isinstance(value, int) and not isinstance(value, bool):
        return value
    if isinstance(value, bool):
        return value
    text = str(value).strip()
    if re.fullmatch(r"-?\d+", text):
        try:
            return int(text)
        except Exception:
            return text
    if re.fullmatch(r"-?\d+\.\d+", text):
        try:
            return round(float(text), 8)
        except Exception:
            return text
    return text


def _jsonable_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
