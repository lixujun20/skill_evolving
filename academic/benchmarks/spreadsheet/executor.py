"""SpreadsheetBench task execution."""
from __future__ import annotations

import asyncio
import json
import os
import re
import shutil
import subprocess
import tempfile
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

import openpyxl

from academic.benchmarks.core.artifacts import ArtifactStore
from academic.benchmarks.core.cost_accounting import make_cost_event
from academic.benchmarks.core.llm_text import ask_text_llm
from academic.benchmarks.core.skill_injector import (
    injector_llm_config,
    injector_model_name,
    render_skill_prompt_blocks,
    select_skill_context_with_llm,
)
from academic.benchmarks.core.types import BenchmarkResult, BenchmarkTask
from academic.benchmarks.spreadsheet.models import SpreadsheetTrace
from academic.benchmarks.spreadsheet.prompts import (
    SPREADSHEET_BASH_SYSTEM,
    SPREADSHEET_DONE_PATTERN,
    SPREADSHEET_NOTEBOOK_SYSTEM,
    SPREADSHEET_SYSTEM,
)
from academic.benchmarks.spreadsheet.skill_runtime import (
    NotebookPythonSession,
    called_spreadsheet_skill_code_reads,
    called_spreadsheet_skill_functions,
    extract_code,
    run_code,
    write_spreadsheet_skill_library,
    write_spreadsheet_skill_packages,
)
from academic.benchmarks.spreadsheet.verifier import jsonable_cell_value, verify_spreadsheet_output
from academic.config import CODE_EXEC_TIMEOUT


async def run_spreadsheet_task(
    task: BenchmarkTask,
    *,
    llm_config: str,
    model_name: Optional[str] = None,
    artifact_store: Optional[ArtifactStore] = None,
    top_k_skills: int = 5,
    min_skill_score: float = 0.01,
    skill_injector_mode: str | None = None,
    skill_context_budget_chars: int | None = None,
    callable_disclosure_mode: str | None = None,
    pending_skill_fraction: float = 0.0,
    llm_request_timeout_s: float | None = None,
    work_dir: Optional[Path] = None,
) -> BenchmarkResult:
    t0 = time.monotonic()
    trace = SpreadsheetTrace(task_id=task.task_id)
    query = str(task.question)
    retrieved, retrieval_event = _retrieve_spreadsheet_skills(
        artifact_store,
        query=query,
        top_k=top_k_skills,
        min_score=min_skill_score,
        pending_skill_fraction=pending_skill_fraction,
    )
    trace.retrieved_skills = [skill.name for skill in retrieved]
    if retrieval_event:
        trace.injector_events.append(retrieval_event)
    presentation_mode = _spreadsheet_presentation_mode(
        skill_injector_mode
        or os.environ.get("SPREADSHEET_SKILL_INJECTOR_MODE")
        or os.environ.get("SKILL_INJECTOR_MODE"),
    )
    injected = list(retrieved)
    skill_prompt = artifact_store.build_prompt(injected) if artifact_store else "(none)"
    if artifact_store and not _spreadsheet_direct_skill_context_enabled(presentation_mode):
        injection = await select_skill_context_with_llm(
            retrieved,
            query=query,
            llm_config=llm_config,
            model_name=model_name,
            presentation_mode=presentation_mode,
            allowed_injection_types={"informational", "workflow", "functional"},
            max_selected=top_k_skills,
            budget_chars=int(skill_context_budget_chars or os.environ.get("SPREADSHEET_SKILL_CONTEXT_BUDGET_CHARS", os.environ.get("SKILL_INJECTOR_BUDGET_CHARS", "2200")) or "2200"),
            compact_chars_per_skill=int(os.environ.get("SPREADSHEET_SKILL_COMPACT_CHARS_PER_SKILL", os.environ.get("SKILL_INJECTOR_COMPACT_CHARS_PER_SKILL", "900")) or "900"),
            benchmark="spreadsheet",
            task_id=task.task_id,
            phase="executor",
        )
        injected = injection.artifacts
        skill_prompt = injection.prompt()
        trace.filtered_skills = list(injection.filtered)
        injection_event = injection.as_event()
        trace.injector_events.append(injection_event)
        trace.cost_events.append(
            make_cost_event(
                role="injector",
                phase="executor",
                benchmark="spreadsheet",
                task_id=task.task_id,
                model=injector_model_name(model_name, benchmark="spreadsheet") or "",
                llm_config=injector_llm_config(llm_config, benchmark="spreadsheet"),
                skill_prompt_chars=int(injection_event.get("prompt_chars") or 0),
                metadata={
                    "mode": presentation_mode,
                    "gate": "llm",
                    "injected_count": injection_event.get("injected_count"),
                    "filtered_count": injection_event.get("filtered_count"),
                    "execution_mode": "single",
                },
            )
        )
    elif artifact_store:
        injected = list(retrieved)
        skill_prompt = render_skill_prompt_blocks(
            injected,
            mode="full",
            budget_chars=int(skill_context_budget_chars or os.environ.get("SPREADSHEET_SKILL_CONTEXT_BUDGET_CHARS", os.environ.get("SKILL_INJECTOR_BUDGET_CHARS", "2200")) or "2200"),
            compact_chars_per_skill=int(os.environ.get("SPREADSHEET_SKILL_COMPACT_CHARS_PER_SKILL", os.environ.get("SKILL_INJECTOR_COMPACT_CHARS_PER_SKILL", "900")) or "900"),
        )
        trace.injector_events.append(
            {
                "mode": presentation_mode,
                "gate": "direct",
                "injected_count": len(injected),
                "filtered_count": 0,
                "prompt_chars": len(skill_prompt),
                "injected": [
                    {
                        "skill_name": skill.name,
                        "decision": "direct_inject",
                        "reason": "preselected_bundle_replay_candidate",
                        "kind": getattr(skill, "kind", ""),
                        "injection_type": skill.injection_type(),
                    }
                    for skill in injected
                ],
                "filtered": [],
            }
        )

    base_work_dir = work_dir or Path(tempfile.mkdtemp(prefix="spreadsheetbench_"))
    base_work_dir.mkdir(parents=True, exist_ok=True)
    callable_prompt = write_spreadsheet_skill_library(
        injected,
        base_work_dir,
        disclosure_mode=callable_disclosure_mode or os.environ.get("SPREADSHEET_CALLABLE_DISCLOSURE_MODE") or "full",
    )
    package_prompt = write_spreadsheet_skill_packages(
        injected,
        base_work_dir,
        disclosure_mode=os.environ.get("SPREADSHEET_PACKAGE_DISCLOSURE_MODE") or "progressive",
    )
    trace.callable_skills = callable_prompt["skills"]
    trace.package_skills = package_prompt["skills"]
    trace.prompt_injected_skills = [skill.name for skill in injected]
    trace.skill_code_reads = []
    system = SPREADSHEET_SYSTEM.format(
        skills=skill_prompt,
        callable_skills="\n\n".join(
            item
            for item in [
                callable_prompt["prompt"] or "(no callable function skills available)",
                package_prompt["prompt"],
            ]
            if item
        ),
    )
    input_copy = base_work_dir / f"{task.task_id}_input.xlsx"
    output_path = base_work_dir / f"{task.task_id}_output.xlsx"
    shutil.copyfile(task.input_artifacts["input_xlsx"], input_copy)
    shutil.copyfile(input_copy, output_path)
    prompt = build_spreadsheet_prompt(task, input_copy, output_path)
    trace.prompt = prompt
    try:
        response = await ask_text_llm(
            llm_config=llm_config,
            model_name=model_name,
            system=system,
            prompt=prompt,
            max_request_wall_s=llm_request_timeout_s,
        )
        trace.input_tokens += response.prompt_tokens
        trace.cache_input_tokens += response.cache_input_tokens
        trace.completion_tokens += response.completion_tokens
        trace.total_tokens += response.prompt_tokens + response.cache_input_tokens + response.completion_tokens
        trace.cost_events.append(
            make_cost_event(
                role="executor",
                phase="task_rollout",
                benchmark="spreadsheet",
                task_id=task.task_id,
                model=response.model_name or model_name or "",
                llm_config=llm_config,
                input_tokens=response.prompt_tokens,
                cache_input_tokens=response.cache_input_tokens,
                output_tokens=response.completion_tokens,
                prompt_chars=len(prompt),
                skill_prompt_chars=len(skill_prompt) + len(callable_prompt["prompt"] or "") + len(package_prompt["prompt"] or ""),
                system_prompt_chars=len(system),
                final_conversation_chars=len(system) + len(prompt) + len(response.content or ""),
                metadata={
                    "api_style": response.api_style,
                    "skill_injector_mode": presentation_mode,
                    "prompt_injected_skills": list(trace.prompt_injected_skills),
                    "callable_skills": list(trace.callable_skills),
                    "package_skills": list(trace.package_skills),
                    "callable_disclosure_mode": callable_prompt.get("disclosure_mode"),
                    "execution_mode": "single",
                },
            )
        )
        trace.code = extract_code(response.content)
        if not trace.code:
            trace.elapsed_s = round(time.monotonic() - t0, 3)
            return BenchmarkResult(
                benchmark="spreadsheet",
                task_id=task.task_id,
                success=False,
                score=0.0,
                metrics={"reason": "no_python_code"},
                trace=trace.as_dict(),
            )
        trace.called_skill_functions = called_spreadsheet_skill_functions(
            trace.code,
            callable_prompt["skills"],
        )
        trace.skill_code_reads.extend(
            called_spreadsheet_skill_code_reads(trace.code, callable_prompt["skills"])
        )
        stdout, stderr, returncode = await asyncio.to_thread(
            run_code, trace.code, input_copy, output_path, base_work_dir
        )
        trace.stdout = stdout
        trace.stderr = stderr
        verify = verify_spreadsheet_output(
            predicted_xlsx=output_path,
            golden_xlsx=Path(task.expected["golden_xlsx"]),
            sheet_name=task.expected.get("answer_sheet"),
            answer_range=task.expected.get("answer_position"),
        )
        verify["returncode"] = returncode
        verify["execution_ok"] = returncode == 0
        success = bool(returncode == 0 and verify["pass"])
    except Exception as exc:
        trace.elapsed_s = round(time.monotonic() - t0, 3)
        return BenchmarkResult(
            benchmark="spreadsheet",
            task_id=task.task_id,
            success=False,
            score=0.0,
            metrics={"exception": type(exc).__name__},
            trace=trace.as_dict(),
            error=str(exc),
        )

    trace.elapsed_s = round(time.monotonic() - t0, 3)
    verify["total_tokens"] = trace.total_tokens
    verify["input_tokens"] = trace.input_tokens
    verify["cache_input_tokens"] = trace.cache_input_tokens
    verify["completion_tokens"] = trace.completion_tokens
    verify["cost_events"] = trace.cost_events
    verify["elapsed_s"] = trace.elapsed_s
    verify["retrieved_skills"] = trace.retrieved_skills
    verify["prompt_injected_skills"] = trace.prompt_injected_skills
    verify["called_skill_functions"] = trace.called_skill_functions
    verify["skill_code_reads"] = trace.skill_code_reads
    verify["package_skills"] = trace.package_skills
    verify["filtered_skills"] = trace.filtered_skills
    verify["injector_events"] = trace.injector_events
    verify["skill_injector_mode"] = presentation_mode
    verify["model_name"] = response.model_name
    verify["llm_api_style"] = response.api_style
    return BenchmarkResult(
        benchmark="spreadsheet",
        task_id=task.task_id,
        success=success,
        score=float(verify["cell_accuracy"]),
        metrics=verify,
        trace=trace.as_dict(),
    )


async def run_spreadsheet_task_notebook(
    task: BenchmarkTask,
    *,
    llm_config: str,
    model_name: Optional[str] = None,
    artifact_store: Optional[ArtifactStore] = None,
    top_k_skills: int = 5,
    min_skill_score: float = 0.01,
    skill_injector_mode: str | None = None,
    skill_context_budget_chars: int | None = None,
    callable_disclosure_mode: str | None = None,
    pending_skill_fraction: float = 0.0,
    max_turns: int = 5,
    llm_request_timeout_s: float | None = None,
    work_dir: Optional[Path] = None,
) -> BenchmarkResult:
    t0 = time.monotonic()
    trace = SpreadsheetTrace(task_id=task.task_id, execution_mode="notebook")
    presentation_mode = _spreadsheet_presentation_mode(
        skill_injector_mode
        or os.environ.get("SPREADSHEET_SKILL_INJECTOR_MODE")
        or os.environ.get("SKILL_INJECTOR_MODE"),
    )
    base_work_dir = work_dir or Path(tempfile.mkdtemp(prefix="spreadsheetbench_nb_"))
    base_work_dir.mkdir(parents=True, exist_ok=True)
    injected: List[Any] = []
    callable_prompt: Dict[str, Any] = {"prompt": "", "skills": [], "disclosure_mode": "full"}
    package_prompt: Dict[str, Any] = {"prompt": "", "skills": [], "disclosure_mode": "progressive"}
    skill_prompt = "(no reusable skill artifacts retrieved)"
    skill_prompt_for_system = skill_prompt
    trace.skill_code_reads = []
    input_copy = base_work_dir / f"{task.task_id}_input.xlsx"
    output_path = base_work_dir / f"{task.task_id}_output.xlsx"
    shutil.copyfile(task.input_artifacts["input_xlsx"], input_copy)
    shutil.copyfile(input_copy, output_path)

    base_prompt = build_spreadsheet_notebook_prompt(task, input_copy, output_path, max_turns=max_turns)
    trace.prompt = base_prompt
    system = SPREADSHEET_NOTEBOOK_SYSTEM.format(
        skills="(runtime skill retrieval updates, if any, are appended in user messages during the notebook conversation)",
        callable_skills="(callable and folder-style skill updates, if any, are appended in user messages during the notebook conversation)",
        done_pattern=SPREADSHEET_DONE_PATTERN,
    )
    session: NotebookPythonSession | None = None
    stopped_by_done = False
    try:
        session = NotebookPythonSession(base_work_dir)
        session.run_cell(
            f"INPUT_XLSX = Path({str(input_copy)!r})\nOUTPUT_XLSX = Path({str(output_path)!r})",
            timeout=CODE_EXEC_TIMEOUT,
        )
        history: List[Dict[str, Any]] = []
        skill_context_messages: List[Dict[str, str]] = []
        response_model = model_name
        response_api_style = ""
        for turn_index in range(max(1, int(max_turns or 5))):
            turn_query = _spreadsheet_notebook_retrieval_query(
                task=task,
                base_prompt=base_prompt,
                history=history,
                turn_index=turn_index,
            )
            retrieved, retrieval_event = _retrieve_spreadsheet_skills(
                artifact_store,
                query=turn_query,
                top_k=top_k_skills,
                min_score=min_skill_score,
                pending_skill_fraction=pending_skill_fraction,
            )
            retrieved_names = [skill.name for skill in retrieved]
            trace.retrieved_skills = list(dict.fromkeys([*trace.retrieved_skills, *retrieved_names]))
            if retrieval_event:
                retrieval_event = {**retrieval_event, "turn_index": turn_index, "phase": "notebook_turn_start"}
                trace.injector_events.append(retrieval_event)
            prior_injected_names = {str(getattr(skill, "name", "")) for skill in injected}
            if artifact_store and not _spreadsheet_direct_skill_context_enabled(presentation_mode):
                injection = await select_skill_context_with_llm(
                    retrieved,
                    query=turn_query,
                    llm_config=llm_config,
                    model_name=model_name,
                    presentation_mode=presentation_mode,
                    allowed_injection_types={"informational", "workflow", "functional"},
                    max_selected=top_k_skills,
                    budget_chars=int(skill_context_budget_chars or os.environ.get("SPREADSHEET_SKILL_CONTEXT_BUDGET_CHARS", os.environ.get("SKILL_INJECTOR_BUDGET_CHARS", "2200")) or "2200"),
                    compact_chars_per_skill=int(os.environ.get("SPREADSHEET_SKILL_COMPACT_CHARS_PER_SKILL", os.environ.get("SKILL_INJECTOR_COMPACT_CHARS_PER_SKILL", "900")) or "900"),
                    benchmark="spreadsheet",
                    task_id=task.task_id,
                    phase="executor_step_update",
                    metadata={"turn_index": turn_index, "execution_mode": "notebook"},
                )
                injected = _merge_spreadsheet_artifacts(injected, injection.artifacts)
                skill_prompt = render_skill_prompt_blocks(
                    injected,
                    mode=presentation_mode,
                    budget_chars=int(skill_context_budget_chars or os.environ.get("SPREADSHEET_SKILL_CONTEXT_BUDGET_CHARS", os.environ.get("SKILL_INJECTOR_BUDGET_CHARS", "2200")) or "2200"),
                    compact_chars_per_skill=int(os.environ.get("SPREADSHEET_SKILL_COMPACT_CHARS_PER_SKILL", os.environ.get("SKILL_INJECTOR_COMPACT_CHARS_PER_SKILL", "900")) or "900"),
                )
                trace.filtered_skills.extend(
                    [{**row, "turn_index": turn_index} for row in list(injection.filtered)]
                )
                injection_event = injection.as_event()
                injection_event["turn_index"] = turn_index
                trace.injector_events.append(injection_event)
                trace.cost_events.append(
                    make_cost_event(
                        role="injector",
                        phase="executor_step_update",
                        benchmark="spreadsheet",
                        task_id=task.task_id,
                        turn_index=turn_index,
                        model=injector_model_name(model_name, benchmark="spreadsheet") or "",
                        llm_config=injector_llm_config(llm_config, benchmark="spreadsheet"),
                        skill_prompt_chars=int(injection_event.get("prompt_chars") or 0),
                        metadata={
                            "mode": presentation_mode,
                            "gate": "llm",
                            "injected_count": injection_event.get("injected_count"),
                            "filtered_count": injection_event.get("filtered_count"),
                            "execution_mode": "notebook",
                        },
                        )
                    )
            elif artifact_store:
                injected = _merge_spreadsheet_artifacts(injected, retrieved)
                skill_prompt = render_skill_prompt_blocks(
                    injected,
                    mode="full",
                    budget_chars=int(skill_context_budget_chars or os.environ.get("SPREADSHEET_SKILL_CONTEXT_BUDGET_CHARS", os.environ.get("SKILL_INJECTOR_BUDGET_CHARS", "2200")) or "2200"),
                    compact_chars_per_skill=int(os.environ.get("SPREADSHEET_SKILL_COMPACT_CHARS_PER_SKILL", os.environ.get("SKILL_INJECTOR_COMPACT_CHARS_PER_SKILL", "900")) or "900"),
                )
                injection_event = {
                    "mode": presentation_mode,
                    "gate": "direct",
                    "turn_index": turn_index,
                    "injected_count": len(retrieved),
                    "filtered_count": 0,
                    "prompt_chars": len(skill_prompt),
                    "injected": [
                        {
                            "skill_name": skill.name,
                            "decision": "direct_inject",
                            "reason": "preselected_bundle_replay_candidate",
                            "kind": getattr(skill, "kind", ""),
                            "injection_type": skill.injection_type(),
                        }
                        for skill in retrieved
                    ],
                    "filtered": [],
                }
                trace.injector_events.append(injection_event)
            else:
                injected = _merge_spreadsheet_artifacts(injected, retrieved)
                skill_prompt = artifact_store.build_prompt(injected) if artifact_store else "(no reusable skill artifacts retrieved)"
            trace.prompt_injected_skills = [skill.name for skill in injected]
            new_injected_names = [
                str(getattr(skill, "name", ""))
                for skill in injected
                if str(getattr(skill, "name", "")) not in prior_injected_names
            ]
            callable_prompt = write_spreadsheet_skill_library(
                injected,
                base_work_dir,
                disclosure_mode=callable_disclosure_mode or os.environ.get("SPREADSHEET_CALLABLE_DISCLOSURE_MODE") or "full",
            )
            package_prompt = write_spreadsheet_skill_packages(
                injected,
                base_work_dir,
                disclosure_mode=os.environ.get("SPREADSHEET_PACKAGE_DISCLOSURE_MODE") or "progressive",
            )
            trace.callable_skills = _merge_callable_rows(trace.callable_skills, callable_prompt["skills"])
            trace.package_skills = _merge_package_rows(trace.package_skills, package_prompt["skills"])
            if new_injected_names:
                added_artifacts = [
                    skill
                    for skill in injected
                    if str(getattr(skill, "name", "")) in set(new_injected_names)
                ]
                added_skill_prompt = (
                    render_skill_prompt_blocks(
                        added_artifacts,
                        mode=presentation_mode,
                        budget_chars=int(skill_context_budget_chars or os.environ.get("SPREADSHEET_SKILL_CONTEXT_BUDGET_CHARS", os.environ.get("SKILL_INJECTOR_BUDGET_CHARS", "2200")) or "2200"),
                        compact_chars_per_skill=int(os.environ.get("SPREADSHEET_SKILL_COMPACT_CHARS_PER_SKILL", os.environ.get("SKILL_INJECTOR_COMPACT_CHARS_PER_SKILL", "900")) or "900"),
                    )
                    if added_artifacts
                    else "(no reusable skill artifacts retrieved)"
                )
                added_callable_prompt = _spreadsheet_callable_update_prompt(
                    [
                        row
                        for row in callable_prompt["skills"]
                        if str(row.get("skill_name") or "") in set(new_injected_names)
                    ],
                    disclosure_mode=str(callable_prompt.get("disclosure_mode") or "full"),
                )
                added_package_prompt = _spreadsheet_package_update_prompt(
                    [
                        row
                        for row in package_prompt["skills"]
                        if str(row.get("skill_name") or "") in set(new_injected_names)
                    ],
                    execution_mode="notebook",
                )
                step_context_msg = _spreadsheet_step_skill_context_message(
                    added_skill_names=new_injected_names,
                    skill_prompt=added_skill_prompt,
                    callable_skill_prompt=added_callable_prompt,
                    package_skill_prompt=added_package_prompt,
                )
                skill_context_messages.append(step_context_msg)
                trace.prompt_context_updates.append(
                    {
                        "turn_index": turn_index,
                        "retrieved_skills": retrieved_names,
                        "new_skills": new_injected_names,
                        "callable_skills": [row.get("skill_name") for row in callable_prompt["skills"]],
                        "package_skills": [row.get("skill_name") for row in package_prompt["skills"]],
                        "message": step_context_msg,
                    }
                )
            skill_prompt_for_system = "\n\n".join(
                str(message.get("content") or "") for message in skill_context_messages
            )
            conversation_messages = build_spreadsheet_notebook_turn_messages(
                base_prompt=base_prompt,
                history=history,
                skill_context_messages=skill_context_messages,
                turn_index=turn_index,
                max_turns=max_turns,
            )
            prompt = render_spreadsheet_notebook_messages(conversation_messages)
            response = await ask_text_llm(
                llm_config=llm_config,
                model_name=model_name,
                system=system,
                prompt=prompt,
                messages=conversation_messages,
                max_request_wall_s=llm_request_timeout_s,
            )
            response_model = response.model_name or model_name
            response_api_style = response.api_style
            trace.input_tokens += response.prompt_tokens
            trace.cache_input_tokens += response.cache_input_tokens
            trace.completion_tokens += response.completion_tokens
            trace.total_tokens += response.prompt_tokens + response.cache_input_tokens + response.completion_tokens
            code = extract_code(response.content)
            done_requested = SPREADSHEET_DONE_PATTERN in (response.content or "")
            turn: Dict[str, Any] = {
                "turn_index": turn_index,
                "response": (response.content or "")[-4000:],
                "code": code,
                "done_requested": done_requested,
                "prompt_tokens": response.prompt_tokens,
                "cache_input_tokens": response.cache_input_tokens,
                "completion_tokens": response.completion_tokens,
                "retrieved_skills": retrieved_names,
                "prompt_injected_skills": list(trace.prompt_injected_skills),
                "new_prompt_injected_skills": new_injected_names,
                "callable_skills": list(callable_prompt["skills"]),
                "package_skills": list(package_prompt["skills"]),
                "skill_context_updated": bool(new_injected_names),
            }
            if code:
                called_this_turn = called_spreadsheet_skill_functions(
                    code,
                    callable_prompt["skills"],
                )
                trace.skill_code_reads.extend(
                    called_spreadsheet_skill_code_reads(code, callable_prompt["skills"])
                )
                exec_result = session.run_cell(code, timeout=CODE_EXEC_TIMEOUT)
                turn.update(exec_result)
                trace.code = (trace.code + "\n\n# %% notebook turn " + str(turn_index) + "\n" + code).strip()
                history.append(
                    {
                        "turn_index": turn_index,
                        "code": code,
                        "response": response.content or "",
                        "stdout": exec_result.get("stdout", ""),
                        "stderr": exec_result.get("stderr", ""),
                        "returncode": exec_result.get("returncode"),
                        "called_skill_functions": called_this_turn,
                    }
                )
                trace.called_skill_functions = list(
                    dict.fromkeys([*trace.called_skill_functions, *called_this_turn])
                )
            else:
                turn.update({"stdout": "", "stderr": "", "returncode": None})
                history.append(
                    {
                        "turn_index": turn_index,
                        "code": "",
                        "response": response.content or "",
                        "stdout": "",
                        "stderr": "No fenced python code was returned.",
                        "returncode": None,
                    }
                )
            trace.notebook_turns.append(turn)
            trace.cost_events.append(
                make_cost_event(
                    role="executor",
                    phase="task_rollout",
                    benchmark="spreadsheet",
                    task_id=task.task_id,
                    model=response.model_name or model_name or "",
                    llm_config=llm_config,
                    input_tokens=response.prompt_tokens,
                    cache_input_tokens=response.cache_input_tokens,
                    output_tokens=response.completion_tokens,
                    prompt_chars=len(prompt),
                    skill_prompt_chars=len(skill_prompt_for_system),
                    system_prompt_chars=len(system),
                    final_conversation_chars=len(system) + len(prompt) + len(response.content or ""),
                    metadata={
                        "api_style": response.api_style,
                    "skill_injector_mode": presentation_mode,
                        "prompt_injected_skills": list(trace.prompt_injected_skills),
                        "callable_skills": list(trace.callable_skills),
                        "package_skills": list(trace.package_skills),
                        "callable_disclosure_mode": callable_prompt.get("disclosure_mode"),
                        "execution_mode": "notebook",
                        "turn_index": turn_index,
                    },
                )
            )
            if done_requested:
                stopped_by_done = True
                break
        trace.stdout = "\n".join(str(turn.get("stdout") or "") for turn in trace.notebook_turns)[-4000:]
        trace.stderr = "\n".join(str(turn.get("stderr") or "") for turn in trace.notebook_turns)[-4000:]
        verify = verify_spreadsheet_output(
            predicted_xlsx=output_path,
            golden_xlsx=Path(task.expected["golden_xlsx"]),
            sheet_name=task.expected.get("answer_sheet"),
            answer_range=task.expected.get("answer_position"),
        )
        last_returncode = trace.notebook_turns[-1].get("returncode") if trace.notebook_turns else None
        verify["returncode"] = last_returncode
        verify["execution_ok"] = not any((turn.get("returncode") not in {0, None}) for turn in trace.notebook_turns)
        success = bool(verify["pass"])
    except Exception as exc:
        trace.elapsed_s = round(time.monotonic() - t0, 3)
        return BenchmarkResult(
            benchmark="spreadsheet",
            task_id=task.task_id,
            success=False,
            score=0.0,
            metrics={"exception": type(exc).__name__, "execution_mode": "notebook"},
            trace=trace.as_dict(),
            error=str(exc),
        )
    finally:
        if session is not None:
            session.close()

    trace.elapsed_s = round(time.monotonic() - t0, 3)
    verify["total_tokens"] = trace.total_tokens
    verify["input_tokens"] = trace.input_tokens
    verify["cache_input_tokens"] = trace.cache_input_tokens
    verify["completion_tokens"] = trace.completion_tokens
    verify["cost_events"] = trace.cost_events
    verify["elapsed_s"] = trace.elapsed_s
    verify["retrieved_skills"] = trace.retrieved_skills
    verify["prompt_injected_skills"] = trace.prompt_injected_skills
    verify["called_skill_functions"] = trace.called_skill_functions
    verify["skill_code_reads"] = trace.skill_code_reads
    verify["package_skills"] = trace.package_skills
    verify["filtered_skills"] = trace.filtered_skills
    verify["injector_events"] = trace.injector_events
    verify["skill_injector_mode"] = presentation_mode
    verify["execution_mode"] = "notebook"
    verify["notebook_turn_count"] = len(trace.notebook_turns)
    verify["notebook_stopped_by_done"] = stopped_by_done
    verify["model_name"] = response_model
    verify["llm_api_style"] = response_api_style
    return BenchmarkResult(
        benchmark="spreadsheet",
        task_id=task.task_id,
        success=success,
        score=float(verify["cell_accuracy"]),
        metrics=verify,
        trace=trace.as_dict(),
    )


async def run_spreadsheet_task_bash_react(
    task: BenchmarkTask,
    *,
    llm_config: str,
    model_name: Optional[str] = None,
    artifact_store: Optional[ArtifactStore] = None,
    top_k_skills: int = 5,
    min_skill_score: float = 0.01,
    skill_injector_mode: str | None = None,
    skill_context_budget_chars: int | None = None,
    callable_disclosure_mode: str | None = None,
    pending_skill_fraction: float = 0.0,
    max_turns: int = 20,
    llm_request_timeout_s: float | None = None,
    work_dir: Optional[Path] = None,
) -> BenchmarkResult:
    t0 = time.monotonic()
    trace = SpreadsheetTrace(task_id=task.task_id, execution_mode="bash_react")
    presentation_mode = _spreadsheet_presentation_mode(
        skill_injector_mode
        or os.environ.get("SPREADSHEET_SKILL_INJECTOR_MODE")
        or os.environ.get("SKILL_INJECTOR_MODE"),
    )
    base_work_dir = work_dir or Path(tempfile.mkdtemp(prefix="spreadsheetbench_bash_"))
    base_work_dir.mkdir(parents=True, exist_ok=True)
    injected: List[Any] = []
    callable_prompt: Dict[str, Any] = {"prompt": "", "skills": [], "disclosure_mode": "full"}
    package_prompt: Dict[str, Any] = {"prompt": "", "skills": [], "disclosure_mode": "progressive"}
    skill_prompt = "(no reusable skill artifacts retrieved)"
    skill_prompt_for_system = skill_prompt
    trace.skill_code_reads = []
    input_copy = base_work_dir / "input.xlsx"
    output_path = base_work_dir / "output.xlsx"
    shutil.copyfile(task.input_artifacts["input_xlsx"], input_copy)
    shutil.copyfile(input_copy, output_path)

    base_prompt = build_spreadsheet_bash_prompt(task, input_copy, output_path, max_turns=max_turns)
    trace.prompt = base_prompt
    system = SPREADSHEET_BASH_SYSTEM.format(
        skills="(runtime skill retrieval updates, if any, are appended in user messages during the bash conversation)",
        callable_skills="(callable and folder-style skill updates, if any, are appended in user messages during the bash conversation)",
        done_pattern=SPREADSHEET_DONE_PATTERN,
    )
    stopped_by_done = False
    try:
        history: List[Dict[str, Any]] = []
        skill_context_messages: List[Dict[str, str]] = []
        response_model = model_name
        response_api_style = ""
        for turn_index in range(max(1, int(max_turns or 20))):
            turn_query = _spreadsheet_notebook_retrieval_query(
                task=task,
                base_prompt=base_prompt,
                history=history,
                turn_index=turn_index,
            )
            retrieved, retrieval_event = _retrieve_spreadsheet_skills(
                artifact_store,
                query=turn_query,
                top_k=top_k_skills,
                min_score=min_skill_score,
                pending_skill_fraction=pending_skill_fraction,
            )
            retrieved_names = [skill.name for skill in retrieved]
            trace.retrieved_skills = list(dict.fromkeys([*trace.retrieved_skills, *retrieved_names]))
            if retrieval_event:
                retrieval_event = {**retrieval_event, "turn_index": turn_index, "phase": "bash_turn_start"}
                trace.injector_events.append(retrieval_event)
            prior_injected_names = {str(getattr(skill, "name", "")) for skill in injected}
            if artifact_store and not _spreadsheet_direct_skill_context_enabled(presentation_mode):
                injection = await select_skill_context_with_llm(
                    retrieved,
                    query=turn_query,
                    llm_config=llm_config,
                    model_name=model_name,
                    presentation_mode=presentation_mode,
                    allowed_injection_types={"informational", "workflow", "functional"},
                    max_selected=top_k_skills,
                    budget_chars=int(skill_context_budget_chars or os.environ.get("SPREADSHEET_SKILL_CONTEXT_BUDGET_CHARS", os.environ.get("SKILL_INJECTOR_BUDGET_CHARS", "2200")) or "2200"),
                    compact_chars_per_skill=int(os.environ.get("SPREADSHEET_SKILL_COMPACT_CHARS_PER_SKILL", os.environ.get("SKILL_INJECTOR_COMPACT_CHARS_PER_SKILL", "900")) or "900"),
                    benchmark="spreadsheet",
                    task_id=task.task_id,
                    phase="executor_step_update",
                    metadata={"turn_index": turn_index, "execution_mode": "bash_react"},
                )
                injected = _merge_spreadsheet_artifacts(injected, injection.artifacts)
                skill_prompt = render_skill_prompt_blocks(
                    injected,
                    mode=presentation_mode,
                    budget_chars=int(skill_context_budget_chars or os.environ.get("SPREADSHEET_SKILL_CONTEXT_BUDGET_CHARS", os.environ.get("SKILL_INJECTOR_BUDGET_CHARS", "2200")) or "2200"),
                    compact_chars_per_skill=int(os.environ.get("SPREADSHEET_SKILL_COMPACT_CHARS_PER_SKILL", os.environ.get("SKILL_INJECTOR_COMPACT_CHARS_PER_SKILL", "900")) or "900"),
                )
                trace.filtered_skills.extend(
                    [{**row, "turn_index": turn_index} for row in list(injection.filtered)]
                )
                injection_event = injection.as_event()
                injection_event["turn_index"] = turn_index
                trace.injector_events.append(injection_event)
                trace.cost_events.append(
                    make_cost_event(
                        role="injector",
                        phase="executor_step_update",
                        benchmark="spreadsheet",
                        task_id=task.task_id,
                        turn_index=turn_index,
                        model=injector_model_name(model_name, benchmark="spreadsheet") or "",
                        llm_config=injector_llm_config(llm_config, benchmark="spreadsheet"),
                        skill_prompt_chars=int(injection_event.get("prompt_chars") or 0),
                        metadata={
                            "mode": presentation_mode,
                            "gate": "llm",
                            "injected_count": injection_event.get("injected_count"),
                            "filtered_count": injection_event.get("filtered_count"),
                            "execution_mode": "bash_react",
                        },
                        )
                    )
            elif artifact_store:
                injected = _merge_spreadsheet_artifacts(injected, retrieved)
                skill_prompt = render_skill_prompt_blocks(
                    injected,
                    mode="full",
                    budget_chars=int(skill_context_budget_chars or os.environ.get("SPREADSHEET_SKILL_CONTEXT_BUDGET_CHARS", os.environ.get("SKILL_INJECTOR_BUDGET_CHARS", "2200")) or "2200"),
                    compact_chars_per_skill=int(os.environ.get("SPREADSHEET_SKILL_COMPACT_CHARS_PER_SKILL", os.environ.get("SKILL_INJECTOR_COMPACT_CHARS_PER_SKILL", "900")) or "900"),
                )
                injection_event = {
                    "mode": presentation_mode,
                    "gate": "direct",
                    "turn_index": turn_index,
                    "injected_count": len(retrieved),
                    "filtered_count": 0,
                    "prompt_chars": len(skill_prompt),
                    "injected": [
                        {
                            "skill_name": skill.name,
                            "decision": "direct_inject",
                            "reason": "preselected_bundle_replay_candidate",
                            "kind": getattr(skill, "kind", ""),
                            "injection_type": skill.injection_type(),
                        }
                        for skill in retrieved
                    ],
                    "filtered": [],
                }
                trace.injector_events.append(injection_event)
            else:
                injected = _merge_spreadsheet_artifacts(injected, retrieved)
                skill_prompt = artifact_store.build_prompt(injected) if artifact_store else "(no reusable skill artifacts retrieved)"
            trace.prompt_injected_skills = [skill.name for skill in injected]
            new_injected_names = [
                str(getattr(skill, "name", ""))
                for skill in injected
                if str(getattr(skill, "name", "")) not in prior_injected_names
            ]
            callable_prompt = write_spreadsheet_skill_library(
                injected,
                base_work_dir,
                disclosure_mode=callable_disclosure_mode or os.environ.get("SPREADSHEET_CALLABLE_DISCLOSURE_MODE") or "full",
            )
            package_prompt = write_spreadsheet_skill_packages(
                injected,
                base_work_dir,
                disclosure_mode=os.environ.get("SPREADSHEET_PACKAGE_DISCLOSURE_MODE") or "progressive",
            )
            trace.callable_skills = _merge_callable_rows(trace.callable_skills, callable_prompt["skills"])
            trace.package_skills = _merge_package_rows(trace.package_skills, package_prompt["skills"])
            if new_injected_names:
                added_artifacts = [
                    skill
                    for skill in injected
                    if str(getattr(skill, "name", "")) in set(new_injected_names)
                ]
                added_skill_prompt = (
                    render_skill_prompt_blocks(
                        added_artifacts,
                        mode=presentation_mode,
                        budget_chars=int(skill_context_budget_chars or os.environ.get("SPREADSHEET_SKILL_CONTEXT_BUDGET_CHARS", os.environ.get("SKILL_INJECTOR_BUDGET_CHARS", "2200")) or "2200"),
                        compact_chars_per_skill=int(os.environ.get("SPREADSHEET_SKILL_COMPACT_CHARS_PER_SKILL", os.environ.get("SKILL_INJECTOR_COMPACT_CHARS_PER_SKILL", "900")) or "900"),
                    )
                    if added_artifacts
                    else "(no reusable skill artifacts retrieved)"
                )
                added_callable_prompt = _spreadsheet_callable_update_prompt(
                    [
                        row
                        for row in callable_prompt["skills"]
                        if str(row.get("skill_name") or "") in set(new_injected_names)
                    ],
                    disclosure_mode=str(callable_prompt.get("disclosure_mode") or "full"),
                    execution_mode="bash_react",
                )
                added_package_prompt = _spreadsheet_package_update_prompt(
                    [
                        row
                        for row in package_prompt["skills"]
                        if str(row.get("skill_name") or "") in set(new_injected_names)
                    ],
                    execution_mode="bash_react",
                )
                step_context_msg = _spreadsheet_step_skill_context_message(
                    added_skill_names=new_injected_names,
                    skill_prompt=added_skill_prompt,
                    callable_skill_prompt=added_callable_prompt,
                    package_skill_prompt=added_package_prompt,
                    execution_mode="bash_react",
                )
                skill_context_messages.append(step_context_msg)
                trace.prompt_context_updates.append(
                    {
                        "turn_index": turn_index,
                        "retrieved_skills": retrieved_names,
                        "new_skills": new_injected_names,
                        "callable_skills": [row.get("skill_name") for row in callable_prompt["skills"]],
                        "package_skills": [row.get("skill_name") for row in package_prompt["skills"]],
                        "message": step_context_msg,
                    }
                )
            skill_prompt_for_system = "\n\n".join(
                str(message.get("content") or "") for message in skill_context_messages
            )
            conversation_messages = build_spreadsheet_bash_turn_messages(
                base_prompt=base_prompt,
                history=history,
                skill_context_messages=skill_context_messages,
                turn_index=turn_index,
                max_turns=max_turns,
            )
            prompt = render_spreadsheet_notebook_messages(conversation_messages)
            response = await ask_text_llm(
                llm_config=llm_config,
                model_name=model_name,
                system=system,
                prompt=prompt,
                messages=conversation_messages,
                max_request_wall_s=llm_request_timeout_s,
            )
            response_model = response.model_name or model_name
            response_api_style = response.api_style
            trace.input_tokens += response.prompt_tokens
            trace.cache_input_tokens += response.cache_input_tokens
            trace.completion_tokens += response.completion_tokens
            trace.total_tokens += response.prompt_tokens + response.cache_input_tokens + response.completion_tokens
            command = extract_bash_command(response.content)
            done_requested = SPREADSHEET_DONE_PATTERN in (response.content or "")
            turn: Dict[str, Any] = {
                "turn_index": turn_index,
                "response": (response.content or "")[-4000:],
                "code": command,
                "command": command,
                "done_requested": done_requested,
                "prompt_tokens": response.prompt_tokens,
                "cache_input_tokens": response.cache_input_tokens,
                "completion_tokens": response.completion_tokens,
                "retrieved_skills": retrieved_names,
                "prompt_injected_skills": list(trace.prompt_injected_skills),
                "new_prompt_injected_skills": new_injected_names,
                "callable_skills": list(callable_prompt["skills"]),
                "package_skills": list(package_prompt["skills"]),
                "skill_context_updated": bool(new_injected_names),
            }
            if command:
                called_this_turn = called_spreadsheet_skill_functions_from_bash(
                    command,
                    callable_prompt["skills"],
                )
                trace.skill_code_reads.extend(
                    called_spreadsheet_skill_code_reads(command, callable_prompt["skills"])
                )
                exec_result = await asyncio.to_thread(
                    run_bash_command,
                    command,
                    base_work_dir,
                    input_copy,
                    output_path,
                    CODE_EXEC_TIMEOUT,
                )
                turn.update(exec_result)
                trace.code = (trace.code + "\n\n# %% bash turn " + str(turn_index) + "\n" + command).strip()
                history.append(
                    {
                        "turn_index": turn_index,
                        "code": command,
                        "command": command,
                        "response": response.content or "",
                        "stdout": exec_result.get("stdout", ""),
                        "stderr": exec_result.get("stderr", ""),
                        "returncode": exec_result.get("returncode"),
                        "called_skill_functions": called_this_turn,
                    }
                )
                trace.called_skill_functions = list(
                    dict.fromkeys([*trace.called_skill_functions, *called_this_turn])
                )
            else:
                turn.update({"stdout": "", "stderr": "No fenced bash command was returned.", "returncode": None})
                history.append(
                    {
                        "turn_index": turn_index,
                        "code": "",
                        "command": "",
                        "response": response.content or "",
                        "stdout": "",
                        "stderr": "No fenced bash command was returned.",
                        "returncode": None,
                    }
                )
            trace.notebook_turns.append(turn)
            trace.cost_events.append(
                make_cost_event(
                    role="executor",
                    phase="task_rollout",
                    benchmark="spreadsheet",
                    task_id=task.task_id,
                    model=response.model_name or model_name or "",
                    llm_config=llm_config,
                    input_tokens=response.prompt_tokens,
                    cache_input_tokens=response.cache_input_tokens,
                    output_tokens=response.completion_tokens,
                    prompt_chars=len(prompt),
                    skill_prompt_chars=len(skill_prompt_for_system),
                    system_prompt_chars=len(system),
                    final_conversation_chars=len(system) + len(prompt) + len(response.content or ""),
                    metadata={
                        "api_style": response.api_style,
                    "skill_injector_mode": presentation_mode,
                        "prompt_injected_skills": list(trace.prompt_injected_skills),
                        "callable_skills": list(trace.callable_skills),
                        "package_skills": list(trace.package_skills),
                        "callable_disclosure_mode": callable_prompt.get("disclosure_mode"),
                        "execution_mode": "bash_react",
                        "turn_index": turn_index,
                    },
                )
            )
            if done_requested:
                stopped_by_done = True
                break
        trace.stdout = "\n".join(str(turn.get("stdout") or "") for turn in trace.notebook_turns)[-4000:]
        trace.stderr = "\n".join(str(turn.get("stderr") or "") for turn in trace.notebook_turns)[-4000:]
        verify = verify_spreadsheet_output(
            predicted_xlsx=output_path,
            golden_xlsx=Path(task.expected["golden_xlsx"]),
            sheet_name=task.expected.get("answer_sheet"),
            answer_range=task.expected.get("answer_position"),
        )
        last_returncode = trace.notebook_turns[-1].get("returncode") if trace.notebook_turns else None
        verify["returncode"] = last_returncode
        verify["execution_ok"] = not any((turn.get("returncode") not in {0, None}) for turn in trace.notebook_turns)
        success = bool(verify["pass"])
    except Exception as exc:
        trace.elapsed_s = round(time.monotonic() - t0, 3)
        return BenchmarkResult(
            benchmark="spreadsheet",
            task_id=task.task_id,
            success=False,
            score=0.0,
            metrics={"exception": type(exc).__name__, "execution_mode": "bash_react"},
            trace=trace.as_dict(),
            error=str(exc),
        )

    trace.elapsed_s = round(time.monotonic() - t0, 3)
    verify["total_tokens"] = trace.total_tokens
    verify["input_tokens"] = trace.input_tokens
    verify["cache_input_tokens"] = trace.cache_input_tokens
    verify["completion_tokens"] = trace.completion_tokens
    verify["cost_events"] = trace.cost_events
    verify["elapsed_s"] = trace.elapsed_s
    verify["retrieved_skills"] = trace.retrieved_skills
    verify["prompt_injected_skills"] = trace.prompt_injected_skills
    verify["called_skill_functions"] = trace.called_skill_functions
    verify["skill_code_reads"] = trace.skill_code_reads
    verify["package_skills"] = trace.package_skills
    verify["filtered_skills"] = trace.filtered_skills
    verify["injector_events"] = trace.injector_events
    verify["skill_injector_mode"] = presentation_mode
    verify["execution_mode"] = "bash_react"
    verify["bash_turn_count"] = len(trace.notebook_turns)
    verify["bash_stopped_by_done"] = stopped_by_done
    verify["model_name"] = response_model
    verify["llm_api_style"] = response_api_style
    return BenchmarkResult(
        benchmark="spreadsheet",
        task_id=task.task_id,
        success=success,
        score=float(verify["cell_accuracy"]),
        metrics=verify,
        trace=trace.as_dict(),
    )


def _retrieve_spreadsheet_skills(
    artifact_store: Optional[ArtifactStore],
    *,
    query: str,
    top_k: int,
    min_score: float = 0.01,
    pending_skill_fraction: float = 0.0,
) -> tuple[List[Any], Dict[str, Any] | None]:
    if not artifact_store or top_k <= 0:
        return [], None
    min_score = max(0.0, float(min_score or 0.0))
    fraction = max(0.0, min(1.0, float(pending_skill_fraction or 0.0)))
    if fraction <= 0.0:
        selected = artifact_store.retrieve(query, top_k=top_k, min_score=min_score)
        return selected, {
            "type": "spreadsheet_retrieval",
            "top_k": top_k,
            "min_score": min_score,
            "selected": [{"name": getattr(artifact, "name", "")} for artifact in selected],
        }
    pending_k = min(top_k, int(round(top_k * fraction)))
    if pending_k <= 0:
        pending_k = 1
    active_k = max(0, top_k - pending_k)
    active_audit = (
        artifact_store.retrieve_audit(
            query,
            top_k=active_k,
            min_score=min_score,
            predicate=lambda artifact: str(getattr(artifact, "status", "")) != "pending"
            and not bool((getattr(artifact, "metadata", {}) or {}).get("is_pending_skill")),
        )
        if active_k
        else {"selected": []}
    )
    pending_audit = artifact_store.retrieve_audit(
        query,
        top_k=pending_k,
        min_score=min_score,
        predicate=lambda artifact: str(getattr(artifact, "status", "")) == "pending"
        or bool((getattr(artifact, "metadata", {}) or {}).get("is_pending_skill")),
        include_pending=True,
    )
    active = [
        artifact_store.get(str(item.get("name") or ""))
        for item in active_audit.get("selected", [])
    ]
    pending = [
        artifact_store.get(str(item.get("name") or ""))
        for item in pending_audit.get("selected", [])
    ]
    active = [artifact for artifact in active if artifact is not None]
    pending = [artifact for artifact in pending if artifact is not None]
    merged: List[Any] = []
    seen: set[str] = set()
    for artifact in [*active, *pending]:
        name = str(getattr(artifact, "name", ""))
        if name and name not in seen:
            merged.append(artifact)
            seen.add(name)
    return merged[:top_k], {
        "type": "spreadsheet_pending_mixed_retrieval",
        "top_k": top_k,
        "min_score": min_score,
        "pending_skill_fraction": fraction,
        "active_k": active_k,
        "pending_k": pending_k,
        "active_selected": active_audit.get("selected", []),
        "pending_selected": pending_audit.get("selected", []),
    }


def _spreadsheet_presentation_mode(mode: str | None) -> str:
    value = str(mode or "").strip().lower()
    if value:
        return value
    return "full"


def _spreadsheet_direct_skill_context_enabled(mode: str | None) -> bool:
    return str(mode or "").strip().lower() in {
        "direct",
        "direct_full",
        "preselected",
        "no_llm",
        "none",
        "off",
    }


def _spreadsheet_notebook_retrieval_query(
    *,
    task: BenchmarkTask,
    base_prompt: str,
    history: Sequence[Dict[str, Any]],
    turn_index: int,
) -> str:
    parts = [
        "SpreadsheetBench notebook step retrieval query.",
        f"Task id: {task.task_id}",
        f"Turn index: {turn_index}",
        f"Instruction: {task.question}",
        f"Answer sheet: {task.expected.get('answer_sheet')}",
        f"Answer range: {task.expected.get('answer_position')}",
        f"Data position: {task.metadata.get('data_position')}",
    ]
    if not history:
        parts.append("Trace prefix: no notebook code has executed yet.")
        parts.append("Initial workbook/task preview:")
        parts.append(clip_notebook_text(base_prompt, 2400))
        return "\n".join(parts)

    last = history[-1]
    stderr = str(last.get("stderr") or "")
    stdout = str(last.get("stdout") or "")
    returncode = last.get("returncode")
    if stderr or returncode not in {0, None}:
        parts.append(
            "Previous step had an execution error; retrieve repair, API-usage, or workbook-shape skills "
            "that directly address the traceback and the current task."
        )
    else:
        parts.append(
            "Previous step executed; retrieve skills relevant to the inspected workbook state and the next solve step."
        )
    parts.extend(
        [
            "Latest executed code:",
            clip_notebook_text(last.get("code") or "", 1800),
            f"Latest return code: {returncode}",
            "Latest stdout:",
            clip_notebook_text(stdout, 1400),
            "Latest stderr:",
            clip_notebook_text(stderr, 1800),
        ]
    )
    if len(history) > 1:
        brief = []
        for row in history[-4:-1]:
            brief.append(
                f"turn={int(row.get('turn_index', 0))} returncode={row.get('returncode')} "
                f"called={row.get('called_skill_functions') or []}"
            )
        parts.append("Earlier trace summary: " + "; ".join(brief))
    return "\n".join(parts)


def _merge_spreadsheet_artifacts(current: Sequence[Any], retrieved: Sequence[Any]) -> List[Any]:
    merged: List[Any] = []
    seen: set[str] = set()
    for artifact in [*list(current or []), *list(retrieved or [])]:
        name = str(getattr(artifact, "name", "") or "")
        if not name or name in seen:
            continue
        seen.add(name)
        merged.append(artifact)
    return merged


def _merge_callable_rows(
    current_rows: Sequence[Dict[str, Any]],
    new_rows: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    merged: List[Dict[str, Any]] = []
    seen: set[str] = set()
    for row in [*list(current_rows or []), *list(new_rows or [])]:
        key = str(row.get("skill_name") or row.get("function_name") or "")
        if not key or key in seen:
            continue
        seen.add(key)
        merged.append(dict(row))
    return merged


def _merge_package_rows(
    current_rows: Sequence[Dict[str, Any]],
    new_rows: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    merged: List[Dict[str, Any]] = []
    seen: set[str] = set()
    for row in [*list(current_rows or []), *list(new_rows or [])]:
        key = str(row.get("skill_name") or row.get("package_name") or "")
        if not key or key in seen:
            continue
        seen.add(key)
        merged.append(dict(row))
    return merged


def build_spreadsheet_prompt(task: BenchmarkTask, input_copy: Path, output_path: Path) -> str:
    preview = workbook_preview(input_copy)
    return (
        f"Instruction:\n{task.question}\n\n"
        f"Input workbook path: {input_copy}\n"
        f"Output workbook path: {output_path}\n"
        f"Answer sheet: {task.expected.get('answer_sheet')}\n"
        f"Answer range: {task.expected.get('answer_position')}\n"
        f"Data position: {task.metadata.get('data_position')}\n\n"
        f"Workbook preview:\n{preview}\n\n"
        "Write robust openpyxl code. It may inspect workbook sheets/cells, then save to OUTPUT_XLSX."
    )


def build_spreadsheet_notebook_prompt(
    task: BenchmarkTask,
    input_copy: Path,
    output_path: Path,
    *,
    max_turns: int,
) -> str:
    preview = workbook_preview(input_copy)
    return (
        f"Instruction:\n{task.question}\n\n"
        f"Input workbook path: {input_copy}\n"
        f"Output workbook path: {output_path}\n"
        f"Answer sheet: {task.expected.get('answer_sheet')}\n"
        f"Answer range: {task.expected.get('answer_position')}\n"
        f"Data position: {task.metadata.get('data_position')}\n\n"
        f"Workbook preview:\n{preview}\n\n"
        "Notebook protocol:\n"
        f"- You have at most {max_turns} turns.\n"
        "- Each fenced python block executes in the same process and can reuse previous variables.\n"
        "- Use print(...) to inspect workbook state when needed.\n"
        "- If execution fails, use the stderr in the next turn to repair the code.\n"
        f"- Save the final answer workbook to OUTPUT_XLSX, then output {SPREADSHEET_DONE_PATTERN}.\n"
    )


def build_spreadsheet_bash_prompt(
    task: BenchmarkTask,
    input_copy: Path,
    output_path: Path,
    *,
    max_turns: int,
) -> str:
    preview = workbook_preview(input_copy)
    return (
        f"Instruction:\n{task.question}\n\n"
        f"Working directory input file: {input_copy.name}\n"
        f"Working directory output file: {output_path.name}\n"
        f"Absolute input workbook path: {input_copy}\n"
        f"Absolute output workbook path: {output_path}\n"
        f"Answer sheet: {task.expected.get('answer_sheet')}\n"
        f"Answer range: {task.expected.get('answer_position')}\n"
        f"Data position: {task.metadata.get('data_position')}\n\n"
        f"Workbook preview:\n{preview}\n\n"
        "Bash ReAct protocol:\n"
        f"- You have at most {max_turns} turns.\n"
        "- Return exactly one fenced bash block per turn.\n"
        "- Commands run in the same task directory; files persist across turns.\n"
        "- Python variables do not persist across turns, so rerun imports or save scripts/files as needed.\n"
        "- Use python/openpyxl from bash to inspect and edit the workbook.\n"
        f"- Save the final answer workbook to output.xlsx, then output {SPREADSHEET_DONE_PATTERN}.\n"
    )


def build_spreadsheet_notebook_turn_prompt(
    *,
    base_prompt: str,
    history: Sequence[Dict[str, Any]],
    skill_context_messages: Sequence[Dict[str, str]] | None = None,
    turn_index: int,
    max_turns: int,
) -> str:
    return render_spreadsheet_notebook_messages(
        build_spreadsheet_notebook_turn_messages(
            base_prompt=base_prompt,
            history=history,
            skill_context_messages=skill_context_messages,
            turn_index=turn_index,
            max_turns=max_turns,
        )
    )


def build_spreadsheet_notebook_turn_messages(
    *,
    base_prompt: str,
    history: Sequence[Dict[str, Any]],
    skill_context_messages: Sequence[Dict[str, str]] | None = None,
    turn_index: int,
    max_turns: int,
) -> List[Dict[str, str]]:
    messages: List[Dict[str, str]] = []
    first_turn_instruction = (
        "This is turn 1. Return a fenced python code block to inspect or solve the workbook. "
        f"When the workbook is saved and final, include {SPREADSHEET_DONE_PATTERN}."
    )
    if not history:
        messages.append({"role": "user", "content": base_prompt + "\n" + first_turn_instruction})
    else:
        messages.append({"role": "user", "content": base_prompt + "\n" + first_turn_instruction})
        for row in history[-4:]:
            response = str(row.get("response") or "").strip()
            if response:
                assistant_content = clip_notebook_text(response, 2600)
            else:
                assistant_content = "\n".join(
                    [
                        "```python",
                        clip_notebook_text(row.get("code") or "", 1800),
                        "```",
                    ]
                )
            messages.append({"role": "assistant", "content": assistant_content})
            messages.append(
                {
                    "role": "user",
                    "content": "\n".join(
                        [
                            f"Execution observation for notebook turn {int(row.get('turn_index', 0)) + 1}:",
                            f"Return code: {row.get('returncode')}",
                            f"stdout:\n{clip_notebook_text(row.get('stdout') or '', 1200)}",
                            f"stderr:\n{clip_notebook_text(row.get('stderr') or '', 1600)}",
                        ]
                    ),
                }
            )
    for message in skill_context_messages or []:
        messages.append(
            {
                "role": "user",
                "content": str(message.get("content") or ""),
            }
        )
    if history:
        messages.append(
            {
                "role": "user",
                "content": (
                    f"This is turn {turn_index + 1} of {max_turns}. "
                    "Return the next fenced python code block. "
                    f"If the workbook is already correct and saved, include {SPREADSHEET_DONE_PATTERN}."
                ),
            }
        )
    return messages


def build_spreadsheet_bash_turn_messages(
    *,
    base_prompt: str,
    history: Sequence[Dict[str, Any]],
    skill_context_messages: Sequence[Dict[str, str]] | None = None,
    turn_index: int,
    max_turns: int,
) -> List[Dict[str, str]]:
    messages: List[Dict[str, str]] = []
    first_turn_instruction = (
        "This is turn 1. Return a fenced bash code block to inspect or solve the workbook. "
        f"When output.xlsx is saved and final, include {SPREADSHEET_DONE_PATTERN}."
    )
    if not history:
        messages.append({"role": "user", "content": base_prompt + "\n" + first_turn_instruction})
    else:
        messages.append({"role": "user", "content": base_prompt + "\n" + first_turn_instruction})
        for row in history[-4:]:
            response = str(row.get("response") or "").strip()
            if response:
                assistant_content = clip_notebook_text(response, 2600)
            else:
                assistant_content = "\n".join(
                    [
                        "```bash",
                        clip_notebook_text(row.get("command") or row.get("code") or "", 1800),
                        "```",
                    ]
                )
            messages.append({"role": "assistant", "content": assistant_content})
            messages.append(
                {
                    "role": "user",
                    "content": "\n".join(
                        [
                            f"Execution observation for bash turn {int(row.get('turn_index', 0)) + 1}:",
                            f"Return code: {row.get('returncode')}",
                            f"stdout:\n{clip_notebook_text(row.get('stdout') or '', 1200)}",
                            f"stderr:\n{clip_notebook_text(row.get('stderr') or '', 1600)}",
                        ]
                    ),
                }
            )
    for message in skill_context_messages or []:
        messages.append({"role": "user", "content": str(message.get("content") or "")})
    if history:
        messages.append(
            {
                "role": "user",
                "content": (
                    f"This is turn {turn_index + 1} of {max_turns}. "
                    "Return the next fenced bash command. "
                    f"If output.xlsx is already correct and saved, include {SPREADSHEET_DONE_PATTERN}."
                ),
            }
        )
    return messages


def render_spreadsheet_notebook_messages(messages: Sequence[Dict[str, str]]) -> str:
    chunks: List[str] = []
    for message in messages:
        chunks.append(
            "\n".join(
                [
                    f"{str(message.get('role') or 'user').upper()} MESSAGE:",
                    str(message.get("content") or ""),
                ]
            )
        )
    return "\n\n".join(chunks)


def _render_spreadsheet_skill_context_messages(messages: Sequence[Dict[str, str]]) -> str:
    if not messages:
        return ""
    chunks = ["\nRuntime skill retrieval updates already appended to this notebook conversation:"]
    for idx, message in enumerate(messages, start=1):
        chunks.append(
            "\n".join(
                [
                    f"User message {idx}:",
                    str(message.get("content") or ""),
                ]
            )
        )
    return "\n\n".join(chunks) + "\n\n"


def _spreadsheet_step_skill_context_message(
    *,
    added_skill_names: Sequence[str],
    skill_prompt: str,
    callable_skill_prompt: str,
    package_skill_prompt: str = "",
    execution_mode: str = "notebook",
    ) -> Dict[str, str]:
    names = ", ".join(str(name) for name in added_skill_names)
    if execution_mode == "bash_react":
        mode_name = "bash task"
        inspect_instruction = (
            "and calling it from `skill_library.py` or running its `skills/<function_name>.py` wrapper; "
            "read `skills.md` first if deciding whether a skill applies, then inspect `skill_library.py` or "
            "`skills/<function_name>.py` before adapting or rewriting.\n\n"
        )
        protocol_header = (
            "[NEW SKILLS]\n"
            "source: runtime skill injector after the previous bash turn\n"
            "status: actionable candidates for the next bash action\n"
            "decision_required: before your next bash command, decide for each listed skill as USE_NOW, USE_LATER, or SKIP\n"
            "decision_rules:\n"
            "- USE_NOW: the skill directly solves the next operation; import it or run its wrapper instead of rewriting that operation.\n"
            "- USE_LATER: the skill directly solves a later sub-operation; name that exact sub-operation and use the skill when you reach it.\n"
            "- SKIP: the skill is irrelevant or unsafe for this workbook, answer range, or user intent.\n\n"
        )
        protocol_footer = "\n[/NEW SKILLS]"
    else:
        mode_name = "notebook task"
        inspect_instruction = (
            "and calling it from `skill_library`; if unsure, inspect its code object before adapting or rewriting.\n\n"
        )
        protocol_header = (
            "[NEW SKILLS]\n"
            "source: runtime skill injector after the previous notebook turn\n"
            "status: actionable candidates for the next notebook cell\n"
            "decision_required: before your next code cell, decide for each listed skill as USE_NOW, USE_LATER, or SKIP\n"
            "decision_rules:\n"
            "- USE_NOW: the skill directly solves the next operation; import it instead of rewriting that operation.\n"
            "- USE_LATER: the skill directly solves a later sub-operation; name that exact sub-operation and use the skill when you reach it.\n"
            "- SKIP: the skill is irrelevant or unsafe for this workbook, answer range, or user intent.\n\n"
        )
        protocol_footer = "\n[/NEW SKILLS]"
    callable_block = (
        "\n\nNew callable Spreadsheet functions:\n" + callable_skill_prompt
        if callable_skill_prompt
        else ""
    )
    package_block = (
        "\n\nNew folder-style Spreadsheet skill packages:\n" + package_skill_prompt
        if package_skill_prompt
        else ""
    )
    return {
        "role": "user",
        "content": (
            f"{protocol_header}"
            f"Runtime skill retrieval update for this same Spreadsheet {mode_name}.\n"
            f"Newly retrieved local rules/functions: {names}.\n"
            "Use these only if they directly match the workbook, answer range, and user intent; "
            "ignore them if they are irrelevant. If a callable function matches, prefer importing "
            f"{inspect_instruction}"
            f"{skill_prompt}"
            f"{callable_block}"
            f"{package_block}"
            f"{protocol_footer}"
        ),
    }


def _spreadsheet_callable_update_prompt(
    rows: Sequence[Dict[str, Any]],
    *,
    disclosure_mode: str,
    execution_mode: str = "notebook",
) -> str:
    if not rows:
        return ""
    disclosure = str(disclosure_mode or "full").strip().lower()
    is_bash = execution_mode == "bash_react"
    if is_bash:
        direct_rows = "\n\n".join(
            "\n".join(
                [
                    "```bash",
                    "python - <<'PY'",
                    "import os",
                    f"from skill_library import {row['function_name']}",
                    f"{row['function_name']}(os.environ['INPUT_XLSX'], os.environ['OUTPUT_XLSX'])",
                    "PY",
                    "```",
                ]
            )
            for row in rows
        )
        inspect_rows = "\n\n".join(
            "\n".join(
                [
                    "```bash",
                    "sed -n '1,220p' skills.md",
                    f"sed -n '1,220p' {row.get('script_path') or 'skills/' + row['function_name'] + '.py'}",
                    "sed -n '1,260p' skill_library.py",
                    "python - <<'PY'",
                    f"from skill_library import {row['function_name']}_skill",
                    f"print({row['function_name']}_skill.code)",
                    "PY",
                    "```",
                ]
            )
            for row in rows
        )
        signature_rows = "\n".join(
            f"- `{row['signature']}` from skill `{row['skill_name']}`: {str(row.get('description') or '')[:220]} "
            f"Files: `{row.get('manifest_path') or 'skills.md'}`, `{row.get('library_path') or 'skill_library.py'}`, `{row.get('script_path') or ''}`."
            for row in rows
        )
        script_rows = "\n\n".join(
            "\n".join(
                [
                    "```bash",
                    f"python {row.get('script_path') or 'skills/' + row['function_name'] + '.py'} \"$INPUT_XLSX\" \"$OUTPUT_XLSX\"",
                    "```",
                ]
            )
            for row in rows
        )
        if disclosure in {"progressive", "signature", "signature_only"}:
            return (
                "Executable Spreadsheet skills are available as readable/writable files in the current task directory. "
                "`skills.md` is the local skill manifest; read it when deciding whether a skill applies. "
                "`skill_library.py` exports the functions and `skills/*.py` contains runnable wrappers. "
                "Prefer calling a matching function directly, or run its wrapper, inspect the files, or edit/adapt them before use.\n"
                "Direct bash call examples:\n"
                + direct_rows
                + "\nScript wrapper examples:\n"
                + script_rows
                + "\n"
                "Available callable signatures:\n"
                + signature_rows
                + "\nInspect implementation examples:\n"
                + inspect_rows
            )
        full_rows = "\n".join(
            "\n".join(
                [
                    f"- `{row['signature']}` from skill `{row['skill_name']}`: {str(row.get('description') or '')[:220]}",
                    "  Full implementation:",
                    "  ```python",
                    "\n".join("  " + line for line in str(row.get("code") or "").splitlines()),
                    "  ```",
                ]
            )
            for row in rows
        )
        return (
            "Executable Spreadsheet skills are available as readable/writable files in the current task directory. "
            "`skills.md` is the local skill manifest; `skill_library.py` exports the functions and `skills/*.py` contains runnable wrappers. "
            "Prefer calling or running a matching skill from bash over rewriting its implementation.\n"
            "Direct bash call examples:\n"
            + direct_rows
            + "\n"
            + full_rows
        )
    if disclosure in {"progressive", "signature", "signature_only"}:
        signature_rows = "\n".join(
            f"- `{row['signature']}` from skill `{row['skill_name']}`: {str(row.get('description') or '')[:220]} "
            f"Files: `{row.get('manifest_path') or 'skills.md'}`, `{row.get('library_path') or 'skill_library.py'}`, `{row.get('script_path') or ''}`."
            for row in rows
        )
        direct_rows = "\n\n".join(
            "\n".join(
                [
                    "```python",
                    f"from skill_library import {row['function_name']}",
                    f"{row['function_name']}(INPUT_XLSX, OUTPUT_XLSX)",
                    "```",
                ]
            )
            for row in rows
        )
        object_rows = "\n".join(
            f"- `{row['function_name']}_skill.code`, `skill_library.py`, and `{row.get('script_path') or ''}` expose "
            f"the full implementation for `{row['function_name']}`."
            for row in rows
        )
        return (
            "Executable Spreadsheet skills are available through readable/writable task files: "
            "`skills.md` lists local skill cards, `skill_library.py` exports the functions, and `skills/*.py` provides wrappers. "
            "Prefer direct reuse when the signature matches, and inspect the implementation only if needed.\n"
            "Direct call examples:\n"
            + direct_rows
            + "\n"
            "Available callable signatures:\n"
            + signature_rows
            + "\nInspectable skill code objects:\n"
            + object_rows
        )
    direct_rows = "\n\n".join(
        "\n".join(
            [
                "```python",
                f"from skill_library import {row['function_name']}",
                f"{row['function_name']}(INPUT_XLSX, OUTPUT_XLSX)",
                "```",
            ]
        )
        for row in rows
    )
    return (
        "Executable Spreadsheet skills are already importable from `skill_library`. "
        "Prefer direct reuse in returned code when the signature matches. Direct call examples:\n"
        + direct_rows
        + "\n"
        + "\n".join(
            "\n".join(
                [
                    f"- `{row['signature']}` from skill `{row['skill_name']}`: {str(row.get('description') or '')[:220]}",
                    "  Full implementation:",
                    "  ```python",
                    "\n".join("  " + line for line in str(row.get("code") or "").splitlines()),
                    "  ```",
                ]
            )
            for row in rows
        )
    )


def _spreadsheet_package_update_prompt(
    rows: Sequence[Dict[str, Any]],
    *,
    execution_mode: str = "notebook",
) -> str:
    if not rows:
        return ""
    row_text = "\n".join(
        "\n".join(
            [
                f"- `{row['skill_name']}` folder: `{row['skill_dir']}`",
                f"  Entry manifest: `{row['skill_md_path']}`",
                f"  Use when: {str(row.get('description') or '')[:260]}",
                "  Runnable scripts: "
                + (", ".join(f"`{path}`" for path in row.get("script_paths") or []) or "(no scripts listed)"),
                "  Reference notes: "
                + (", ".join(f"`{path}`" for path in row.get("reference_paths") or []) or "(no references listed)"),
            ]
        )
        for row in rows
    )
    if execution_mode == "bash_react":
        return (
            "Folder-style skills are local directories with `SKILL.md`, optional `scripts/`, and optional `references/`. "
            "Read `SKILL.md` first; if it matches, inspect/run/import a script instead of rewriting the whole operation.\n"
            "Inspection/run examples:\n"
            "```bash\n"
            "find skills/<skill_name> -maxdepth 3 -type f -print\n"
            "sed -n '1,220p' skills/<skill_name>/SKILL.md\n"
            "sed -n '1,240p' skills/<skill_name>/scripts/<script>.py\n"
            "python skills/<skill_name>/scripts/<script>.py \"$INPUT_XLSX\" \"$OUTPUT_XLSX\"\n"
            "```\n"
            + row_text
        )
    return (
        "Folder-style skills are local directories with `SKILL.md`, optional `scripts/`, and optional `references/`. "
        "Read `SKILL.md` before importing or adapting package scripts.\n"
        + row_text
    )


def extract_bash_command(text: str) -> str:
    if not text:
        return ""
    preferred = re.search(r"```(?:bash|sh|shell)\s*\n?", text, re.I)
    if preferred:
        closing = text.rfind("```")
        if closing > preferred.end():
            return text[preferred.end():closing].strip()
    python_block = re.search(r"```python\s*(.*?)```", text, re.S | re.I)
    if python_block:
        return "python - <<'PY'\n" + python_block.group(1).strip() + "\nPY"
    generic = re.search(r"```\s*(.*?)```", text, re.S)
    if generic:
        return generic.group(1).strip()
    python = extract_code(text)
    if python:
        return "python - <<'PY'\n" + python.strip() + "\nPY"
    return ""


def run_bash_command(
    command: str,
    work_dir: Path,
    input_xlsx: Path,
    output_xlsx: Path,
    timeout: float | int | None,
) -> Dict[str, Any]:
    env = os.environ.copy()
    env["INPUT_XLSX"] = str(input_xlsx)
    env["OUTPUT_XLSX"] = str(output_xlsx)
    try:
        proc = subprocess.run(
            ["bash", "-lc", command],
            cwd=str(work_dir),
            env=env,
            text=True,
            capture_output=True,
            timeout=float(timeout or CODE_EXEC_TIMEOUT),
        )
        return {
            "stdout": proc.stdout[-4000:],
            "stderr": proc.stderr[-4000:],
            "returncode": proc.returncode,
        }
    except subprocess.TimeoutExpired as exc:
        stdout = exc.stdout.decode() if isinstance(exc.stdout, bytes) else str(exc.stdout or "")
        stderr = exc.stderr.decode() if isinstance(exc.stderr, bytes) else str(exc.stderr or "")
        return {
            "stdout": stdout[-4000:],
            "stderr": (stderr + f"\nCommand timed out after {timeout or CODE_EXEC_TIMEOUT} seconds.")[-4000:],
            "returncode": 124,
        }


def called_spreadsheet_skill_functions_from_bash(
    command: str,
    callable_rows: Sequence[Dict[str, Any]],
) -> List[str]:
    called: List[str] = []
    for row in callable_rows or []:
        function_name = str(row.get("function_name") or "")
        skill_name = str(row.get("skill_name") or function_name)
        if not function_name:
            continue
        if re.search(rf"\b{re.escape(function_name)}\s*\(", str(command or "")):
            called.append(skill_name)
    return list(dict.fromkeys(called))


def clip_notebook_text(value: Any, limit: int) -> str:
    text = str(value or "")
    if len(text) <= limit:
        return text
    return text[-limit:]


def workbook_preview(path: Path, max_rows: int = 8, max_cols: int = 8) -> str:
    wb = openpyxl.load_workbook(path, data_only=False)
    parts = []
    for sheet in wb.sheetnames[:5]:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, max_rows),
            min_col=1,
            max_col=min(ws.max_column, max_cols),
            values_only=True,
        ):
            rows.append([jsonable_cell_value(value) for value in row])
        parts.append(f"## {sheet} ({ws.max_row}x{ws.max_column})\n{json.dumps(rows, ensure_ascii=False)}")
    return "\n\n".join(parts)
