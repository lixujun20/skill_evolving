"""SpreadsheetBench maintenance adapter and compatibility exports."""
from __future__ import annotations

import copy
import ast
import json
import os
import re
import tempfile
import time
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from academic.benchmarks.core.artifacts import ArtifactStore
from academic.benchmarks.core.bundle_cases import apply_credit_bundle_suggestions
from academic.benchmarks.core.bundle_policy import default_bundle_case_priority, trim_bundle_cases_to_budget
from academic.benchmarks.core.credit_events import (
    apply_credit_evidence,
    is_strong_harmful_credit,
    normalize_credit_events,
)
from academic.benchmarks.core.credit_scope import (
    credit_candidate_skill_names,
    retrieved_only_skill_names,
    skill_exposure_flags,
)
from academic.benchmarks.core.maintenance_adapter import MaintenanceRunConfig, NoOpMaintenanceAdapter
from academic.benchmarks.core.maintenance_utils import now_iso, stable_id
from academic.benchmarks.core.micro_maintenance import MicroMaintenanceHooks, run_generic_micro_maintenance
from academic.benchmarks.core.types import (
    BenchmarkResult,
    BenchmarkTask,
    SkillArtifact,
    SkillBundleCase,
    SkillEvidence,
    SkillInterface,
    SkillLineage,
    SkillTestCaseRun,
    SkillTestResult,
)
import academic.benchmarks.spreadsheet.executor as _spreadsheet_executor
from academic.benchmarks.core.llm_text import ask_text_llm
from academic.benchmarks.spreadsheet.executor import (
    build_spreadsheet_notebook_prompt as _build_spreadsheet_notebook_prompt,
    build_spreadsheet_notebook_turn_prompt as _build_spreadsheet_notebook_turn_prompt,
    build_spreadsheet_prompt as _build_spreadsheet_prompt,
    clip_notebook_text as _clip_notebook_text,
    run_spreadsheet_task_bash_react as _run_spreadsheet_task_bash_react_impl,
    run_spreadsheet_task as _run_spreadsheet_task_impl,
    run_spreadsheet_task_notebook as _run_spreadsheet_task_notebook_impl,
    workbook_preview as _workbook_preview,
)
from academic.benchmarks.spreadsheet.loader import ensure_spreadsheetbench, load_spreadsheet_tasks
from academic.benchmarks.spreadsheet.models import SpreadsheetTrace
from academic.benchmarks.spreadsheet.prompts import (
    SPREADSHEET_CREDIT_SYSTEM,
    SPREADSHEET_DONE_PATTERN,
    SPREADSHEET_EXTRACT_SYSTEM,
)
from academic.benchmarks.spreadsheet.skill_runtime import (
    NotebookPythonSession as _NotebookPythonSession,
    called_spreadsheet_skill_functions as _called_spreadsheet_skill_functions,
    called_spreadsheet_skill_functions_from_text as _called_spreadsheet_skill_functions_from_text,
    extract_code as _extract_code,
    is_spreadsheet_callable_skill as _is_spreadsheet_callable_skill,
    looks_like_spreadsheet_python as _looks_like_spreadsheet_python,
    render_spreadsheet_skill_function as _render_spreadsheet_skill_function,
    run_code as _run_code,
    spreadsheet_callable_description as _spreadsheet_callable_description,
    spreadsheet_callable_kwargs as _spreadsheet_callable_kwargs,
    spreadsheet_column_value as _spreadsheet_column_value,
    spreadsheet_skill_code as _spreadsheet_skill_code,
    spreadsheet_skill_function_name as _spreadsheet_skill_function_name,
    wrap_spreadsheet_skill_snippet as _wrap_spreadsheet_skill_snippet,
    write_spreadsheet_skill_library as _write_spreadsheet_skill_library,
    is_spreadsheet_package_skill as _is_spreadsheet_package_skill,
    safe_package_relative_path as _safe_package_relative_path,
    spreadsheet_bundle_files as _spreadsheet_bundle_files,
    spreadsheet_package_dir_name as _spreadsheet_package_dir_name,
    spreadsheet_package_files as _spreadsheet_package_files,
    write_spreadsheet_skill_packages as _write_spreadsheet_skill_packages,
)
from academic.benchmarks.spreadsheet.trace_projection import (
    compact_spreadsheet_skill_card as _compact_spreadsheet_skill_card,
    spreadsheet_code_snippet as _spreadsheet_code_snippet,
    spreadsheet_result_projection as _spreadsheet_result_projection,
    spreadsheet_skill_projection as _spreadsheet_skill_projection,
    spreadsheet_task_fragment as _spreadsheet_task_fragment,
    spreadsheet_trace_projection as _spreadsheet_trace_projection,
)
from academic.benchmarks.spreadsheet.verifier import (
    answer_range_refs as _answer_range_refs,
    cells_in_range as _cells_in_range,
    first_sheet_name as _first_sheet_name,
    jsonable_cell_value as _jsonable_cell_value,
    normalize_answer_range_text as _normalize_answer_range_text,
    normalize_cell_value as _normalize_cell_value,
    split_answer_range_list as _split_answer_range_list,
    split_sheet_range as _split_sheet_range,
    verify_spreadsheet_output,
)
from academic.skill_repository.llm_maintenance import (
    _ask_json,
    _extractor_rule_suffix,
    _record_maintenance_token_event,
    _refiner_rule_suffix,
    _role_json_block,
    apply_refine_payload,
    normalize_skill_name,
    refine_skill_artifact_llm,
    summarize_dependency_context,
    update_role_rules_from_feedback_llm,
)


_SPREADSHEET_EXTRACT_PRESERVE_KEYS = {
    "answer_position",
    "answer_sheet",
    "golden_xlsx",
    "input_artifacts",
    "input_xlsx",
    "package_files",
    "bundle_files",
    "prompt_txt_preview",
    "spreadsheet_path",
}


def _spreadsheet_maintenance_llm_config(default: str) -> str:
    for key in ("SPREADSHEET_MAINTENANCE_LLM_CONFIG", "SPREADSHEET_MAINTENANCE_MODEL_LLM_CONFIG"):
        value = str(os.environ.get(key) or "").strip()
        if value:
            return value
    return default


def _spreadsheet_maintenance_model_name(default: str | None) -> str | None:
    for key in ("SPREADSHEET_MAINTENANCE_MODEL_NAME", "SPREADSHEET_MAINTENANCE_INJECTOR_MODEL_NAME"):
        value = str(os.environ.get(key) or "").strip()
        if value:
            return value
    return default


def _spreadsheet_role_feedback(config: MaintenanceRunConfig) -> Dict[str, Any]:
    payload = config.extra.setdefault("role_feedback", {})
    if not isinstance(payload, dict):
        payload = {}
        config.extra["role_feedback"] = payload
    return payload


def _spreadsheet_role_rules(config: MaintenanceRunConfig, role_name: str) -> List[Dict[str, Any]]:
    role_feedback = _spreadsheet_role_feedback(config)
    role_payload = role_feedback.get(role_name) or {}
    if not isinstance(role_payload, dict):
        return []
    return [dict(item) for item in (role_payload.get("rules") or []) if isinstance(item, dict)]


def _compat_module() -> Any:
    import academic.benchmarks.spreadsheet.adapter as facade

    return facade


async def _compat_ask_json(**kwargs: Any) -> Dict[str, Any]:
    if "llm_config" in kwargs:
        kwargs["llm_config"] = _spreadsheet_maintenance_llm_config(str(kwargs["llm_config"]))
    if "model_name" in kwargs:
        kwargs["model_name"] = _spreadsheet_maintenance_model_name(kwargs.get("model_name"))
    return await _compat_module()._ask_json(**kwargs)


async def _compat_refine_skill_artifact_llm(*args: Any, **kwargs: Any) -> Dict[str, Any]:
    if "llm_config" in kwargs:
        kwargs["llm_config"] = _spreadsheet_maintenance_llm_config(str(kwargs["llm_config"]))
    if "model_name" in kwargs:
        kwargs["model_name"] = _spreadsheet_maintenance_model_name(kwargs.get("model_name"))
    return await _compat_module().refine_skill_artifact_llm(*args, **kwargs)


async def run_spreadsheet_task(*args: Any, **kwargs: Any) -> BenchmarkResult:
    _spreadsheet_executor.ask_text_llm = ask_text_llm
    return await _run_spreadsheet_task_impl(*args, **kwargs)


async def run_spreadsheet_task_notebook(*args: Any, **kwargs: Any) -> BenchmarkResult:
    _spreadsheet_executor.ask_text_llm = ask_text_llm
    return await _run_spreadsheet_task_notebook_impl(*args, **kwargs)


async def run_spreadsheet_task_bash_react(*args: Any, **kwargs: Any) -> BenchmarkResult:
    _spreadsheet_executor.ask_text_llm = ask_text_llm
    return await _run_spreadsheet_task_bash_react_impl(*args, **kwargs)


async def _run_spreadsheet_task_for_config(
    task: BenchmarkTask,
    *,
    config: MaintenanceRunConfig,
    store: ArtifactStore,
    top_k_skills: int,
) -> BenchmarkResult:
    execution_mode = str(config.extra.get("spreadsheet_execution_mode") or "single").strip().lower()
    common_kwargs = {
        "llm_config": config.llm_config,
        "model_name": config.model_name,
        "artifact_store": store,
        "top_k_skills": top_k_skills,
        "min_skill_score": float(config.extra.get("min_skill_score", 0.01)),
        "skill_injector_mode": config.extra.get("skill_injector_mode"),
        "skill_context_budget_chars": config.extra.get("skill_context_budget_chars"),
        "callable_disclosure_mode": config.extra.get("spreadsheet_callable_disclosure_mode"),
        "pending_skill_fraction": float(config.extra.get("spreadsheet_pending_skill_fraction") or 0.0),
        "llm_request_timeout_s": config.extra.get("llm_request_timeout_s") or config.max_task_seconds,
    }
    if execution_mode == "notebook":
        return await _compat_module().run_spreadsheet_task_notebook(
            task,
            **common_kwargs,
            max_turns=int(config.extra.get("spreadsheet_max_turns") or 5),
        )
    if execution_mode in {"bash_react", "bash", "cli"}:
        return await _compat_module().run_spreadsheet_task_bash_react(
            task,
            **common_kwargs,
            max_turns=int(config.extra.get("spreadsheet_max_turns") or 20),
        )
    return await _compat_module().run_spreadsheet_task(task, **common_kwargs)


class SpreadsheetMaintenanceAdapter(NoOpMaintenanceAdapter):
    """SpreadsheetBench adapter for the benchmark-agnostic evolution runner.

    Spreadsheet evolution is benchmark-native: successful traces can extract
    openpyxl/workflow skills, retrieved skills receive per-task credit, credit
    produces focused replayable SpreadsheetBench bundle cases, micro
    maintenance refines or filters implicated skills, and macro maintenance
    promotes/deduplicates/filters repository candidates at the window level.
    """

    benchmark = "spreadsheet"

    async def run_task(
        self,
        task: BenchmarkTask,
        *,
        store: ArtifactStore,
        config: MaintenanceRunConfig,
        phase: str,
        task_index: int,
        run_idx: int,
    ) -> BenchmarkResult:
        del phase, task_index, run_idx
        return await _run_spreadsheet_task_for_config(
            task,
            config=config,
            store=store,
            top_k_skills=config.top_k_skills,
        )

    async def assign_credit(
        self,
        *,
        detail: Dict[str, Any],
        store: ArtifactStore,
        config: MaintenanceRunConfig,
        round_index: int,
        task_index: int,
    ) -> List[Dict[str, Any]]:
        del round_index
        projection = _spreadsheet_trace_projection(detail)
        candidate_names = _spreadsheet_credit_candidate_names(projection)
        candidate_artifacts = [
            artifact for name in candidate_names for artifact in [store.get(str(name))] if artifact
        ]
        if not candidate_artifacts:
            return []
        try:
            payload = await _compat_ask_json(
                system=SPREADSHEET_CREDIT_SYSTEM,
                user=_role_json_block(
                    {
                        "task": _spreadsheet_task_fragment(detail),
                        "result": {
                            key: projection.get(key)
                            for key in (
                                "success",
                                "score",
                                "answer_sheet",
                                "answer_position",
                                "checked_cells",
                                "mismatched_cells",
                                "execution_ok",
                                "stderr_tail",
                                "notebook_step_count",
                                "notebook_steps",
                            )
                            if projection.get(key) not in (None, "", [], {})
                        },
                        "skill_exposure": {
                            "retrieved": projection.get("retrieved_skills") or [],
                            "prompt_injected": projection.get("prompt_injected_skills") or [],
                            "called": projection.get("called_skill_functions") or [],
                            "candidate_policy": "prompt_injected_or_called_plus_retrieved_only_relevance_gate",
                        },
                        "retrieval_audit": {
                            "retrieved_only_skills": projection.get("retrieved_only_skills") or [],
                        },
                        "candidate_skills": [
                            _spreadsheet_skill_projection(artifact, projection=projection)
                            for artifact in candidate_artifacts
                        ],
                    }
                ),
                llm_config=config.llm_config,
                model_name=config.model_name,
                role="spreadsheet_credit_assigner",
                metadata={
                    "phase": "spreadsheet_credit",
                    "benchmark": "spreadsheet",
                    "task_id": detail.get("task_id"),
                    "task_index": task_index,
                },
            )
            events = _normalize_spreadsheet_credit_events(
                payload,
                detail=detail,
                candidate_artifacts=candidate_artifacts,
                projection=projection,
            )
        except Exception as exc:
            events = _heuristic_spreadsheet_credit_events(
                detail=detail,
                candidate_artifacts=candidate_artifacts,
                projection=projection,
                reason=f"credit_llm_failed:{type(exc).__name__}",
            )
        apply_credit_evidence(store=store, credit_events=events)
        for event in events:
            artifact = store.get(event["skill_name"])
            if artifact is None:
                continue
            if event.get("judgment") == "helpful":
                artifact.success_count += 1
            if not _spreadsheet_credit_event_is_retrieved_only(event):
                artifact.usage_count += 1
        return events

    async def apply_credit_bundle_cases(
        self,
        *,
        detail: Dict[str, Any],
        credit_events: Sequence[Dict[str, Any]],
        store: ArtifactStore,
        config: MaintenanceRunConfig,
        round_index: int,
        task_index: int,
    ) -> List[Dict[str, Any]]:
        del config, round_index, task_index

        def build_case(case_detail: Dict[str, Any], event: Dict[str, Any], suggestion: Dict[str, Any]) -> SkillBundleCase | None:
            artifact = store.get(str(suggestion.get("skill_name") or event.get("skill_name") or ""))
            if artifact is None:
                return None
            return _spreadsheet_case_from_credit_suggestion(
                detail=case_detail,
                artifact=artifact,
                event=event,
                suggestion=dict(suggestion or {}),
            )

        rows = apply_credit_bundle_suggestions(
            store=store,
            detail=detail,
            credit_events=credit_events,
            build_case=build_case,
            trim_cases=True,
        )
        created = []
        for row in rows:
            if not row.get("created"):
                continue
            artifact = store.get(str(row.get("skill_name") or ""))
            if artifact is not None:
                artifact.bundle.bundle_version += 1
            created.append(
                {
                    "skill_name": row.get("skill_name"),
                    "case_id": row.get("case_id"),
                    "polarity": row.get("polarity"),
                    "source_task_id": detail.get("task_id"),
                    "reason": row.get("reason"),
                }
            )
        return created

    async def run_micro_maintenance(self, **kwargs: Any) -> Dict[str, Any]:
        detail = kwargs.get("detail") or {}
        credit_events = list(kwargs.get("credit_events") or [])
        credit_bundle_cases = list(kwargs.get("credit_bundle_cases") or [])
        store: ArtifactStore = kwargs["store"]
        config: MaintenanceRunConfig = kwargs["config"]
        task_index = int(kwargs.get("task_index") or 0)
        extracted = await _extract_spreadsheet_skills_from_detail(
            detail,
            store=store,
            config=config,
            task_index=task_index,
        )
        print(
            json.dumps(
                {
                    "progress": "spreadsheet_micro_extraction_done",
                    "task_id": detail.get("task_id"),
                    "task_index": task_index,
                    "n_extracted": len(extracted),
                },
                ensure_ascii=False,
            ),
            flush=True,
        )
        extraction_reports: List[Dict[str, Any]] = []
        accepted_extracted: List[SkillArtifact] = []
        for artifact in extracted:
            existing_artifact = store.get(artifact.name)
            if existing_artifact is not None and (
                existing_artifact.status == "pending"
                or bool(existing_artifact.metadata.get("is_pending_skill"))
                or existing_artifact.status == "active"
            ):
                extraction_reports.append(
                    {
                        "skill_name": artifact.name,
                        "status": "skipped_existing",
                        "description": artifact.description,
                        "source_task_ids": artifact.metadata.get("source_task_ids") or [],
                    }
                )
                continue
            print(
                json.dumps(
                    {
                        "progress": "spreadsheet_prestore_gate_start",
                        "task_id": detail.get("task_id"),
                        "task_index": task_index,
                        "skill_name": artifact.name,
                        "skill_kind": artifact.kind,
                    },
                    ensure_ascii=False,
                ),
                flush=True,
            )
            prestore_gate = await _run_spreadsheet_prestore_bundle_gate(
                artifact=artifact,
                store=store,
                config=config,
            )
            print(
                json.dumps(
                    {
                        "progress": "spreadsheet_prestore_gate_done",
                        "task_id": detail.get("task_id"),
                        "task_index": task_index,
                        "skill_name": artifact.name,
                        "passed": bool(prestore_gate.get("passed")),
                        "reason": prestore_gate.get("rejection_reason") or "",
                    },
                    ensure_ascii=False,
                ),
                flush=True,
            )
            artifact = prestore_gate["artifact"]
            prestore_result = prestore_gate.get("final_result")
            if not prestore_gate.get("passed"):
                extraction_reports.append(
                    {
                        "skill_name": artifact.name,
                        "status": "rejected",
                        "description": artifact.description,
                        "source_task_ids": artifact.metadata.get("source_task_ids") or [],
                        "rejection_reason": prestore_gate.get("rejection_reason") or "prestore_bundle_test_failed",
                        "prestore_test_result": prestore_result.as_dict() if prestore_result is not None else {},
                        "prestore_test_results": copy.deepcopy(prestore_gate.get("test_results") or []),
                        "prestore_refine_decisions": copy.deepcopy(prestore_gate.get("refine_decisions") or []),
                    }
                )
                continue
            store.add_pending(artifact)
            if prestore_result is not None:
                store.add_test_result(prestore_result)
            accepted_extracted.append(artifact)
            extraction_reports.append(
                {
                    "skill_name": artifact.name,
                    "status": artifact.status,
                    "description": artifact.description,
                    "source_task_ids": artifact.metadata.get("source_task_ids") or [],
                    "prestore_test_result": prestore_result.as_dict() if prestore_result is not None else {},
                    "prestore_test_results": copy.deepcopy(prestore_gate.get("test_results") or []),
                    "prestore_refine_decisions": copy.deepcopy(prestore_gate.get("refine_decisions") or []),
                }
            )
        async def refine_skill(**hook_kwargs: Any) -> Dict[str, Any]:
            artifact = hook_kwargs["artifact"]
            stage = str(hook_kwargs.get("stage") or "")
            if stage == "post_bundle_failure" and hook_kwargs.get("failed_bundle_result") is not None:
                raw_result = hook_kwargs["failed_bundle_result"]
                test_result = raw_result if isinstance(raw_result, SkillTestResult) else _spreadsheet_test_result_from_dict(raw_result)
                decision = await _refine_spreadsheet_skill_from_bundle(
                    artifact=artifact,
                    test_result=test_result,
                    credit_context=hook_kwargs.get("credit_events") or [],
                    store=store,
                    config=config,
                )
            else:
                decision = await _refine_spreadsheet_skill_from_credit(
                    artifact=artifact,
                    credit_context=hook_kwargs.get("credit_events") or [],
                    detail=detail,
                    store=store,
                    config=config,
                )
            if decision.get("updated_artifact"):
                store.add(decision["updated_artifact"])
            return decision

        async def run_bundle_test(**hook_kwargs: Any) -> Dict[str, Any]:
            artifact = hook_kwargs["artifact"]
            if not artifact.bundle.all_cases() and not _spreadsheet_package_has_bundle_tests(artifact):
                return {
                    "skill_name": artifact.name,
                    "passed": True,
                    "aggregate": {"passed": True, "n_cases": 0, "reason": "no_bundle_cases"},
                }
            result = await _compat_module()._execute_spreadsheet_bundle_tests(
                artifact=artifact,
                config=config,
            )
            store.add_test_result(result)
            return result.as_dict()

        micro_target_reasons = _spreadsheet_micro_target_reasons(
            credit_events=credit_events,
            credit_bundle_cases=credit_bundle_cases,
            extracted=accepted_extracted,
        )
        maintenance_credit_events = _spreadsheet_micro_credit_events(credit_events)
        maintenance_credit_bundle_cases = _spreadsheet_micro_credit_bundle_cases(credit_bundle_cases)
        report = await run_generic_micro_maintenance(
            detail=detail,
            credit_events=maintenance_credit_events,
            credit_bundle_cases=maintenance_credit_bundle_cases,
            store=store,
            config=config,
            hooks=MicroMaintenanceHooks(refine_skill=refine_skill, run_bundle_test=run_bundle_test),
            round_index=int(kwargs.get("round_index") or 0),
            task_index=task_index,
            relevant_skill_names=list(micro_target_reasons),
        )
        report["refine_decisions"] = [
            {k: v for k, v in item.items() if k != "updated_artifact"}
            for item in report.get("refine_decisions", [])
        ]
        report["extraction_reports"] = extraction_reports
        report["reason"] = "spreadsheet_micro_maintenance"
        report["trace_projection"] = _spreadsheet_trace_projection(detail)
        report["micro_target_reasons"] = micro_target_reasons
        report["micro_maintenance_credit_events"] = copy.deepcopy(maintenance_credit_events)
        report["micro_maintenance_credit_bundle_cases"] = copy.deepcopy(maintenance_credit_bundle_cases)
        report["credit_bundle_cases"] = copy.deepcopy(credit_bundle_cases)
        return report

    def maintenance_lock_names(
        self,
        *,
        detail: Dict[str, Any],
        store: ArtifactStore,
        config: MaintenanceRunConfig,
    ) -> List[str]:
        del store, config
        projection = _spreadsheet_trace_projection(detail)
        return _spreadsheet_credit_candidate_names(projection)

    async def run_macro_maintenance(
        self,
        *,
        window_details: Sequence[Dict[str, Any]],
        all_train_details: Sequence[Dict[str, Any]],
        credit_events: Sequence[Dict[str, Any]],
        store: ArtifactStore,
        config: MaintenanceRunConfig,
        round_index: int,
        window_index: int,
        final_window: bool = False,
    ) -> Dict[str, Any]:
        promotions = _promote_spreadsheet_pending_from_window(store, window_details)
        dedupe = _dedupe_spreadsheet_skills(store)
        filtered = _filter_spreadsheet_harmful_skills(store, credit_events)
        trl_feedback = await _update_spreadsheet_role_feedback_from_window(
            window_details=window_details,
            all_train_details=all_train_details,
            credit_events=credit_events,
            store=store,
            config=config,
            round_index=round_index,
            window_index=window_index,
            final_window=final_window,
            promotions=promotions,
            dedupe=dedupe,
            filtered=filtered,
        )
        store.refresh_all_dependencies()
        return {
            "phase": "macro_final" if final_window else "macro",
            "window_index": window_index,
            "task_ids": [str(item.get("task_id") or "") for item in window_details],
            "maintenance_targets": sorted(set(promotions + dedupe + filtered)),
            "maintenance_test_results": [],
            "refine_decisions": [],
            "overlap_refactor": {
                "benchmark": "spreadsheet",
                "attempts": [],
                "promoted_pending_skills": promotions,
                "deduplicated_skills": dedupe,
                "filtered_skills": filtered,
                "trl_feedback": trl_feedback,
                "window_trace_segments": [
                    _spreadsheet_trace_projection(item) for item in window_details
                ],
                "all_train_count": len(all_train_details),
            },
            "trl_feedback": trl_feedback,
            "run_overlap_refactor": True,
            "reason": "spreadsheet_macro_promote_dedupe_filter_trl",
        }

    def store_snapshot(self, store: ArtifactStore) -> Dict[str, Any]:
        artifacts = store.all()
        return {
            "n_skills": len(artifacts),
            "n_active": sum(1 for item in artifacts if item.status == "active"),
            "n_pending": sum(1 for item in artifacts if item.status == "pending"),
            "n_disabled": sum(1 for item in artifacts if item.is_disabled()),
            "n_archived": sum(1 for item in artifacts if item.status == "archived"),
            "skill_names": [artifact.name for artifact in artifacts],
            "skill_versions": {artifact.name: artifact.version for artifact in artifacts},
            "skill_status": {artifact.name: artifact.status for artifact in artifacts},
            "bundle_case_counts": {
                artifact.name: len(artifact.bundle.all_cases()) for artifact in artifacts
            },
        }


def _spreadsheet_credit_candidate_names(projection: Dict[str, Any]) -> List[str]:
    """Credit Spreadsheet retrieved-only skills through a strict relevance gate."""

    ordered: List[str] = []
    seen: set[str] = set()
    for name in [
        *credit_candidate_skill_names(projection),
        *retrieved_only_skill_names(projection),
    ]:
        value = str(name or "").strip()
        if value and value not in seen:
            ordered.append(value)
            seen.add(value)
    return ordered


def _spreadsheet_credit_event_is_retrieved_only(event: Dict[str, Any]) -> bool:
    evidence = dict(event.get("evidence") or {})
    if bool(evidence.get("retrieved_only")):
        return True
    exposure = dict(event.get("projection") or {})
    return str(event.get("skill_name") or "") in set(exposure.get("retrieved_only_skills") or [])


def _spreadsheet_extract_limit(name: str, default: int) -> int:
    try:
        return max(1, int(os.environ.get(f"SPREADSHEET_EXTRACT_{name}", str(default)) or str(default)))
    except Exception:
        return default


def _spreadsheet_extract_limits() -> Dict[str, int]:
    return {
        "max_body_words": _spreadsheet_extract_limit("MAX_BODY_WORDS", 220),
        "max_body_lines": _spreadsheet_extract_limit("MAX_BODY_LINES", 60),
        "max_code_lines": _spreadsheet_extract_limit("MAX_CODE_LINES", 35),
        "max_artifacts": _spreadsheet_extract_limit("MAX_ARTIFACTS", 6),
        "max_rewrite_rounds": _spreadsheet_extract_limit("MAX_REWRITE_ROUNDS", 2),
    }


def _word_count(text: Any) -> int:
    return len(re.findall(r"\b\w+\b", str(text or "")))


def _nonempty_lines(text: Any) -> List[str]:
    return [line for line in str(text or "").splitlines() if line.strip()]


def _spreadsheet_code_blocks(text: str) -> List[str]:
    return [match.group(1).strip() for match in re.finditer(r"```(?:python)?\s*(.*?)```", str(text or ""), re.S | re.I)]


def _spreadsheet_artifact_rubric_failures(raw: Dict[str, Any], *, limits: Dict[str, int]) -> List[str]:
    failures: List[str] = []
    body = _spreadsheet_body_from_structured_raw(raw)
    description = str(raw.get("description") or "")
    kind = str(raw.get("kind") or "").strip().lower()
    metadata = dict(raw.get("metadata") or {})
    if _word_count(description) > 40:
        failures.append(f"description has {_word_count(description)} words > 40")
    body_words = _word_count(body)
    body_lines = len(_nonempty_lines(body))
    if body_words > limits["max_body_words"]:
        failures.append(f"body has {body_words} words > {limits['max_body_words']}")
    if body_lines > limits["max_body_lines"]:
        failures.append(f"body has {body_lines} non-empty lines > {limits['max_body_lines']}")
    code_blocks = _spreadsheet_code_blocks(body)
    for idx, code in enumerate(code_blocks):
        code_lines = len(_nonempty_lines(code))
        if code_lines > limits["max_code_lines"]:
            failures.append(f"code block {idx} has {code_lines} non-empty lines > {limits['max_code_lines']}")
    if kind in {"executable_tool", "function_tool", "script_tool"}:
        probe = SkillArtifact(
            name=normalize_skill_name(str(raw.get("name") or "spreadsheet_candidate")),
            kind=kind,
            description=description,
            body=body,
        )
        if not _is_spreadsheet_callable_skill(probe):
            failures.append("executable_tool does not contain parseable reusable Python code")
    if kind == "skill_package":
        package_files = metadata.get("package_files")
        bundle_files = metadata.get("bundle_files")
        if not isinstance(package_files, dict):
            failures.append("skill_package metadata.package_files must be an object")
        else:
            if "SKILL.md" not in package_files:
                failures.append("skill_package metadata.package_files must include SKILL.md")
            for path, content in package_files.items():
                rel_path = _safe_package_relative_path(path)
                if rel_path is None:
                    failures.append(f"invalid package file path: {path}")
                    continue
                text = str(content or "")
                if str(rel_path) == "SKILL.md" and _word_count(text) > 180:
                    failures.append(f"SKILL.md has {_word_count(text)} words > 180")
                if str(rel_path).endswith(".py") and len(_nonempty_lines(text)) > 80:
                    failures.append(f"package script {rel_path} has {len(_nonempty_lines(text))} non-empty lines > 80")
        if bundle_files is not None and not isinstance(bundle_files, dict):
            failures.append("skill_package metadata.bundle_files must be an object when present")
        elif isinstance(bundle_files, dict):
            for path, content in bundle_files.items():
                rel_path = _safe_package_relative_path(path)
                if rel_path is None:
                    failures.append(f"invalid bundle file path: {path}")
                    continue
                if str(rel_path).endswith(".py") and len(_nonempty_lines(str(content or ""))) > 80:
                    failures.append(f"bundle test {rel_path} has {len(_nonempty_lines(str(content or '')))} non-empty lines > 80")
    return failures


def _spreadsheet_extraction_rubric_failures(
    raw_artifacts: Sequence[Dict[str, Any]],
    *,
    limits: Dict[str, int],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for index, raw in enumerate(raw_artifacts):
        failures = _spreadsheet_artifact_rubric_failures(dict(raw or {}), limits=limits)
        if failures:
            rows.append(
                {
                    "index": index,
                    "name": str((raw or {}).get("name") or ""),
                    "failures": failures,
                }
            )
    return rows


def _spreadsheet_extraction_user_payload(
    *,
    existing: Sequence[Dict[str, Any]],
    detail: Dict[str, Any],
    limits: Dict[str, int],
    skill_format: str = "auto",
    previous_artifacts: Sequence[Dict[str, Any]] | None = None,
    rubric_failures: Sequence[Dict[str, Any]] | None = None,
) -> Dict[str, Any]:
    result = _spreadsheet_result_projection(detail)
    success = bool(result.get("success")) and float(result.get("score") or 0.0) >= 0.9
    edit_steps = _spreadsheet_reusable_edit_steps_from_detail(detail) or _spreadsheet_reusable_edit_steps(result)
    if success and edit_steps:
        trace = dict(result.get("trace") or {})
        trace.pop("code_snippet", None)
        trace["notebook_steps"] = [
            {
                "stdout_tail": str((step or {}).get("stdout_tail") or "")[-300:],
                "stderr_tail": str((step or {}).get("stderr_tail") or "")[-200:],
                "exception": str((step or {}).get("exception") or ""),
            }
            for step in list(trace.get("notebook_steps") or [])[-3:]
        ]
        result = {**result, "trace": trace}
    payload: Dict[str, Any] = {
        "existing_artifacts": list(existing)[-40:],
        "result": result,
        "extraction_policy": {
            "success_trace": success,
            "skill_format": _spreadsheet_skill_format(skill_format),
            "preferred_for_success": (
                "If reusable_edit_steps is non-empty, extract at least one narrow, generalized "
                "skill from the reusable edit operation unless an existing artifact already covers it. "
                "Parameterize workbook-specific sheets, ranges, headers, keywords, thresholds, and markers."
            ),
            "bash_react_note": (
                "The trace may contain bash heredocs. Treat bash as an execution wrapper and extract "
                "the Python/openpyxl logic inside the heredoc into a callable snippet that uses "
                "INPUT_XLSX and OUTPUT_XLSX."
            ),
            "skill_package_note": (
                "For bash_react traces, prefer skill_package when a later executor should read a concise "
                "SKILL.md and run/import or copy/adapt a script from skills/<name>/scripts/. Put exact files in "
                "metadata.package_files and tests in metadata.bundle_files. SKILL.md should state when to use, "
                "when not to use, configurable inputs, workbook assumptions, and copy/adapt guidance."
            ),
            "generalization_policy": (
                "Extract reusable operations, not task answers. Good candidates include keyword/regex marking, "
                "header-based column move/copy with style preservation, cross-sheet key matching, row insertion "
                "around detected markers, and date/threshold filtering. Bad candidates hardcode one workbook's "
                "literal constants without parameters or applicability limits."
            ),
            "format_constraint": _spreadsheet_skill_format_instruction(skill_format),
            "failed_trace_policy": (
                "For failed or partial-score traces, extract only narrow repair/contract skills "
                "grounded in verifier mismatches or stderr."
            ),
        },
        "reusable_edit_steps": edit_steps,
        "rubric": {
            "max_body_words": limits["max_body_words"],
            "max_body_nonempty_lines": limits["max_body_lines"],
            "max_code_nonempty_lines_per_block": limits["max_code_lines"],
            "instruction": (
                "Every artifact must satisfy the limits. If a candidate is too broad or too long, "
                "split it into multiple narrower skills or compress it to the reusable contract. "
                "Do not output oversized artifacts. Executable code blocks must be complete; "
                "never truncate code mid-block."
            ),
        },
    }
    if previous_artifacts is not None:
        payload["previous_artifacts"] = list(previous_artifacts)
        payload["rubric_failures"] = list(rubric_failures or [])
        payload["rewrite_instruction"] = (
            "Rewrite the extraction output so every artifact passes the rubric. "
            "For each oversized executable code block, either replace it with a shorter parameterized helper "
            "or split the behavior into narrower skills. Drop candidates that cannot be made concise and reusable."
        )
    return payload


def _spreadsheet_skill_format(value: Any) -> str:
    raw = str(value or "auto").strip().lower()
    return raw if raw in {"auto", "function", "folder"} else "auto"


def _spreadsheet_skill_format_instruction(value: Any) -> str:
    fmt = _spreadsheet_skill_format(value)
    if fmt == "function":
        return "Return only executable_tool/function_tool/script_tool artifacts. Do not return skill_package artifacts."
    if fmt == "folder":
        return "Return only skill_package artifacts with metadata.package_files and metadata.bundle_files. Do not return executable_tool artifacts."
    return "Either executable_tool or skill_package is allowed; choose the narrower reusable format."


def _spreadsheet_extraction_role_json_block(value: Any) -> str:
    return _role_json_block(value, preserve_keys=_SPREADSHEET_EXTRACT_PRESERVE_KEYS)


def _spreadsheet_reusable_edit_steps(result: Dict[str, Any], *, max_steps: int = 3) -> List[Dict[str, Any]]:
    steps = list(((result.get("trace") or {}).get("notebook_steps") or []))
    rows: List[Dict[str, Any]] = []
    for idx, step in enumerate(steps):
        code = str((step or {}).get("code_snippet") or "")
        if not code.strip():
            continue
        python = _spreadsheet_python_from_bash_heredoc(code) or code
        if not _looks_like_spreadsheet_python(python):
            continue
        lowered = python.lower()
        is_edit = (
            "save(" in lowered
            or "insert_rows(" in lowered
            or "delete_rows(" in lowered
            or "insert_cols(" in lowered
            or "delete_cols(" in lowered
            or re.search(r"(?m)^\s*ws(?:heet)?\s*(?:\[|\.cell\()", python)
            or re.search(r"(?m)^\s*[^#\n]*(?:\.value|\[[^\]]+\])\s*=", python)
        )
        if not is_edit:
            continue
        rows.append(
            {
                "turn_index": idx,
                "python_code": _normalize_spreadsheet_skill_source(python),
                "stdout_tail": str((step or {}).get("stdout_tail") or "")[-500:],
                "stderr_tail": str((step or {}).get("stderr_tail") or "")[-300:],
            }
        )
    return rows[-max_steps:]


def _spreadsheet_reusable_edit_steps_from_detail(detail: Dict[str, Any], *, max_steps: int = 3) -> List[Dict[str, Any]]:
    runs = detail.get("runs") or []
    run = runs[0] if runs else {}
    trace = run.get("trace") or {}
    synthetic = {
        "trace": {
            "notebook_steps": [
                {
                    "code_snippet": str(turn.get("code") or ""),
                    "stdout_tail": str(turn.get("stdout") or "")[-800:],
                    "stderr_tail": str(turn.get("stderr") or "")[-800:],
                }
                for turn in trace.get("notebook_turns") or []
                if isinstance(turn, dict)
            ]
        }
    }
    return _spreadsheet_reusable_edit_steps(synthetic, max_steps=max_steps)


def _spreadsheet_python_from_bash_heredoc(command: str) -> str:
    text = str(command or "")
    match = re.search(r"python(?:3)?(?:\s+-)?\s+<<\s*['\"]?([A-Za-z_][A-Za-z0-9_]*)['\"]?\s*\n(.*?)\n\s*\1\b", text, re.S)
    if match:
        return match.group(2).strip()
    return ""


def _normalize_spreadsheet_skill_source(code: str, *, max_chars: int = 2600) -> str:
    text = str(code or "").strip()
    text = re.sub(r"openpyxl\.load_workbook\(\s*(['\"])(?:/tmp/[^'\"]*?|[^'\"]*?_input\.xlsx|[^'\"]*?/input\.xlsx)\1\s*\)", "openpyxl.load_workbook(INPUT_XLSX)", text)
    text = re.sub(r"load_workbook\(\s*(['\"])(?:/tmp/[^'\"]*?|[^'\"]*?_input\.xlsx|[^'\"]*?/input\.xlsx)\1\s*\)", "load_workbook(INPUT_XLSX)", text)
    text = re.sub(r"wb\.save\(\s*(['\"])(?:/tmp/[^'\"]*?|[^'\"]*?_output\.xlsx|[^'\"]*?/output\.xlsx)\1\s*\)", "wb.save(OUTPUT_XLSX)", text)
    text = re.sub(r"openpyxl\.load_workbook\(\s*os\.environ\[['\"]INPUT_XLSX['\"]\]\s*\)", "openpyxl.load_workbook(INPUT_XLSX)", text)
    text = re.sub(r"openpyxl\.load_workbook\(\s*input_path\s*\)", "openpyxl.load_workbook(INPUT_XLSX)", text)
    text = re.sub(r"load_workbook\(\s*input_path\s*\)", "load_workbook(INPUT_XLSX)", text)
    text = re.sub(r"wb\.save\(\s*os\.environ\[['\"]OUTPUT_XLSX['\"]\]\s*\)", "wb.save(OUTPUT_XLSX)", text)
    text = re.sub(r"wb\.save\(\s*output_path\s*\)", "wb.save(OUTPUT_XLSX)", text)
    text = re.sub(r"(?m)^\s*(input_path|input_file)\s*=\s*os\.environ(?:\.get)?\([^)]*INPUT_XLSX[^)]*\)\s*$", "", text)
    text = re.sub(r"(?m)^\s*output_path\s*=\s*os\.environ(?:\.get)?\([^)]*OUTPUT_XLSX[^)]*\)\s*$", "", text)
    text = re.sub(r"(?m)^\s*input_path\s*=\s*INPUT_XLSX\s*$", "", text)
    text = re.sub(r"(?m)^\s*output_path\s*=\s*OUTPUT_XLSX\s*$", "", text)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    if len(text) > max_chars:
        compact = _spreadsheet_strip_nonessential_python(text)
        if compact and len(compact) <= max_chars:
            text = compact
        else:
            text = _spreadsheet_code_snippet(text, limit=max_chars)
    return text


class _SpreadsheetEssentialCodeTransformer(ast.NodeTransformer):
    def visit_Module(self, node: ast.Module) -> ast.Module:
        node = self.generic_visit(node)
        used_names = {
            child.id
            for child in ast.walk(node)
            if isinstance(child, ast.Name) and isinstance(child.ctx, ast.Load)
        }
        body: List[ast.stmt] = []
        for stmt in node.body:
            if isinstance(stmt, ast.ImportFrom):
                stmt.names = [alias for alias in stmt.names if (alias.asname or alias.name) in used_names]
                if not stmt.names:
                    continue
            elif isinstance(stmt, ast.Import):
                stmt.names = [alias for alias in stmt.names if (alias.asname or alias.name.split(".")[0]) in used_names]
                if not stmt.names:
                    continue
            body.append(stmt)
        node.body = body
        return node

    def visit_Expr(self, node: ast.Expr) -> ast.AST | None:
        if (
            isinstance(node.value, ast.Call)
            and isinstance(node.value.func, ast.Name)
            and node.value.func.id == "print"
        ):
            return None
        return self.generic_visit(node)

    def visit_Assign(self, node: ast.Assign) -> ast.AST | None:
        if (
            len(node.targets) == 1
            and isinstance(node.targets[0], ast.Name)
            and node.targets[0].id in {"input_path", "input_file", "output_path"}
        ):
            return None
        return self.generic_visit(node)

    def visit_Import(self, node: ast.Import) -> ast.AST | None:
        node.names = [alias for alias in node.names if alias.name != "os"]
        return node if node.names else None

    def visit_If(self, node: ast.If) -> ast.AST | None:
        node = self.generic_visit(node)
        if isinstance(node, ast.If):
            node.body = node.body or [ast.Pass()]
            node.orelse = node.orelse or []
        return node

    def visit_For(self, node: ast.For) -> ast.AST | None:
        node = self.generic_visit(node)
        if isinstance(node, ast.For):
            node.body = node.body or [ast.Pass()]
            node.orelse = node.orelse or []
        return node


def _compress_style_copy_idiom(text: str) -> str:
    pattern = (
        r"(?P<indent>[ \t]*)if data\['font'\]:\n"
        r"(?P=indent)    cell\.font = data\['font'\]\n"
        r"(?P=indent)if data\['fill'\]:\n"
        r"(?P=indent)    cell\.fill = data\['fill'\]\n"
        r"(?P=indent)if data\['border'\]:\n"
        r"(?P=indent)    cell\.border = data\['border'\]\n"
        r"(?P=indent)if data\['alignment'\]:\n"
        r"(?P=indent)    cell\.alignment = data\['alignment'\]"
    )

    def repl(match: re.Match[str]) -> str:
        indent = match.group("indent")
        return (
            f"{indent}for attr in ['font', 'fill', 'border', 'alignment']:\n"
            f"{indent}    if data[attr]:\n"
            f"{indent}        setattr(cell, attr, data[attr])"
        )

    return re.sub(pattern, repl, text)


def _contains_spreadsheet_write_node(node: ast.AST) -> bool:
    for child in ast.walk(node):
        if isinstance(child, ast.Call):
            if isinstance(child.func, ast.Attribute) and child.func.attr in {
                "save",
                "insert_rows",
                "delete_rows",
                "insert_cols",
                "delete_cols",
                "append",
                "merge_cells",
                "unmerge_cells",
            }:
                return True
        if isinstance(child, (ast.Assign, ast.AugAssign, ast.AnnAssign)):
            targets: List[ast.AST] = []
            if isinstance(child, ast.Assign):
                targets = list(child.targets)
            elif isinstance(child, ast.AugAssign):
                targets = [child.target]
            elif isinstance(child, ast.AnnAssign):
                targets = [child.target]
            if any(isinstance(target, (ast.Subscript, ast.Attribute)) for target in targets):
                return True
    return False


def _drop_trailing_verification_nodes(tree: ast.Module) -> ast.Module:
    save_index = -1
    for idx, node in enumerate(tree.body):
        if "save(" in ast.unparse(node):
            save_index = idx
    if save_index < 0:
        return tree
    last_write_before_save = -1
    for idx, node in enumerate(tree.body[:save_index]):
        if _contains_spreadsheet_write_node(node):
            last_write_before_save = idx
    kept: List[ast.stmt] = []
    for idx, node in enumerate(tree.body):
        if (
            idx <= last_write_before_save
            or idx >= save_index
            or _contains_spreadsheet_write_node(node)
        ):
            kept.append(node)
    tree.body = kept
    return tree


def _spreadsheet_strip_nonessential_python(code: str) -> str:
    text = str(code or "").strip()
    if not text:
        return ""
    try:
        tree = ast.parse(text)
    except SyntaxError:
        return ""
    tree = _SpreadsheetEssentialCodeTransformer().visit(tree)
    if isinstance(tree, ast.Module):
        tree = _drop_trailing_verification_nodes(tree)
    ast.fix_missing_locations(tree)
    try:
        compact = ast.unparse(tree)
        compact = _compress_style_copy_idiom(compact)
        ast.parse(compact)
    except Exception:
        return ""
    return compact.strip()


def _spreadsheet_make_skill_name(*parts: str, max_len: int = 80) -> str:
    tokens: List[str] = []
    seen = set()
    for part in parts:
        for token in re.findall(r"[A-Za-z0-9]+", str(part or "").lower()):
            if not token or token in seen:
                continue
            tokens.append(token)
            seen.add(token)
    return normalize_skill_name("_".join(tokens))[:max_len]


def _trim_text_words(text: str, *, max_words: int) -> str:
    words = re.findall(r"\S+", str(text or ""))
    if len(words) <= max_words:
        return str(text or "").strip()
    return " ".join(words[:max_words]).rstrip() + " ..."


def _trim_spreadsheet_body_fields(body: str) -> str:
    text = str(body or "").strip()
    if not text:
        return ""
    limits = _spreadsheet_extract_limits()
    if (
        _word_count(text) <= limits["max_body_words"]
        and len(_nonempty_lines(text)) <= limits["max_body_lines"]
    ):
        return text
    code_blocks = _spreadsheet_code_blocks(text)
    code_block = ""
    if code_blocks:
        first = code_blocks[0]
        if len(_nonempty_lines(first)) <= limits["max_code_lines"]:
            code_block = "```python\n" + first.strip() + "\n```"
    section_labels = [
        "Applicability",
        "Corrective rule",
        "Reusable openpyxl idiom",
        "Openpyxl idiom",
        "Failure evidence to avoid",
        "Non-applicability",
    ]
    sections: List[str] = []
    for label in section_labels:
        pattern = rf"(?is)\b{re.escape(label)}\s*:\s*(.*?)(?=\n\s*\n(?:{'|'.join(re.escape(item) for item in section_labels)})\s*:|\n```|\Z)"
        match = re.search(pattern, text)
        if not match:
            continue
        content = _trim_text_words(match.group(1).strip(), max_words=45)
        if content:
            sections.append(f"{label}: {content}")
    if code_block:
        insert_at = 1 if sections else 0
        sections.insert(insert_at, "Reusable code:\n" + code_block)
    if not sections:
        return _trim_text_words(text, max_words=limits["max_body_words"])
    return "\n\n".join(sections)


def _spreadsheet_body_from_structured_raw(raw: Dict[str, Any]) -> str:
    body = str(raw.get("body") or "").strip()
    if str(raw.get("kind") or "").strip().lower() == "skill_package":
        metadata = dict(raw.get("metadata") or {})
        package_files = metadata.get("package_files")
        if isinstance(package_files, dict):
            skill_md = str(package_files.get("SKILL.md") or package_files.get("skill.md") or "").strip()
            if skill_md:
                if not body or body == skill_md:
                    return skill_md
                return body + "\n\nSKILL.md:\n" + skill_md if body else skill_md
        return body
    code_lines = raw.get("code_lines")
    if code_lines is None:
        code_lines = raw.get("code")
    if isinstance(code_lines, str):
        code = code_lines.strip()
    elif isinstance(code_lines, Sequence) and not isinstance(code_lines, (bytes, bytearray)):
        code = "\n".join(str(line).rstrip() for line in code_lines if str(line).strip()).strip()
    else:
        code = ""
    if code:
        code = _normalize_spreadsheet_skill_source(code, max_chars=3200)
        compact = _spreadsheet_strip_nonessential_python(code)
        if compact:
            code = compact
        if body:
            return body + "\n\nReusable code:\n```python\n" + code.strip() + "\n```"
        return "Reusable code:\n```python\n" + code.strip() + "\n```"
    return body


def _coerce_spreadsheet_artifact(raw: Dict[str, Any], *, detail: Dict[str, Any]) -> SkillArtifact | None:
    name = normalize_skill_name(str(raw.get("name") or ""))
    description = str(raw.get("description") or "").strip()
    body = _trim_spreadsheet_body_fields(_spreadsheet_body_from_structured_raw(raw))
    if not name or not description or not body:
        return None
    kind = str(raw.get("kind") or "executable_tool")
    kind_lower = kind.strip().lower()
    if kind_lower == "executable_tool" and "openpyxl" not in body.lower():
        body = "Use openpyxl with INPUT_XLSX and OUTPUT_XLSX.\n\n" + body
    raw_interface = dict(raw.get("interface") or {})
    raw_invocation = dict(raw_interface.get("invocation_contract") or {})
    requested_injection = str(
        (raw.get("metadata") or {}).get("injection_type")
        or raw_invocation.get("injection_type")
        or ""
    ).strip().lower()
    probe = SkillArtifact(name=name, kind=kind, description=description, body=body)
    if kind_lower == "skill_package":
        injection_type = "functional"
    elif kind_lower in {"executable_tool", "function_tool", "script_tool"} and _is_spreadsheet_callable_skill(probe):
        injection_type = "functional"
    elif "workflow" in kind_lower:
        injection_type = "workflow"
    elif requested_injection in {"functional", "workflow", "informational"}:
        injection_type = requested_injection
    else:
        injection_type = "informational"
    task = _spreadsheet_task_fragment(detail)
    result = _spreadsheet_result_projection(detail)
    source_success = bool(result.get("success")) and float(result.get("score") or 0.0) >= 0.9
    metadata = dict(raw.get("metadata") or {})
    if kind_lower == "skill_package":
        package_files = metadata.get("package_files")
        if not isinstance(package_files, dict):
            return None
        cleaned_package_files = {
            str(_safe_package_relative_path(path)): str(content or "")
            for path, content in package_files.items()
            if _safe_package_relative_path(path) is not None
        }
        if "SKILL.md" not in cleaned_package_files:
            return None
        metadata["package_format"] = metadata.get("package_format") or "skills_md"
        metadata["package_files"] = cleaned_package_files
        bundle_files = metadata.get("bundle_files")
        if isinstance(bundle_files, dict):
            metadata["bundle_files"] = {
                str(_safe_package_relative_path(path)): str(content or "")
                for path, content in bundle_files.items()
                if _safe_package_relative_path(path) is not None
            }
    domains = [str(item).strip() for item in (metadata.get("domains") or []) if str(item).strip()]
    if "SpreadsheetBench" not in domains:
        domains.insert(0, "SpreadsheetBench")
    instruction_type = str((task.get("metadata") or {}).get("instruction_type") or "").strip()
    if instruction_type and instruction_type not in domains:
        domains.append(instruction_type)
    metadata.update(
        {
            "domains": domains,
            "allowed_tools": ["openpyxl", "bash"] if kind_lower == "skill_package" else ["openpyxl"],
            "source": metadata.get("source") or "spreadsheet_llm_trace_extraction",
            "source_task_ids": list(dict.fromkeys([*(metadata.get("source_task_ids") or []), str(task.get("task_id") or "")])),
            "benchmark": "spreadsheet",
            "instruction_type": instruction_type,
            "version_kind": "seed",
            "injection_type": injection_type,
            "source_success": source_success,
            "source_score": float(result.get("score") or 0.0),
        }
    )
    tags = [
        "domain:SpreadsheetBench",
        "tool:openpyxl",
        *(f"intent:{item}" for item in metadata.get("intent_keywords") or []),
    ]
    artifact = SkillArtifact(
        name=name,
        kind=kind,
        description=description[:240],
        body=body,
        metadata=metadata,
        tags=list(dict.fromkeys(str(tag) for tag in tags if str(tag).strip())),
        interface=SkillInterface(
            summary=str(((raw.get("interface") or {}).get("summary")) or description),
            usage=str(((raw.get("interface") or {}).get("usage")) or "Use when the SpreadsheetBench instruction and workbook shape match this skill scope."),
            input_contract=dict(((raw.get("interface") or {}).get("input_contract")) or {"benchmark": "SpreadsheetBench"}),
            output_contract=dict(((raw.get("interface") or {}).get("output_contract")) or {"tool": "openpyxl"}),
            invocation_contract={**raw_invocation, "injection_type": injection_type},
            compatibility_notes=str(((raw.get("interface") or {}).get("compatibility_notes")) or metadata.get("non_applicability") or ""),
        ),
        evidence=SkillEvidence(
            source_traces=[_spreadsheet_result_projection(detail)],
        ),
        lineage=SkillLineage(version_kind="seed"),
        dependencies=[str(item).strip() for item in (raw.get("dependencies") or []) if str(item).strip()],
        status="pending",
    )
    if source_success:
        artifact.bundle.positive_cases.append(
            _spreadsheet_case_from_task(
                detail=detail,
                skill_name=artifact.name,
                polarity="positive",
                reason="successful source trace used to bootstrap the skill contract",
                source="distilled_success",
                confidence=0.7,
            )
        )
    else:
        artifact.bundle.negative_cases.append(
            _spreadsheet_case_from_task(
                detail=detail,
                skill_name=artifact.name,
                polarity="negative",
                reason="failure-derived repair evidence must pass validation before promotion",
                source="distilled_failure_repair",
                confidence=0.7,
            )
        )
    return _normalize_spreadsheet_injection_contract(artifact)


def _normalize_spreadsheet_injection_contract(artifact: SkillArtifact) -> SkillArtifact:
    kind_lower = str(artifact.kind or "").strip().lower()
    if kind_lower == "skill_package" or _is_spreadsheet_package_skill(artifact):
        injection_type = "functional"
    elif kind_lower in {"executable_tool", "function_tool", "script_tool"} and _is_spreadsheet_callable_skill(artifact):
        injection_type = "functional"
    elif "workflow" in kind_lower:
        injection_type = "workflow"
    else:
        requested = str(artifact.metadata.get("injection_type") or "").strip().lower()
        invocation_requested = str((artifact.interface.invocation_contract or {}).get("injection_type") or "").strip().lower()
        injection_type = requested if requested in {"workflow", "informational"} else invocation_requested
        if injection_type == "functional" and not _is_spreadsheet_callable_skill(artifact):
            injection_type = "informational"
        if injection_type not in {"functional", "workflow", "informational"}:
            injection_type = "informational"
    artifact.metadata["injection_type"] = injection_type
    artifact.interface.invocation_contract = {
        **dict(artifact.interface.invocation_contract or {}),
        "injection_type": injection_type,
    }
    return artifact


async def _extract_spreadsheet_skills_from_detail(
    detail: Dict[str, Any],
    *,
    store: ArtifactStore,
    config: MaintenanceRunConfig,
    task_index: int,
) -> List[SkillArtifact]:
    projection = _spreadsheet_trace_projection(detail)
    has_success_evidence = bool(projection.get("success")) and float(projection.get("score") or 0.0) >= 0.9
    has_repair_evidence = _spreadsheet_has_repair_evidence(projection)
    if not has_success_evidence and not has_repair_evidence:
        return []
    existing = [
        {
            "name": artifact.name,
            "description": artifact.description,
            "status": artifact.status,
            "domains": artifact.metadata.get("domains") or [],
            "intent_keywords": artifact.metadata.get("intent_keywords") or [],
        }
        for artifact in store.all()
        if artifact.status in {"active", "pending", "stale"}
    ]
    limits = _spreadsheet_extract_limits()
    if _spreadsheet_skill_format(config.extra.get("spreadsheet_skill_format", "auto")) == "folder":
        artifacts = await _extract_spreadsheet_folder_skills_terminal(
            detail=detail,
            existing=existing,
            limits=limits,
            config=config,
            task_index=task_index,
        )
        if artifacts:
            return artifacts[: limits["max_artifacts"]]
        if not bool(config.extra.get("disable_spreadsheet_heuristic_extraction")):
            raw = _heuristic_spreadsheet_artifact_payload(detail, skill_format="folder")
            artifact = _coerce_spreadsheet_artifact(raw, detail=detail)
            return [artifact] if artifact is not None else []
        return []
    raw_artifacts: List[Dict[str, Any]] = []
    try:
        previous_artifacts: List[Dict[str, Any]] | None = None
        rubric_failures: List[Dict[str, Any]] | None = None
        for rewrite_round in range(max(1, limits["max_rewrite_rounds"] + 1)):
            payload = await _compat_ask_json(
                system=SPREADSHEET_EXTRACT_SYSTEM
                + _extractor_rule_suffix(_spreadsheet_role_rules(config, "extractor")),
                user=_spreadsheet_extraction_role_json_block(
                    _spreadsheet_extraction_user_payload(
                        existing=existing,
                        detail=detail,
                        limits=limits,
                        skill_format=config.extra.get("spreadsheet_skill_format", "auto"),
                        previous_artifacts=previous_artifacts,
                        rubric_failures=rubric_failures,
                    )
                ),
                llm_config=config.llm_config,
                model_name=config.model_name,
                role="spreadsheet_extractor" if rewrite_round == 0 else "spreadsheet_extractor_rubric_rewrite",
                metadata={
                    "phase": "spreadsheet_extract" if rewrite_round == 0 else "spreadsheet_extract_rubric_rewrite",
                    "benchmark": "spreadsheet",
                    "task_id": detail.get("task_id"),
                    "task_index": task_index,
                    "rewrite_round": rewrite_round,
                    "n_extractor_rules": len(_spreadsheet_role_rules(config, "extractor")),
                },
            )
            raw_artifacts = [dict(item or {}) for item in list(payload.get("artifacts") or [])]
            rubric_failures = _spreadsheet_extraction_rubric_failures(raw_artifacts, limits=limits)
            if not rubric_failures:
                break
            previous_artifacts = raw_artifacts
    except Exception:
        raw_artifacts = [
            _heuristic_spreadsheet_artifact_payload(
                detail,
                skill_format=config.extra.get("spreadsheet_skill_format", "auto"),
            )
            if has_success_evidence
            else _heuristic_spreadsheet_repair_artifact_payload(detail)
        ]
    if (
        not raw_artifacts
        and has_success_evidence
        and not bool(config.extra.get("disable_spreadsheet_heuristic_extraction"))
    ):
        raw_artifacts = [
            _heuristic_spreadsheet_artifact_payload(
                detail,
                skill_format=config.extra.get("spreadsheet_skill_format", "auto"),
            )
        ]
    if (
        not raw_artifacts
        and has_repair_evidence
        and not bool(config.extra.get("disable_spreadsheet_heuristic_extraction"))
        and _spreadsheet_skill_format(config.extra.get("spreadsheet_skill_format", "auto")) != "folder"
    ):
        raw_artifacts = [_heuristic_spreadsheet_repair_artifact_payload(detail)]
    artifacts: List[SkillArtifact] = []
    raw_artifacts = [
        raw
        for raw in raw_artifacts
        if _spreadsheet_raw_artifact_allowed_by_format(
            raw,
            skill_format=config.extra.get("spreadsheet_skill_format", "auto"),
        )
    ]
    final_failures = _spreadsheet_extraction_rubric_failures(raw_artifacts, limits=limits)
    failed_indexes = {int(row.get("index") or 0) for row in final_failures}
    for index, raw in enumerate(raw_artifacts[: limits["max_artifacts"]]):
        if index in failed_indexes:
            continue
        artifact = _coerce_spreadsheet_artifact(dict(raw or {}), detail=detail)
        if artifact is None:
            continue
        artifacts.append(artifact)
    return artifacts


def _spreadsheet_raw_artifact_allowed_by_format(raw: Dict[str, Any], *, skill_format: Any) -> bool:
    fmt = _spreadsheet_skill_format(skill_format)
    if fmt == "auto":
        return True
    kind = str((raw or {}).get("kind") or "").strip().lower()
    if fmt == "folder":
        return kind == "skill_package"
    if fmt == "function":
        return kind in {"executable_tool", "function_tool", "script_tool"}
    return True


async def _ask_spreadsheet_text_role(
    *,
    system: str,
    prompt: str,
    config: MaintenanceRunConfig,
    role: str,
    metadata: Dict[str, Any] | None = None,
    max_request_wall_s: Any = None,
) -> Any:
    started = time.monotonic()
    llm_config = _spreadsheet_maintenance_llm_config(config.llm_config)
    model = _spreadsheet_maintenance_model_name(config.model_name)
    print(
        json.dumps(
            {
                "progress": "maintenance_llm_start",
                "role": role,
                "llm_config": llm_config,
                "model": model,
                "user_chars": len(prompt),
                "system_chars": len(system),
                "attempt": 1,
                "max_attempts": 1,
                "api": "text",
            },
            ensure_ascii=False,
        ),
        flush=True,
    )
    try:
        response = await ask_text_llm(
            llm_config=llm_config,
            model_name=model,
            system=system,
            prompt=prompt,
            max_request_wall_s=max_request_wall_s,
        )
    except Exception as exc:
        duration_ms = int((time.monotonic() - started) * 1000)
        print(
            json.dumps(
                {
                    "progress": "maintenance_llm_error",
                    "role": role,
                    "llm_config": llm_config,
                    "model": model,
                    "duration_ms": duration_ms,
                    "attempt": 1,
                    "max_attempts": 1,
                    "api": "text",
                    "error_type": type(exc).__name__,
                    "error": str(exc)[-1000:],
                },
                ensure_ascii=False,
            ),
            flush=True,
        )
        raise
    duration_ms = int((time.monotonic() - started) * 1000)
    usage = {
        "prompt_tokens": int(getattr(response, "prompt_tokens", 0) or 0),
        "cache_input_tokens": int(getattr(response, "cache_input_tokens", 0) or 0),
        "completion_tokens": int(getattr(response, "completion_tokens", 0) or 0),
    }
    _record_maintenance_token_event(
        role=role,
        llm_config=llm_config,
        model_name=model,
        usage=usage,
        metadata=metadata,
        duration_ms=duration_ms,
        system_chars=len(system),
        user_chars=len(prompt),
    )
    print(
        json.dumps(
            {
                "progress": "maintenance_llm_done",
                "role": role,
                "duration_ms": duration_ms,
                "response_chars": len(str(getattr(response, "content", "") or "")),
                "prompt_tokens": usage["prompt_tokens"],
                "cache_input_tokens": usage["cache_input_tokens"],
                "completion_tokens": usage["completion_tokens"],
                "total_tokens": sum(usage.values()),
                "attempt": 1,
                "api": "text",
            },
            ensure_ascii=False,
        ),
        flush=True,
    )
    return response


async def _extract_spreadsheet_folder_skills_terminal(
    *,
    detail: Dict[str, Any],
    existing: Sequence[Dict[str, Any]],
    limits: Dict[str, int],
    config: MaintenanceRunConfig,
    task_index: int,
) -> List[SkillArtifact]:
    """Extract folder-style skills by writing files, avoiding JSON-escaped code."""

    max_turns = max(
        1,
        int(config.extra.get("spreadsheet_folder_extractor_max_turns", os.environ.get("SPREADSHEET_FOLDER_EXTRACTOR_MAX_TURNS", "3")) or 3),
    )
    with tempfile.TemporaryDirectory(prefix="spreadsheet_folder_extract_") as tmp:
        work_dir = Path(tmp)
        input_path = work_dir / "input.json"
        payload = _spreadsheet_extraction_user_payload(
            existing=existing,
            detail=detail,
            limits=limits,
            skill_format="folder",
        )
        input_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        current_output = ""
        validation_feedback = (
            "INSPECT_FIRST\n"
            "Turn 0 must inspect input.json and print a concise summary of the task, reusable_edit_steps, "
            "answer range, and any candidate skill idea. Do not create skills on turn 0."
        )
        for turn_index in range(max_turns):
            allow_write = turn_index > 0 or bool(str(current_output or "").strip())
            instruction = {
                "input_path": "input.json",
                "task_summary": _spreadsheet_folder_extractor_task_summary(payload),
                "workspace_contract": {
                    "skill_dir_pattern": "skills/<skill_name>/",
                    "bundle_dir_pattern": "bundles/<skill_name>/",
                    "report_pattern": "reports/<skill_name>.txt",
                    "required_skill_files": ["SKILL.md", "scripts/apply.py"],
                    "required_bundle_files": ["run_tests.py"],
                },
                "report_format": (
                    "Use plain text headings exactly: # NAME, # DESCRIPTION, # INTENT_KEYWORDS, "
                    "# NON_APPLICABILITY, # SOURCE_TASK_IDS. Put one skill per report file."
                ),
                "limits": {
                    "max_artifacts": limits["max_artifacts"],
                    "max_skill_md_words": 180,
                    "recommended_skill_md_words": 160,
                    "max_report_description_words": 40,
                    "max_skill_md_nonempty_lines": limits["max_body_lines"],
                    "max_script_nonempty_lines": 80,
                },
                "validation_feedback": validation_feedback,
                "previous_command_output": current_output[-3000:],
                "allow_writing_skill_files_this_turn": allow_write,
                "instruction": (
                    "Return exactly one fenced bash command. Do not cd out of the current workspace. "
                    "If allow_writing_skill_files_this_turn is false, only inspect input.json and print a concise summary. "
                    "If true, create or edit folder-style SpreadsheetBench skills on disk. Do not print JSON. "
                    "Keep scripts concise and reusable; use INPUT_XLSX/OUTPUT_XLSX or argv paths. "
                    "Tests should import or run scripts from the sibling skill folder. "
                    "The skill must be grounded in reusable_edit_steps or verifier mismatches; do not create unrelated generic Excel skills. "
                    "Extract a reusable operation, not the solved task transcript. Parameterize workbook-specific "
                    "sheet names, ranges, headers, literal keywords, thresholds, markers, and destination columns "
                    "unless the package states a narrow applicability contract. SKILL.md must include: When to use, "
                    "When not to use, Inputs to configure, Workbook assumptions, and How to copy/adapt or run. "
                    "Keep SKILL.md under 160 words when possible and never over 180 words. Keep the report "
                    "description under 40 words and put details in the script comments only when necessary. "
                    "When writing SKILL.md or reports inside a bash heredoc, do not include triple-backtick "
                    "Markdown code fences; use indented examples or plain text command lines instead. "
                    "Good packages include keyword/regex marking, header-based column move/copy with style "
                    "preservation, cross-sheet key filtering/deletion, marker-based row insertion, and date/threshold "
                    "filtering. Bad packages hardcode one workbook's constants, such as one project name, one fixed "
                    "keyword list, or one exact D-to-F edit, without configuration."
                ),
            }
            system = (
                "You are the folder-skill extractor for SpreadsheetBench. "
                "You work in a temporary terminal workspace. Read input.json, then write skill package files "
                "instead of returning JSON. Prefer one narrow, generalized package grounded in the trace. "
                "The package should be useful to a future executor as a callable script or as code to copy/adapt. "
                "Hard constraints: SKILL.md <= 180 words, preferably <= 160; report description <= 40 words; "
                "no triple-backtick Markdown fences inside heredoc-written files; parameterize sheet/range/header/"
                "keyword/threshold/marker constants; bad packages hardcode one workbook's constants without "
                "configuration. "
                "Return only one fenced bash block."
            )
            prompt = _role_json_block(instruction, preserve_keys={"previous_command_output", "validation_feedback"})
            try:
                response = await _ask_spreadsheet_text_role(
                    system=system,
                    prompt=prompt,
                    config=config,
                    role="spreadsheet_folder_extractor",
                    metadata={
                        "phase": "spreadsheet_folder_extract",
                        "benchmark": "spreadsheet",
                        "task_id": detail.get("task_id"),
                        "task_index": task_index,
                        "turn_index": turn_index,
                        "allow_write": allow_write,
                    },
                    max_request_wall_s=config.extra.get("llm_request_timeout_s") or config.max_task_seconds,
                )
            except Exception:
                return []
            command = _spreadsheet_executor.extract_bash_command(response.content)
            if not command:
                validation_feedback = "VALIDATION_FAILED\n- No fenced bash command was returned."
                continue
            if not allow_write and re.search(r"\b(?:mkdir|cat\s*>|tee\s+|touch|python\s+-)\b", command):
                validation_feedback = "VALIDATION_FAILED\n- Turn 0 must inspect input.json only; do not create files until the next turn."
                continue
            result = _spreadsheet_executor.run_bash_command(
                command,
                work_dir,
                work_dir / "input.xlsx",
                work_dir / "output.xlsx",
                min(float(config.max_task_seconds or 60), 60),
            )
            current_output = "\n".join(
                [
                    f"returncode={result.get('returncode')}",
                    "stdout:",
                    str(result.get("stdout") or ""),
                    "stderr:",
                    str(result.get("stderr") or ""),
                ]
            )
            raw_artifacts, failures = _spreadsheet_folder_extractor_raw_artifacts(
                work_dir=work_dir,
                detail=detail,
                limits=limits,
            )
            if not allow_write:
                validation_feedback = (
                    "Now create exactly one narrow folder skill grounded in the inspected task. "
                    "Use the workspace paths from the contract and keep SKILL.md under the limits."
                )
                continue
            if raw_artifacts and not failures:
                artifacts: List[SkillArtifact] = []
                bundle_failures: List[str] = []
                for raw in raw_artifacts[: limits["max_artifacts"]]:
                    artifact = _coerce_spreadsheet_artifact(raw, detail=detail)
                    if artifact is not None:
                        bundle_result = await _compat_module()._execute_spreadsheet_bundle_tests(
                            artifact=artifact,
                            config=config,
                        )
                        bundle_dict = bundle_result.as_dict() if hasattr(bundle_result, "as_dict") else dict(bundle_result or {})
                        aggregate = dict(bundle_dict.get("aggregate") or {})
                        passed = bool(aggregate.get("passed") or aggregate.get("pass_all_tests"))
                        if not passed:
                            case_runs = list(bundle_dict.get("unit_case_runs") or [])
                            failure_text = ""
                            for case_run in case_runs:
                                if not bool((case_run or {}).get("passed")):
                                    failure_text = str((case_run or {}).get("failure_summary") or "")
                                    if not failure_text:
                                        failure_text = json.dumps((case_run or {}).get("metadata") or {}, ensure_ascii=False)
                                    break
                            bundle_failures.append(
                                f"{artifact.name}: {(failure_text or 'bundle test failed')[:1200]}"
                            )
                            continue
                        artifacts.append(artifact)
                if artifacts and not bundle_failures:
                    return artifacts
                if bundle_failures:
                    validation_feedback = "VALIDATION_FAILED\n" + "\n".join(
                        f"- {failure}" for failure in bundle_failures[:12]
                    )
                    if raw_artifacts:
                        validation_feedback += (
                            "\nCurrent detected skill packages: "
                            + ", ".join(str(raw.get("name") or "") for raw in raw_artifacts)
                        )
                    current_output = current_output + "\n" + validation_feedback
                    continue
            validation_feedback = _spreadsheet_folder_extractor_feedback(failures, raw_artifacts)
        raw_artifacts, failures = _spreadsheet_folder_extractor_raw_artifacts(
            work_dir=work_dir,
            detail=detail,
            limits=limits,
        )
        if failures:
            return []
        artifacts = []
        for raw in raw_artifacts[: limits["max_artifacts"]]:
            if _spreadsheet_extraction_rubric_failures([raw], limits=limits):
                continue
            artifact = _coerce_spreadsheet_artifact(raw, detail=detail)
            if artifact is not None:
                artifacts.append(artifact)
        return artifacts


def _spreadsheet_folder_extractor_feedback(
    failures: Sequence[str],
    raw_artifacts: Sequence[Dict[str, Any]],
) -> str:
    if not failures:
        return "VALIDATION_OK"
    lines = ["VALIDATION_FAILED"]
    lines.extend(f"- {failure}" for failure in list(failures)[:12])
    if raw_artifacts:
        lines.append("Current detected skill packages: " + ", ".join(str(raw.get("name") or "") for raw in raw_artifacts))
    else:
        lines.append("No complete skill package was detected on disk.")
    return "\n".join(lines)


def _spreadsheet_folder_extractor_task_summary(payload: Dict[str, Any]) -> Dict[str, Any]:
    result = dict(payload.get("result") or {})
    task = dict((result.get("task") or payload.get("task") or {}) or {})
    if not task:
        # _spreadsheet_extraction_user_payload nests task fields under result->task_fragment in some projections.
        task = dict(result.get("task_fragment") or {})
    trace = dict(result.get("trace") or {})
    steps = list(payload.get("reusable_edit_steps") or [])
    return {
        "task_id": result.get("task_id") or task.get("task_id"),
        "success": result.get("success"),
        "score": result.get("score"),
        "question": _trim_text_words(str(task.get("question") or result.get("question") or ""), max_words=80),
        "answer_sheet": result.get("answer_sheet") or task.get("answer_sheet"),
        "answer_position": result.get("answer_position") or task.get("answer_position"),
        "instruction_type": (task.get("metadata") or {}).get("instruction_type"),
        "reusable_edit_step_count": len(steps),
        "last_reusable_edit_step": {
            "turn_index": (steps[-1] or {}).get("turn_index"),
            "python_code": str((steps[-1] or {}).get("python_code") or "")[:1800],
            "stderr_tail": str((steps[-1] or {}).get("stderr_tail") or "")[-300:],
        }
        if steps
        else {},
        "mismatched_cells": list(result.get("mismatched_cells") or [])[:5],
        "trace_code_hint": str(trace.get("code_snippet") or "")[:1200],
    }


def _spreadsheet_folder_extractor_raw_artifacts(
    *,
    work_dir: Path,
    detail: Dict[str, Any],
    limits: Dict[str, int],
) -> Tuple[List[Dict[str, Any]], List[str]]:
    skills_root = work_dir / "skills"
    bundles_root = work_dir / "bundles"
    reports_root = work_dir / "reports"
    raw_artifacts: List[Dict[str, Any]] = []
    failures: List[str] = []
    if not skills_root.exists():
        return [], ["missing skills/ directory"]
    for skill_dir in sorted(path for path in skills_root.iterdir() if path.is_dir()):
        safe_name = _safe_package_relative_path(skill_dir.name)
        if safe_name is None or str(safe_name) != skill_dir.name:
            failures.append(f"invalid skill directory name: {skill_dir.name}")
            continue
        package_files: Dict[str, str] = {}
        for path in sorted(skill_dir.rglob("*")):
            if (
                not path.is_file()
                or path.name == "__init__.py"
                or "__pycache__" in path.parts
                or path.suffix in {".pyc", ".pyo"}
            ):
                continue
            rel = path.relative_to(skill_dir)
            safe_rel = _safe_package_relative_path(str(rel))
            if safe_rel is None:
                failures.append(f"{skill_dir.name}: invalid package path {rel}")
                continue
            package_files[str(safe_rel)] = path.read_text(errors="replace")
        bundle_files: Dict[str, str] = {}
        bundle_dir = bundles_root / skill_dir.name
        if bundle_dir.exists():
            for path in sorted(bundle_dir.rglob("*")):
                if not path.is_file():
                    continue
                rel = path.relative_to(bundle_dir)
                safe_rel = _safe_package_relative_path(str(rel))
                if safe_rel is None:
                    failures.append(f"{skill_dir.name}: invalid bundle path {rel}")
                    continue
                bundle_files[str(safe_rel)] = path.read_text(errors="replace")
        report = _parse_spreadsheet_folder_report(reports_root / f"{skill_dir.name}.txt")
        if "SKILL.md" not in package_files:
            failures.append(f"{skill_dir.name}: missing SKILL.md")
        elif isinstance(package_files.get("SKILL.md"), str):
            package_files["SKILL.md"] = _compact_spreadsheet_skill_md(package_files["SKILL.md"], max_words=180)
        script_items = [
            (path, content)
            for path, content in package_files.items()
            if path.startswith("scripts/") and path.endswith(".py")
        ]
        if not script_items:
            failures.append(f"{skill_dir.name}: missing scripts/*.py")
        for rel_path, content in script_items:
            script = str(content or "").strip()
            if not script:
                failures.append(f"{skill_dir.name}: {rel_path} is empty")
                continue
            try:
                ast.parse(script)
            except SyntaxError as exc:
                failures.append(f"{skill_dir.name}: {rel_path} has Python syntax error: {exc.msg}")
            lowered = script.lower()
            if "openpyxl" not in lowered:
                failures.append(f"{skill_dir.name}: {rel_path} must use openpyxl")
            if "def " not in script and "__main__" not in script:
                failures.append(f"{skill_dir.name}: {rel_path} must expose a function or __main__ entrypoint")
        if "run_tests.py" not in bundle_files:
            failures.append(f"{skill_dir.name}: missing bundles/{skill_dir.name}/run_tests.py")
        elif not str(bundle_files.get("run_tests.py") or "").strip():
            failures.append(f"{skill_dir.name}: bundles/{skill_dir.name}/run_tests.py is empty")
        raw = {
            "name": normalize_skill_name(report.get("name") or skill_dir.name),
            "kind": "skill_package",
            "description": str(report.get("description") or _spreadsheet_folder_description_from_skill_md(package_files.get("SKILL.md", "")) or f"Folder skill package {skill_dir.name}.")[:240],
            "body": package_files.get("SKILL.md", ""),
            "interface": {"invocation_contract": {"injection_type": "functional"}},
            "metadata": {
                "domains": ["SpreadsheetBench"],
                "intent_keywords": _split_report_list(report.get("intent_keywords")),
                "source_task_ids": _split_report_list(report.get("source_task_ids")) or [str(detail.get("task_id") or "")],
                "non_applicability": report.get("non_applicability", ""),
                "package_format": "skills_md",
                "package_files": package_files,
                "bundle_files": bundle_files,
            },
            "dependencies": [],
        }
        raw_failures = _spreadsheet_extraction_rubric_failures([raw], limits=limits)
        failures.extend(f"{skill_dir.name}: {failure}" for row in raw_failures for failure in (row.get("failures") or []))
        raw_artifacts.append(raw)
    if len(raw_artifacts) > limits["max_artifacts"]:
        failures.append(f"too many skill packages: {len(raw_artifacts)} > {limits['max_artifacts']}")
    return raw_artifacts, failures


def _compact_spreadsheet_skill_md(text: str, *, max_words: int = 180) -> str:
    """Shorten generated SKILL.md without changing executable package files."""

    if _word_count(text) <= max_words:
        return text
    compact_lines: List[str] = []
    previous_blank = False
    for line in str(text or "").splitlines():
        stripped = line.strip()
        if not stripped:
            if compact_lines and not previous_blank:
                compact_lines.append("")
            previous_blank = True
            continue
        previous_blank = False
        if stripped.startswith("#"):
            compact_lines.append(stripped)
            continue
        words = stripped.split()
        limit = 18 if stripped.startswith(("-", "*")) else 24
        compact_lines.append(" ".join(words[:limit]))
    compact = "\n".join(compact_lines).strip()
    if _word_count(compact) <= max_words:
        return compact
    kept: List[str] = []
    total = 0
    for line in compact.splitlines():
        words = re.findall(r"\b\w+\b", line)
        if kept and total + len(words) > max_words:
            break
        kept.append(line)
        total += len(words)
    return "\n".join(kept).strip()


def _parse_spreadsheet_folder_report(path: Path) -> Dict[str, str]:
    if not path.exists():
        return {}
    result: Dict[str, str] = {}
    current = ""
    lines: List[str] = []
    for line in path.read_text(errors="replace").splitlines():
        match = re.match(r"^#\s+([A-Za-z_ -]+)\s*$", line.strip())
        if match:
            if current:
                result[current] = "\n".join(lines).strip()
            current = re.sub(r"[^a-z0-9]+", "_", match.group(1).strip().lower()).strip("_")
            lines = []
            continue
        if current:
            lines.append(line)
    if current:
        result[current] = "\n".join(lines).strip()
    return result


def _split_report_list(value: Any) -> List[str]:
    return [item.strip() for item in re.split(r"[,;\n]+", str(value or "")) if item.strip()]


def _spreadsheet_folder_description_from_skill_md(text: str) -> str:
    match = re.search(r"(?im)^description:\s*(.+)$", str(text or ""))
    if match:
        return match.group(1).strip()
    for line in str(text or "").splitlines():
        clean = line.strip("# ").strip()
        if clean:
            return clean[:160]
    return ""


def _heuristic_spreadsheet_artifact_payload(
    detail: Dict[str, Any],
    *,
    skill_format: str = "auto",
) -> Dict[str, Any]:
    task = _spreadsheet_task_fragment(detail)
    result = _spreadsheet_result_projection(detail)
    instruction_type = str((task.get("metadata") or {}).get("instruction_type") or "spreadsheet_task").strip()
    answer_position = str((task.get("expected") or {}).get("answer_position") or "").strip()
    question = str(task.get("question") or "")
    keywords = _spreadsheet_keywords(question, instruction_type)
    name = _spreadsheet_make_skill_name("spreadsheet", instruction_type, "_".join(keywords[:4]), max_len=80)
    edit_steps = _spreadsheet_reusable_edit_steps_from_detail(detail) or _spreadsheet_reusable_edit_steps(result)
    snippet = _spreadsheet_compact_success_code(edit_steps[-1]["python_code"] if edit_steps else result["trace"]["code_snippet"])
    if _spreadsheet_skill_format(skill_format) == "folder":
        return _heuristic_spreadsheet_package_artifact_payload(
            task=task,
            result=result,
            instruction_type=instruction_type,
            answer_position=answer_position,
            keywords=keywords,
            name=name,
            snippet=snippet,
        )
    body = (
        f"Applicability: SpreadsheetBench {instruction_type}; terms {keywords[:4]}; range `{answer_position}`. "
        "Inspect layout; preserve unrelated content.\n\n"
        "Reusable code:\n"
        "```python\n"
        f"{snippet}\n"
        "```\n\n"
        "Non-applicability: require matching inspected layout."
    )
    return {
        "name": name,
        "kind": "executable_tool",
        "description": f"Openpyxl pattern for {instruction_type} spreadsheet tasks.",
        "body": body,
        "interface": {
            "summary": f"SpreadsheetBench {instruction_type} openpyxl pattern.",
            "usage": "Retrieve for similar SpreadsheetBench instructions before writing Python code.",
            "input_contract": {"benchmark": "SpreadsheetBench", "instruction_type": instruction_type},
            "output_contract": {"tool": "openpyxl", "save_to": "OUTPUT_XLSX"},
            "invocation_contract": {"injection_type": "functional"},
            "compatibility_notes": "Validate workbook layout dynamically; avoid hard-coded values from the source task.",
        },
        "metadata": {
            "domains": ["SpreadsheetBench", instruction_type],
            "allowed_tools": ["openpyxl"],
            "intent_keywords": keywords,
            "source_task_ids": [str(task.get("task_id") or "")],
            "evidence_span": f"successful task score={result.get('score')} answer_position={answer_position}",
            "scope": f"SpreadsheetBench {instruction_type}",
            "non_applicability": "Current workbook layout or instruction type differs materially from the source evidence.",
            "maintenance_action": "new_skill",
            "injection_type": "functional",
        },
        "dependencies": [],
    }


def _heuristic_spreadsheet_package_artifact_payload(
    *,
    task: Dict[str, Any],
    result: Dict[str, Any],
    instruction_type: str,
    answer_position: str,
    keywords: Sequence[str],
    name: str,
    snippet: str,
) -> Dict[str, Any]:
    script = (
        "from __future__ import annotations\n"
        "import sys\n"
        "import openpyxl\n\n"
        "def apply(input_xlsx, output_xlsx):\n"
        "    INPUT_XLSX = input_xlsx\n"
        "    OUTPUT_XLSX = output_xlsx\n"
        + "\n".join("    " + line for line in str(snippet or "pass").splitlines())
        + "\n\n"
        "if __name__ == '__main__':\n"
        "    apply(sys.argv[1], sys.argv[2])\n"
    )
    return {
        "name": name,
        "kind": "skill_package",
        "description": f"Folder skill for {instruction_type} spreadsheet tasks.",
        "body": f"Use when SpreadsheetBench {instruction_type} terms {list(keywords[:4])} and range `{answer_position}` match.",
        "interface": {
            "summary": f"SpreadsheetBench {instruction_type} folder skill.",
            "usage": "Read SKILL.md, inspect scripts/apply.py, then run or adapt the script.",
            "input_contract": {"benchmark": "SpreadsheetBench", "instruction_type": instruction_type},
            "output_contract": {"tool": "openpyxl", "save_to": "OUTPUT_XLSX"},
            "invocation_contract": {"injection_type": "functional"},
            "compatibility_notes": "Validate workbook layout dynamically; avoid hard-coded source-task values.",
        },
        "metadata": {
            "domains": ["SpreadsheetBench", instruction_type],
            "allowed_tools": ["openpyxl", "bash"],
            "intent_keywords": list(keywords),
            "source_task_ids": [str(task.get("task_id") or "")],
            "evidence_span": f"successful task score={result.get('score')} answer_position={answer_position}",
            "scope": f"SpreadsheetBench {instruction_type}",
            "non_applicability": "Current workbook layout or instruction type differs materially from source evidence.",
            "maintenance_action": "new_skill",
            "injection_type": "functional",
            "package_format": "skills_md",
            "package_files": {
                "SKILL.md": (
                    f"---\nname: {name}\ndescription: {instruction_type} SpreadsheetBench reusable script.\n---\n\n"
                    f"# {instruction_type} Spreadsheet Skill\n\n"
                    f"Use only when the current workbook layout, instruction terms `{list(keywords[:4])}`, "
                    f"and answer range `{answer_position}` match the source evidence. "
                    "Inspect `scripts/apply.py` before running or adapting it."
                ),
                "scripts/apply.py": script,
                "references/source.md": f"Source task id: {task.get('task_id')}\nAnswer range: {answer_position}\n",
            },
            "bundle_files": {
                "run_tests.py": (
                    "from pathlib import Path\n"
                    "import sys\n"
                    "ROOT = Path(__file__).resolve().parents[2]\n"
                    f"sys.path.insert(0, str(ROOT / 'skills' / '{name}' / 'scripts'))\n"
                    "import apply\n"
                    "assert hasattr(apply, 'apply')\n"
                )
            },
        },
        "dependencies": [],
    }


def _spreadsheet_has_repair_evidence(projection: Dict[str, Any]) -> bool:
    if projection.get("execution_ok") is False and projection.get("stderr_tail"):
        return True
    mismatches = list(projection.get("mismatched_cells") or [])
    if not mismatches:
        return False
    expected_text = " ".join(str(item.get("expected") or "") for item in mismatches if isinstance(item, dict))
    predicted_text = " ".join(str(item.get("predicted") or "") for item in mismatches if isinstance(item, dict))
    return bool(expected_text.strip() or predicted_text.strip())


def _heuristic_spreadsheet_repair_artifact_payload(detail: Dict[str, Any]) -> Dict[str, Any]:
    task = _spreadsheet_task_fragment(detail)
    result = _spreadsheet_result_projection(detail)
    projection = _spreadsheet_trace_projection(detail)
    instruction_type = str((task.get("metadata") or {}).get("instruction_type") or "spreadsheet_repair").strip()
    mismatches = list(projection.get("mismatched_cells") or [])
    expected_examples = [
        {"cell": item.get("cell"), "expected": item.get("expected")}
        for item in mismatches
        if isinstance(item, dict)
    ][:4]
    predicted_examples = [
        {"cell": item.get("cell"), "predicted": item.get("predicted")}
        for item in mismatches
        if isinstance(item, dict)
    ][:4]
    question = str(task.get("question") or "")
    keywords = _spreadsheet_keywords(question, instruction_type)
    formula_tokens = _spreadsheet_formula_tokens(expected_examples)
    name = _spreadsheet_make_skill_name(
        "spreadsheet",
        "repair",
        instruction_type,
        "_".join((formula_tokens or keywords)[:4]),
        max_len=90,
    )
    body = (
        f"Applicability: SpreadsheetBench {instruction_type} tasks with similar instruction terms "
        f"{keywords[:8]} and verifier evidence matching these expected formula/value patterns: "
        f"{json.dumps(expected_examples, ensure_ascii=False)}.\n\n"
        "Corrective rule: when generating formulas, preserve the official spreadsheet semantics "
        "visible in the task and workbook preview, including blank-cell guards, lookup ranges, "
        "absolute references, and formula strings. Prefer writing the formula pattern itself when "
        "the answer range expects formulas, not computed display values.\n\n"
        "Failure evidence to avoid:\n"
        f"predicted={json.dumps(predicted_examples, ensure_ascii=False)}\n"
        f"expected={json.dumps(expected_examples, ensure_ascii=False)}\n\n"
        "Openpyxl idiom:\n"
        "```python\n"
        "import openpyxl\n"
        "wb = openpyxl.load_workbook(INPUT_XLSX)\n"
        "# inspect workbook layout, then write formulas/values to the requested answer range\n"
        "wb.save(OUTPUT_XLSX)\n"
        "```\n\n"
        "Non-applicability: do not hard-code this exact formula unless the current task has the "
        "same workbook layout, answer range, and instruction semantics."
    )
    return {
        "name": name,
        "kind": "workflow_guardrail_card",
        "description": f"Repair pattern for {instruction_type} SpreadsheetBench formula/value mismatches.",
        "body": body,
        "interface": {
            "summary": f"SpreadsheetBench {instruction_type} mismatch repair pattern.",
            "usage": "Retrieve for similar formula/value mismatch-prone spreadsheet tasks.",
            "input_contract": {"benchmark": "SpreadsheetBench", "instruction_type": instruction_type},
            "output_contract": {"tool": "openpyxl", "save_to": "OUTPUT_XLSX", "preserve_formula_semantics": True},
            "invocation_contract": {"injection_type": "workflow"},
            "compatibility_notes": "Validate workbook layout and answer range before applying any source formula pattern.",
        },
        "metadata": {
            "domains": ["SpreadsheetBench", instruction_type],
            "allowed_tools": ["openpyxl"],
            "intent_keywords": list(dict.fromkeys([*keywords, *formula_tokens])),
            "source_task_ids": [str(task.get("task_id") or "")],
            "evidence_span": f"failed trace score={result.get('score')} mismatches={expected_examples}",
            "scope": f"SpreadsheetBench {instruction_type} mismatch repair",
            "non_applicability": "Do not apply when formula/value semantics, sheet layout, or answer range differ.",
            "maintenance_action": "new_skill",
            "extracted_from_failure": True,
        },
        "dependencies": [],
    }


def _spreadsheet_compact_success_code(code: str) -> str:
    text = _normalize_spreadsheet_skill_source(code, max_chars=3200)
    compact = _spreadsheet_strip_nonessential_python(text)
    if compact and len(_nonempty_lines(compact)) <= _spreadsheet_extract_limits()["max_code_lines"]:
        return compact
    if compact:
        text = compact
    return text or "import openpyxl\nwb = openpyxl.load_workbook(INPUT_XLSX)\n# inspect workbook layout, edit requested answer range\nwb.save(OUTPUT_XLSX)"


def _spreadsheet_formula_tokens(examples: Sequence[Dict[str, Any]]) -> List[str]:
    text = " ".join(str(item.get("expected") or "") for item in examples)
    tokens = []
    for word in re.findall(r"[A-Za-z_][A-Za-z0-9_]+", text):
        lowered = word.lower()
        if lowered in {"if", "or", "and"} or len(lowered) >= 4:
            if lowered not in tokens:
                tokens.append(lowered)
        if len(tokens) >= 8:
            break
    return tokens


def _spreadsheet_keywords(question: str, instruction_type: str = "") -> List[str]:
    raw = str(question or "")
    words = re.findall(r"[A-Za-z][A-Za-z0-9_]+", raw.lower())
    stop = {
        "the", "and", "for", "with", "into", "from", "this", "that", "workbook",
        "spreadsheet", "sheet", "cell", "cells", "please", "using", "make",
        "level", "manipulation",
    }
    out: List[str] = []
    for word in words:
        if len(word) < 3 or word in stop:
            continue
        if word not in out:
            out.append(word)
        if len(out) >= 12:
            break
    return out or ["spreadsheet", "openpyxl"]


def _normalize_spreadsheet_credit_events(
    payload: Dict[str, Any],
    *,
    detail: Dict[str, Any],
    candidate_artifacts: Sequence[SkillArtifact],
    projection: Dict[str, Any],
) -> List[Dict[str, Any]]:
    known = {artifact.name for artifact in candidate_artifacts}
    raw_events: List[Dict[str, Any]] = []
    for raw in payload.get("skill_judgments") or []:
        row = dict(raw or {})
        skill_name = str(row.get("skill_name") or "").strip()
        if skill_name not in known:
            continue
        judgment = str(row.get("judgment") or "uncertain").strip().lower()
        if judgment not in {"helpful", "harmful", "neutral", "uncertain"}:
            judgment = "uncertain"
        suggestions = []
        for item in row.get("bundle_case_suggestions") or []:
            suggestion = dict(item or {})
            suggestion.setdefault("skill_name", skill_name)
            suggestions.append(_normalize_spreadsheet_bundle_suggestion(suggestion, detail=detail))
        raw_events.append(
            {
                "benchmark": "spreadsheet",
                "task_id": detail.get("task_id"),
                "skill_name": skill_name,
                "judgment": judgment,
                "effect_type": str(row.get("effect_type") or "unknown").strip().lower() or "unknown",
                "confidence": max(0.0, min(1.0, float(row.get("confidence") or 0.0))),
                "reason": str(row.get("reason") or ""),
                "maintenance_actions": [
                    dict(item or {}) for item in (row.get("maintenance_actions") or []) if isinstance(item, dict)
                ],
                "refine_required": bool(row.get("refine_required")),
                "filter_candidate": bool(row.get("filter_candidate")),
                "failure_mode": _normalize_spreadsheet_credit_failure_mode(row.get("failure_mode")),
                "evidence_strength": str(row.get("evidence_strength") or "weak").strip().lower() or "weak",
                "attribution_scope": str(row.get("attribution_scope") or "none").strip().lower() or "none",
                "bundle_case_suggestions": suggestions,
                "evidence": copy.deepcopy(dict(row.get("evidence") or {})),
                "projection": copy.deepcopy(projection),
            }
        )
    return normalize_credit_events(raw_events, task_id=str(detail.get("task_id") or ""), benchmark="spreadsheet")


def _normalize_spreadsheet_credit_failure_mode(value: Any) -> str:
    raw = str(value or "").strip().lower()
    allowed = {
        "irrelevant_retrieval",
        "skill_scope_too_broad",
        "skill_body_wrong_or_incomplete",
        "executor_misuse",
        "insufficient_evidence",
    }
    return raw if raw in allowed else "insufficient_evidence"


def _normalize_spreadsheet_bundle_suggestion(
    suggestion: Dict[str, Any],
    *,
    detail: Dict[str, Any],
) -> Dict[str, Any]:
    polarity = str(suggestion.get("polarity") or "").strip().lower()
    if polarity not in {"positive", "negative", "integration"}:
        polarity = "integration"
    policy = str(suggestion.get("task_fragment_policy") or "reuse_official_fragment").strip()
    return {
        "skill_name": str(suggestion.get("skill_name") or ""),
        "polarity": polarity,
        "reason": str(suggestion.get("reason") or ""),
        "source_task_id": str(suggestion.get("source_task_id") or detail.get("task_id") or ""),
        "focus_turn_indices": [0] if policy != "no_replayable_fragment" else [],
        "required_context_turn_indices": [],
        "state_requirements": copy.deepcopy(dict(suggestion.get("state_requirements") or {})),
        "expected_contract": str(suggestion.get("expected_contract") or "SpreadsheetBench verifier should match the official answer range in the golden workbook."),
        "task_fragment_policy": policy if policy in {"reuse_official_fragment", "no_replayable_fragment"} else "reuse_official_fragment",
    }


def _heuristic_spreadsheet_credit_events(
    *,
    detail: Dict[str, Any],
    candidate_artifacts: Sequence[SkillArtifact],
    projection: Dict[str, Any],
    reason: str,
) -> List[Dict[str, Any]]:
    events: List[Dict[str, Any]] = []
    success = bool(projection.get("success"))
    score = float(projection.get("score") or 0.0)
    stderr = str(projection.get("stderr_tail") or "")
    mismatches = projection.get("mismatched_cells") or []
    for artifact in candidate_artifacts:
        overlap = _spreadsheet_scope_overlap(artifact, detail)
        retrieved_only = artifact.name in set(projection.get("retrieved_only_skills") or [])
        if retrieved_only:
            judgment = "neutral"
            effect = "domain_match" if overlap else "no_material_effect"
            confidence = 0.45
            suggestions = []
        elif success and score >= 0.9 and overlap:
            judgment = "helpful"
            effect = "correctness_gain"
            confidence = 0.55
            suggestions = [
                {
                    "skill_name": artifact.name,
                    "polarity": "positive",
                    "reason": "successful task with retrieved in-scope spreadsheet skill",
                    "source_task_id": detail.get("task_id"),
                    "focus_turn_indices": [0],
                    "required_context_turn_indices": [],
                    "state_requirements": {},
                    "expected_contract": "official SpreadsheetBench answer range should match the golden workbook",
                    "task_fragment_policy": "reuse_official_fragment",
                }
            ]
        elif (stderr or mismatches) and overlap:
            judgment = "uncertain"
            effect = "unknown"
            confidence = 0.35
            suggestions = []
        else:
            judgment = "neutral"
            effect = "no_material_effect"
            confidence = 0.5
            suggestions = []
        events.append(
            {
                "benchmark": "spreadsheet",
                "task_id": detail.get("task_id"),
                "skill_name": artifact.name,
                "judgment": judgment,
                "effect_type": effect,
                "confidence": confidence,
                "reason": reason,
                "maintenance_actions": [],
                "refine_required": False,
                "filter_candidate": False,
                "evidence_strength": "weak",
                "attribution_scope": "prompt_influence" if judgment == "helpful" else "none",
                "bundle_case_suggestions": suggestions,
                "evidence": {
                    **skill_exposure_flags(artifact.name, projection),
                    "trace_signals": [reason],
                },
                "projection": copy.deepcopy(projection),
            }
        )
    return events


def _spreadsheet_scope_overlap(artifact: SkillArtifact, detail: Dict[str, Any]) -> bool:
    task = _spreadsheet_task_fragment(detail)
    instruction_type = str((task.get("metadata") or {}).get("instruction_type") or "").lower()
    question = str(task.get("question") or "").lower()
    keywords = [str(item).lower() for item in (artifact.metadata.get("intent_keywords") or [])]
    domains = [str(item).lower() for item in (artifact.metadata.get("domains") or [])]
    if instruction_type and instruction_type in domains:
        return True
    return any(keyword and keyword in question for keyword in keywords)


def _spreadsheet_case_from_credit_suggestion(
    *,
    detail: Dict[str, Any],
    artifact: SkillArtifact,
    event: Dict[str, Any],
    suggestion: Dict[str, Any],
) -> SkillBundleCase | None:
    policy = str(suggestion.get("task_fragment_policy") or "").strip()
    if policy == "no_replayable_fragment":
        artifact.evidence.repeated_evidence.append(
            {
                "task_id": detail.get("task_id"),
                "skill_name": artifact.name,
                "reason": suggestion.get("reason") or event.get("reason"),
                "not_replayable": True,
            }
        )
        return None
    polarity = str(suggestion.get("polarity") or "").strip().lower()
    confidence = float(event.get("confidence") or 0.0)
    judgment = str(event.get("judgment") or "")
    used = bool((event.get("evidence") or {}).get("used"))
    if polarity == "negative" and not (judgment == "harmful" and (confidence >= 0.65 or used)):
        return None
    if polarity == "positive" and str(event.get("effect_type") or "") not in {
        "token_saving",
        "schema_help",
        "workflow_alignment",
        "correctness_gain",
    }:
        return None
    return _spreadsheet_case_from_task(
        detail=detail,
        skill_name=artifact.name,
        polarity=polarity,
        reason=str(suggestion.get("reason") or event.get("reason") or ""),
        source=f"credit_assigner_{polarity}",
        confidence=confidence,
        credit_event=event,
    )


def _spreadsheet_case_from_task(
    *,
    detail: Dict[str, Any],
    skill_name: str,
    polarity: str,
    reason: str,
    source: str,
    confidence: float,
    credit_event: Dict[str, Any] | None = None,
) -> SkillBundleCase:
    task = _spreadsheet_task_fragment(detail)
    case_id = f"{skill_name}:{polarity}:{stable_id(task.get('task_id'), reason, source)}"
    return SkillBundleCase(
        case_id=case_id,
        source=source,
        prompt=str(task.get("question") or "")[:240],
        expected={
            "answer_sheet": (task.get("expected") or {}).get("answer_sheet"),
            "answer_position": (task.get("expected") or {}).get("answer_position"),
            "verifier": "spreadsheet_golden_range",
        },
        context={
            "task_fragment": copy.deepcopy(task),
            "source_task_id": task.get("task_id"),
            "focus_turns": [0],
            "focus_tools": ["openpyxl"],
            "reason": reason,
            "credit_event": copy.deepcopy(credit_event or {}),
            "confidence": confidence,
        },
        tags=["spreadsheet", polarity],
        polarity=polarity,
        contrast_protocol={"with_skill": True, "without_skill": polarity != "negative"},
    )


def _trim_spreadsheet_bundle_cases(artifact: SkillArtifact) -> None:
    trim_bundle_cases_to_budget(artifact, priority_fn=_spreadsheet_case_priority)


def _spreadsheet_case_priority(case: SkillBundleCase, index: int = 0) -> tuple[Any, ...]:
    return default_bundle_case_priority(case, index)


def _spreadsheet_micro_targets(
    *,
    credit_events: Sequence[Dict[str, Any]],
    credit_bundle_cases: Sequence[Dict[str, Any]],
    extracted: Sequence[SkillArtifact],
) -> List[str]:
    return list(
        _spreadsheet_micro_target_reasons(
            credit_events=credit_events,
            credit_bundle_cases=credit_bundle_cases,
            extracted=extracted,
        )
    )


def _spreadsheet_micro_target_reasons(
    *,
    credit_events: Sequence[Dict[str, Any]],
    credit_bundle_cases: Sequence[Dict[str, Any]],
    extracted: Sequence[SkillArtifact],
) -> Dict[str, List[str]]:
    reasons: Dict[str, List[str]] = {}

    def add(name: Any, reason: str) -> None:
        value = str(name or "").strip()
        if not value:
            return
        rows = reasons.setdefault(value, [])
        if reason not in rows:
            rows.append(reason)

    for event in credit_events or []:
        name = event.get("skill_name")
        if event.get("filter_candidate"):
            add(name, "filter_candidate")
        if event.get("refine_required"):
            add(name, "refine_required")
        if _spreadsheet_is_retrieved_only_scope_refine(event):
            add(name, "retrieved_only_scope_refine")
        if is_strong_harmful_credit(event):
            add(name, "strong_harmful_credit")
    for row in credit_bundle_cases or []:
        polarity = str(row.get("polarity") or "").strip().lower()
        if polarity in {"negative", "integration"}:
            add(row.get("skill_name"), f"{polarity}_bundle_case")
    for artifact in extracted or []:
        add(getattr(artifact, "name", ""), "new_extracted_candidate")
    return reasons


def _spreadsheet_micro_credit_events(
    credit_events: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    return [
        copy.deepcopy(event)
        for event in credit_events or []
        if event.get("filter_candidate")
        or event.get("refine_required")
        or _spreadsheet_is_retrieved_only_scope_refine(event)
        or is_strong_harmful_credit(event)
    ]


def _spreadsheet_micro_credit_bundle_cases(
    credit_bundle_cases: Sequence[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    return [
        copy.deepcopy(row)
        for row in credit_bundle_cases or []
        if str(row.get("polarity") or "").strip().lower() in {"negative", "integration"}
    ]


def _spreadsheet_bundle_test_artifact(artifact: SkillArtifact) -> SkillArtifact:
    candidate = copy.deepcopy(artifact)
    if not candidate.is_disabled():
        candidate.status = "active"
        candidate.metadata["is_pending_skill"] = False
        candidate.metadata.pop("retrieval_disabled_reason", None)
    return candidate


async def _run_spreadsheet_prestore_bundle_gate(
    *,
    artifact: SkillArtifact,
    store: ArtifactStore,
    config: MaintenanceRunConfig,
) -> Dict[str, Any]:
    """Run test/refine/test before a freshly extracted Spreadsheet skill enters the store."""

    if not artifact.bundle.all_cases() and not _spreadsheet_package_has_bundle_tests(artifact):
        return {
            "artifact": artifact,
            "passed": False,
            "final_result": None,
            "test_results": [],
            "refine_decisions": [],
            "rejection_reason": "prestore_bundle_test_unavailable:missing_bundle_cases",
        }
    repair_limit = max(
        0,
        int(
            config.extra.get(
                "spreadsheet_prestore_refine_max_rounds",
                config.extra.get("micro_refine_max_repair_rounds", 1),
            )
            or 0
        ),
    )
    current = artifact
    test_results: List[Dict[str, Any]] = []
    refine_decisions: List[Dict[str, Any]] = []
    final_result: SkillTestResult | None = None
    last_error = ""
    for attempt in range(repair_limit + 1):
        try:
            final_result = await _compat_module()._execute_spreadsheet_bundle_tests(
                artifact=current,
                config=config,
            )
        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"
            break
        test_results.append(final_result.as_dict())
        aggregate = final_result.aggregate or {}
        if bool(aggregate.get("pass_all_tests") or aggregate.get("passed")):
            return {
                "artifact": current,
                "passed": True,
                "final_result": final_result,
                "test_results": test_results,
                "refine_decisions": refine_decisions,
                "rejection_reason": "",
            }
        if attempt >= repair_limit:
            break
        decision = await _refine_spreadsheet_skill_from_bundle(
            artifact=current,
            test_result=final_result,
            credit_context=[],
            store=store,
            config=config,
        )
        public_decision = {key: value for key, value in decision.items() if key != "updated_artifact"}
        public_decision["stage"] = "prestore_post_bundle_failure"
        public_decision["repair_round"] = attempt + 1
        refine_decisions.append(public_decision)
        updated = decision.get("updated_artifact")
        if isinstance(updated, SkillArtifact) and str(decision.get("action") or "") not in {"disable", "rollback"}:
            current = updated
            continue
        break
    reason = "prestore_bundle_test_failed"
    if final_result is None:
        reason = f"prestore_bundle_test_unavailable:{last_error or 'unknown'}"
    elif refine_decisions:
        reason = "prestore_bundle_test_failed_after_refine"
    return {
        "artifact": current,
        "passed": False,
        "final_result": final_result,
        "test_results": test_results,
        "refine_decisions": refine_decisions,
        "rejection_reason": reason,
    }


async def _refine_spreadsheet_skill_from_credit(
    *,
    artifact: SkillArtifact,
    credit_context: Sequence[Dict[str, Any]],
    detail: Dict[str, Any],
    store: ArtifactStore,
    config: MaintenanceRunConfig,
) -> Dict[str, Any]:
    strong = [
        event for event in credit_context
        if event.get("refine_required")
        or event.get("filter_candidate")
        or _spreadsheet_is_retrieved_only_scope_refine(event)
        or (event.get("judgment") == "harmful" and float(event.get("confidence") or 0.0) >= 0.65)
    ]
    if not strong:
        return {"skill_name": artifact.name, "action": "keep", "reason": "no_strong_credit_signal"}
    test_result = {
        "result_id": f"{artifact.name}:credit:{stable_id(detail.get('task_id'), now_iso())}",
        "aggregate": {"passed": False, "reason": "pre_bundle_credit_refine"},
        "failed_cases": [_spreadsheet_result_projection(detail)],
    }
    return await _run_spreadsheet_refiner(
        artifact=artifact,
        test_result=test_result,
        credit_context=list(strong),
        store=store,
        config=config,
        phase="spreadsheet_credit_pre_refine",
    )


async def _refine_spreadsheet_skill_from_bundle(
    *,
    artifact: SkillArtifact,
    test_result: SkillTestResult,
    credit_context: Sequence[Dict[str, Any]],
    store: ArtifactStore,
    config: MaintenanceRunConfig,
) -> Dict[str, Any]:
    return await _run_spreadsheet_refiner(
        artifact=artifact,
        test_result=test_result.as_dict(),
        credit_context=list(credit_context),
        store=store,
        config=config,
        phase="spreadsheet_bundle_refine",
    )


async def _run_spreadsheet_refiner(
    *,
    artifact: SkillArtifact,
    test_result: Dict[str, Any],
    credit_context: Sequence[Dict[str, Any]],
    store: ArtifactStore,
    config: MaintenanceRunConfig,
    phase: str,
) -> Dict[str, Any]:
    if _is_spreadsheet_package_skill(artifact):
        package_decision = await _run_spreadsheet_package_refiner(
            artifact=artifact,
            test_result=test_result,
            credit_context=credit_context,
            config=config,
            phase=phase,
        )
        if package_decision is not None:
            return package_decision
    prompt_artifact = _spreadsheet_refiner_prompt_artifact(artifact)
    try:
        payload = await _compat_refine_skill_artifact_llm(
            prompt_artifact,
            test_result=test_result,
            credit_context=list(credit_context),
            refinement_history=artifact.history[-3:],
            dependency_summaries=summarize_dependency_context(store.all()),
            refiner_rules=_spreadsheet_role_rules(config, "refiner"),
            llm_config=config.llm_config,
            model_name=config.model_name,
            audit_context={"phase": phase, "benchmark": "spreadsheet"},
        )
    except Exception as exc:
        if any(event.get("filter_candidate") for event in credit_context) or _spreadsheet_has_high_confidence_harmful_credit(credit_context):
            updated = _normalize_spreadsheet_injection_contract(copy.deepcopy(artifact))
            updated.status = "disabled"
            updated.metadata["disabled"] = True
            reason_kind = (
                "filter_credit"
                if any(event.get("filter_candidate") for event in credit_context)
                else "high_confidence_harmful_credit"
            )
            updated.metadata["disabled_reason"] = f"spreadsheet_refiner_unavailable_after_{reason_kind}:{type(exc).__name__}"
            return {
                "skill_name": artifact.name,
                "action": "disable",
                "reason": updated.metadata["disabled_reason"],
                "updated_artifact": updated,
            }
        return {
            "skill_name": artifact.name,
            "action": "keep",
            "reason": f"refiner_failed:{type(exc).__name__}",
        }
    decision = dict(payload.get("decision") or {})
    action = str(decision.get("action") or "keep").strip()
    if action == "keep":
        return {"skill_name": artifact.name, "action": "keep", "reason": decision.get("reason") or ""}
    updated = _normalize_spreadsheet_injection_contract(apply_refine_payload(artifact, payload))
    if _artifact_semantic_signature(updated) == _artifact_semantic_signature(artifact):
        return {"skill_name": artifact.name, "action": "keep", "reason": "refiner_returned_no_semantic_change"}
    return {
        "skill_name": artifact.name,
        "action": action,
        "reason": decision.get("reason") or "",
        "updated_artifact": updated,
    }


def _spreadsheet_has_high_confidence_harmful_credit(credit_context: Sequence[Dict[str, Any]]) -> bool:
    for event in credit_context or []:
        if str(event.get("judgment") or "").strip().lower() != "harmful":
            continue
        if float(event.get("confidence") or 0.0) >= 0.85:
            return True
    return False


def _spreadsheet_is_retrieved_only_scope_refine(event: Dict[str, Any]) -> bool:
    if not _spreadsheet_credit_event_is_retrieved_only(event):
        return False
    if not bool(event.get("refine_required")):
        return False
    judgment = str(event.get("judgment") or "").strip().lower()
    if judgment not in {"neutral", "uncertain", "harmful"}:
        return False
    failure_mode = str(event.get("failure_mode") or "").strip().lower()
    if failure_mode not in {"skill_scope_too_broad", "skill_body_wrong_or_incomplete", "executor_misuse"}:
        return False
    attribution_scope = str(event.get("attribution_scope") or "").strip().lower()
    return attribution_scope in {"retrieval_noise", "integration_context", "prompt_influence", "none"}


def _spreadsheet_refiner_prompt_artifact(artifact: SkillArtifact) -> SkillArtifact:
    view = copy.deepcopy(artifact)
    compact = _compact_spreadsheet_skill_card(artifact)
    view.metadata = {
        **dict(view.metadata or {}),
        "spreadsheet_compact_projection": {
            "projection_kind": compact.get("projection_kind"),
            "body_chars": compact.get("body_chars"),
            "body_truncated": False,
            "prompt_body_preserved": True,
            "executable_code_chars": compact.get("executable_code_chars"),
            "executable_code_preserved": compact.get("executable_code_preserved"),
        },
    }
    return view


def _artifact_semantic_signature(artifact: SkillArtifact) -> str:
    payload = {
        "kind": artifact.kind,
        "description": artifact.description,
        "body": artifact.body,
        "metadata": artifact.metadata,
        "interface": artifact.interface.as_dict(),
        "status": artifact.status,
        "dependencies": artifact.dependencies,
    }
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


async def _run_spreadsheet_package_refiner(
    *,
    artifact: SkillArtifact,
    test_result: Dict[str, Any],
    credit_context: Sequence[Dict[str, Any]],
    config: MaintenanceRunConfig,
    phase: str,
) -> Dict[str, Any] | None:
    max_turns = int(config.extra.get("spreadsheet_package_refiner_max_turns", 2) or 2)
    if max_turns <= 0:
        return None
    with tempfile.TemporaryDirectory(prefix="spreadsheet_package_refine_") as tmp:
        work_dir = Path(tmp)
        _write_spreadsheet_skill_packages([artifact], work_dir)
        result_path = work_dir / "failure_report.json"
        result_path.write_text(json.dumps(test_result, ensure_ascii=False, indent=2))
        current_output = ""
        command_history: List[Dict[str, Any]] = []
        for turn_index in range(max_turns):
            payload = {
                "phase": phase,
                "skill_name": artifact.name,
                "skill_dir": str(Path("skills") / _spreadsheet_package_dir_name(artifact.name)),
                "bundle_dir": str(Path("bundles") / _spreadsheet_package_dir_name(artifact.name)),
                "failure_report_path": "failure_report.json",
                "credit_context": list(credit_context),
                "runtime_informed_refiner_rules": _spreadsheet_role_rules(config, "refiner"),
                "previous_command_output": current_output[-3000:],
                "instruction": (
                    "Return exactly one fenced bash command. Inspect or edit files under the skill_dir or bundle_dir, "
                    "then run the package unit test if possible. Do not edit files outside this temp workspace."
                ),
            }
            system = (
                "You are a lightweight terminal refiner for a SpreadsheetBench folder-style skill package. "
                "Repair SKILL.md, scripts, or bundle tests so the package is reusable and its tests pass. "
                "Use concise shell commands and Python snippets. Return only one fenced bash block."
            ) + _refiner_rule_suffix(_spreadsheet_role_rules(config, "refiner"))
            prompt = _role_json_block(payload, preserve_keys={"previous_command_output"})
            response = await _ask_spreadsheet_text_role(
                system=system,
                prompt=prompt,
                config=config,
                role="spreadsheet_package_refiner",
                metadata={
                    "phase": phase,
                    "benchmark": "spreadsheet",
                    "skill_name": artifact.name,
                    "turn_index": turn_index,
                    "n_refiner_rules": len(_spreadsheet_role_rules(config, "refiner")),
                },
                max_request_wall_s=config.extra.get("llm_request_timeout_s") or config.max_task_seconds,
            )
            command = _spreadsheet_executor.extract_bash_command(response.content)
            if not command:
                break
            result = _spreadsheet_executor.run_bash_command(
                command,
                work_dir,
                work_dir / "input.xlsx",
                work_dir / "output.xlsx",
                min(float(config.max_task_seconds or 60), 60),
            )
            command_history.append({"turn_index": turn_index, "command": command[-2000:], **result})
            current_output = "\n".join(
                [
                    f"Return code: {result.get('returncode')}",
                    "stdout:",
                    str(result.get("stdout") or ""),
                    "stderr:",
                    str(result.get("stderr") or ""),
                ]
            )
            if int(result.get("returncode") or 0) == 0 and (
                "pytest" in command + str(result.get("stdout") or "")
                or "run_tests.py" in command
                or "bundles/" in command
            ):
                break
        updated = copy.deepcopy(artifact)
        _collect_spreadsheet_package_files_from_disk(updated, work_dir)
        if _artifact_semantic_signature(updated) == _artifact_semantic_signature(artifact):
            return {
                "skill_name": artifact.name,
                "action": "keep",
                "reason": "package_refiner_no_file_change",
                "terminal_refiner_history": command_history,
            }
        updated = _normalize_spreadsheet_injection_contract(updated)
        updated.lineage.version_kind = "minor"
        updated.lineage.refined_from_result_ids.append(str(test_result.get("result_id") or "spreadsheet_package_refiner"))
        return {
            "skill_name": artifact.name,
            "action": "update",
            "reason": "package_refiner_updated_files",
            "updated_artifact": updated,
            "terminal_refiner_history": command_history,
        }


def _collect_spreadsheet_package_files_from_disk(artifact: SkillArtifact, work_dir: Path) -> None:
    safe_name = _spreadsheet_package_dir_name(artifact.name)
    skill_dir = work_dir / "skills" / safe_name
    bundle_dir = work_dir / "bundles" / safe_name
    package_files: Dict[str, str] = {}
    bundle_files: Dict[str, str] = {}
    if skill_dir.exists():
        for path in sorted(skill_dir.rglob("*")):
            if not path.is_file():
                continue
            rel = path.relative_to(skill_dir)
            if rel.name == "__init__.py":
                continue
            safe_rel = _safe_package_relative_path(str(rel))
            if safe_rel is not None:
                package_files[str(safe_rel)] = path.read_text(errors="replace")
    if bundle_dir.exists():
        for path in sorted(bundle_dir.rglob("*")):
            if not path.is_file():
                continue
            rel = path.relative_to(bundle_dir)
            safe_rel = _safe_package_relative_path(str(rel))
            if safe_rel is not None:
                bundle_files[str(safe_rel)] = path.read_text(errors="replace")
    if package_files:
        artifact.metadata["package_format"] = "skills_md"
        artifact.metadata["package_files"] = package_files
        artifact.body = package_files.get("SKILL.md") or artifact.body
    if bundle_files:
        artifact.metadata["bundle_files"] = bundle_files


def _spreadsheet_test_result_from_dict(payload: Dict[str, Any]) -> SkillTestResult:
    return SkillTestResult(
        result_id=str(payload.get("result_id") or "spreadsheet_bundle_result"),
        skill_name=str(payload.get("skill_name") or ""),
        skill_version=int(payload.get("skill_version") or 1),
        bundle_id=str(payload.get("bundle_id") or ""),
        bundle_version=int(payload.get("bundle_version") or 1),
        run_label=str(payload.get("run_label") or "spreadsheet_bundle_unit"),
        aggregate=copy.deepcopy(dict(payload.get("aggregate") or {"passed": bool(payload.get("passed"))})),
        counterfactual=copy.deepcopy(dict(payload.get("counterfactual") or {})),
        integration_failures=copy.deepcopy(list(payload.get("integration_failures") or [])),
        created_at=str(payload.get("created_at") or now_iso()),
    )


async def _execute_spreadsheet_bundle_tests(
    *,
    artifact: SkillArtifact,
    config: MaintenanceRunConfig,
) -> SkillTestResult:
    case_runs: List[SkillTestCaseRun] = []
    integration_failures: List[Dict[str, Any]] = []
    with_skill_valid_by_task: Dict[str, bool | None] = {}
    without_skill_valid_by_task: Dict[str, bool | None] = {}
    comparable_case_count = 0
    improved = 0
    regressed = 0
    tokens_delta = 0
    if _is_spreadsheet_package_skill(artifact):
        package_run = _run_spreadsheet_package_bundle_unit(artifact, config=config)
        if package_run is not None:
            case_runs.append(package_run)
            comparable_case_count += 1
            if not package_run.passed:
                regressed += 1
                integration_failures.append(
                    {
                        "task_id": artifact.name,
                        "case_id": package_run.case_id,
                        "error": package_run.failure_summary,
                        "contract_failures": copy.deepcopy(package_run.metadata.get("contract_failures") or []),
                    }
                )
    for case in artifact.bundle.all_cases():
        gate = _spreadsheet_bundle_strict_gate(case)
        task = gate.get("task")
        task_id = str((task.task_id if isinstance(task, BenchmarkTask) else "") or case.case_id)
        if gate["failures"]:
            failure_summary = json.dumps(gate["failures"], ensure_ascii=False)
            run = _spreadsheet_bundle_gate_failure_run(
                case=case,
                task=task if isinstance(task, BenchmarkTask) else None,
                variant="strict_gate",
                failures=gate["failures"],
                artifact=artifact,
            )
            case_runs.append(run)
            integration_failures.append(
                {
                    "task_id": task_id,
                    "case_id": case.case_id,
                    "error": failure_summary,
                    "contract_failures": copy.deepcopy(gate["failures"]),
                }
            )
            regressed += 1
            comparable_case_count += 1
            with_skill_valid_by_task[task_id] = False
            continue

        assert isinstance(task, BenchmarkTask)
        test_artifact = _spreadsheet_bundle_test_artifact(artifact)
        with_config = copy.deepcopy(config)
        with_config.extra = {
            **dict(config.extra or {}),
            "skill_injector_mode": "direct",
            "spreadsheet_pending_skill_fraction": 0.0,
            "min_skill_score": 0.0,
        }
        with_result = await _run_spreadsheet_task_for_config(
            task,
            config=with_config,
            store=ArtifactStore([test_artifact]),
            top_k_skills=1,
        )
        with_run = _spreadsheet_bundle_case_run(
            case=case,
            task=task,
            artifact=artifact,
            result=with_result,
            variant="with_skill",
            top_k_skills=1,
            skill_injection_mode="direct",
        )
        case_runs.append(with_run)
        with_skill_valid_by_task[task.task_id] = bool(with_result.success)
        if not with_run.passed:
            integration_failures.append(
                {
                    "task_id": task.task_id,
                    "case_id": case.case_id,
                    "metrics": copy.deepcopy(with_result.metrics or {}),
                    "trace": copy.deepcopy(with_result.trace or {}),
                    "error": with_result.error or with_run.failure_summary,
                    "contract_failures": copy.deepcopy(with_run.metadata.get("contract_failures") or []),
                }
            )
            regressed += 1
            comparable_case_count += 1
            continue

        should_run_without = bool((case.contrast_protocol or {}).get("without_skill", True))
        if not should_run_without:
            comparable_case_count += 1
            if case.polarity in {"positive", "integration"}:
                improved += 1
            continue

        without_result = await _run_spreadsheet_task_for_config(
            task,
            config=config,
            store=ArtifactStore([]),
            top_k_skills=0,
        )
        without_run = _spreadsheet_bundle_case_run(
            case=case,
            task=task,
            artifact=artifact,
            result=without_result,
            variant="without_skill",
            top_k_skills=0,
            skill_injection_mode="none",
        )
        case_runs.append(without_run)
        without_skill_valid_by_task[task.task_id] = bool(without_result.success)
        comparable_case_count += 1
        before_valid = bool(without_result.success)
        after_valid = bool(with_result.success)
        if after_valid and not before_valid:
            improved += 1
        if before_valid and not after_valid:
            regressed += 1
        tokens_delta += int((with_result.metrics or {}).get("total_tokens") or 0) - int((without_result.metrics or {}).get("total_tokens") or 0)

    strict_failures = [
        {
            "case_id": run.case_id,
            "variant": run.variant,
            "contract_failures": copy.deepcopy(run.metadata.get("contract_failures") or []),
        }
        for run in case_runs
        if not run.passed and run.metadata.get("contract_failures")
    ]
    with_runs = [run for run in case_runs if run.variant == "with_skill"]
    passed = bool(case_runs) and all(run.passed for run in case_runs if run.variant != "without_skill")
    avg_accuracy = round(
        sum(float(run.accuracy or 0.0) for run in with_runs) / max(len(with_runs), 1),
        4,
    )
    result = SkillTestResult(
        result_id=f"{artifact.name}:spreadsheet_bundle:{stable_id(artifact.version, now_iso())}",
        skill_name=artifact.name,
        skill_version=artifact.version,
        bundle_id=artifact.bundle.bundle_id or f"{artifact.name}.bundle",
        bundle_version=artifact.bundle.bundle_version,
        run_label="spreadsheet_bundle_unit",
        unit_case_runs=case_runs,
        aggregate={
            "passed": passed,
            "pass_all_tests": passed and not strict_failures and not integration_failures,
            "n_cases": len(artifact.bundle.all_cases()) + (1 if _spreadsheet_package_has_bundle_tests(artifact) else 0),
            "n_case_runs": len(case_runs),
            "n_comparable_cases": comparable_case_count,
            "n_passed": sum(1 for run in case_runs if run.passed),
            "n_improved": improved,
            "n_regressed": regressed,
            "n_strict_failures": len(strict_failures),
            "strict_failures": strict_failures,
            "strict_contract_gate": True,
            "with_before_without": True,
            "avg_accuracy": avg_accuracy,
            "unit_utility_report": {
                "delta_accuracy": round((improved - regressed) / max(comparable_case_count, 1), 4),
                "delta_tokens": tokens_delta,
            },
        },
        counterfactual={
            "without_skill_valid_by_task": without_skill_valid_by_task,
            "with_skill_valid_by_task": with_skill_valid_by_task,
            "with_without_delta": {
                "n_improved": improved,
                "n_regressed": regressed,
            },
        },
        integration_failures=integration_failures,
        created_at=now_iso(),
    )
    return result


def _spreadsheet_bundle_strict_gate(case: SkillBundleCase) -> Dict[str, Any]:
    failures: List[Dict[str, Any]] = []
    task_fragment = dict((case.context or {}).get("task_fragment") or {})
    if not task_fragment:
        failures.append({"type": "missing_task_fragment", "reason": "bundle case has no replayable task_fragment"})
    expected = copy.deepcopy(dict(task_fragment.get("expected") or {}))
    input_artifacts = copy.deepcopy(dict(task_fragment.get("input_artifacts") or {}))
    metadata = copy.deepcopy(dict(task_fragment.get("metadata") or {}))
    task = BenchmarkTask(
        benchmark="spreadsheet",
        task_id=str(task_fragment.get("task_id") or case.case_id),
        question=task_fragment.get("question") or case.prompt,
        expected=expected,
        input_artifacts=input_artifacts,
        metadata=metadata,
    )
    if not input_artifacts.get("input_xlsx"):
        failures.append({"type": "missing_input_xlsx", "reason": "bundle case task_fragment lacks input workbook path"})
    if not expected.get("golden_xlsx"):
        failures.append({"type": "missing_golden_xlsx", "reason": "bundle case task_fragment lacks golden workbook path"})
    if not expected.get("answer_position"):
        failures.append({"type": "missing_answer_position", "reason": "bundle case task_fragment lacks answer range"})
    for label, path_value in (
        ("input_xlsx", input_artifacts.get("input_xlsx")),
        ("golden_xlsx", expected.get("golden_xlsx")),
    ):
        if path_value and not Path(str(path_value)).exists():
            failures.append({"type": f"{label}_not_found", "path": str(path_value)})
    if expected.get("answer_position"):
        try:
            refs = _answer_range_refs(str(expected.get("answer_position")), default_sheet=_first_sheet_name(expected.get("answer_sheet")))
        except Exception as exc:
            failures.append({"type": "invalid_answer_range", "reason": type(exc).__name__})
        else:
            if not refs:
                failures.append({"type": "empty_answer_range", "reason": "answer range produced no replay cells"})
    return {"task": task, "failures": failures}


def _spreadsheet_package_has_bundle_tests(artifact: SkillArtifact) -> bool:
    return bool(_spreadsheet_bundle_files(artifact).get("run_tests.py"))


def _run_spreadsheet_package_bundle_unit(
    artifact: SkillArtifact,
    *,
    config: MaintenanceRunConfig,
) -> SkillTestCaseRun | None:
    if not _spreadsheet_package_has_bundle_tests(artifact):
        return None
    with tempfile.TemporaryDirectory(prefix="spreadsheet_package_bundle_") as tmp:
        work_dir = Path(tmp)
        _write_spreadsheet_skill_packages([artifact], work_dir)
        safe_name = _spreadsheet_package_dir_name(artifact.name)
        entrypoint = work_dir / "bundles" / safe_name / "run_tests.py"
        if not entrypoint.exists():
            return None
        env_timeout = min(float(config.max_task_seconds or 60), 60.0)
        proc = None
        try:
            import subprocess

            env = os.environ.copy()
            env["SKILL_DIR"] = str(work_dir / "skills" / safe_name)
            env["BUNDLE_DIR"] = str(work_dir / "bundles" / safe_name)
            env["PYTHONPATH"] = os.pathsep.join(
                [
                    str(work_dir),
                    str(work_dir / "skills" / safe_name),
                    str(work_dir / "skills" / safe_name / "scripts"),
                    env.get("PYTHONPATH", ""),
                ]
            )
            proc = subprocess.run(
                ["python", str(entrypoint)],
                cwd=str(work_dir),
                env=env,
                text=True,
                capture_output=True,
                timeout=env_timeout,
            )
            stdout = proc.stdout[-4000:]
            stderr = proc.stderr[-4000:]
            returncode = proc.returncode
        except subprocess.TimeoutExpired as exc:
            stdout = (exc.stdout.decode() if isinstance(exc.stdout, bytes) else str(exc.stdout or ""))[-4000:]
            stderr = (
                (exc.stderr.decode() if isinstance(exc.stderr, bytes) else str(exc.stderr or ""))
                + f"\nPackage bundle test timed out after {env_timeout} seconds."
            )[-4000:]
            returncode = 124
        passed = returncode == 0
        failures = [] if passed else [{"type": "package_bundle_failed", "reason": stderr or stdout or f"returncode={returncode}"}]
        return SkillTestCaseRun(
            case_id=f"{artifact.name}:package_unit",
            variant="package_unit",
            passed=passed,
            accuracy=1.0 if passed else 0.0,
            validity=passed,
            failure_summary="" if passed else (stderr or stdout or f"returncode={returncode}")[-1000:],
            trace_ref=f"{artifact.name}:package_unit",
            input_payload={
                "skill_dir": str(Path("skills") / safe_name),
                "bundle_dir": str(Path("bundles") / safe_name),
                "entrypoint": str(Path("bundles") / safe_name / "run_tests.py"),
            },
            expected_behavior={"entrypoint_returncode": 0},
            actual_output={"stdout": stdout, "stderr": stderr, "returncode": returncode},
            skill_snapshot={
                "name": artifact.name,
                "version": artifact.version,
                "kind": artifact.kind,
                "package_files": sorted(_spreadsheet_package_files(artifact).keys()),
                "bundle_files": sorted(_spreadsheet_bundle_files(artifact).keys()),
            },
            metadata={
                "polarity": "positive",
                "source": "package_bundle_files",
                "contract_passed": passed,
                "contract_failures": failures,
                "package_unit": True,
            },
        )


def _spreadsheet_bundle_gate_failure_run(
    *,
    case: SkillBundleCase,
    task: BenchmarkTask | None,
    variant: str,
    failures: Sequence[Dict[str, Any]],
    artifact: SkillArtifact,
) -> SkillTestCaseRun:
    return SkillTestCaseRun(
        case_id=case.case_id,
        variant=variant,
        passed=False,
        accuracy=0.0,
        validity=False,
        failure_summary=json.dumps(list(failures), ensure_ascii=False),
        trace_ref=(task.task_id if task is not None else case.case_id),
        input_payload={
            "task": task.as_dict() if task is not None else {},
            "variant": variant,
            "llm_test_scope": "spreadsheet_strict_gate",
        },
        expected_behavior={
            "bundle_case_expected": copy.deepcopy(case.expected or {}),
            "contrast_protocol": copy.deepcopy(case.contrast_protocol or {}),
            "polarity": case.polarity,
        },
        skill_snapshot={"name": artifact.name, "version": artifact.version},
        bundle_case_snapshot=case.as_dict(),
        metadata={
            "polarity": case.polarity,
            "source": case.source,
            "contract_passed": False,
            "contract_failures": copy.deepcopy(list(failures)),
            "strict_gate": True,
        },
    )


def _spreadsheet_bundle_case_run(
    *,
    case: SkillBundleCase,
    task: BenchmarkTask,
    artifact: SkillArtifact,
    result: BenchmarkResult,
    variant: str,
    top_k_skills: int,
    skill_injection_mode: str,
) -> SkillTestCaseRun:
    metrics = dict(result.metrics or {})
    passed = bool(result.success)
    return SkillTestCaseRun(
        case_id=case.case_id,
        variant=variant,
        passed=passed,
        accuracy=float(result.score or 0.0),
        validity=passed,
        tokens=int(metrics.get("total_tokens") or 0),
        failure_summary="" if passed else str(result.error or metrics.get("reason") or metrics.get("mismatched_cells") or "failed"),
        trace_ref=result.task_id,
        trace=copy.deepcopy(result.trace),
        input_payload={
            "task": task.as_dict(),
            "variant": variant,
            "top_k_skills": top_k_skills,
            "skill_injection_mode": skill_injection_mode,
            "llm_test_scope": "spreadsheet_with_before_without_counterfactual",
        },
        expected_behavior={
            "bundle_case_expected": copy.deepcopy(case.expected or {}),
            "task_expected": copy.deepcopy(task.expected or {}),
            "contrast_protocol": copy.deepcopy(case.contrast_protocol or {}),
            "polarity": case.polarity,
        },
        actual_output={"metrics": metrics, "score": result.score, "success": result.success},
        trace_summary=_spreadsheet_trace_projection(
            {"task_id": result.task_id, "task": task.as_dict(), "runs": [result.as_dict()]}
        ),
        skill_snapshot={"name": artifact.name, "version": artifact.version} if variant == "with_skill" else {},
        bundle_case_snapshot=case.as_dict(),
        metadata={
            "polarity": case.polarity,
            "source": case.source,
            "metrics": copy.deepcopy(metrics),
            "contract_passed": passed,
            "contract_failures": [] if passed else [{"type": "spreadsheet_replay_failed", "reason": str(result.error or metrics.get("reason") or metrics.get("mismatched_cells") or "failed")}],
        },
    )


def _promote_spreadsheet_pending_from_window(
    store: ArtifactStore,
    window_details: Sequence[Dict[str, Any]],
) -> List[str]:
    window_task_ids = {str(item.get("task_id") or "") for item in window_details}
    successful_window_task_ids = {
        str(item.get("task_id") or "")
        for item in window_details
        if item.get("n_success", 0) or float(item.get("avg_score") or 0.0) >= 0.9
    }
    promoted: List[str] = []
    for artifact in store.pending_artifacts():
        source_ids = {str(item) for item in (artifact.metadata.get("source_task_ids") or [])}
        helpful_task_ids = {
            str(item.get("task_id") or item.get("source_task_id") or "")
            for item in artifact.evidence.helpful_cases
            if str(item.get("task_id") or item.get("source_task_id") or "")
        }
        if not (source_ids & window_task_ids or helpful_task_ids & window_task_ids):
            continue
        source_success = bool(source_ids & successful_window_task_ids) or bool(artifact.metadata.get("source_success"))
        has_passing_test = any(
            bool((result.aggregate or {}).get("pass_all_tests") or (result.aggregate or {}).get("passed"))
            for result in store.test_results()
            if result.skill_name == artifact.name
        )
        helpful_credit_count = len({task_id for task_id in helpful_task_ids if task_id not in source_ids})
        has_posterior_helpful_credit = helpful_credit_count >= _spreadsheet_promotion_helpful_credit_threshold()
        if not artifact.bundle.positive_cases or not has_passing_test or not has_posterior_helpful_credit:
            artifact.metadata.setdefault("promotion_state", "pending")
            artifact.metadata["promotion_blocked_reason"] = (
                "requires_successful_source_posterior_helpful_credit_and_passing_bundle_test"
                if source_success
                else "requires_successful_source_or_passing_bundle_test"
            )
            artifact.metadata["promotion_helpful_credit_count"] = helpful_credit_count
            artifact.metadata["promotion_helpful_credit_threshold"] = _spreadsheet_promotion_helpful_credit_threshold()
            continue
        if store.promote_pending(
            artifact.name,
            reason="spreadsheet_successful_source_trace_in_macro_window",
            refactor_group_id=f"spreadsheet_macro:{stable_id(*sorted(window_task_ids))}",
        ):
            promoted.append(artifact.name)
    return promoted


def _spreadsheet_promotion_helpful_credit_threshold() -> int:
    try:
        return max(1, int(os.environ.get("SPREADSHEET_PROMOTION_HELPFUL_CREDIT_THRESHOLD", "1") or "1"))
    except Exception:
        return 1


def _dedupe_spreadsheet_skills(store: ArtifactStore) -> List[str]:
    active = [artifact for artifact in store.all() if artifact.status == "active" and not artifact.is_disabled()]
    groups: Dict[str, List[SkillArtifact]] = {}
    for artifact in active:
        key = _spreadsheet_dedupe_key(artifact)
        groups.setdefault(key, []).append(artifact)
    archived: List[str] = []
    for group in groups.values():
        if len(group) <= 1:
            continue
        group.sort(
            key=lambda item: (
                item.success_count,
                item.usage_count,
                len(item.evidence.helpful_cases),
                -len(item.evidence.harmful_cases),
            ),
            reverse=True,
        )
        keeper = group[0]
        for duplicate in group[1:]:
            duplicate.status = "archived"
            duplicate.metadata["archived_reason"] = f"spreadsheet_duplicate_of:{keeper.name}"
            duplicate.metadata["duplicate_keeper"] = keeper.name
            keeper.evidence.repeated_evidence.append(
                {
                    "merged_duplicate": duplicate.name,
                    "source_task_ids": duplicate.metadata.get("source_task_ids") or [],
                }
            )
            archived.append(duplicate.name)
    return archived


def _spreadsheet_dedupe_key(artifact: SkillArtifact) -> str:
    instruction_type = str(artifact.metadata.get("instruction_type") or "").lower()
    keywords = sorted(str(item).lower() for item in (artifact.metadata.get("intent_keywords") or [])[:6])
    text = " ".join([instruction_type, artifact.kind, *keywords])
    return stable_id(text, length=12)


def _filter_spreadsheet_harmful_skills(
    store: ArtifactStore,
    credit_events: Sequence[Dict[str, Any]],
) -> List[str]:
    harmful_by_skill: Dict[str, List[Dict[str, Any]]] = {}
    helpful_by_skill: Dict[str, int] = {}
    filter_candidate_by_skill: Dict[str, List[Dict[str, Any]]] = {}
    for event in credit_events:
        skill_name = str(event.get("skill_name") or "")
        if not skill_name:
            continue
        if event.get("filter_candidate"):
            filter_candidate_by_skill.setdefault(skill_name, []).append(dict(event))
        if event.get("judgment") == "harmful" and float(event.get("confidence") or 0.0) >= 0.65:
            harmful_by_skill.setdefault(skill_name, []).append(dict(event))
        if event.get("judgment") == "helpful":
            helpful_by_skill[skill_name] = helpful_by_skill.get(skill_name, 0) + 1
    filtered: List[str] = []
    candidate_names = set(harmful_by_skill) | set(filter_candidate_by_skill)
    for skill_name in candidate_names:
        rows = [*harmful_by_skill.get(skill_name, []), *filter_candidate_by_skill.get(skill_name, [])]
        has_filter_candidate = bool(filter_candidate_by_skill.get(skill_name))
        has_repeated_harm = len(harmful_by_skill.get(skill_name, [])) >= 2
        if (not has_filter_candidate and not has_repeated_harm) or helpful_by_skill.get(skill_name, 0) > 0:
            continue
        artifact = store.get(skill_name)
        if artifact is None or artifact.is_disabled():
            continue
        artifact.status = "disabled"
        artifact.metadata["disabled"] = True
        artifact.metadata["disabled_reason"] = (
            "spreadsheet_macro_filter_candidate_credit"
            if has_filter_candidate
            else "spreadsheet_macro_repeated_high_confidence_harmful_credit"
        )
        artifact.metadata["disabled_credit_events"] = rows[-5:]
        filtered.append(skill_name)
    return filtered


async def _update_spreadsheet_role_feedback_from_window(
    *,
    window_details: Sequence[Dict[str, Any]],
    all_train_details: Sequence[Dict[str, Any]],
    credit_events: Sequence[Dict[str, Any]],
    store: ArtifactStore,
    config: MaintenanceRunConfig,
    round_index: int,
    window_index: int,
    final_window: bool,
    promotions: Sequence[str],
    dedupe: Sequence[str],
    filtered: Sequence[str],
) -> Dict[str, Any]:
    enabled = bool(config.extra.get("spreadsheet_trl_enabled"))
    feedback_rows = _spreadsheet_trl_feedback_rows(
        window_details=window_details,
        credit_events=credit_events,
        store=store,
        promotions=promotions,
        dedupe=dedupe,
        filtered=filtered,
    )
    report: Dict[str, Any] = {
        "enabled": enabled,
        "ran": False,
        "round_index": round_index,
        "window_index": window_index,
        "final_window": final_window,
        "n_feedback_rows": len(feedback_rows),
        "feedback_rows": copy.deepcopy(feedback_rows),
        "role_updates": {},
        "role_feedback": copy.deepcopy(_spreadsheet_role_feedback(config)),
        "n_train_seen": len(all_train_details),
    }
    if not enabled or not feedback_rows:
        return report
    role_feedback = _spreadsheet_role_feedback(config)
    max_rules = int(config.extra.get("spreadsheet_trl_max_rules", os.environ.get("SPREADSHEET_TRL_MAX_RULES", "5")) or 5)
    for role_name in ("extractor", "refiner"):
        role_rows = [
            row for row in feedback_rows
            if role_name in set(row.get("target_roles") or [])
        ]
        if not role_rows:
            continue
        update = await update_role_rules_from_feedback_llm(
            role_name=role_name,
            current_rules=(role_feedback.get(role_name) or {}).get("rules") if isinstance(role_feedback.get(role_name), dict) else [],
            feedback_rows=role_rows,
            llm_config=config.llm_config,
            model_name=config.model_name,
            max_rules=max_rules,
            audit_context={
                "phase": "spreadsheet_macro_trl_feedback",
                "benchmark": "spreadsheet",
                "round_index": round_index,
                "window_index": window_index,
                "role": role_name,
            },
        )
        existing_role_feedback = (
            dict(role_feedback.get(role_name) or {})
            if isinstance(role_feedback.get(role_name), dict)
            else {}
        )
        prior_history = list(existing_role_feedback.get("history") or [])[-4:]
        role_feedback[role_name] = {
            **existing_role_feedback,
            "rules": copy.deepcopy(update.get("rules") or []),
            "last_update_summary": update.get("summary"),
            "last_updated_at": update.get("updated_at"),
            "history": [
                *prior_history,
                {
                    "summary": update.get("summary"),
                    "n_feedback_rows": len(role_rows),
                    "rules": copy.deepcopy(update.get("rules") or []),
                    "round_index": round_index,
                    "window_index": window_index,
                },
            ],
        }
        report["role_updates"][role_name] = copy.deepcopy(update)
    report["ran"] = bool(report["role_updates"])
    report["role_feedback"] = copy.deepcopy(role_feedback)
    return report


def _spreadsheet_trl_feedback_rows(
    *,
    window_details: Sequence[Dict[str, Any]],
    credit_events: Sequence[Dict[str, Any]],
    store: ArtifactStore,
    promotions: Sequence[str],
    dedupe: Sequence[str],
    filtered: Sequence[str],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    window_task_ids = {str(item.get("task_id") or "") for item in window_details}
    credit_by_skill: Dict[str, List[Dict[str, Any]]] = {}
    for event in credit_events:
        task_id = str(event.get("task_id") or event.get("source_task_id") or "")
        if window_task_ids and task_id and task_id not in window_task_ids:
            continue
        skill_name = str(event.get("skill_name") or "")
        if skill_name:
            credit_by_skill.setdefault(skill_name, []).append(dict(event))
    for skill_name, events in sorted(credit_by_skill.items()):
        judgments = [str(item.get("judgment") or "") for item in events]
        helpful = judgments.count("helpful")
        harmful = judgments.count("harmful")
        neutral = judgments.count("neutral")
        retrieved_only = sum(1 for item in events if item.get("scope") == "retrieved_only" or item.get("retrieved_but_unused"))
        filter_candidates = [item for item in events if item.get("filter_candidate")]
        refine_required = [item for item in events if item.get("refine_required")]
        if not (harmful or retrieved_only or filter_candidates or refine_required):
            continue
        artifact = store.get(skill_name)
        rows.append(
            {
                "feedback_type": "spreadsheet_skill_runtime_feedback",
                "skill_name": skill_name,
                "target_roles": ["extractor", "refiner"],
                "task_ids": sorted({str(item.get("task_id") or item.get("source_task_id") or "") for item in events if str(item.get("task_id") or item.get("source_task_id") or "")}),
                "helpful_count": helpful,
                "harmful_count": harmful,
                "neutral_count": neutral,
                "retrieved_but_unused_count": retrieved_only,
                "filter_candidate_count": len(filter_candidates),
                "refine_required_count": len(refine_required),
                "skill_status": getattr(artifact, "status", None),
                "skill_kind": getattr(artifact, "kind", None),
                "skill_description": getattr(artifact, "description", "")[:240] if artifact else "",
                "lesson": _spreadsheet_trl_lesson(
                    harmful=harmful,
                    retrieved_only=retrieved_only,
                    filter_candidate_count=len(filter_candidates),
                    refine_required_count=len(refine_required),
                ),
                "credit_event_samples": [
                    {
                        key: item.get(key)
                        for key in ("task_id", "judgment", "scope", "confidence", "reason", "refine_required", "filter_candidate")
                        if item.get(key) not in (None, "", [], {})
                    }
                    for item in events[-5:]
                ],
            }
        )
    for skill_name in sorted(set(promotions) | set(dedupe) | set(filtered)):
        rows.append(
            {
                "feedback_type": "spreadsheet_macro_decision",
                "skill_name": skill_name,
                "target_roles": ["extractor", "refiner"],
                "promoted": skill_name in set(promotions),
                "deduplicated": skill_name in set(dedupe),
                "filtered": skill_name in set(filtered),
                "lesson": "Use macro outcomes to prefer reusable, tested, repeatedly helpful skills and narrow or drop duplicated/harmful ones.",
            }
        )
    return rows[: int(os.environ.get("SPREADSHEET_TRL_MAX_FEEDBACK_ROWS", "24") or "24")]


def _spreadsheet_trl_lesson(
    *,
    harmful: int,
    retrieved_only: int,
    filter_candidate_count: int,
    refine_required_count: int,
) -> str:
    if filter_candidate_count or harmful:
        return "Future extraction/refinement should narrow applicability and add explicit non-applicability notes for skills that harm or distract executor behavior."
    if retrieved_only:
        return "Skills that are repeatedly retrieved but unused should be made more actionable, renamed/described for the right trigger, or split from overbroad operations."
    if refine_required_count:
        return "Refiner should preserve useful behavior while repairing the exact failing precondition shown by credit or bundle evidence."
    return "Use runtime evidence to improve skill scope, trigger, and testability."


def _split_sheet_range(answer_range: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    if not answer_range:
        return None, answer_range
    text = str(answer_range).strip()
    if "!" not in text:
        return None, text.strip().strip("'").strip('"')
    sheet, cell_range = text.rsplit("!", 1)
    sheet = sheet.strip().strip("'").strip('"')
    cell_range = cell_range.strip().strip("'").strip('"')
    return sheet or None, cell_range


def _answer_range_refs(
    answer_range: Optional[str],
    *,
    default_sheet: Optional[str],
) -> List[Tuple[Optional[str], Optional[str]]]:
    parts = _split_answer_range_list(answer_range)
    refs: List[Tuple[Optional[str], Optional[str]]] = []
    for part in parts:
        sheet, cell_range = _split_sheet_range(_normalize_answer_range_text(part))
        refs.append((sheet or default_sheet, cell_range))
    if not refs and default_sheet:
        refs.append((default_sheet, None))
    return refs


def _split_answer_range_list(answer_range: Optional[str]) -> List[str]:
    text = str(answer_range or "").strip()
    if not text:
        return []
    parts: List[str] = []
    current: List[str] = []
    in_quote = False
    for char in text:
        if char == "'":
            in_quote = not in_quote
        if char == "," and not in_quote:
            part = "".join(current).strip()
            if part:
                parts.append(part)
            current = []
            continue
        current.append(char)
    part = "".join(current).strip()
    if part:
        parts.append(part)
    return parts


def _normalize_answer_range_text(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return text
    # Some SpreadsheetBench rows encode ranges as "'Sheet1!'A1:B2".
    text = re.sub(r"^'([^']+!)'([A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?)$", r"'\1\2", text)
    text = text.strip().strip('"')
    if text.count("'") == 1 and text.startswith("'") and "!" in text:
        text = text[1:]
    return text


def _first_sheet_name(sheet_name: Optional[str]) -> Optional[str]:
    if not sheet_name:
        return None
    return str(sheet_name).split(",", 1)[0].strip().strip("'").strip('"') or None


def _normalize_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 8)
    if isinstance(value, int) and not isinstance(value, bool):
        return value
    if isinstance(value, bool):
        return value
    text = str(value).strip()
    if re.fullmatch(r"-?\d+", text):
        try:
            return int(text)
        except Exception:
            return text
    if re.fullmatch(r"-?\d+\.\d+", text):
        try:
            return round(float(text), 8)
        except Exception:
            return text
    return text


def _jsonable_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
