"""SpreadsheetBench executable skill and Python runtime helpers."""
from __future__ import annotations

import ast
import json
import queue
import re
import stat
import subprocess
import threading
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

import openpyxl
from openpyxl.utils import column_index_from_string

from academic.benchmarks.core.types import SkillArtifact
from academic.config import CODE_EXEC_TIMEOUT

SPREADSHEET_PACKAGE_KINDS = {"skill_package", "script_package", "folder_skill"}
SPREADSHEET_PACKAGE_FORMATS = {"skills_md", "skill_package", "folder_skill"}


def extract_code(text: str) -> str:
    match = re.search(r"```(?:python)?\s*(.*?)```", text or "", re.S | re.I)
    return match.group(1).strip() if match else ""


def write_spreadsheet_skill_library(
    skills: Sequence[SkillArtifact],
    work_dir: Path,
    *,
    disclosure_mode: str = "full",
) -> Dict[str, Any]:
    callable_rows: List[Dict[str, Any]] = []
    skill_objects: List[str] = []
    skill_scripts_dir = work_dir / "skills"
    chunks = [
        "from pathlib import Path\n",
        "from types import SimpleNamespace\n",
        "import openpyxl\n",
        "from openpyxl import load_workbook\n",
        "from openpyxl.utils import column_index_from_string\n",
        "from datetime import datetime, date, time, timedelta\n",
        "import math\n",
        "import re\n\n",
        "def _spreadsheet_column_value(value):\n",
        "    if isinstance(value, str) and re.fullmatch(r\"[A-Za-z]+\", value.strip()):\n",
        "        return column_index_from_string(value.strip().upper())\n",
        "    return value\n\n",
        "def _spreadsheet_callable_kwargs(kwargs, ws):\n",
        "    env = dict(kwargs or {})\n",
        "    for key, value in list(env.items()):\n",
        "        if key.endswith(\"_column\"):\n",
        "            base = key[: -len(\"_column\")]\n",
        "            env.setdefault(f\"{base}_col\", _spreadsheet_column_value(value))\n",
        "            if base.endswith(\"ies\"):\n",
        "                env.setdefault(f\"{base[:-3]}y_col\", _spreadsheet_column_value(value))\n",
        "        elif key.endswith(\"_col\"):\n",
        "            env[key] = _spreadsheet_column_value(value)\n",
        "            env.setdefault(f\"{key[: -len('_col')]}_column\", value)\n",
        "    env.setdefault(\"value_start_row\", 1)\n",
        "    env.setdefault(\"value_end_row\", ws.max_row)\n",
        "    env.setdefault(\"result_start_row\", 1)\n",
        "    env.setdefault(\"result_end_row\", ws.max_row)\n",
        "    return env\n\n",
    ]
    for skill in skills:
        if not is_spreadsheet_callable_skill(skill):
            continue
        func_name = spreadsheet_skill_function_name(skill.name)
        code = spreadsheet_skill_code(skill)
        if not code:
            continue
        rendered = render_spreadsheet_skill_function(func_name, code)
        if not rendered:
            continue
        try:
            ast.parse(rendered)
        except SyntaxError:
            continue
        chunks.append(f"\n# Skill: {skill.name}\n")
        chunks.append(rendered)
        chunks.append("\n")
        callable_rows.append(
            {
                "skill_name": skill.name,
                "function_name": func_name,
                "description": spreadsheet_callable_description(skill),
                "signature": f"{func_name}(INPUT_XLSX, OUTPUT_XLSX, **kwargs)",
                "code": code,
                "code_preview": spreadsheet_skill_code_preview(code),
                "manifest_path": "skills.md",
                "library_path": "skill_library.py",
                "script_path": str(Path("skills") / f"{func_name}.py"),
            }
        )
        skill_objects.append(
            f"{func_name}_skill = SimpleNamespace("
            f"name={skill.name!r}, "
            f"function={func_name}, "
            f"signature={f'{func_name}(INPUT_XLSX, OUTPUT_XLSX, **kwargs)'!r}, "
            f"description={spreadsheet_callable_description(skill)!r}, "
            f"code={code!r})\n"
        )
    if skill_objects:
        chunks.append("\n# Skill object metadata for progressive disclosure.\n")
        chunks.extend(skill_objects)
    if callable_rows:
        work_dir.mkdir(parents=True, exist_ok=True)
        (work_dir / "skill_library.py").write_text("".join(chunks))
        skill_scripts_dir.mkdir(parents=True, exist_ok=True)
        for row in callable_rows:
            script_path = work_dir / str(row["script_path"])
            script_path.write_text(render_spreadsheet_skill_script(str(row["function_name"])))
            script_path.chmod(script_path.stat().st_mode | stat.S_IXUSR)
        (work_dir / "skills.md").write_text(render_spreadsheet_skills_manifest(callable_rows))
    disclosure = str(disclosure_mode or "full").strip().lower()
    prompt = _spreadsheet_callable_prompt(callable_rows, disclosure_mode=disclosure)
    return {"prompt": prompt, "skills": callable_rows, "disclosure_mode": disclosure}


def is_spreadsheet_package_skill(skill: SkillArtifact) -> bool:
    kind = str(skill.kind or "").strip().lower()
    metadata = dict(skill.metadata or {})
    package_format = str(metadata.get("package_format") or "").strip().lower()
    package_files = metadata.get("package_files")
    return kind in SPREADSHEET_PACKAGE_KINDS or package_format in SPREADSHEET_PACKAGE_FORMATS or isinstance(package_files, dict)


def write_spreadsheet_skill_packages(
    skills: Sequence[SkillArtifact],
    work_dir: Path,
    *,
    disclosure_mode: str = "progressive",
) -> Dict[str, Any]:
    package_rows: List[Dict[str, Any]] = []
    for skill in skills:
        if not is_spreadsheet_package_skill(skill):
            continue
        files = spreadsheet_package_files(skill)
        if not files:
            continue
        safe_name = spreadsheet_package_dir_name(skill.name)
        skill_dir = work_dir / "skills" / safe_name
        bundle_dir = work_dir / "bundles" / safe_name
        skill_dir.mkdir(parents=True, exist_ok=True)
        (skill_dir / "scripts").mkdir(parents=True, exist_ok=True)
        (skill_dir / "references").mkdir(parents=True, exist_ok=True)
        (skill_dir / "__init__.py").write_text("")
        (skill_dir / "scripts" / "__init__.py").write_text("")
        script_paths: List[str] = []
        reference_paths: List[str] = []
        for rel_path, content in files.items():
            target_rel = safe_package_relative_path(rel_path)
            if not target_rel:
                continue
            target = skill_dir / target_rel
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_text(str(content or ""))
            if target.suffix == ".py":
                target.chmod(target.stat().st_mode | stat.S_IXUSR)
            display_path = str(Path("skills") / safe_name / target_rel)
            if str(target_rel).startswith("scripts/") and target.suffix == ".py":
                script_paths.append(display_path)
            elif str(target_rel).startswith("references/"):
                reference_paths.append(display_path)
        bundle_files = spreadsheet_bundle_files(skill)
        bundle_paths: List[str] = []
        if bundle_files:
            bundle_dir.mkdir(parents=True, exist_ok=True)
            for rel_path, content in bundle_files.items():
                target_rel = safe_package_relative_path(rel_path)
                if not target_rel:
                    continue
                target = bundle_dir / target_rel
                target.parent.mkdir(parents=True, exist_ok=True)
                target.write_text(str(content or ""))
                if target.suffix == ".py":
                    target.chmod(target.stat().st_mode | stat.S_IXUSR)
                bundle_paths.append(str(Path("bundles") / safe_name / target_rel))
        row = {
            "skill_name": skill.name,
            "package_name": safe_name,
            "description": spreadsheet_callable_description(skill),
            "skill_dir": str(Path("skills") / safe_name),
            "skill_md_path": str(Path("skills") / safe_name / "SKILL.md"),
            "script_paths": sorted(script_paths),
            "reference_paths": sorted(reference_paths),
            "bundle_dir": str(Path("bundles") / safe_name),
            "bundle_paths": sorted(bundle_paths),
            "test_entrypoint": str(Path("bundles") / safe_name / "run_tests.py")
            if str(Path("bundles") / safe_name / "run_tests.py") in bundle_paths
            else "",
        }
        package_rows.append(row)
    disclosure = str(disclosure_mode or "progressive").strip().lower()
    return {
        "prompt": render_spreadsheet_package_prompt(package_rows, disclosure_mode=disclosure),
        "skills": package_rows,
        "disclosure_mode": disclosure,
    }


def spreadsheet_package_files(skill: SkillArtifact) -> Dict[str, str]:
    metadata = dict(skill.metadata or {})
    raw_files = metadata.get("package_files")
    files: Dict[str, str] = {}
    if isinstance(raw_files, dict):
        files = {str(path): str(content or "") for path, content in raw_files.items() if str(path).strip()}
    skill_md = files.get("SKILL.md") or files.get("skill.md")
    if not skill_md:
        files["SKILL.md"] = synthesize_spreadsheet_skill_md(skill)
    if skill.body and "SKILL.md" in files and str(skill.body).strip() not in files["SKILL.md"]:
        files["SKILL.md"] = files["SKILL.md"].rstrip() + "\n\n## Repository Body\n" + str(skill.body).strip() + "\n"
    return {path: content for path, content in files.items() if safe_package_relative_path(path) is not None}


def spreadsheet_bundle_files(skill: SkillArtifact) -> Dict[str, str]:
    raw_files = dict(skill.metadata or {}).get("bundle_files")
    if not isinstance(raw_files, dict):
        return {}
    return {
        str(path): str(content or "")
        for path, content in raw_files.items()
        if str(path).strip() and safe_package_relative_path(path) is not None
    }


def synthesize_spreadsheet_skill_md(skill: SkillArtifact) -> str:
    description = str(skill.description or "").strip() or str(skill.name)
    body = str(skill.body or "").strip()
    return (
        "---\n"
        f"name: {spreadsheet_package_dir_name(skill.name)}\n"
        f"description: {description[:240]}\n"
        "---\n\n"
        "# Spreadsheet Skill\n\n"
        f"{description}\n\n"
        "Use this folder only when the task, workbook shape, and answer range match the description.\n"
        "Read scripts under `scripts/` before calling or adapting them.\n"
        + ("\n## Notes\n\n" + body + "\n" if body else "")
    )


def render_spreadsheet_package_prompt(package_rows: Sequence[Dict[str, Any]], *, disclosure_mode: str) -> str:
    if not package_rows:
        return ""
    rows: List[str] = []
    for row in package_rows:
        scripts = ", ".join(f"`{path}`" for path in row.get("script_paths") or []) or "(no scripts listed)"
        refs = ", ".join(f"`{path}`" for path in row.get("reference_paths") or []) or "(no references listed)"
        tests = ", ".join(f"`{path}`" for path in row.get("bundle_paths") or []) or "(no bundle tests listed)"
        rows.append(
            "\n".join(
                [
                    f"- Skill package `{row['skill_name']}` at `{row['skill_dir']}`",
                    f"  - Entry manifest: `{row['skill_md_path']}`",
                    f"  - Use when: {str(row.get('description') or '')[:260]}",
                    f"  - Runnable scripts: {scripts}",
                    f"  - Reference notes: {refs}",
                    f"  - Bundle tests: {tests}",
                ]
            )
        )
    return (
        "Folder-style Spreadsheet skills are local readable/writable executable directories. "
        "Treat each folder as a code package: read `SKILL.md` first, then inspect or run `scripts/` when the package "
        "matches the workbook and target range. You may import or execute the scripts directly, and you may edit them "
        "in place if the task requires an adaptation.\n"
        "Common bash usage patterns:\n"
        "```bash\n"
        "sed -n '1,220p' skills/<skill_name>/SKILL.md\n"
        "sed -n '1,240p' skills/<skill_name>/scripts/<script>.py\n"
        "python skills/<skill_name>/scripts/<script>.py \"$INPUT_XLSX\" \"$OUTPUT_XLSX\"\n"
        "```\n"
        "Available folder skill packages:\n"
        + "\n".join(rows)
    )


def spreadsheet_package_dir_name(name: str) -> str:
    return spreadsheet_skill_function_name(name)


def safe_package_relative_path(path_value: Any) -> Path | None:
    raw = str(path_value or "").strip().replace("\\", "/")
    if not raw:
        return None
    path = Path(raw)
    if path.is_absolute():
        return None
    parts = path.parts
    if any(part in {"", ".", ".."} for part in parts):
        return None
    if parts[0] in {"skills", "bundles"}:
        return None
    return Path(*parts)


def _spreadsheet_callable_prompt(callable_rows: Sequence[Dict[str, Any]], *, disclosure_mode: str) -> str:
    if not callable_rows:
        return ""
    if disclosure_mode in {"progressive", "signature", "signature_only"}:
        rows = "\n".join(
            f"- `{row['signature']}` from skill `{row['skill_name']}`: {row['description'][:220]} "
            f"Files: `{row.get('library_path') or 'skill_library.py'}`, `{row.get('script_path') or ''}`."
            for row in callable_rows
        )
        object_rows = "\n".join(
            f"- `{row['function_name']}_skill.code`, `sed -n '1,220p' skill_library.py`, or "
            f"`sed -n '1,220p' {row.get('script_path') or ''}` shows the implementation for `{row['function_name']}`."
            for row in callable_rows
        )
        direct_rows = "\n\n".join(
            "\n".join(
                [
                    "```python",
                    f"from skill_library import {row['function_name']}",
                    f"{row['function_name']}(INPUT_XLSX, OUTPUT_XLSX)",
                    "```",
                ]
            )
            for row in callable_rows
        )
        return (
            "Executable Spreadsheet skills are available as local callable code. "
            "Read `skills.md` first when deciding whether a skill applies. `skill_library.py` exports functions and "
            "`skills/*.py` provides runnable wrappers. If a signature matches the task, prefer a direct import and "
            "call. You may inspect the code object or wrapper script to verify details, then either call the function, "
            "run the wrapper, or adapt the file.\n"
            "Direct call examples:\n"
            + direct_rows
            + "\n"
            "Available callable signatures:\n"
            + rows
            + "\nInspectable skill code objects:\n"
            + object_rows
        )
    rows = "\n".join(
        "\n".join(
            [
                f"- `{row['signature']}` from skill `{row['skill_name']}`: {row['description'][:220]}",
                "  Full implementation:",
                "  ```python",
                "\n".join("  " + line for line in str(row.get("code") or "").splitlines()),
                "  ```",
            ]
        )
        for row in callable_rows
    )
    direct_rows = "\n\n".join(
        "\n".join(
            [
                "```python",
                f"from skill_library import {row['function_name']}",
                f"{row['function_name']}(INPUT_XLSX, OUTPUT_XLSX)",
                "```",
            ]
        )
        for row in callable_rows
    )
    if rows:
        return (
            "Executable Spreadsheet skills are already importable from `skill_library`. "
            "Prefer direct reuse in returned code when the signature matches. Direct call examples:\n"
            + direct_rows
            + "\n"
            "Prefer importing a callable over rewriting its implementation when its scope matches the workbook and requested answer range. "
            "Pass explicit keyword arguments for sheet names, ranges, columns, and row bounds when the description requires them.\n"
            + rows
        )
    return ""


def render_spreadsheet_skills_manifest(callable_rows: Sequence[Dict[str, Any]]) -> str:
    chunks = [
        "# Spreadsheet Skills",
        "",
        "This file lists local executable Spreadsheet skills available in this task directory.",
        "Use a skill only when its scope matches the workbook, target sheet/range, and requested operation.",
        "You may inspect or edit the referenced Python files before calling a skill.",
        "",
        "Common invocation patterns:",
        "```bash",
        "python - <<'PY'",
        "import os",
        "from skill_library import <function_name>",
        "<function_name>(os.environ['INPUT_XLSX'], os.environ['OUTPUT_XLSX'])",
        "PY",
        "```",
        "```bash",
        "python skills/<function_name>.py \"$INPUT_XLSX\" \"$OUTPUT_XLSX\" '{\"sheet_name\": \"Sheet1\"}'",
        "```",
    ]
    for row in callable_rows:
        chunks.extend(
            [
                "",
                f"## {row['skill_name']}",
                "",
                f"- Signature: `{row['signature']}`",
                f"- Function: `{row['function_name']}`",
                f"- Use when: {str(row.get('description') or '').strip()}",
                f"- Import: `from skill_library import {row['function_name']}`",
                f"- Wrapper: `python {row.get('script_path') or 'skills/' + row['function_name'] + '.py'} \"$INPUT_XLSX\" \"$OUTPUT_XLSX\" '{{}}'`",
                f"- Implementation files: `{row.get('library_path') or 'skill_library.py'}`, `{row.get('script_path') or ''}`",
                f"- Code object: `{row['function_name']}_skill.code`",
            ]
        )
        preview = str(row.get("code_preview") or "").strip()
        if preview:
            chunks.append(f"- Code preview: `{preview}`")
    return "\n".join(chunks).rstrip() + "\n"


def render_spreadsheet_skill_script(function_name: str) -> str:
    return (
        "#!/usr/bin/env python3\n"
        "\"\"\"Runnable wrapper for an evolved Spreadsheet callable skill.\"\"\"\n"
        "from __future__ import annotations\n\n"
        "import json\n"
        "import os\n"
        "import sys\n"
        "from pathlib import Path\n\n"
        "ROOT = Path(__file__).resolve().parents[1]\n"
        "if str(ROOT) not in sys.path:\n"
        "    sys.path.insert(0, str(ROOT))\n\n"
        f"from skill_library import {function_name}\n\n\n"
        "def main() -> None:\n"
        "    input_xlsx = sys.argv[1] if len(sys.argv) > 1 else os.environ.get('INPUT_XLSX')\n"
        "    output_xlsx = sys.argv[2] if len(sys.argv) > 2 else os.environ.get('OUTPUT_XLSX')\n"
        "    kwargs = json.loads(sys.argv[3]) if len(sys.argv) > 3 and sys.argv[3].strip() else {}\n"
        "    if not input_xlsx or not output_xlsx:\n"
        "        raise SystemExit('Usage: python skills/"
        + function_name
        + ".py INPUT_XLSX OUTPUT_XLSX [json_kwargs]')\n"
        f"    {function_name}(input_xlsx, output_xlsx, **kwargs)\n\n\n"
        "if __name__ == '__main__':\n"
        "    main()\n"
    )


def spreadsheet_callable_description(skill: SkillArtifact, *, limit: int = 520) -> str:
    parts: List[str] = []
    if skill.description:
        parts.append(str(skill.description).strip())
    body = str(skill.body or "")
    applicability = re.search(
        r"(?is)(Applicability\s*:\s*.*?)(?:\n\n|Reusable openpyxl idiom|```|Non-applicability\s*:)",
        body,
    )
    if applicability:
        parts.append(applicability.group(1).strip())
    required = []
    for line in body.splitlines()[:12]:
        stripped = line.strip().lstrip("#").strip()
        if re.search(r"(?i)\b(required variables|assumes|parameters|kwargs)\b", stripped):
            required.append(stripped)
    if required:
        parts.append(" ".join(required[:3]))
    text = " ".join(part for part in parts if part).strip()
    text = re.sub(r"\s+", " ", text)
    if len(text) > limit:
        text = text[: limit - 3].rstrip() + "..."
    return text or str(skill.description or skill.name)


def spreadsheet_skill_code_preview(code: str, *, limit: int = 320) -> str:
    lines: List[str] = []
    for raw in str(code or "").splitlines():
        stripped = raw.strip()
        if not stripped:
            continue
        if stripped.startswith("#"):
            continue
        lines.append(stripped)
        if len(" ".join(lines)) >= limit:
            break
    preview = " ".join(lines)
    preview = re.sub(r"\s+", " ", preview).strip()
    if len(preview) > limit:
        preview = preview[: limit - 3].rstrip() + "..."
    return preview


def is_spreadsheet_callable_skill(skill: SkillArtifact) -> bool:
    """Return whether a Spreadsheet skill should be exposed as importable code.

    Older evolved stores sometimes set ``metadata.injection_type`` to
    ``informational`` even for ``executable_tool`` artifacts.  Callable exposure
    should follow the artifact's semantic kind and code availability, not only
    the prompt injection label.  Workflow and knowledge/interface cards remain
    prompt-only.
    """
    kind = str(skill.kind or "").strip().lower()
    if kind not in {"executable_tool", "function_tool", "script_tool"}:
        return False
    code = spreadsheet_skill_code(skill)
    if not code:
        return False
    try:
        ast.parse(code)
    except SyntaxError:
        return False
    return True


def called_spreadsheet_skill_functions(
    code: str,
    callable_rows: Sequence[Dict[str, Any]],
) -> List[str]:
    function_to_skill = {
        str(row.get("function_name") or ""): str(row.get("skill_name") or "")
        for row in callable_rows
        if str(row.get("function_name") or "") and str(row.get("skill_name") or "")
    }
    if not function_to_skill or not str(code or "").strip():
        return []
    try:
        tree = ast.parse(code)
    except SyntaxError:
        return called_spreadsheet_skill_functions_from_text(code, function_to_skill)
    imported_aliases: Dict[str, str] = {}
    module_aliases: set[str] = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.ImportFrom) and node.module == "skill_library":
            for alias in node.names:
                imported = str(alias.name)
                if imported in function_to_skill:
                    imported_aliases[str(alias.asname or alias.name)] = imported
        elif isinstance(node, ast.Import):
            for alias in node.names:
                if alias.name == "skill_library":
                    module_aliases.add(str(alias.asname or alias.name))
    called: List[str] = []
    for node in ast.walk(tree):
        func = getattr(node, "func", None)
        function_name = ""
        if isinstance(func, ast.Name):
            function_name = imported_aliases.get(func.id, func.id if func.id in function_to_skill else "")
        elif isinstance(func, ast.Attribute) and isinstance(func.value, ast.Name):
            if func.value.id in module_aliases and func.attr in function_to_skill:
                function_name = func.attr
        if function_name:
            skill_name = function_to_skill.get(function_name)
            if skill_name and skill_name not in called:
                called.append(skill_name)
    return called


def called_spreadsheet_skill_functions_from_text(
    code: str,
    function_to_skill: Dict[str, str],
) -> List[str]:
    called: List[str] = []
    for function_name, skill_name in function_to_skill.items():
        if re.search(rf"\b{re.escape(function_name)}\s*\(", str(code or "")):
            if skill_name not in called:
                called.append(skill_name)
    return called


def called_spreadsheet_skill_code_reads(
    text: str,
    callable_rows: Sequence[Dict[str, Any]],
) -> List[Dict[str, str]]:
    reads: List[Dict[str, str]] = []
    if not str(text or "").strip():
        return reads
    for row in callable_rows or []:
        skill_name = str(row.get("skill_name") or "")
        function_name = str(row.get("function_name") or "")
        script_path = str(row.get("script_path") or "")
        if not skill_name and not function_name:
            continue
        if function_name and re.search(rf"\b{re.escape(function_name)}_skill\.code\b", text):
            reads.append(
                {
                    "skill_name": skill_name or function_name,
                    "read_type": "code_object",
                    "target": f"{function_name}_skill.code",
                }
            )
        if function_name and re.search(rf"\bsed\s+-n\s+'1,220p'\s+skill_library\.py\b", text):
            reads.append(
                {
                    "skill_name": skill_name or function_name,
                    "read_type": "library_file",
                    "target": "skill_library.py",
                }
            )
        if script_path and re.search(rf"\bsed\s+-n\s+'1,220p'\s+{re.escape(script_path)}\b", text):
            reads.append(
                {
                    "skill_name": skill_name or function_name,
                    "read_type": "script_file",
                    "target": script_path,
                }
            )
        if script_path and re.search(rf"\bpython\s+{re.escape(script_path)}\b", text):
            reads.append(
                {
                    "skill_name": skill_name or function_name,
                    "read_type": "script_run",
                    "target": script_path,
                }
            )
    deduped: List[Dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    for item in reads:
        key = (item["skill_name"], item["read_type"], item["target"])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped


def spreadsheet_skill_function_name(name: str) -> str:
    value = re.sub(r"\W+", "_", str(name or "").strip()).strip("_").lower()
    if not value:
        value = "spreadsheet_skill"
    if value[0].isdigit():
        value = f"skill_{value}"
    return value


def spreadsheet_skill_code(skill: SkillArtifact) -> str:
    body = str(skill.body or "")
    match = re.search(r"```(?:python)?\s*(.*?)```", body, re.S | re.I)
    if match:
        code = match.group(1).strip()
    else:
        if "```" in body:
            return ""
        code = body.strip()
        if not looks_like_spreadsheet_python(code):
            return ""
    code = re.sub(r"load_workbook\((['\"])[^'\"]*?_input\\.xlsx\1\)", "load_workbook(INPUT_XLSX)", code)
    code = re.sub(r"\\.save\((['\"])[^'\"]*?_output\\.xlsx\1\)", ".save(OUTPUT_XLSX)", code)
    code = re.sub(r"(?m)^\s*INPUT_XLSX\s*=\s*(['\"])[^'\"]*?_input\\.xlsx\1\s*$", "", code)
    code = re.sub(r"(?m)^\s*OUTPUT_XLSX\s*=\s*(['\"])[^'\"]*?_output\\.xlsx\1\s*$", "", code)
    return code


def looks_like_spreadsheet_python(code: str) -> bool:
    if not code.strip():
        return False
    python_start = re.search(
        r"(?m)^\s*(?:from|import|def|class|for|while|if|try|with)\s+\w+|^\s*(?:wb|ws|worksheet|workbook)\b",
        code,
    )
    if python_start:
        return True
    spreadsheet_markers = [
        "load_workbook(",
        "openpyxl.",
        "wb.save(",
        ".cell(",
        "ws[",
    ]
    return any(marker in code for marker in spreadsheet_markers)


def render_spreadsheet_skill_function(func_name: str, code: str) -> str:
    try:
        tree = ast.parse(code)
    except SyntaxError:
        return wrap_spreadsheet_skill_snippet(func_name, code)
    functions = [node.name for node in tree.body if isinstance(node, ast.FunctionDef)]
    if functions:
        first = functions[0]
        return (
            code
            + "\n\n"
            + f"def {func_name}(INPUT_XLSX, OUTPUT_XLSX, **kwargs):\n"
            + "    try:\n"
            + f"        return {first}(INPUT_XLSX, OUTPUT_XLSX, **kwargs)\n"
            + "    except TypeError:\n"
            + f"        return {first}(INPUT_XLSX, OUTPUT_XLSX)\n"
        )
    return wrap_spreadsheet_skill_snippet(func_name, code)


def wrap_spreadsheet_skill_snippet(func_name: str, code: str) -> str:
    code_literal = repr(code.strip() or "pass")
    return (
        f"def {func_name}(INPUT_XLSX, OUTPUT_XLSX, **kwargs):\n"
        "    INPUT_XLSX = Path(INPUT_XLSX)\n"
        "    OUTPUT_XLSX = Path(OUTPUT_XLSX)\n"
        "    wb = load_workbook(INPUT_XLSX)\n"
        "    ws = wb.active\n"
        "    env = dict(globals())\n"
        "    env.update({\n"
        "        'INPUT_XLSX': INPUT_XLSX,\n"
        "        'OUTPUT_XLSX': OUTPUT_XLSX,\n"
        "        'wb': wb,\n"
        "        'ws': ws,\n"
        "        'worksheet': ws,\n"
        "        'workbook': wb,\n"
        "    })\n"
        "    env.update(_spreadsheet_callable_kwargs(kwargs, ws))\n"
        f"    exec({code_literal}, env, env)\n"
        "    out_wb = env.get('wb', wb)\n"
        "    if hasattr(out_wb, 'save'):\n"
        "        out_wb.save(OUTPUT_XLSX)\n"
        "    return OUTPUT_XLSX\n"
    )


def spreadsheet_callable_kwargs(kwargs: Any, ws: Any) -> Dict[str, Any]:
    env = dict(kwargs or {})
    for key, value in list(env.items()):
        if key.endswith("_column"):
            base = key[: -len("_column")]
            env.setdefault(f"{base}_col", spreadsheet_column_value(value))
            if base.endswith("ies"):
                env.setdefault(f"{base[:-3]}y_col", spreadsheet_column_value(value))
        elif key.endswith("_col"):
            env[key] = spreadsheet_column_value(value)
            env.setdefault(f"{key[: -len('_col')]}_column", value)
    env.setdefault("value_start_row", 1)
    env.setdefault("value_end_row", ws.max_row)
    env.setdefault("result_start_row", 1)
    env.setdefault("result_end_row", ws.max_row)
    return env


def spreadsheet_column_value(value: Any) -> Any:
    if isinstance(value, str) and re.fullmatch(r"[A-Za-z]+", value.strip()):
        return column_index_from_string(value.strip().upper())
    return value


def run_code(code: str, input_xlsx: Path, output_xlsx: Path, work_dir: Path) -> Tuple[str, str, int]:
    script = work_dir / "run_spreadsheet_solution.py"
    script.write_text(
        "from pathlib import Path\n"
        f"INPUT_XLSX = Path({str(input_xlsx)!r})\n"
        f"OUTPUT_XLSX = Path({str(output_xlsx)!r})\n\n"
        + code
        + "\n"
    )
    proc = subprocess.run(
        ["python", str(script)],
        cwd=str(work_dir),
        text=True,
        capture_output=True,
        timeout=CODE_EXEC_TIMEOUT,
    )
    return proc.stdout[-4000:], proc.stderr[-4000:], proc.returncode


class NotebookPythonSession:
    def __init__(self, work_dir: Path) -> None:
        self.work_dir = work_dir.resolve()
        self.work_dir.mkdir(parents=True, exist_ok=True)
        self.driver_path = self.work_dir / "spreadsheet_notebook_driver.py"
        self.driver_path.write_text(NOTEBOOK_DRIVER_CODE)
        self.proc = subprocess.Popen(
            ["python", str(self.driver_path.resolve())],
            cwd=str(self.work_dir),
            text=True,
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            bufsize=1,
        )
        self._stderr_tail: List[str] = []
        self._stderr_thread = threading.Thread(target=self._drain_stderr, daemon=True)
        self._stderr_thread.start()

    def _drain_stderr(self) -> None:
        if self.proc.stderr is None:
            return
        for line in self.proc.stderr:
            self._stderr_tail.append(line)
            if len(self._stderr_tail) > 40:
                self._stderr_tail = self._stderr_tail[-40:]

    def run_cell(self, code: str, *, timeout: float) -> Dict[str, Any]:
        if self.proc.poll() is not None:
            return {
                "stdout": "",
                "stderr": "Notebook Python session is not running.\n" + "".join(self._stderr_tail)[-2000:],
                "returncode": self.proc.returncode,
                "timed_out": False,
            }
        assert self.proc.stdin is not None
        assert self.proc.stdout is not None
        payload = json.dumps({"code": code}, ensure_ascii=False)
        self.proc.stdin.write(payload + "\n")
        self.proc.stdin.flush()
        out_queue: queue.Queue[str] = queue.Queue(maxsize=1)

        def read_one() -> None:
            try:
                out_queue.put(self.proc.stdout.readline())
            except Exception as exc:
                out_queue.put(json.dumps({"stdout": "", "stderr": str(exc), "returncode": 1}))

        reader = threading.Thread(target=read_one, daemon=True)
        reader.start()
        try:
            line = out_queue.get(timeout=max(1.0, float(timeout or CODE_EXEC_TIMEOUT)))
        except queue.Empty:
            self.close(kill=True)
            return {"stdout": "", "stderr": "Notebook cell timed out.", "returncode": 124, "timed_out": True}
        if not line:
            return {
                "stdout": "",
                "stderr": "Notebook Python session exited.\n" + "".join(self._stderr_tail)[-2000:],
                "returncode": self.proc.returncode,
                "timed_out": False,
            }
        try:
            result = json.loads(line)
        except Exception:
            result = {"stdout": "", "stderr": f"Malformed notebook driver response: {line[-1000:]}", "returncode": 1}
        result["stdout"] = str(result.get("stdout") or "")[-4000:]
        result["stderr"] = str(result.get("stderr") or "")[-4000:]
        result["returncode"] = int(result.get("returncode") or 0)
        result["timed_out"] = bool(result.get("timed_out", False))
        return result

    def close(self, *, kill: bool = False) -> None:
        if self.proc.poll() is not None:
            return
        if kill:
            self.proc.kill()
            return
        try:
            if self.proc.stdin is not None:
                self.proc.stdin.write(json.dumps({"shutdown": True}) + "\n")
                self.proc.stdin.flush()
            self.proc.wait(timeout=2)
        except Exception:
            self.proc.kill()


NOTEBOOK_DRIVER_CODE = r'''
import contextlib
import io
import json
import sys
import traceback
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

ns = {
    "__name__": "__spreadsheet_notebook__",
    "Path": Path,
    "openpyxl": openpyxl,
    "load_workbook": load_workbook,
}

for raw in sys.stdin:
    try:
        payload = json.loads(raw)
    except Exception as exc:
        print(json.dumps({"stdout": "", "stderr": f"bad json: {exc}", "returncode": 1}), flush=True)
        continue
    if payload.get("shutdown"):
        break
    stdout = io.StringIO()
    stderr = io.StringIO()
    rc = 0
    try:
        with contextlib.redirect_stdout(stdout), contextlib.redirect_stderr(stderr):
            exec(str(payload.get("code") or ""), ns, ns)
    except Exception:
        rc = 1
        stderr.write(traceback.format_exc())
    print(
        json.dumps(
            {
                "stdout": stdout.getvalue()[-4000:],
                "stderr": stderr.getvalue()[-4000:],
                "returncode": rc,
            },
            ensure_ascii=False,
        ),
        flush=True,
    )
'''
