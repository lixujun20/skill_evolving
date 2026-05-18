"""SpreadsheetBench workbook verification and range helpers."""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


def verify_spreadsheet_output(
    *,
    predicted_xlsx: Path,
    golden_xlsx: Path,
    sheet_name: Optional[str],
    answer_range: Optional[str],
) -> Dict[str, Any]:
    pred_wb = openpyxl.load_workbook(predicted_xlsx, data_only=False)
    gold_wb = openpyxl.load_workbook(golden_xlsx, data_only=False)
    requested_sheet = first_sheet_name(sheet_name)
    refs = answer_range_refs(answer_range, default_sheet=requested_sheet)
    if not refs:
        refs = [(requested_sheet, None)]
    mismatches = []
    checked = 0
    for ref_sheet, cell_range in refs:
        sheet = ref_sheet if ref_sheet in gold_wb.sheetnames else gold_wb.sheetnames[0]
        if sheet not in pred_wb.sheetnames:
            mismatches.append({"cell": "__sheet__", "predicted": None, "expected": sheet})
            continue
        pred_ws = pred_wb[sheet]
        gold_ws = gold_wb[sheet]
        cells = cells_in_range(cell_range, gold_ws)
        checked += len(cells)
        for cell in cells:
            pv = normalize_cell_value(pred_ws[cell].value)
            gv = normalize_cell_value(gold_ws[cell].value)
            if pv != gv:
                mismatches.append({"cell": f"{sheet}!{cell}", "predicted": pv, "expected": gv})
    correct = checked - len(mismatches)
    return {
        "pass": len(mismatches) == 0 and checked > 0,
        "cell_accuracy": round(correct / max(checked, 1), 4),
        "checked_cells": checked,
        "mismatched_cells": mismatches[:20],
        "answer_sheet": requested_sheet,
        "answer_position": answer_range,
    }


def cells_in_range(answer_range: Optional[str], ws: Any) -> List[str]:
    if not answer_range:
        return [
            cell.coordinate
            for row in ws.iter_rows()
            for cell in row
            if cell.value is not None
        ]
    target = ws[answer_range]
    if hasattr(target, "coordinate"):
        return [target.coordinate]
    cells = []
    for row in target:
        if isinstance(row, tuple):
            cells.extend(cell.coordinate for cell in row)
        else:
            cells.append(row.coordinate)
    return cells


def split_sheet_range(answer_range: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    if not answer_range:
        return None, answer_range
    text = str(answer_range).strip()
    if "!" not in text:
        return None, text.strip().strip("'").strip('"')
    sheet, cell_range = text.rsplit("!", 1)
    sheet = sheet.strip().strip("'").strip('"')
    cell_range = cell_range.strip().strip("'").strip('"')
    return sheet or None, cell_range


def answer_range_refs(
    answer_range: Optional[str],
    *,
    default_sheet: Optional[str],
) -> List[Tuple[Optional[str], Optional[str]]]:
    parts = split_answer_range_list(answer_range)
    refs: List[Tuple[Optional[str], Optional[str]]] = []
    for part in parts:
        sheet, cell_range = split_sheet_range(normalize_answer_range_text(part))
        refs.append((sheet or default_sheet, cell_range))
    if not refs and default_sheet:
        refs.append((default_sheet, None))
    return refs


def split_answer_range_list(answer_range: Optional[str]) -> List[str]:
    text = str(answer_range or "").strip()
    if not text:
        return []
    parts: List[str] = []
    current: List[str] = []
    in_quote = False
    for char in text:
        if char == "'":
            in_quote = not in_quote
        if char == "," and not in_quote:
            part = "".join(current).strip()
            if part:
                parts.append(part)
            current = []
            continue
        current.append(char)
    part = "".join(current).strip()
    if part:
        parts.append(part)
    return parts


def normalize_answer_range_text(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return text
    # Some SpreadsheetBench rows encode ranges as "'Sheet1!'A1:B2".
    text = re.sub(r"^'([^']+!)'([A-Z]+[0-9]+(?::[A-Z]+[0-9]+)?)$", r"'\1\2", text)
    text = text.strip().strip('"')
    if text.count("'") == 1 and text.startswith("'") and "!" in text:
        text = text[1:]
    return text


def first_sheet_name(sheet_name: Optional[str]) -> Optional[str]:
    if not sheet_name:
        return None
    return str(sheet_name).split(",", 1)[0].strip().strip("'").strip('"') or None


def normalize_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 8)
    if isinstance(value, int) and not isinstance(value, bool):
        return value
    if isinstance(value, bool):
        return value
    text = str(value).strip()
    if re.fullmatch(r"-?\d+", text):
        try:
            return int(text)
        except Exception:
            return text
    if re.fullmatch(r"-?\d+\.\d+", text):
        try:
            return round(float(text), 8)
        except Exception:
            return text
    return text


def jsonable_cell_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
