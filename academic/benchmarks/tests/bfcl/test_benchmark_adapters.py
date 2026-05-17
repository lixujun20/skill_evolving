import asyncio
from pathlib import Path
from types import SimpleNamespace
from typing import Any, List
from unittest.mock import AsyncMock

import openpyxl
import pytest

from academic.benchmarks.bfcl import (
    BFCLToolCall,
    _ToolModelResponse,
    _RetrievalObservation,
    _RetrievalPolicy,
    _bfcl_retrieval_context,
    _bfcl_skill_matches_task,
    _bfcl_skill_matches_turn,
    _effective_llm_timeout,
    _native_skill_system,
    _turn_skill_constraints,
    load_bfcl_tasks,
    load_bfcl_tools,
    run_bfcl_task,
    score_bfcl_calls,
)
from academic.benchmarks.core.artifacts import ArtifactStore
from academic.benchmarks.core.runner import (
    _build_bfcl_skill_bundles,
    _build_maintenance_test_results,
    _build_skill_test_result,
    _load_saved_details,
    _run_bfcl_baseline,
    _run_spreadsheet_baseline,
    _result_from_dict,
    _refine_bfcl_skill_store,
    _validate_refine_output_consistency,
)
from academic.benchmarks.bfcl.maintenance.adapter import (
    _bundle_test_signature,
    _latest_matching_test_result,
    _run_contract_assertions,
    _task_from_case,
    _validate_expected_call_schema,
    execute_bfcl_bundle_tests,
    trim_bundle_cases,
)
from academic.benchmarks.bfcl.skills import extract_bfcl_skills_from_results
from academic.benchmarks.core.types import BenchmarkResult
from academic.benchmarks.core.types import SkillArtifact, SkillBundleCase, SkillTestResult
from academic.benchmarks.core.types import BenchmarkTask
from academic.benchmarks.core.registry import BENCHMARK_REGISTRY
from academic.benchmarks.spreadsheet.adapter import load_spreadsheet_tasks, verify_spreadsheet_output
from academic.skill_repository.debug_events import DebugEventSink


def test_registry_has_selected_benchmarks() -> None:
    assert {"bfcl_v3", "appworld", "officeqa", "spreadsheet", "tir_bench"} <= set(BENCHMARK_REGISTRY)


def test_bfcl_loader_and_scorer_contract() -> None:
    train, test = load_bfcl_tasks(cache_dir=Path("data/benchmarks/bfcl_v3"), n_train=2, n_test=2)
    tools = load_bfcl_tools(Path("data/benchmarks/bfcl_v3"))
    assert len(train) == 2
    assert len(test) == 2
    assert any(t["function"]["name"] == "cd" for t in tools)

    from academic.benchmarks.bfcl import _parse_call

    task = train[0]
    calls = []
    for turn_index, turn in enumerate(task.expected):
        for raw in turn:
            name, args = _parse_call(raw)
            calls.append(BFCLToolCall(name=name, arguments=args, turn_index=turn_index))
    score = score_bfcl_calls(calls, task.expected)
    assert score["task_success"] is True
    assert score["call_f1"] == 1.0


def test_bfcl_scorer_argument_mismatch_cannot_receive_full_match_credit() -> None:
    expected = [["create_ticket(title='Cancellation Issue', description='Error encountered during flight cancellation process.')"]]
    calls = [
        BFCLToolCall(
            name="create_ticket",
            arguments={
                "title": "Cancellation Issue",
                "description": "Error encountered during flight cancellation process.",
                "priority": 3,
            },
            turn_index=0,
        )
    ]
    score = score_bfcl_calls(calls, expected)
    assert score["call_f1"] == 0.0
    assert score["n_matched_calls"] == 0
    assert score["task_success"] is False
    assert any(
        item["type"] == "argument_mismatch" and item["name"] == "create_ticket"
        for item in score["call_errors"]
    )


def test_bfcl_scorer_returns_structured_error_for_invalid_expected_call() -> None:
    calls = [
        BFCLToolCall(
            name="place_order",
            arguments={"order_type": "Buy", "symbol": "TSLA", "price": 667.92, "amount": 75},
            turn_index=0,
        )
    ]
    expected = [["place_order(order_type='Buy', symbol='TSLA', price=<market_price>, amount=75)"]]
    score = score_bfcl_calls(calls, expected)
    assert score["task_success"] is False
    assert score["call_f1"] == 0.0
    assert any(item["type"] == "invalid_expected" for item in score["call_errors"])


def test_turn_watchdog_breaks_on_repeated_call() -> None:
    from academic.benchmarks.bfcl import _TurnWatchdog

    watchdog = _TurnWatchdog()
    assert watchdog.observe([BFCLToolCall(name="a", arguments={"x": 1}, turn_index=0)]) is None
    # exact duplicate of a previously seen call → break
    assert (
        watchdog.observe([BFCLToolCall(name="a", arguments={"x": 1}, turn_index=0)])
        == "repeated_call"
    )


def test_turn_watchdog_does_not_use_expected_coverage_to_stop_extras() -> None:
    from academic.benchmarks.bfcl import _TurnWatchdog

    watchdog = _TurnWatchdog(expected_names=["a", "b"])
    assert (
        watchdog.observe(
            [
                BFCLToolCall(name="a", arguments={"x": 1}, turn_index=0),
                BFCLToolCall(name="b", arguments={"y": 1}, turn_index=0),
            ]
        )
        is None
    )
    assert (
        watchdog.observe([BFCLToolCall(name="c", arguments={"z": 1}, turn_index=0)])
        is None
    )
    assert (
        watchdog.observe([BFCLToolCall(name="d", arguments={"z": 2}, turn_index=0)])
        is None
    )
    assert watchdog.early_stop_reason is None


def test_turn_watchdog_only_repeated_calls_break_even_with_expected_names() -> None:
    from academic.benchmarks.bfcl import _TurnWatchdog

    watchdog = _TurnWatchdog(expected_names=["a"])
    assert watchdog.observe([BFCLToolCall(name="a", arguments={}, turn_index=0)]) is None
    assert watchdog.observe([BFCLToolCall(name="c", arguments={"z": 1}, turn_index=0)]) is None
    assert watchdog.observe([BFCLToolCall(name="d", arguments={"z": 2}, turn_index=0)]) is None
    assert watchdog.observe([BFCLToolCall(name="d", arguments={"z": 2}, turn_index=0)]) == "repeated_call"


def test_turn_watchdog_no_break_without_expected() -> None:
    from academic.benchmarks.bfcl import _TurnWatchdog

    watchdog = _TurnWatchdog(expected_names=[])
    # Without expected names, only repetition can trigger a break
    assert watchdog.observe([BFCLToolCall(name="a", arguments={"x": 1}, turn_index=0)]) is None
    assert watchdog.observe([BFCLToolCall(name="b", arguments={"y": 2}, turn_index=0)]) is None
    assert watchdog.observe([BFCLToolCall(name="c", arguments={"z": 3}, turn_index=0)]) is None
    assert watchdog.early_stop_reason is None


def test_effective_llm_timeout_reserves_task_budget_buffer() -> None:
    assert _effective_llm_timeout(None) >= 1
    assert _effective_llm_timeout(600.0) <= 600
    assert _effective_llm_timeout(30.0) == 15
    assert _effective_llm_timeout(10.0) == 1


def test_spreadsheet_loader_and_verifier_contract() -> None:
    train, test = load_spreadsheet_tasks(cache_dir=Path("data/benchmarks/spreadsheet"), n_train=2, n_test=1)
    assert len(train) == 2
    assert len(test) == 1
    task = train[0]
    result = verify_spreadsheet_output(
        predicted_xlsx=Path(task.expected["golden_xlsx"]),
        golden_xlsx=Path(task.expected["golden_xlsx"]),
        sheet_name=task.expected["answer_sheet"],
        answer_range=task.expected["answer_position"],
    )
    assert result["pass"] is True
    assert result["cell_accuracy"] == 1.0


def test_spreadsheet_verifier_accepts_single_cell_range(tmp_path: Path) -> None:
    pred = tmp_path / "pred.xlsx"
    gold = tmp_path / "gold.xlsx"
    for path in (pred, gold):
        wb = openpyxl.Workbook()
        wb.active["B1"] = 8
        wb.save(path)

    result = verify_spreadsheet_output(
        predicted_xlsx=pred,
        golden_xlsx=gold,
        sheet_name=None,
        answer_range="B1",
    )

    assert result["pass"] is True
    assert result["checked_cells"] == 1
    assert result["cell_accuracy"] == 1.0


async def test_spreadsheet_runner_forwards_model_name(monkeypatch) -> None:
    calls = []

    async def fake_run_spreadsheet_task(task, **kwargs):
        calls.append(kwargs)
        return BenchmarkResult(
            benchmark="spreadsheet",
            task_id=task.task_id,
            success=True,
            score=1.0,
            metrics={},
            trace={},
        )

    monkeypatch.setattr(
        "academic.benchmarks.core.runner.run_spreadsheet_task",
        fake_run_spreadsheet_task,
    )
    task = SimpleNamespace(task_id="sheet_1")

    await _run_spreadsheet_baseline(
        [task],
        1,
        "local_claude_proxy",
        ArtifactStore(),
        model_name="claude-sonnet-4-5",
    )

    assert calls[0]["llm_config"] == "local_claude_proxy"
    assert calls[0]["model_name"] == "claude-sonnet-4-5"


async def test_spreadsheet_task_passes_model_override_to_llm(monkeypatch, tmp_path: Path) -> None:
    from academic.benchmarks.spreadsheet.adapter import run_spreadsheet_task
    from academic.benchmarks.core.llm_text import TextLLMResponse

    captured = {}

    async def fake_ask_text_llm(**kwargs):
        captured.update(kwargs)
        return TextLLMResponse(
            content="""```python
import openpyxl
wb = openpyxl.load_workbook(INPUT_XLSX)
ws = wb.active
ws["B1"] = 8
wb.save(OUTPUT_XLSX)
```""",
            prompt_tokens=10,
            completion_tokens=5,
            model_name=kwargs["model_name"],
            api_style="anthropic_direct",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)
    input_xlsx = tmp_path / "input.xlsx"
    golden_xlsx = tmp_path / "golden.xlsx"
    for path in (input_xlsx, golden_xlsx):
        wb = openpyxl.Workbook()
        wb.active["A1"] = 1
        wb.active["B1"] = 8 if path == golden_xlsx else None
        wb.save(path)
    task = BenchmarkTask(
        benchmark="spreadsheet",
        task_id="sheet_1",
        question="Write 8 to B1.",
        expected={"golden_xlsx": str(golden_xlsx), "answer_sheet": None, "answer_position": "B1"},
        input_artifacts={"input_xlsx": str(input_xlsx)},
        metadata={},
    )

    result = await run_spreadsheet_task(
        task,
        llm_config="local_claude_proxy",
        model_name="claude-sonnet-4-5",
        artifact_store=ArtifactStore(),
        top_k_skills=0,
        work_dir=tmp_path / "work",
    )

    assert result.success is True
    assert captured["llm_config"] == "local_claude_proxy"
    assert captured["model_name"] == "claude-sonnet-4-5"
    assert result.metrics["model_name"] == "claude-sonnet-4-5"
    assert result.metrics["llm_api_style"] == "anthropic_direct"


def test_bfcl_refine_disables_harmful_auto_skill() -> None:
    harmful = SkillArtifact(
        name="bfcl_avoid_extra_startEngine",
        kind="negative_rule_card",
        description="Avoid unnecessary startEngine calls unless required.",
        body="Do not use startEngine as a speculative lookup.",
        metadata={"source": "evolve_rollouts", "injection_type": "informational", "source_task_ids": ["t_bad"]},
    )
    helpful = SkillArtifact(
        name="bfcl_params_cp",
        kind="atomic_tool_rule_card",
        description="Observed parameter names for cp.",
        body="For `cp`, use source and destination.",
        metadata={"source": "evolve_rollouts", "injection_type": "informational", "source_task_ids": ["t_good"]},
    )
    store = ArtifactStore([harmful, helpful])
    train_details = [
        {
            "task_id": "t_bad",
            "runs": [
                {
                    "metrics": {
                        "official_valid": True,
                    }
                }
            ],
        },
        {
            "task_id": "t_good",
            "runs": [
                {
                    "metrics": {
                        "official_valid": False,
                    }
                }
            ],
        },
    ]
    replay_details = [
        {
            "task_id": "t_bad",
            "runs": [
                {
                    "metrics": {
                        "official_valid": False,
                        "retrieved_skills": ["bfcl_avoid_extra_startEngine"],
                        "prompt_injected_skills": ["bfcl_avoid_extra_startEngine"],
                        "tool_injected_skills": [],
                        "called_skill_tools": [],
                        "used_skills": [],
                    }
                }
            ],
        },
        {
            "task_id": "t_good",
            "runs": [
                {
                    "metrics": {
                        "official_valid": True,
                        "retrieved_skills": ["bfcl_params_cp"],
                        "prompt_injected_skills": ["bfcl_params_cp"],
                        "tool_injected_skills": [],
                        "called_skill_tools": [],
                        "used_skills": [],
                    }
                }
            ],
        },
    ]
    _build_bfcl_skill_bundles(store, train_details=train_details, replay_details=replay_details)
    maintenance = _build_maintenance_test_results(
        store,
        train_details=train_details,
        replay_details=replay_details,
        run_label="unit",
    )
    decisions = _refine_bfcl_skill_store(
        store,
        train_details=train_details,
        replay_details=replay_details,
        maintenance_test_results=maintenance,
        min_fail_count=2,
    )
    harmful_after = next(skill for skill in store.all() if skill.name == "bfcl_avoid_extra_startEngine")
    helpful_after = next(skill for skill in store.all() if skill.name == "bfcl_params_cp")
    assert harmful_after.is_disabled() is True
    assert helpful_after.is_disabled() is False
    harmful_decision = next(row for row in decisions if row["skill_name"] == "bfcl_avoid_extra_startEngine")
    assert harmful_decision["action"] == "disable_on_regression"


def test_bfcl_retrieval_context_does_not_leak_expected_tool_at_turn_start() -> None:
    task = BenchmarkTask(
        benchmark="bfcl_v3",
        task_id="ticket-task",
        question=[[{"role": "user", "content": "Create a high priority support ticket."}]],
        expected=[["create_ticket(title='Issue', description='Help', priority=4)"]],
        metadata={"involved_classes": ["TicketAPI"]},
    )
    ticket_skill = SkillArtifact(
        name="ticket_schema_rule",
        kind="atomic_tool_rule_card",
        description="Use exact schema names for action requests.",
        body="Use exact schema names for action requests.",
        metadata={"domains": ["TicketAPI"], "allowed_tools": ["create_ticket"], "intent_keywords": ["schema"]},
    )
    travel_skill = SkillArtifact(
        name="travel_schema_rule",
        kind="atomic_tool_rule_card",
        description="Use exact schema names for action requests.",
        body="Use exact schema names for action requests.",
        metadata={"domains": ["TravelAPI"], "allowed_tools": ["book_flight"], "intent_keywords": ["schema"]},
    )
    store = ArtifactStore([travel_skill, ticket_skill])
    audit = store.retrieve_audit(
        "Use exact schema names for action requests.",
        top_k=2,
        debug_context=_bfcl_retrieval_context(task, phase="unit", turn_index=0),
    )
    assert audit["selected"][0]["name"] == "ticket_schema_rule"
    assert "domain:TicketAPI" in audit["selected"][0]["tag_matches"]
    assert "tool:create_ticket" not in audit["selected"][0]["tag_matches"]
    assert "tool:create_ticket" not in audit["context"]["query_tags"]


def test_bfcl_retrieval_context_can_use_runtime_failed_tool_on_previous_observation() -> None:
    task = BenchmarkTask(
        benchmark="bfcl_v3",
        task_id="ticket-task",
        question=[[{"role": "user", "content": "Create a high priority support ticket."}]],
        expected=[["create_ticket(title='Issue', description='Help', priority=4)"]],
        metadata={"involved_classes": ["TicketAPI"]},
    )
    context = _bfcl_retrieval_context(
        task,
        phase="previous_observation",
        turn_index=0,
        tool_name="create_ticket",
    )
    assert "domain:TicketAPI" in context["query_tags"]
    assert "tool:create_ticket" in context["query_tags"]
    assert context["runtime_tools"] == ["create_ticket"]
    assert "expected_tools" not in context


def test_bfcl_retrieval_policy_uses_previous_observation_without_expected_leak() -> None:
    task = BenchmarkTask(
        benchmark="bfcl_v3",
        task_id="ticket-task",
        question=[[{"role": "user", "content": "Create a high priority support ticket."}]],
        expected=[["create_ticket(title='Issue', description='Help', priority=4)"]],
        metadata={"involved_classes": ["TicketAPI"]},
    )
    schema_skill = SkillArtifact(
        name="ticket_priority_schema_rule",
        kind="interface_contract_card",
        description="Use exact create_ticket priority schema names after argument errors.",
        body="If create_ticket reports a parameter error, use title, description, and priority exactly.",
        metadata={
            "domains": ["TicketAPI"],
            "allowed_tools": ["create_ticket"],
            "intent_keywords": ["schema", "argument"],
        },
    )
    store = ArtifactStore([schema_skill])
    policy = _RetrievalPolicy(store=store, task=task, top_k_skills=2, min_skill_score=0.0)

    skills, audit, query = policy.retrieve(
        turn_index=0,
        user_messages=task.question[0],
        observation=_RetrievalObservation(
            tool_name="create_ticket",
            arguments={"issue": "Help"},
            error="missing required parameter title",
        ),
    )

    assert [skill.name for skill in skills] == ["ticket_priority_schema_rule"]
    assert audit["context"]["phase"] == "previous_observation"
    assert "tool:create_ticket" in audit["context"]["query_tags"]
    assert "expected_tools" not in audit["context"]
    assert "tool_error tool=create_ticket" in query


@pytest.mark.asyncio
async def test_run_bfcl_task_error_feedback_uses_step_start_context_update(monkeypatch) -> None:
    class FakeToolClient:
        prompt_tokens = 0
        completion_tokens = 0

        def __init__(self) -> None:
            self.calls = 0
            self.messages_at_ask: list[list[dict[str, Any]]] = []
            self.systems: list[str] = []

        async def ask(self, *, messages, system, tools, temperature, max_request_wall_s):
            self.calls += 1
            self.messages_at_ask.append(list(messages))
            self.systems.append(system)
            if self.calls == 1:
                return _ToolModelResponse(
                    content="",
                    tool_calls=[{"id": "call_1", "name": "create_ticket", "arguments": '{"issue": "Help"}'}],
                    assistant_msg={
                        "role": "assistant",
                        "content": "",
                        "tool_calls": [
                            {
                                "id": "call_1",
                                "type": "function",
                                "function": {"name": "create_ticket", "arguments": '{"issue": "Help"}'},
                            }
                        ],
                    },
                )
            return _ToolModelResponse(
                content="done",
                tool_calls=[],
                assistant_msg={"role": "assistant", "content": "done"},
            )

        def tool_result_message(self, tool_call_id: str, content: str):
            return {"role": "tool", "tool_call_id": tool_call_id, "content": content}

        def total_tokens(self) -> int:
            return 0

    fake_client = FakeToolClient()

    class FakeEnv:
        available = True
        backend_name = "official"

        def __init__(self, *args, **kwargs) -> None:
            pass

        def call(self, name, args):
            return {"error": "missing required parameter title"}, "missing required parameter title"

    monkeypatch.setattr("academic.benchmarks.bfcl.adapter._make_tool_api_client", lambda **kwargs: fake_client)
    monkeypatch.setattr("academic.benchmarks.bfcl.adapter.BFCLOfficialEnvironment", FakeEnv)

    task = BenchmarkTask(
        benchmark="bfcl_v3",
        task_id="ticket-task",
        question=[[{"role": "user", "content": "Create a high priority support ticket."}]],
        expected=[["create_ticket(title='Issue', description='Help', priority=4)"]],
        metadata={"involved_classes": ["TicketAPI"]},
        input_artifacts={"initial_config": {}},
    )
    initial_skill = SkillArtifact(
        name="initial_ticket_hint",
        kind="interface_contract_card",
        description="Create support tickets when asked.",
        body="For TicketAPI support requests, create a ticket.",
        metadata={
            "domains": ["TicketAPI"],
            "allowed_tools": ["create_ticket"],
            "intent_keywords": ["support", "ticket"],
        },
    )
    skill = SkillArtifact(
        name="ticket_priority_schema_rule",
        kind="interface_contract_card",
        description="Use exact create_ticket title and priority schema.",
        body="After create_ticket argument errors, use title, description, and priority exactly.",
        metadata={
            "domains": ["TicketAPI"],
            "allowed_tools": ["create_ticket"],
            "intent_keywords": ["schema", "argument"],
        },
    )
    store = ArtifactStore([initial_skill, skill])
    original_retrieve = store.retrieve_audit

    def fake_retrieve_audit(query, *args, **kwargs):
        audit = original_retrieve(query, *args, **kwargs)
        selected_name = "ticket_priority_schema_rule" if "tool_error" in str(query) else "initial_ticket_hint"
        audit["selected"] = [
            row for row in audit.get("candidates", []) if row.get("name") == selected_name
        ][:1]
        return audit

    monkeypatch.setattr(store, "retrieve_audit", fake_retrieve_audit)

    result = await run_bfcl_task(
        task,
        llm_config="unused",
        model_name="fake",
        tools=[
            {
                "type": "function",
                "function": {
                    "name": "create_ticket",
                    "description": "Create a ticket.",
                    "parameters": {
                        "type": "object",
                        "properties": {"title": {"type": "string"}, "description": {"type": "string"}, "priority": {"type": "integer"}},
                        "required": ["title", "description"],
                    },
                    "x_bfcl_source_file": "ticket_api.json",
                },
            }
        ],
        artifact_store=store,
        top_k_skills=1,
        min_skill_score=0.0,
        max_steps_per_turn=2,
        adapter_mode="official",
        execution_backend="official",
        prompt_style="native",
        tool_api_style="auto",
        skill_injection_mode="prompt_only",
        debug_sink=DebugEventSink(collect_events=True),
    )

    assert result.error is None
    assert fake_client.calls == 2
    assert len(fake_client.messages_at_ask) == 2
    second_messages = fake_client.messages_at_ask[1]
    assert any(
        msg.get("role") == "user" and "Runtime skill retrieval update" in str(msg.get("content", ""))
        and "ticket_priority_schema_rule" in str(msg.get("content", ""))
        for msg in second_messages
    )
    trace_events = result.trace["debug_events"]
    event_types = [event["event_type"] for event in trace_events]
    assert "prompt_reinjection" not in event_types
    assert "prompt_context_update" in event_types
    retrieval_events = [event for event in trace_events if event["event_type"] == "retrieval"]
    assert any(event.get("trigger") == "step_start" for event in retrieval_events)
    assert all(event.get("trigger") != "tool_error_retry" for event in retrieval_events)
    step_start = next(event for event in retrieval_events if event.get("trigger") == "step_start")
    assert step_start["output"]["context"]["phase"] == "previous_observation"
    assert "expected_tools" not in step_start["output"]["context"]
    ordered = [
        (event["event_type"], event.get("trigger"))
        for event in trace_events
        if event["event_type"] in {"executor_step", "tool_call", "tool_result", "retrieval"}
    ]
    expected_subsequence = [
        ("retrieval", "turn_start"),
        ("executor_step", None),
        ("tool_call", None),
        ("tool_result", None),
        ("retrieval", "step_start"),
        ("executor_step", None),
    ]
    cursor = 0
    for item in ordered:
        if item == expected_subsequence[cursor]:
            cursor += 1
            if cursor == len(expected_subsequence):
                break
    assert cursor == len(expected_subsequence)


@pytest.mark.asyncio
async def test_bfcl_baseline_concurrency_runs_tasks_in_parallel(monkeypatch) -> None:
    active = 0
    max_active = 0

    async def fake_run_bfcl_task(task, **kwargs):
        nonlocal active, max_active
        active += 1
        max_active = max(max_active, active)
        await asyncio.sleep(0.02)
        active -= 1
        return BenchmarkResult(
            benchmark="bfcl_v3",
            task_id=task.task_id,
            success=True,
            score=1.0,
            metrics={"official_valid": True, "call_f1": 1.0},
            trace={"task_id": task.task_id},
        )

    monkeypatch.setattr("academic.benchmarks.core.runner.run_bfcl_task", fake_run_bfcl_task)
    tasks = [
        type("Task", (), {"task_id": f"task_{idx}", "as_dict": lambda self: {"task_id": self.task_id}})()
        for idx in range(4)
    ]

    details = await _run_bfcl_baseline(
        tasks,
        1,
        "unused",
        [],
        ArtifactStore(),
        adapter_mode="official",
        concurrency=2,
    )

    assert [row["task_id"] for row in details] == ["task_0", "task_1", "task_2", "task_3"]
    assert max_active == 2


def test_bfcl_skill_predicate_rejects_cross_domain_and_cross_tool_noise() -> None:
    vehicle_task = BenchmarkTask(
        benchmark="bfcl_v3",
        task_id="vehicle-task",
        question=[[{"role": "user", "content": "Fill the fuel tank before my road trip."}]],
        expected=[["setFuelLevel(fuelLevel=100)"]],
        metadata={
            "involved_classes": ["VehicleControlAPI"],
            "path": ["VehicleControlAPI.setFuelLevel"],
        },
    )
    trading_skill = SkillArtifact(
        name="avoid_get_current_time_for_market_status",
        kind="atomic_tool_rule_card",
        description="When verifying market status before trading, skip get_current_time.",
        body="Use get_stock_info and place_order instead of get_current_time.",
        metadata={
            "domains": ["all", "stock_market", "trading"],
            "allowed_tools": ["get_stock_info", "place_order", "get_current_time"],
            "intent_keywords": ["market status", "operational status"],
        },
    )
    vehicle_skill = SkillArtifact(
        name="set_fuel_level_directly",
        kind="atomic_tool_rule_card",
        description="Use setFuelLevel for fuel tank requests.",
        body="Call setFuelLevel when the user asks to fill the tank.",
        metadata={
            "domains": ["VehicleControlAPI"],
            "allowed_tools": ["setFuelLevel"],
            "intent_keywords": ["fuel", "tank"],
        },
    )
    assert _bfcl_skill_matches_task(trading_skill, vehicle_task) is False
    assert _bfcl_skill_matches_turn(
        trading_skill,
        vehicle_task,
        0,
        vehicle_task.question[0],
    ) is False
    assert _bfcl_skill_matches_task(vehicle_skill, vehicle_task) is True


def test_bfcl_skill_predicate_does_not_use_expected_tools_as_filter() -> None:
    task = BenchmarkTask(
        benchmark="bfcl_v3",
        task_id="ticket-task",
        question=[[{"role": "user", "content": "Create a high priority support ticket."}]],
        expected=[["create_ticket(title='Issue', description='Help', priority=4)"]],
        metadata={"involved_classes": ["TicketAPI"]},
    )
    skill = SkillArtifact(
        name="ticket_rule_with_different_tool_provenance",
        kind="atomic_tool_rule_card",
        description="Use precise priority values for tickets.",
        body="Use the priority value requested by the user.",
        metadata={"domains": ["TicketAPI"], "allowed_tools": ["not_the_expected_tool"]},
    )
    assert _bfcl_skill_matches_task(skill, task) is True


def test_turn_skill_constraints_do_not_leak_expected_tools() -> None:
    task = BenchmarkTask(
        benchmark="bfcl_v3",
        task_id="ticket-task",
        question=[[{"role": "user", "content": "Create a high priority support ticket."}]],
        expected=[["create_ticket(title='Issue', description='Help', priority=4)"]],
        metadata={"involved_classes": ["TicketAPI"]},
    )
    skill = SkillArtifact(
        name="ticket_priority_rule",
        kind="atomic_tool_rule_card",
        description="Use precise priority values for tickets.",
        body="Use the priority value requested by the user.",
        metadata={"domains": ["TicketAPI"], "allowed_tools": ["create_ticket"]},
    )
    constraints = _turn_skill_constraints([skill], task, 0)
    assert "Expected tool focus" not in constraints
    assert "create_ticket" not in constraints


def test_native_skill_system_explicitly_tells_model_to_ignore_irrelevant_skills() -> None:
    text = _native_skill_system("### skill_x\nUse x.")
    assert "ignore it completely" in text
    assert "unrelated to the current tool family" in text


def test_bfcl_refine_rolls_back_regressed_skill_version_when_history_exists() -> None:
    store = ArtifactStore()
    original = SkillArtifact(
        name="bfcl_schema_parameter_names",
        kind="atomic_tool_rule_card",
        description="Use exact schema names.",
        body="Use booking_id exactly when the schema requires booking_id.",
        metadata={"source": "evolve_rollouts", "injection_type": "informational"},
    )
    store.add(original)
    broken = SkillArtifact(
        name="bfcl_schema_parameter_names",
        kind="atomic_tool_rule_card",
        description="Broken alias rule.",
        body="For invoice and support calls, use reservation_id instead of booking_id.",
        metadata={"source": "evolve_rollouts", "injection_type": "informational"},
    )
    store.add(broken)
    train_details = [
        {
            "task_id": "task-1",
            "runs": [{"metrics": {"official_valid": True}}],
        }
    ]
    replay_details = [
        {
            "task_id": "task-1",
            "runs": [
                {
                    "metrics": {
                        "official_valid": False,
                        "retrieved_skills": ["bfcl_schema_parameter_names"],
                        "prompt_injected_skills": ["bfcl_schema_parameter_names"],
                        "tool_injected_skills": [],
                        "called_skill_tools": [],
                        "used_skills": ["bfcl_schema_parameter_names"],
                    }
                }
            ],
        }
    ]
    _build_bfcl_skill_bundles(store, train_details=train_details, replay_details=replay_details)
    maintenance = _build_maintenance_test_results(
        store,
        train_details=train_details,
        replay_details=replay_details,
        run_label="unit",
    )
    decisions = _refine_bfcl_skill_store(
        store,
        train_details=train_details,
        replay_details=replay_details,
        maintenance_test_results=maintenance,
        min_fail_count=1,
    )
    repaired = next(skill for skill in store.all() if skill.name == "bfcl_schema_parameter_names")
    assert repaired.version == 1
    assert repaired.is_disabled() is False
    row = next(item for item in decisions if item["skill_name"] == "bfcl_schema_parameter_names")
    assert row["action"] == "rollback_on_regression"


def test_bfcl_refine_rolls_back_repeatedly_harmful_skill_without_help() -> None:
    store = ArtifactStore()
    original = SkillArtifact(
        name="bfcl_multi_action_turn_completion",
        kind="planning_card",
        description="Complete all required tool calls before ending a turn.",
        body="Keep calling tools until all requested actions are complete.",
        metadata={"source": "evolve_rollouts", "injection_type": "workflow"},
    )
    store.add(original)
    broken = SkillArtifact(
        name="bfcl_multi_action_turn_completion",
        kind="planning_card",
        description="Broken stop-early rule.",
        body="End the turn after the first relevant tool call.",
        metadata={"source": "evolve_rollouts", "injection_type": "workflow"},
    )
    store.add(broken)
    train_details = [
        {"task_id": "task-a", "runs": [{"metrics": {"official_valid": False}}]},
        {"task_id": "task-b", "runs": [{"metrics": {"official_valid": False}}]},
    ]
    replay_details = [
        {
            "task_id": "task-a",
            "runs": [
                {
                    "metrics": {
                        "official_valid": False,
                        "retrieved_skills": ["bfcl_multi_action_turn_completion"],
                        "prompt_injected_skills": ["bfcl_multi_action_turn_completion"],
                        "tool_injected_skills": [],
                        "called_skill_tools": [],
                        "used_skills": ["bfcl_multi_action_turn_completion"],
                    }
                }
            ],
        },
        {
            "task_id": "task-b",
            "runs": [
                {
                    "metrics": {
                        "official_valid": False,
                        "retrieved_skills": ["bfcl_multi_action_turn_completion"],
                        "prompt_injected_skills": ["bfcl_multi_action_turn_completion"],
                        "tool_injected_skills": [],
                        "called_skill_tools": [],
                        "used_skills": ["bfcl_multi_action_turn_completion"],
                    }
                }
            ],
        },
    ]
    _build_bfcl_skill_bundles(store, train_details=train_details, replay_details=replay_details)
    maintenance = _build_maintenance_test_results(
        store,
        train_details=train_details,
        replay_details=replay_details,
        run_label="unit",
    )
    decisions = _refine_bfcl_skill_store(
        store,
        train_details=train_details,
        replay_details=replay_details,
        maintenance_test_results=maintenance,
        min_fail_count=1,
    )
    repaired = next(skill for skill in store.all() if skill.name == "bfcl_multi_action_turn_completion")
    assert repaired.version == 1
    row = next(item for item in decisions if item["skill_name"] == "bfcl_multi_action_turn_completion")
    assert row["action"] == "rollback_on_no_help"


def test_bfcl_refine_disables_manual_fault_without_harming_unchanged_copresent_skills() -> None:
    good = SkillArtifact(
        name="bfcl_state_id_reuse",
        kind="functional_workflow_card",
        description="Reuse ids from previous turns.",
        body="Reuse exact ids from tool outputs and prior turns.",
        metadata={"injection_type": "workflow"},
    )
    bad = SkillArtifact(
        name="bad_cancel_order_143",
        kind="workflow_guardrail_card",
        description="Broken task-specific cancel rule.",
        body="Do not call cancel_order for the reviewed order.",
        metadata={
            "manual_fault_injected": True,
            "injection_type": "workflow",
            "intent_keywords": ["cancel", "order", "reviewed"],
            "source_task_ids": ["task-143"],
        },
    )
    store = ArtifactStore([good, bad])
    train_details = [
        {
            "task_id": "task-143",
            "runs": [{"metrics": {"official_valid": True}}],
        }
    ]
    replay_details = [
        {
            "task_id": "task-143",
            "runs": [
                {
                    "metrics": {
                        "official_valid": False,
                        "retrieved_skills": ["bfcl_state_id_reuse", "bad_cancel_order_143"],
                        "prompt_injected_skills": ["bfcl_state_id_reuse", "bad_cancel_order_143"],
                        "tool_injected_skills": [],
                        "called_skill_tools": [],
                        "used_skills": [],
                    }
                }
            ],
        }
    ]
    _build_bfcl_skill_bundles(store, train_details=train_details, replay_details=replay_details)
    maintenance = _build_maintenance_test_results(
        store,
        train_details=train_details,
        replay_details=replay_details,
        run_label="unit",
    )
    decisions = _refine_bfcl_skill_store(
        store,
        train_details=train_details,
        replay_details=replay_details,
        maintenance_test_results=maintenance,
        min_fail_count=1,
    )
    bad_after = next(skill for skill in store.all() if skill.name == "bad_cancel_order_143")
    good_after = next(skill for skill in store.all() if skill.name == "bfcl_state_id_reuse")
    assert bad_after.is_disabled() is True
    assert good_after.is_disabled() is False
    bad_row = next(item for item in decisions if item["skill_name"] == "bad_cancel_order_143")
    good_row = next(item for item in decisions if item["skill_name"] == "bfcl_state_id_reuse")
    assert bad_row["action"] == "disable_on_regression"
    assert good_row["action"] == "keep"


def test_artifact_store_separates_bundle_and_test_results(tmp_path: Path) -> None:
    artifact = SkillArtifact(
        name="bfcl_state_rule",
        kind="workflow_guardrail_card",
        description="Reuse ids",
        body="Reuse ids from earlier turns.",
    )
    artifact.bundle.positive_cases = [
        SkillBundleCase(
            case_id="bfcl_state_rule:positive:0",
            source="manual",
            prompt="reuse booking id",
            expected={"official_valid": True},
        )
    ]
    store = ArtifactStore([artifact])
    store.add_test_result(
        SkillTestResult(
            result_id="res1",
            skill_name="bfcl_state_rule",
            skill_version=1,
            bundle_id=artifact.bundle.bundle_id,
            bundle_version=artifact.bundle.bundle_version,
            run_label="unit",
        )
    )
    skill_path = tmp_path / "skills.json"
    result_path = tmp_path / "results.json"
    store.save(skill_path)
    store.save_test_results(result_path)
    loaded = ArtifactStore.load(skill_path)
    loaded_results = ArtifactStore.load_test_results(result_path)
    assert len(loaded.all()) == 1
    assert loaded.all()[0].bundle.positive_cases[0].case_id == "bfcl_state_rule:positive:0"
    assert loaded.test_results() == []
    assert len(loaded_results) == 1
    assert loaded_results[0].result_id == "res1"


def test_artifact_store_marks_dependents_stale_on_minor_update() -> None:
    upstream = SkillArtifact(
        name="shared_rule",
        kind="rule_card",
        description="Shared rule",
        body="Common rule text.",
        metadata={"version_kind": "seed"},
    )
    downstream = SkillArtifact(
        name="consumer_rule",
        kind="rule_card",
        description="Consumer",
        body="See shared_rule before acting.",
        metadata={"dependencies": ["shared_rule"], "version_kind": "seed"},
    )
    store = ArtifactStore([upstream, downstream])
    updated = SkillArtifact(
        name="shared_rule",
        kind="rule_card",
        description="Shared rule improved",
        body="Common rule text improved.",
        metadata={"version_kind": "minor"},
    )
    store.add(updated)
    refreshed = next(skill for skill in store.all() if skill.name == "consumer_rule")
    assert refreshed.stale is True
    assert refreshed.status == "stale"
    assert refreshed.metadata["stale_due_to"]["version_kind"] == "minor"


def test_minor_update_cannot_drop_existing_cases() -> None:
    artifact = SkillArtifact(
        name="rule",
        kind="rule_card",
        description="Rule",
        body="Rule body",
        metadata={"version_kind": "seed"},
    )
    artifact.bundle.positive_cases = [
        SkillBundleCase(
            case_id="rule:positive:0",
            source="manual",
            prompt="case one",
            expected={"official_valid": True},
        )
    ]
    store = ArtifactStore([artifact])
    updated = SkillArtifact(
        name="rule",
        kind="rule_card",
        description="Rule updated",
        body="Rule body updated",
        metadata={"version_kind": "minor"},
    )
    store.add(updated)
    latest = next(skill for skill in store.all() if skill.name == "rule")
    latest.bundle.positive_cases = []
    try:
        _validate_refine_output_consistency(latest)
    except ValueError as exc:
        assert "removed existing tests" in str(exc)
    else:
        raise AssertionError("Expected minor update consistency check to fail")


@pytest.mark.asyncio
async def test_execute_bfcl_bundle_tests_enforces_strict_contract_gate() -> None:
    artifact = SkillArtifact(
        name="strict_contract_skill",
        kind="atomic_tool_rule_card",
        description="Use exact schema name.",
        body="Call remove_stock_from_watchlist with symbol, not stock.",
    )
    artifact.bundle.positive_cases = [
        SkillBundleCase(
            case_id="strict_contract_skill:positive:0",
            source="manual",
            prompt="remove NVDA from watchlist",
            expected={
                "official_valid": True,
                "tool_calls": [
                    {
                        "name": "remove_stock_from_watchlist",
                        "arguments": {"stock": "NVDA"},
                    }
                ],
            },
            context={
                "task_fragment": {
                    "task_id": "strict-contract-case",
                    "question": [[{"role": "user", "content": "Remove NVDA from my watchlist."}]],
                    "expected": [[{"name": "remove_stock_from_watchlist", "arguments": {"stock": "NVDA"}}]],
                    "input_artifacts": {},
                    "metadata": {},
                }
            },
            polarity="positive",
        )
    ]
    result = await execute_bfcl_bundle_tests(
        artifact,
        tools=[],
        llm_config="dummy",
        model_name=None,
        adapter_mode="official",
        execution_backend="official",
        prompt_style="official",
        tool_api_style="openai",
        max_steps_per_turn=1,
        temperature=0.0,
        synthetic_continue=False,
        explicit_skill_tool=False,
        max_case_seconds=0.01,
    )
    assert result.aggregate["strict_contract_gate"] is True
    assert result.aggregate["n_strict_failures"] >= 1
    assert result.aggregate["pass_all_tests"] is False
    with_run = next(run for run in result.unit_case_runs if run.variant == "with_skill")
    assert with_run.passed is False
    assert with_run.metadata["contract_failures"]


@pytest.mark.asyncio
async def test_execute_bfcl_bundle_tests_runs_with_without_variants_concurrently(monkeypatch) -> None:
    import academic.benchmarks.bfcl.maintenance.adapter as adapter

    artifact = SkillArtifact(
        name="parallel_bundle_skill",
        kind="atomic_tool_rule_card",
        description="parallel bundle",
        body="body",
    )
    artifact.bundle.positive_cases = [
        SkillBundleCase(
            case_id="parallel_bundle_skill:positive:0",
            source="manual",
            prompt="do a thing",
            expected={"official_valid": True},
            context={
                "task_fragment": {
                    "task_id": "parallel-case",
                    "question": [[{"role": "user", "content": "Do a thing."}]],
                    "expected": [[]],
                    "input_artifacts": {"initial_config": {"X": {}}},
                    "metadata": {"involved_classes": ["X"]},
                }
            },
            polarity="positive",
        )
    ]
    monkeypatch.setenv("BFCL_BUNDLE_VARIANT_CONCURRENCY", "2")
    monkeypatch.setattr(adapter, "_validate_bundle_task_fragment", lambda **kwargs: None)
    active = 0
    max_active = 0
    modes: List[str] = []

    async def fake_run_case_with_timeout(task, **kwargs):
        nonlocal active, max_active
        active += 1
        max_active = max(max_active, active)
        modes.append(kwargs["skill_injection_mode"])
        await asyncio.sleep(0.01)
        active -= 1
        return BenchmarkResult(
            benchmark="bfcl_v3",
            task_id=task.task_id,
            success=True,
            score=1.0,
            metrics={"official_valid": True, "call_f1": 1.0, "call_errors": []},
            trace={"tool_calls": []},
        )

    monkeypatch.setattr(adapter, "_run_case_with_timeout", fake_run_case_with_timeout)

    result = await execute_bfcl_bundle_tests(
        artifact,
        tools=[],
        llm_config="dummy",
        model_name=None,
        adapter_mode="official",
        execution_backend="official",
        prompt_style="official",
        tool_api_style="openai",
        max_steps_per_turn=1,
        temperature=0.0,
        synthetic_continue=False,
        explicit_skill_tool=False,
        max_case_seconds=1.0,
    )

    assert max_active == 2
    assert set(modes) == {"none", "prompt_only"}
    assert [run.variant for run in result.unit_case_runs] == ["without_skill", "with_skill"]


def test_trim_bundle_cases_keeps_multiple_cases_up_to_limit() -> None:
    artifact = SkillArtifact(
        name="trim_skill",
        kind="rule_card",
        description="bundle trim",
        body="body",
    )
    artifact.bundle.positive_cases = [
        SkillBundleCase(case_id=f"trim_skill:positive:{idx}", source="manual", prompt=f"p{idx}")
        for idx in range(4)
    ]
    artifact.bundle.negative_cases = [
        SkillBundleCase(case_id=f"trim_skill:negative:{idx}", source="manual", prompt=f"n{idx}")
        for idx in range(3)
    ]
    artifact.bundle.integration_cases = [
        SkillBundleCase(case_id=f"trim_skill:integration:{idx}", source="manual", prompt=f"i{idx}")
        for idx in range(5)
    ]
    changed = trim_bundle_cases(artifact, per_polarity_limit=2)
    assert changed is True
    assert len(artifact.bundle.positive_cases) == 2
    assert len(artifact.bundle.negative_cases) == 2
    assert len(artifact.bundle.integration_cases) == 2


def test_trim_bundle_cases_enforces_total_cap_balanced_across_groups(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("BFCL_BUNDLE_MAX_TOTAL_CASES", "4")
    artifact = SkillArtifact(
        name="trim_total_skill",
        kind="rule_card",
        description="bundle trim total",
        body="body",
    )
    artifact.bundle.positive_cases = [
        SkillBundleCase(case_id=f"trim_total_skill:positive:{idx}", source="manual", prompt=f"p{idx}")
        for idx in range(3)
    ]
    artifact.bundle.negative_cases = [
        SkillBundleCase(case_id=f"trim_total_skill:negative:{idx}", source="manual", prompt=f"n{idx}")
        for idx in range(3)
    ]
    artifact.bundle.integration_cases = [
        SkillBundleCase(case_id=f"trim_total_skill:integration:{idx}", source="manual", prompt=f"i{idx}")
        for idx in range(3)
    ]
    changed = trim_bundle_cases(artifact, per_polarity_limit=3)
    assert changed is True
    assert len(artifact.bundle.all_cases()) == 4
    assert len(artifact.bundle.positive_cases) >= 1
    assert len(artifact.bundle.negative_cases) >= 1
    assert len(artifact.bundle.integration_cases) >= 1
    assert artifact.bundle.fixtures["bundle_trimmed"] is True


def test_latest_matching_test_result_reuses_identical_bundle_signature() -> None:
    artifact = SkillArtifact(
        name="cache_skill",
        kind="rule_card",
        description="desc",
        body="body",
    )
    artifact.bundle.positive_cases = [
        SkillBundleCase(case_id="cache_skill:positive:0", source="manual", prompt="p0")
    ]
    signature = _bundle_test_signature(
        artifact,
        max_steps_per_turn=2,
        adapter_mode="official",
        execution_backend="official",
        prompt_style="official",
        tool_api_style="openai",
        synthetic_continue=False,
        explicit_skill_tool=False,
    )
    result = SkillTestResult(
        result_id="cache_skill:bundle:abc",
        skill_name="cache_skill",
        skill_version=artifact.version,
        bundle_id=artifact.bundle.bundle_id,
        bundle_version=artifact.bundle.bundle_version,
        dependency_versions=artifact.dependency_version_map(),
        run_label="llm_bundle_unit",
        aggregate={"pass_all_tests": True, "test_signature": signature},
    )
    store = ArtifactStore([artifact], test_results=[result])
    reused = _latest_matching_test_result(store, artifact=artifact, test_signature=signature)
    assert reused is not None
    assert reused.result_id == result.result_id


def test_task_from_case_marks_synthetic_bundle_case_invalid() -> None:
    case = SkillBundleCase(
        case_id="bad:0",
        source="manual",
        prompt="bad bundle",
        expected={},
        context={
            "source_task_id": "task_from_trace",
            "task_fragment": {
                "task_id": "task_from_trace",
                "question": [[{"role": "user", "content": "Remove NVDA from my watchlist"}]],
                "expected": [["remove_stock_from_watchlist(stock='NVDA')"]],
                "input_artifacts": {},
                "metadata": {},
            },
        },
        polarity="positive",
    )
    task = _task_from_case(case)
    assert task is not None
    assert task.metadata["_bundle_case_invalid"]["reason"] == "non_replayable_synthetic_task_id"


def test_task_from_case_rejects_schema_mismatched_fragment_against_source_task() -> None:
    case = SkillBundleCase(
        case_id="bad-schema:0",
        source="manual",
        prompt="bad bundle",
        expected={},
        context={
            "source_task_id": "multi_turn_base_116",
            "task_fragment": {
                "task_id": "multi_turn_base_116",
                "question": [[{"role": "user", "content": "I'm inclined to shake things up a bit. Let's take Zeta Corp out of the equation from my watchlist, shall we?"}]],
                "expected": [["remove_stock_from_watchlist(stock='ZETA')"]],
                "input_artifacts": {},
                "metadata": {},
            },
        },
        polarity="positive",
    )
    task = _task_from_case(case)
    assert task is not None
    assert task.metadata["_bundle_case_invalid"]["reason"] == "unknown_expected_tool_argument"


def test_task_from_case_rejects_official_fragment_with_wrong_argument_name() -> None:
    case = SkillBundleCase(
        case_id="bad-arg:0",
        source="manual",
        prompt="bad bundle",
        expected={},
        context={
            "source_task_id": "multi_turn_base_116",
            "task_fragment": {
                "task_id": "multi_turn_base_116",
                "question": [[{"role": "user", "content": "I'm inclined to shake things up a bit. Let's take Zeta Corp out of the equation from my watchlist, shall we?"}]],
                "expected": [["remove_stock_from_watchlist(stock='ZETA')"]],
                "input_artifacts": {"initial_config": {"TradingBot": {"authenticated": True}}},
                "metadata": {"involved_classes": ["TradingBot"]},
            },
        },
        polarity="positive",
    )
    task = _task_from_case(case)
    assert task is not None
    assert task.metadata["_bundle_case_invalid"]["reason"] == "unknown_expected_tool_argument"
    assert task.metadata["_bundle_case_invalid"]["unknown_arguments"] == ["stock"]


def test_task_from_case_rejects_expected_placeholder_even_when_source_shape_matches() -> None:
    case = SkillBundleCase(
        case_id="bad-placeholder:0",
        source="manual",
        prompt="bad bundle",
        expected={},
        context={
            "source_task_id": "multi_turn_base_120",
            "task_fragment": {
                "task_id": "multi_turn_base_120",
                "question": [[{"role": "user", "content": "After confirming the market's operational status, proceed to purchase 100 Apple shares at the prevailing market price."}]],
                "expected": [[
                    "get_stock_info(symbol='AAPL')",
                    "place_order(order_type='Buy',symbol='AAPL',price=<market_price>,amount=100)",
                ]],
                "input_artifacts": {"initial_config": {"TradingBot": {"authenticated": True}}},
                "metadata": {"involved_classes": ["TradingBot"]},
            },
        },
        polarity="positive",
    )
    task = _task_from_case(case)
    assert task is not None
    assert task.metadata["_bundle_case_invalid"]["reason"] == "invalid_expected_call_syntax"


def test_expected_call_schema_rejects_non_literal_placeholder() -> None:
    schema = {
        "place_order": {
            "function": {
                "name": "place_order",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "order_type": {"type": "string"},
                        "symbol": {"type": "string"},
                        "price": {"type": "number"},
                        "amount": {"type": "integer"},
                    },
                },
            }
        }
    }
    invalid = _validate_expected_call_schema(
        [["place_order(order_type='Buy', symbol='AAPL', price=<market_price>, amount=100)"]],
        tool_schemas=schema,
    )
    assert invalid is not None
    assert invalid["reason"] == "invalid_expected_call_syntax"


def test_task_from_case_accepts_official_prefix_fragment() -> None:
    case = SkillBundleCase(
        case_id="prefix-ok:0",
        source="manual",
        prompt="prefix bundle",
        expected={},
        context={
            "source_task_id": "multi_turn_base_116",
            "task_fragment": {
                "task_id": "multi_turn_base_116",
                "question": [[{"role": "user", "content": "Could you peruse my stock watchlist and share what's on my radar right now, please?"}]],
                "expected": [["get_watchlist()"]],
                "input_artifacts": {
                    "initial_config": {
                        "TradingBot": {"authenticated": True}
                    }
                },
                "metadata": {"involved_classes": ["TradingBot"]},
            },
        },
        polarity="positive",
    )
    task = _task_from_case(case)
    assert task is not None
    assert "_bundle_case_invalid" not in task.metadata


def test_bundle_contract_match_task_expected_rejects_extra_calls_even_when_official_valid() -> None:
    task = BenchmarkTask(
        benchmark="bfcl_v3",
        task_id="strict_case",
        question=[[{"role": "user", "content": "Buy Apple shares."}]],
        expected=[[
            "get_stock_info(symbol='AAPL')",
            "place_order(order_type='Buy', symbol='AAPL', price=227.16, amount=100)",
        ]],
        input_artifacts={"initial_config": {"TradingBot": {"authenticated": True}}},
        metadata={"involved_classes": ["TradingBot"]},
    )
    case = SkillBundleCase(
        case_id="strict:negative:0",
        source="credit_assigner_negative",
        prompt="No extra lookup calls.",
        expected={"match_task_expected": True, "official_valid": True},
        context={},
        polarity="negative",
    )
    result = SimpleNamespace(
        metrics={"official_valid": True},
        trace={
            "tool_calls": [
                {"name": "get_current_time", "arguments": {}, "turn_index": 0},
                {"name": "get_stock_info", "arguments": {"symbol": "AAPL"}, "turn_index": 0},
                {
                    "name": "place_order",
                    "arguments": {"order_type": "Buy", "symbol": "AAPL", "price": 227.16, "amount": 100},
                    "turn_index": 0,
                },
            ]
        },
    )

    contract = _run_contract_assertions(case=case, result=result, task=task)

    assert contract["passed"] is False
    assert any(failure["type"] == "task_expected_call_count_mismatch" for failure in contract["failures"])


@pytest.mark.asyncio
async def test_execute_bfcl_bundle_tests_rejects_invalid_bundle_context_before_runtime() -> None:
    artifact = SkillArtifact(
        name="invalid_bundle_skill",
        kind="atomic_tool_rule_card",
        description="invalid bundle",
        body="body",
    )
    artifact.bundle.positive_cases = [
        SkillBundleCase(
            case_id="invalid_bundle_skill:positive:0",
            source="manual",
            prompt="bad case",
            expected={"tool_calls": [{"name": "remove_stock_from_watchlist", "arguments": {"stock": "NVDA"}}]},
            context={
                "source_task_id": "task_from_trace",
                "task_fragment": {
                    "task_id": "task_from_trace",
                    "question": [[{"role": "user", "content": "Remove NVDA from my watchlist"}]],
                    "expected": [["remove_stock_from_watchlist(stock='NVDA')"]],
                    "input_artifacts": {},
                    "metadata": {},
                },
            },
            polarity="positive",
        )
    ]
    result = await execute_bfcl_bundle_tests(
        artifact,
        tools=[],
        llm_config="dummy",
        model_name=None,
        adapter_mode="official",
        execution_backend="official",
        prompt_style="official",
        tool_api_style="openai",
        max_steps_per_turn=1,
        temperature=0.0,
        synthetic_continue=False,
        explicit_skill_tool=False,
        max_case_seconds=0.01,
    )
    assert result.aggregate["pass_all_tests"] is False
    assert result.aggregate["n_strict_failures"] >= 1
    with_run = next(run for run in result.unit_case_runs if run.variant == "with_skill")
    assert with_run.metadata["bundle_case_invalid"]["reason"] == "non_replayable_synthetic_task_id"


@pytest.mark.asyncio
async def test_execute_bfcl_bundle_tests_rejects_placeholder_fragment_before_runtime() -> None:
    artifact = SkillArtifact(
        name="placeholder_bundle_skill",
        kind="atomic_tool_rule_card",
        description="placeholder bundle",
        body="body",
    )
    artifact.bundle.negative_cases = [
        SkillBundleCase(
            case_id="placeholder_bundle_skill:negative:0",
            source="manual",
            prompt="bad placeholder case",
            expected={"official_valid": False},
            context={
                "source_task_id": "manual_negative",
                "task_fragment": {
                    "task_id": "manual_negative",
                    "question": [[{"role": "user", "content": "Purchase 75 shares of Tesla at the current price."}]],
                    "expected": [[
                        "get_symbol_by_name(company_name='Tesla')",
                        "get_stock_info(symbol='TSLA')",
                        "place_order(order_type='Buy', symbol='TSLA', price=<market_price>, amount=75)",
                    ]],
                    "input_artifacts": {"initial_config": {"TradingBot": {"authenticated": True}}},
                    "metadata": {"involved_classes": ["TradingBot"]},
                },
            },
            polarity="negative",
        )
    ]
    result = await execute_bfcl_bundle_tests(
        artifact,
        tools=[],
        llm_config="dummy",
        model_name=None,
        adapter_mode="official",
        execution_backend="official",
        prompt_style="official",
        tool_api_style="openai",
        max_steps_per_turn=1,
        temperature=0.0,
        synthetic_continue=False,
        explicit_skill_tool=False,
        max_case_seconds=0.01,
    )
    assert result.aggregate["pass_all_tests"] is False
    with_run = next(run for run in result.unit_case_runs if run.variant == "with_skill")
    assert with_run.metadata["bundle_case_invalid"]["reason"] == "fragment_mismatch_with_source_task"


def test_validate_refine_output_consistency_rejects_synthetic_bfcl_bundle_case() -> None:
    artifact = SkillArtifact(
        name="bad_bundle_rule",
        kind="rule_card",
        description="Rule",
        body="Rule body",
        metadata={"version_kind": "minor"},
    )
    artifact.history = [{"bundle": {"positive_cases": [], "negative_cases": [], "integration_cases": []}}]
    artifact.lineage.version_kind = "minor"
    artifact.interface.summary = "summary"
    artifact.bundle.bundle_id = "bad_bundle_rule.bundle"
    artifact.bundle.positive_cases = [
        SkillBundleCase(
            case_id="bad_bundle_rule:positive:0",
            source="manual",
            prompt="bad",
            context={
                "source_task_id": "task_from_trace",
                "task_fragment": {
                    "task_id": "task_from_trace",
                    "question": [[{"role": "user", "content": "x"}]],
                    "expected": [["foo()"]],
                    "input_artifacts": {},
                    "metadata": {},
                },
            },
        )
    ]
    try:
        _validate_refine_output_consistency(artifact)
    except ValueError as exc:
        assert "non-replayable synthetic task id" in str(exc)
    else:
        raise AssertionError("Expected synthetic BFCL bundle case to be rejected")


def test_bfcl_maintenance_result_reports_with_without_deltas() -> None:
    artifact = SkillArtifact(
        name="bfcl_state_rule",
        kind="workflow_guardrail_card",
        description="Reuse ids",
        body="Reuse ids from earlier turns.",
        metadata={"source": "evolve_rollouts"},
    )
    artifact.metadata["source_task_ids"] = ["task-1"]
    store = ArtifactStore([artifact])
    train_details = [
        {
            "task_id": "task-1",
            "runs": [
                {
                    "metrics": {
                        "official_valid": False,
                        "total_tokens": 40,
                        "n_model_steps": 3,
                    }
                }
            ],
        }
    ]
    replay_details = [
        {
            "task_id": "task-1",
            "runs": [
                {
                    "metrics": {
                        "official_valid": True,
                        "total_tokens": 32,
                        "n_model_steps": 2,
                        "retrieved_skills": ["bfcl_state_rule"],
                        "prompt_injected_skills": ["bfcl_state_rule"],
                        "tool_injected_skills": [],
                        "called_skill_tools": [],
                        "used_skills": ["bfcl_state_rule"],
                    }
                }
            ],
        }
    ]
    _build_bfcl_skill_bundles(store, train_details=train_details, replay_details=replay_details)
    results = _build_maintenance_test_results(
        store,
        train_details=train_details,
        replay_details=replay_details,
        run_label="train_refine",
    )
    assert len(results) == 1
    aggregate = results[0].aggregate
    assert aggregate["n_improved"] == 1
    assert aggregate["n_regressed"] == 0
    assert aggregate["unit_utility_report"]["delta_accuracy"] == 1.0
    assert aggregate["unit_utility_report"]["delta_tokens"] == -8
    assert aggregate["unit_utility_report"]["delta_steps"] == -1


def test_skill_test_result_uses_train_vs_post_refine_not_heldout_test() -> None:
    artifact = SkillArtifact(
        name="bfcl_state_rule",
        kind="workflow_guardrail_card",
        description="Reuse ids",
        body="Reuse ids from earlier turns.",
        metadata={"source": "evolve_rollouts"},
    )
    artifact.bundle.positive_cases = [
        SkillBundleCase(
            case_id="bfcl_state_rule:positive:0",
            source="train_positive",
            prompt="task-1",
            expected={},
            context={"task_id": "task-1"},
        )
    ]
    train_details = [
        {
            "task_id": "task-1",
            "runs": [
                {
                    "metrics": {
                        "official_valid": False,
                        "total_tokens": 20,
                        "n_model_steps": 2,
                    }
                }
            ],
        }
    ]
    post_refine_details = [
        {
            "task_id": "task-1",
            "runs": [
                {
                    "metrics": {
                        "official_valid": True,
                        "total_tokens": 18,
                        "n_model_steps": 1,
                        "retrieved_skills": ["bfcl_state_rule"],
                        "prompt_injected_skills": ["bfcl_state_rule"],
                        "tool_injected_skills": [],
                        "called_skill_tools": [],
                        "used_skills": ["bfcl_state_rule"],
                    }
                }
            ],
        }
    ]
    heldout_test_details = [
        {
            "task_id": "task-999",
            "runs": [
                {
                    "metrics": {
                        "official_valid": False,
                        "total_tokens": 999,
                        "n_model_steps": 9,
                        "retrieved_skills": ["bfcl_state_rule"],
                        "prompt_injected_skills": ["bfcl_state_rule"],
                        "tool_injected_skills": [],
                        "called_skill_tools": [],
                        "used_skills": ["bfcl_state_rule"],
                    }
                }
            ],
        }
    ]
    result = _build_skill_test_result(
        artifact,
        train_details=train_details,
        replay_details=post_refine_details,
        run_label="post_refine",
    )
    assert result.aggregate["n_improved"] == 1
    assert result.aggregate["unit_utility_report"]["delta_tokens"] == -2
    assert all(run.trace_ref != "task-999" for run in result.unit_case_runs)


def test_historical_bfcl_train_details_produce_nonempty_maintenance_assets() -> None:
    train_path = Path(
        "/home/lixujun/skill_evolving/academic/results/"
        "bfcl_v3_glm47_official_tracecheck_evolve_3x3_partial_train.json"
    )
    if not train_path.exists():
        raise AssertionError(f"Missing historical BFCL fixture: {train_path}")
    details = _load_saved_details(train_path)
    results = [_result_from_dict(run) for item in details for run in item.get("runs", [])]
    tools = load_bfcl_tools(Path("/home/lixujun/skill_evolving/data/benchmarks/bfcl_v3"), data_source="bfcl_eval_bundle")
    artifacts = extract_bfcl_skills_from_results(results, tool_schemas=tools)
    assert artifacts
    store = ArtifactStore(artifacts)
    _build_bfcl_skill_bundles(store, train_details=details, replay_details=details)
    maintenance = _build_maintenance_test_results(
        store,
        train_details=details,
        replay_details=details,
        run_label="historical",
    )
    assert maintenance
    assert any(item.bundle_id for item in maintenance)
    assert any(len(skill.bundle.all_cases()) > 0 for skill in store.all())
