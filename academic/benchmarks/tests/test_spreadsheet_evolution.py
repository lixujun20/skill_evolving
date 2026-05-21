import copy
import os
from pathlib import Path
import subprocess
import sys
from typing import Any, Dict, List

import openpyxl

from academic.benchmarks.core.artifacts import ArtifactStore
from academic.benchmarks.core.evolution import OnlineSkillEvolutionRunner
from academic.benchmarks.core.llm_text import TextLLMResponse
from academic.benchmarks.core.maintenance_adapter import MaintenanceRunConfig
from academic.benchmarks.core.types import BenchmarkResult, BenchmarkTask, SkillArtifact, SkillBundleCase, SkillInterface, SkillTestResult
from academic.benchmarks.spreadsheet.adapter import (
    SPREADSHEET_DONE_PATTERN,
    SpreadsheetMaintenanceAdapter,
    _called_spreadsheet_skill_functions,
    _execute_spreadsheet_bundle_tests,
    _refine_spreadsheet_skill_from_credit,
    _spreadsheet_result_projection,
    _spreadsheet_skill_projection,
    _spreadsheet_trace_projection,
    _answer_range_refs,
    _is_spreadsheet_callable_skill,
    _is_spreadsheet_package_skill,
    _coerce_spreadsheet_artifact,
    _write_spreadsheet_skill_library,
    _write_spreadsheet_skill_packages,
    run_spreadsheet_task,
    run_spreadsheet_task_bash_react,
    run_spreadsheet_task_notebook,
)
from academic.benchmarks.spreadsheet.maintenance.adapter import (
    _ask_spreadsheet_text_role,
    _extract_spreadsheet_skills_from_detail,
    _spreadsheet_reusable_edit_steps,
    _spreadsheet_extraction_role_json_block,
    _spreadsheet_extraction_user_payload,
    _spreadsheet_extract_limits,
    _spreadsheet_extraction_rubric_failures,
)
from academic.benchmarks.spreadsheet.executor import _retrieve_spreadsheet_skills


def test_spreadsheet_adapter_facade_points_to_maintenance_adapter() -> None:
    from academic.benchmarks.spreadsheet.maintenance.adapter import SpreadsheetMaintenanceAdapter as MaintenanceAdapter

    assert SpreadsheetMaintenanceAdapter is MaintenanceAdapter
    assert callable(_execute_spreadsheet_bundle_tests)
    assert callable(run_spreadsheet_task)


def _xlsx(path: Path, *, source: int = 7, answer: int | None = None) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = source
    ws["B1"] = answer
    wb.save(path)


def _task(tmp_path: Path, task_id: str, question: str, *, source: int = 7, answer: int = 14) -> BenchmarkTask:
    input_xlsx = tmp_path / f"{task_id}_init.xlsx"
    golden_xlsx = tmp_path / f"{task_id}_golden.xlsx"
    _xlsx(input_xlsx, source=source, answer=None)
    _xlsx(golden_xlsx, source=source, answer=answer)
    return BenchmarkTask(
        benchmark="spreadsheet",
        task_id=task_id,
        question=question,
        expected={"golden_xlsx": str(golden_xlsx), "answer_sheet": "Sheet1", "answer_position": "B1"},
        input_artifacts={"input_xlsx": str(input_xlsx), "prompt_txt": question},
        metadata={"instruction_type": "formula_generation", "data_position": "Sheet1!A1:B1"},
    )


def test_spreadsheet_answer_range_parser_handles_quoted_sheet_prefix_and_trailing_quote() -> None:
    refs = _answer_range_refs(
        "'Sheet1!'A1:A50,'Sheet2!'A1:E20,'Sheet3!'A1:A50'",
        default_sheet="Sheet1",
    )

    assert refs == [
        ("Sheet1", "A1:A50"),
        ("Sheet2", "A1:E20"),
        ("Sheet3", "A1:A50"),
    ]


def test_spreadsheet_extractor_prompt_preserves_replay_artifacts(tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_extract_prompt", "Double A1 into B1.", source=9, answer=18)
    detail = _detail(
        task,
        success=True,
        score=1.0,
        retrieved=[],
        code="from openpyxl import load_workbook\nwb = load_workbook(OUTPUT_XLSX)\nws = wb['Sheet1']\nws['B1'] = ws['A1'].value * 2\nwb.save(OUTPUT_XLSX)",
    )

    prompt = _spreadsheet_extraction_role_json_block(
        _spreadsheet_extraction_user_payload(
            existing=[],
            detail=detail,
            limits={
                "max_body_words": 120,
                "max_body_lines": 12,
                "max_code_lines": 25,
                "max_artifacts": 3,
                "max_rewrite_rounds": 1,
            },
        )
    )

    assert str(task.input_artifacts["input_xlsx"]) in prompt
    assert str(task.expected["golden_xlsx"]) in prompt
    assert '"answer_position": "B1"' in prompt
    assert '"answer_sheet": "Sheet1"' in prompt
    assert '"extraction_policy"' in prompt
    assert "preferred_for_success" in prompt


def test_spreadsheet_extractor_payload_surfaces_bash_reusable_edit_step(tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_extract_bash_prompt", "Double A1 into B1.", source=9, answer=18)
    detail = _detail(task, success=True, score=1.0, retrieved=[], code="")
    detail["runs"][0]["trace"]["notebook_turns"] = [
        {
            "turn_index": 0,
            "code": """python3 << 'PY'
import openpyxl
import os
wb = openpyxl.load_workbook(os.environ["INPUT_XLSX"])
ws = wb["Sheet1"]
ws["B1"] = ws["A1"].value * 2
wb.save(os.environ["OUTPUT_XLSX"])
PY""",
            "stdout": "saved",
            "stderr": "",
            "returncode": 0,
        }
    ]

    payload = _spreadsheet_extraction_user_payload(
        existing=[],
        detail=detail,
        limits={
            "max_body_words": 120,
            "max_body_lines": 12,
            "max_code_lines": 25,
            "max_artifacts": 3,
            "max_rewrite_rounds": 1,
        },
    )

    assert payload["extraction_policy"]["success_trace"] is True
    assert payload["reusable_edit_steps"]
    edit_code = payload["reusable_edit_steps"][0]["python_code"]
    assert "python3 <<" not in edit_code
    assert "openpyxl.load_workbook(INPUT_XLSX)" in edit_code
    assert "wb.save(OUTPUT_XLSX)" in edit_code


def test_spreadsheet_extractor_payload_uses_full_trace_for_long_bash_edit(tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_extract_long_bash", "Insert rows above X markers.", source=9, answer=18)
    detail = _detail(task, success=True, score=1.0, retrieved=[], code="")
    long_probe = "\n".join(f"print('probe {idx}')" for idx in range(20))
    detail["runs"][0]["trace"]["notebook_turns"] = [
        {
            "turn_index": 0,
            "code": f"""python3 << 'PY'
import openpyxl
import os
wb = openpyxl.load_workbook(os.environ["INPUT_XLSX"])
ws = wb["Sheet1"]
{long_probe}
for row_num in [10, 20]:
    ws.insert_rows(row_num, 1)
wb.save(os.environ["OUTPUT_XLSX"])
PY""",
            "stdout": "saved",
            "stderr": "",
            "returncode": 0,
        }
    ]

    payload = _spreadsheet_extraction_user_payload(
        existing=[],
        detail=detail,
        limits={
            "max_body_words": 120,
            "max_body_lines": 12,
            "max_code_lines": 25,
            "max_artifacts": 3,
            "max_rewrite_rounds": 1,
        },
    )

    edit_code = payload["reusable_edit_steps"][0]["python_code"]
    assert "python3 <<" not in edit_code
    assert "ws.insert_rows" in edit_code
    assert "wb.save(OUTPUT_XLSX)" in edit_code


def test_spreadsheet_extractor_code_lines_are_callable(tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_extract_code_lines", "Double A1 into B1.", source=9, answer=18)
    detail = _detail(task, success=True, score=1.0, retrieved=[], code="")
    raw = {
        "name": "spreadsheet_double_a1_code_lines",
        "kind": "executable_tool",
        "description": "Double A1 into B1.",
        "body": "Applicability: double numeric A1 into B1. Non-applicability: unrelated layouts.",
        "code_lines": [
            "import openpyxl",
            "wb = openpyxl.load_workbook(INPUT_XLSX)",
            "ws = wb.active",
            "ws['B1'] = ws['A1'].value * 2",
            "wb.save(OUTPUT_XLSX)",
        ],
        "interface": {"invocation_contract": {"injection_type": "functional"}},
        "metadata": {"domains": ["SpreadsheetBench"], "allowed_tools": ["openpyxl"]},
        "dependencies": [],
    }

    assert _spreadsheet_extraction_rubric_failures([raw], limits=_spreadsheet_extract_limits()) == []
    artifact = _coerce_spreadsheet_artifact(raw, detail=detail)
    assert artifact is not None
    assert artifact.injection_type() == "functional"
    assert _is_spreadsheet_callable_skill(artifact) is True
    assert "Reusable code:" in artifact.body


async def test_spreadsheet_success_empty_extractor_falls_back_to_callable_from_bash(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_success_fallback", "Double A1 into B1.", source=9, answer=18)
    detail = _detail(task, success=True, score=1.0, retrieved=[], code="")
    detail["runs"][0]["trace"]["notebook_turns"] = [
        {
            "turn_index": 0,
            "code": """python3 << 'PY'
import openpyxl
import os
wb = openpyxl.load_workbook(os.environ["INPUT_XLSX"])
ws = wb["Sheet1"]
ws["B1"] = ws["A1"].value * 2
wb.save(os.environ["OUTPUT_XLSX"])
PY""",
            "stdout": "saved",
            "stderr": "",
            "returncode": 0,
        }
    ]

    async def fake_ask_json(**kwargs: Any) -> Dict[str, Any]:
        assert kwargs["role"] == "spreadsheet_extractor"
        return {"artifacts": []}

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._compat_ask_json", fake_ask_json)
    artifacts = await _extract_spreadsheet_skills_from_detail(
        detail,
        store=ArtifactStore(),
        config=MaintenanceRunConfig(llm_config="mock", model_name="mock-model"),
        task_index=0,
    )

    assert artifacts
    artifact = artifacts[0]
    assert artifact.kind == "executable_tool"
    assert artifact.injection_type() == "functional"
    assert _is_spreadsheet_callable_skill(artifact) is True
    assert "python3 <<" not in artifact.body
    assert "openpyxl.load_workbook(INPUT_XLSX)" in artifact.body
    assert "wb.save(OUTPUT_XLSX)" in artifact.body


async def test_spreadsheet_maintenance_text_role_uses_env_model_override(monkeypatch) -> None:
    captured: Dict[str, Any] = {}

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        captured.update(kwargs)
        return TextLLMResponse(content="```bash\necho ok\n```", prompt_tokens=3, completion_tokens=2)

    monkeypatch.setenv("SPREADSHEET_MAINTENANCE_LLM_CONFIG", "maintenance_proxy")
    monkeypatch.setenv("SPREADSHEET_MAINTENANCE_MODEL_NAME", "claude-haiku-4-5")
    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter.ask_text_llm", fake_ask_text_llm)

    response = await _ask_spreadsheet_text_role(
        system="system",
        prompt="prompt",
        config=MaintenanceRunConfig(llm_config="executor_proxy", model_name="claude-sonnet-4-5"),
        role="spreadsheet_folder_extractor",
    )

    assert response.content.startswith("```bash")
    assert captured["llm_config"] == "maintenance_proxy"
    assert captured["model_name"] == "claude-haiku-4-5"


def _detail(
    task: BenchmarkTask,
    *,
    success: bool,
    score: float,
    retrieved: List[str],
    code: str,
    prompt_injected: List[str] | None = None,
    called: List[str] | None = None,
) -> Dict[str, Any]:
    injected = prompt_injected if prompt_injected is not None else retrieved
    called = called or []
    result = BenchmarkResult(
        benchmark="spreadsheet",
        task_id=task.task_id,
        success=success,
        score=score,
        metrics={
            "answer_sheet": "Sheet1",
            "answer_position": "B1",
            "checked_cells": 1,
            "mismatched_cells": [] if success else [{"cell": "B1", "predicted": 0, "expected": 14}],
            "execution_ok": True,
            "returncode": 0,
            "total_tokens": 120,
            "retrieved_skills": retrieved,
            "prompt_injected_skills": injected,
            "called_skill_functions": called,
        },
        trace={
            "retrieved_skills": retrieved,
            "prompt_injected_skills": injected,
            "called_skill_functions": called,
            "code": code,
            "stderr": "",
            "stdout": "",
        },
    )
    return {
        "task_id": task.task_id,
        "task": task.as_dict(),
        "n_runs": 1,
        "n_success": 1 if success else 0,
        "avg_score": score,
        "runs": [{**result.as_dict(), "run_idx": 0}],
    }


def _passing_skill_test_result(artifact: SkillArtifact) -> SkillTestResult:
    return SkillTestResult(
        result_id=f"{artifact.name}:mock_prestore_pass",
        skill_name=artifact.name,
        skill_version=artifact.version,
        bundle_id=artifact.bundle.bundle_id,
        bundle_version=artifact.bundle.bundle_version,
        run_label="spreadsheet_bundle_unit",
        aggregate={"passed": True, "pass_all_tests": True, "n_cases": len(artifact.bundle.all_cases())},
    )


async def test_spreadsheet_function_skill_can_be_imported_and_called(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_callable", "Double A1 into B1.", source=9, answer=18)
    skill = SkillArtifact(
        name="spreadsheet_double_a1_to_b1",
        kind="executable_tool",
        description="Callable openpyxl helper that doubles A1 into B1.",
        body="""```python
def apply_double(INPUT_XLSX, OUTPUT_XLSX, **kwargs):
    import openpyxl
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb["Sheet1"]
    ws["B1"] = ws["A1"].value * 2
    wb.save(OUTPUT_XLSX)
```""",
        interface=SkillInterface(
            invocation_contract={"injection_type": "functional"},
            input_contract={"benchmark": "SpreadsheetBench"},
        ),
        metadata={
            "domains": ["SpreadsheetBench", "formula_generation"],
            "allowed_tools": ["openpyxl"],
            "intent_keywords": ["double"],
        },
    )
    captured = {}

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        captured["system"] = kwargs["system"]
        assert "from skill_library import" in kwargs["system"]
        assert "spreadsheet_double_a1_to_b1" in kwargs["system"]
        return TextLLMResponse(
            content="""```python
from skill_library import spreadsheet_double_a1_to_b1
spreadsheet_double_a1_to_b1(INPUT_XLSX, OUTPUT_XLSX)
```""",
            prompt_tokens=12,
            completion_tokens=8,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)

    result = await run_spreadsheet_task(
        task,
        llm_config="mock",
        model_name="mock-model",
        artifact_store=ArtifactStore([skill]),
        top_k_skills=1,
        skill_injector_mode="compact",
        skill_context_budget_chars=800,
        work_dir=tmp_path / "work",
    )

    assert result.success is True
    assert result.metrics["prompt_injected_skills"] == ["spreadsheet_double_a1_to_b1"]
    assert result.metrics["called_skill_functions"] == ["spreadsheet_double_a1_to_b1"]
    assert result.trace["callable_skills"][0]["function_name"] == "spreadsheet_double_a1_to_b1"
    assert (tmp_path / "work" / "skill_library.py").exists()


async def test_spreadsheet_callable_progressive_disclosure_exposes_signature_and_code_object(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_callable_progressive", "Double A1 into B1.", source=5, answer=10)
    skill = SkillArtifact(
        name="spreadsheet_double_progressive",
        kind="executable_tool",
        description="Callable helper that doubles A1 into B1.",
        body="""```python
def apply_double(INPUT_XLSX, OUTPUT_XLSX, **kwargs):
    import openpyxl
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb["Sheet1"]
    ws["B1"] = ws["A1"].value * 2
    wb.save(OUTPUT_XLSX)
```""",
        metadata={"domains": ["SpreadsheetBench"], "allowed_tools": ["openpyxl"], "intent_keywords": ["double"]},
    )
    captured: Dict[str, str] = {}

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        captured["system"] = kwargs["system"]
        return TextLLMResponse(
            content="""```python
from skill_library import spreadsheet_double_progressive
spreadsheet_double_progressive(INPUT_XLSX, OUTPUT_XLSX)
```""",
            prompt_tokens=12,
            completion_tokens=8,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)

    result = await run_spreadsheet_task(
        task,
        llm_config="mock",
        model_name="mock-model",
        artifact_store=ArtifactStore([skill]),
        top_k_skills=1,
        callable_disclosure_mode="progressive",
        work_dir=tmp_path / "work_progressive",
    )

    assert result.success is True
    assert "Available callable signatures" in captured["system"]
    assert "Direct call examples:" in captured["system"]
    assert "prefer a direct import and call" in captured["system"]
    assert "from skill_library import spreadsheet_double_progressive" in captured["system"]
    assert "spreadsheet_double_progressive(INPUT_XLSX, OUTPUT_XLSX, **kwargs)" in captured["system"]
    assert "spreadsheet_double_progressive_skill.code" in captured["system"]
    assert "skills.md" in captured["system"]
    assert "skills/spreadsheet_double_progressive.py" in captured["system"]
    assert "Implementation gist:" not in captured["system"]
    library = (tmp_path / "work_progressive" / "skill_library.py").read_text()
    assert "spreadsheet_double_progressive_skill = SimpleNamespace" in library
    manifest = (tmp_path / "work_progressive" / "skills.md").read_text()
    assert "# Spreadsheet Skills" in manifest
    assert "## spreadsheet_double_progressive" in manifest
    assert "from skill_library import spreadsheet_double_progressive" in manifest
    assert (tmp_path / "work_progressive" / "skills" / "spreadsheet_double_progressive.py").exists()
    assert result.metrics["called_skill_functions"] == ["spreadsheet_double_progressive"]


async def test_spreadsheet_callable_full_disclosure_includes_complete_code(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_callable_full_code", "Double A1 into B1.", source=5, answer=10)
    skill = SkillArtifact(
        name="spreadsheet_double_full_code",
        kind="executable_tool",
        description="Callable helper that doubles A1 into B1.",
        body="""```python
def apply_double(INPUT_XLSX, OUTPUT_XLSX, **kwargs):
    import openpyxl
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb["Sheet1"]
    ws["B1"] = ws["A1"].value * 2
    wb.save(OUTPUT_XLSX)
```""",
        metadata={"domains": ["SpreadsheetBench"], "allowed_tools": ["openpyxl"], "intent_keywords": ["double"]},
    )
    captured: Dict[str, str] = {}

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        captured["system"] = kwargs["system"]
        return TextLLMResponse(
            content="""```python
from skill_library import spreadsheet_double_full_code
spreadsheet_double_full_code(INPUT_XLSX, OUTPUT_XLSX)
```""",
            prompt_tokens=12,
            completion_tokens=8,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)
    result = await run_spreadsheet_task(
        task,
        llm_config="mock",
        model_name="mock-model",
        artifact_store=ArtifactStore([skill]),
        top_k_skills=1,
        callable_disclosure_mode="full",
        work_dir=tmp_path / "work_full_code",
    )

    assert result.success is True
    assert "Full implementation:" in captured["system"]
    assert 'ws["B1"] = ws["A1"].value * 2' in captured["system"]
    assert "spreadsheet_double_full_code_skill.code" not in captured["system"]


async def test_spreadsheet_executable_tool_marked_informational_still_callable(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_callable_legacy", "Double A1 into B1.", source=6, answer=12)
    skill = SkillArtifact(
        name="spreadsheet_legacy_double_callable",
        kind="executable_tool",
        description="Legacy executable skill whose metadata was incorrectly stored as informational.",
        body="""Applicability: double a source cell.
```python
from openpyxl import load_workbook
wb = load_workbook(INPUT_XLSX)
ws = wb["Sheet1"]
ws["B1"] = ws["A1"].value * 2
wb.save(OUTPUT_XLSX)
```
""",
        metadata={
            "injection_type": "informational",
            "domains": ["SpreadsheetBench", "formula_generation"],
            "allowed_tools": ["openpyxl"],
            "intent_keywords": ["double"],
        },
    )
    assert skill.injection_type() == "informational"
    assert _is_spreadsheet_callable_skill(skill) is True

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        assert "Callable function skills:" in kwargs["system"]
        assert "spreadsheet_legacy_double_callable" in kwargs["system"]
        return TextLLMResponse(
            content="""```python
from skill_library import spreadsheet_legacy_double_callable
spreadsheet_legacy_double_callable(INPUT_XLSX, OUTPUT_XLSX)
```""",
            prompt_tokens=12,
            completion_tokens=8,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)
    result = await run_spreadsheet_task(
        task,
        llm_config="mock",
        model_name="mock-model",
        artifact_store=ArtifactStore([skill]),
        top_k_skills=1,
        skill_injector_mode="compact",
        skill_context_budget_chars=800,
        work_dir=tmp_path / "work",
    )

    assert result.success is True
    assert result.trace["callable_skills"][0]["skill_name"] == "spreadsheet_legacy_double_callable"
    assert result.metrics["called_skill_functions"] == ["spreadsheet_legacy_double_callable"]


async def test_spreadsheet_workflow_skill_is_not_exported_as_callable(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_workflow_only", "Double A1 into B1.", source=5, answer=10)
    skill = SkillArtifact(
        name="spreadsheet_workflow_openpyxl_guardrail",
        kind="workflow_guardrail_card",
        description="Prompt-only workflow guidance mentioning openpyxl.",
        body="Use openpyxl only after inspecting workbook shape; preserve unrelated cells.",
        metadata={
            "injection_type": "informational",
            "domains": ["SpreadsheetBench"],
            "allowed_tools": ["openpyxl"],
            "intent_keywords": ["double", "inspect"],
        },
    )
    assert _is_spreadsheet_callable_skill(skill) is False

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        assert "spreadsheet_workflow_openpyxl_guardrail" not in kwargs["system"].split("Callable function skills:", 1)[1]
        return TextLLMResponse(
            content="""```python
import openpyxl
wb = openpyxl.load_workbook(INPUT_XLSX)
ws = wb["Sheet1"]
ws["B1"] = ws["A1"].value * 2
wb.save(OUTPUT_XLSX)
```""",
            prompt_tokens=12,
            completion_tokens=8,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)
    result = await run_spreadsheet_task(
        task,
        llm_config="mock",
        model_name="mock-model",
        artifact_store=ArtifactStore([skill]),
        top_k_skills=1,
        skill_injector_mode="compact",
        skill_context_budget_chars=800,
        work_dir=tmp_path / "work",
    )

    assert result.success is True
    assert result.trace["callable_skills"] == []
    assert not (tmp_path / "work" / "skill_library.py").exists()


def test_spreadsheet_called_skill_function_parser_handles_alias_and_module_import() -> None:
    rows = [
        {"skill_name": "spreadsheet_double_a1_to_b1", "function_name": "spreadsheet_double_a1_to_b1"},
        {"skill_name": "spreadsheet_sum_column", "function_name": "spreadsheet_sum_column"},
    ]

    direct = _called_spreadsheet_skill_functions(
        "from skill_library import spreadsheet_double_a1_to_b1 as dbl\n"
        "dbl(INPUT_XLSX, OUTPUT_XLSX)\n",
        rows,
    )
    module = _called_spreadsheet_skill_functions(
        "import skill_library as skills\n"
        "skills.spreadsheet_sum_column(INPUT_XLSX, OUTPUT_XLSX)\n",
        rows,
    )

    assert direct == ["spreadsheet_double_a1_to_b1"]
    assert module == ["spreadsheet_sum_column"]


async def test_spreadsheet_notebook_mode_returns_errors_and_reuses_variables(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_notebook", "Double A1 into B1.", source=9, answer=18)
    prompts: List[str] = []

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        prompts.append(kwargs["prompt"])
        if len(prompts) == 1:
            return TextLLMResponse(
                content="""```python
import openpyxl
wb = openpyxl.load_workbook(INPUT_XLSX)
ws = wb["Sheet1"]
print("source", ws["A1"].value)
missing_name
```""",
                prompt_tokens=20,
                completion_tokens=15,
                model_name="mock-model",
                api_style="mock",
            )
        return TextLLMResponse(
            content=f"""```python
ws["B1"] = ws["A1"].value * 2
wb.save(OUTPUT_XLSX)
```
{SPREADSHEET_DONE_PATTERN}""",
            prompt_tokens=25,
            completion_tokens=12,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)
    result = await run_spreadsheet_task_notebook(
        task,
        llm_config="mock",
        model_name="mock-model",
        artifact_store=ArtifactStore(),
        max_turns=5,
        work_dir=tmp_path / "work",
    )

    assert result.success is True
    assert result.metrics["execution_mode"] == "notebook"
    assert result.metrics["notebook_turn_count"] == 2
    assert result.metrics["notebook_stopped_by_done"] is True
    assert "NameError" in prompts[1]
    assert "source 9" in prompts[1]
    assert result.trace["notebook_turns"][0]["returncode"] == 1
    assert "NameError" in result.trace["notebook_turns"][0]["stderr"]
    assert result.trace["notebook_turns"][1]["returncode"] == 0
    assert result.metrics["total_tokens"] == 72


async def test_spreadsheet_bash_react_runs_commands_in_persistent_workdir(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_bash_react", "Double A1 into B1.", source=6, answer=12)
    prompts: List[str] = []

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        prompts.append(str(kwargs.get("prompt") or ""))
        if len(prompts) == 1:
            return TextLLMResponse(
                content="""```bash
python - <<'PY'
import openpyxl
wb = openpyxl.load_workbook("input.xlsx")
ws = wb["Sheet1"]
print("source", ws["A1"].value)
PY
```""",
                prompt_tokens=20,
                completion_tokens=15,
                model_name="mock-model",
                api_style="mock",
            )
        return TextLLMResponse(
            content=f"""```bash
python - <<'PY'
import openpyxl
wb = openpyxl.load_workbook("output.xlsx")
ws = wb["Sheet1"]
ws["B1"] = ws["A1"].value * 2
wb.save("output.xlsx")
PY
```
{SPREADSHEET_DONE_PATTERN}""",
            prompt_tokens=25,
            completion_tokens=12,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)
    result = await run_spreadsheet_task_bash_react(
        task,
        llm_config="mock",
        model_name="mock-model",
        artifact_store=ArtifactStore(),
        max_turns=5,
        work_dir=tmp_path / "bash_work",
    )

    assert result.success is True
    assert result.metrics["execution_mode"] == "bash_react"
    assert result.metrics["bash_turn_count"] == 2
    assert result.metrics["bash_stopped_by_done"] is True
    assert "source 6" in prompts[1]
    assert result.trace["notebook_turns"][0]["returncode"] == 0
    assert result.trace["notebook_turns"][1]["returncode"] == 0
    assert result.metrics["total_tokens"] == 72


async def test_spreadsheet_bash_react_skill_update_uses_bash_callable_examples(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_bash_step_skill", "Double A1 into B1.", source=7, answer=14)
    repair_skill = SkillArtifact(
        name="spreadsheet_bash_double_skill",
        kind="executable_tool",
        description="Double A1 into B1.",
        body="""```python
def apply_double(INPUT_XLSX, OUTPUT_XLSX, **kwargs):
    import openpyxl
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb["Sheet1"]
    ws["B1"] = ws["A1"].value * 2
    wb.save(OUTPUT_XLSX)
```""",
        interface=SkillInterface(invocation_contract={"injection_type": "functional"}),
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
        status="pending",
    )

    class StepStore(ArtifactStore):
        def __init__(self) -> None:
            super().__init__([repair_skill])

        def retrieve(self, query: str, top_k: int = 5, **kwargs: Any) -> List[SkillArtifact]:
            if "NameError" in query:
                return [repair_skill]
            return []

    prompts: List[str] = []

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        prompts.append(str(kwargs.get("prompt") or ""))
        if len(prompts) == 1:
            return TextLLMResponse(
                content="""```bash
python - <<'PY'
raise NameError("need skill")
PY
```""",
                prompt_tokens=10,
                completion_tokens=5,
                model_name="mock-model",
                api_style="mock",
            )
        assert "Runtime skill retrieval update for this same Spreadsheet bash task." in kwargs["prompt"]
        assert "[NEW SKILLS]" in kwargs["prompt"]
        assert "[/NEW SKILLS]" in kwargs["prompt"]
        assert "USE_NOW" in kwargs["prompt"]
        assert "USE_LATER" in kwargs["prompt"]
        assert "SKIP" in kwargs["prompt"]
        assert "decision_required" in kwargs["prompt"]
        assert "Direct bash call examples:" in kwargs["prompt"]
        assert "import os" in kwargs["prompt"]
        assert "os.environ['INPUT_XLSX']" in kwargs["prompt"]
        assert "os.environ['OUTPUT_XLSX']" in kwargs["prompt"]
        assert "from skill_library import spreadsheet_bash_double_skill" in kwargs["prompt"]
        assert "from skill_library import spreadsheet_bash_double_skill_skill" in kwargs["prompt"]
        assert "sed -n '1,220p' skills.md" in kwargs["prompt"]
        assert "skills/spreadsheet_bash_double_skill.py" in kwargs["prompt"]
        assert "Prefer calling a matching function directly" in kwargs["prompt"]
        assert "readable/writable files" in kwargs["prompt"]
        assert "Implementation gist:" not in kwargs["prompt"]
        assert "in a notebook cell" not in kwargs["prompt"]
        return TextLLMResponse(
            content=f"""```bash
python - <<'PY'
import os
from skill_library import spreadsheet_bash_double_skill
spreadsheet_bash_double_skill(os.environ["INPUT_XLSX"], os.environ["OUTPUT_XLSX"])
PY
```
{SPREADSHEET_DONE_PATTERN}""",
            prompt_tokens=20,
            completion_tokens=10,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)
    result = await run_spreadsheet_task_bash_react(
        task,
        llm_config="mock",
        model_name="mock-model",
        artifact_store=StepStore(),
        top_k_skills=1,
        pending_skill_fraction=1.0,
        callable_disclosure_mode="progressive",
        max_turns=3,
        work_dir=tmp_path / "bash_skill_work",
    )

    assert result.success is True
    assert result.metrics["called_skill_functions"] == ["spreadsheet_bash_double_skill"]
    assert any(turn["skill_context_updated"] for turn in result.trace["notebook_turns"])


async def test_spreadsheet_notebook_retrieves_and_injects_new_skills_per_turn(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_notebook_step_retrieval", "Double A1 into B1.", source=8, answer=16)
    repair_skill = SkillArtifact(
        name="spreadsheet_repair_after_name_error",
        kind="executable_tool",
        description="Repair notebook after a NameError and write double of A1 to B1.",
        body="""```python
def apply_repair(INPUT_XLSX, OUTPUT_XLSX, **kwargs):
    import openpyxl
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb["Sheet1"]
    ws["B1"] = ws["A1"].value * 2
    wb.save(OUTPUT_XLSX)
```""",
        interface=SkillInterface(invocation_contract={"injection_type": "functional"}),
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["NameError", "double"]},
        status="pending",
    )

    class StepStore(ArtifactStore):
        def __init__(self) -> None:
            super().__init__([repair_skill])
            self.queries: List[str] = []

        def retrieve(self, query: str, top_k: int = 5, **kwargs: Any) -> List[SkillArtifact]:
            self.queries.append(query)
            if "NameError" in query:
                return [repair_skill]
            return []

    store = StepStore()
    systems: List[str] = []
    prompts: List[str] = []

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        systems.append(kwargs["system"])
        prompts.append(kwargs["prompt"])
        if len(systems) == 1:
            assert "spreadsheet_repair_after_name_error" not in kwargs["system"]
            assert "spreadsheet_repair_after_name_error" not in kwargs["prompt"]
            return TextLLMResponse(
                content="""```python
missing_name
```""",
                prompt_tokens=10,
                completion_tokens=6,
                model_name="mock-model",
                api_style="mock",
            )
        assert "spreadsheet_repair_after_name_error" not in kwargs["system"]
        assert "Runtime skill retrieval update for this same Spreadsheet notebook task." in kwargs["prompt"]
        assert "[NEW SKILLS]" in kwargs["prompt"]
        assert "[/NEW SKILLS]" in kwargs["prompt"]
        assert "USE_NOW" in kwargs["prompt"]
        assert "USE_LATER" in kwargs["prompt"]
        assert "SKIP" in kwargs["prompt"]
        assert "decision_required" in kwargs["prompt"]
        assert "### spreadsheet_repair_after_name_error" in kwargs["prompt"]
        assert "(no reusable skill artifacts retrieved)" not in kwargs["prompt"]
        assert "from `skill_library`" in kwargs["prompt"]
        assert "Direct call examples:" in kwargs["prompt"]
        assert "from skill_library import spreadsheet_repair_after_name_error" in kwargs["prompt"]
        assert "spreadsheet_repair_after_name_error(INPUT_XLSX, OUTPUT_XLSX, **kwargs)" in kwargs["prompt"]
        return TextLLMResponse(
            content=f"""```python
from skill_library import spreadsheet_repair_after_name_error
spreadsheet_repair_after_name_error(INPUT_XLSX, OUTPUT_XLSX)
```
{SPREADSHEET_DONE_PATTERN}""",
            prompt_tokens=14,
            completion_tokens=8,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)
    result = await run_spreadsheet_task_notebook(
        task,
        llm_config="mock",
        model_name="mock-model",
        artifact_store=store,
        top_k_skills=1,
        max_turns=3,
        work_dir=tmp_path / "work_step_retrieval",
    )

    assert result.success is True
    assert len(store.queries) == 2
    assert "Trace prefix: no notebook code has executed yet." in store.queries[0]
    assert "NameError" in store.queries[1]
    assert len(result.trace["prompt_context_updates"]) == 1
    update = result.trace["prompt_context_updates"][0]
    assert update["turn_index"] == 1
    assert update["retrieved_skills"] == ["spreadsheet_repair_after_name_error"]
    assert update["new_skills"] == ["spreadsheet_repair_after_name_error"]
    assert update["callable_skills"] == ["spreadsheet_repair_after_name_error"]
    assert update["message"]["role"] == "user"
    assert "Runtime skill retrieval update" in update["message"]["content"]
    assert "[NEW SKILLS]" in update["message"]["content"]
    assert "[/NEW SKILLS]" in update["message"]["content"]
    assert "USE_NOW" in update["message"]["content"]
    assert "USE_LATER" in update["message"]["content"]
    assert "SKIP" in update["message"]["content"]
    assert result.metrics["called_skill_functions"] == ["spreadsheet_repair_after_name_error"]
    assert "skill_code_reads" in result.metrics
    assert result.trace["notebook_turns"][1]["skill_context_updated"] is True


def test_spreadsheet_notebook_trace_projection_keeps_steps_and_final_traceback_frame(tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_notebook_projection", "Double A1 into B1.", source=9, answer=18)
    long_traceback = (
        "debug line\n"
        "Traceback (most recent call last):\n"
        '  File "/tmp/driver.py", line 10, in <module>\n'
        "    run_cell()\n"
        + "\n".join(f'  File "/tmp/noisy_{idx}.py", line {idx}, in helper\n    value += {idx}' for idx in range(80))
        + '\n  File "/tmp/final_cell.py", line 4, in <module>\n'
        "    missing_name\n"
        "NameError: name 'missing_name' is not defined\n"
    )
    detail = _detail(
        task,
        success=False,
        score=0.0,
        retrieved=[],
        code="",
    )
    detail["runs"][0]["metrics"]["execution_ok"] = False
    detail["runs"][0]["trace"].update(
        {
            "stderr": long_traceback,
            "notebook_turns": [
                {
                    "turn_index": 0,
                    "code": "print('inspect')",
                    "stdout": "inspect\n",
                    "stderr": "",
                    "returncode": 0,
                    "done_requested": False,
                    "changed_variables": {"wb": "Workbook"},
                },
                {
                    "turn_index": 1,
                    "code": "missing_name",
                    "stdout": "",
                    "stderr": long_traceback,
                    "returncode": 1,
                    "done_requested": False,
                    "changed_variables": {"error": "NameError"},
                },
                "raw malformed notebook step",
                {
                    "turn_index": 2,
                    "code": "ws['B1'] = ws['A1'].value * 2",
                    "stdout": "",
                    "stderr": "",
                    "returncode": 0,
                    "done_requested": True,
                    "changed_variables": {"saved": True},
                },
            ],
        }
    )

    result_projection = _spreadsheet_result_projection(detail)
    trace_projection = _spreadsheet_trace_projection(detail)
    steps = result_projection["trace"]["notebook_steps"]

    assert result_projection["trace"]["notebook_step_count"] == 4
    assert [set(step) for step in steps] == [
        {"code_snippet", "stdout_tail", "stderr_tail", "exception"},
        {"code_snippet", "stdout_tail", "stderr_tail", "exception"},
        {"code_snippet", "stdout_tail", "stderr_tail", "exception"},
        {"code_snippet", "stdout_tail", "stderr_tail", "exception"},
    ]
    assert "final_cell.py" in steps[1]["stderr_tail"]
    assert "NameError: name 'missing_name' is not defined" in steps[1]["stderr_tail"]
    assert "/tmp/noisy_0.py" not in steps[1]["stderr_tail"]
    assert steps[1]["exception"] == "NameError: name 'missing_name' is not defined"
    assert steps[2]["stderr_tail"] == "raw malformed notebook step"
    assert all("changed_variables" not in step for step in steps)
    assert trace_projection["notebook_step_count"] == 4
    assert "final_cell.py" in trace_projection["stderr_tail"]


def test_spreadsheet_callable_snippet_receives_kwargs_and_column_aliases(tmp_path: Path) -> None:
    input_xlsx = tmp_path / "input.xlsx"
    output_xlsx = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 7
    ws["D1"] = 7
    wb.save(input_xlsx)

    skill = SkillArtifact(
        name="spreadsheet_snippet_with_kwargs",
        kind="executable_tool",
        description="Remove one matching value from a value column.",
        body="""```python
values = []
for row in range(value_start_row, value_end_row + 1):
    values.append(ws.cell(row=row, column=value_col).value)
result_values = []
for row in range(result_start_row, result_end_row + 1):
    result_values.append(ws.cell(row=row, column=result_col).value)
for item in result_values:
    if item in values:
        values.remove(item)
for idx, item in enumerate(values, start=value_start_row):
    ws.cell(row=idx, column=value_col).value = item
```""",
        metadata={"injection_type": "informational", "domains": ["SpreadsheetBench"]},
    )
    info = _write_spreadsheet_skill_library([skill], tmp_path)
    assert info["skills"][0]["function_name"] == "spreadsheet_snippet_with_kwargs"

    import importlib.util

    spec = importlib.util.spec_from_file_location("skill_library", tmp_path / "skill_library.py")
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    spec.loader.exec_module(module)
    module.spreadsheet_snippet_with_kwargs(
        input_xlsx,
        output_xlsx,
        value_column="A",
        result_column="D",
        value_start_row=1,
        value_end_row=2,
        result_start_row=1,
        result_end_row=1,
    )

    out = openpyxl.load_workbook(output_xlsx)
    assert out["Sheet1"]["A1"].value == 1


def test_spreadsheet_callable_script_wrapper_executes_skill(tmp_path: Path) -> None:
    input_xlsx = tmp_path / "input.xlsx"
    output_xlsx = tmp_path / "output.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 7
    wb.save(input_xlsx)

    skill = SkillArtifact(
        name="spreadsheet_script_double",
        kind="executable_tool",
        description="Double A1 into B1.",
        body="""```python
def apply_double(INPUT_XLSX, OUTPUT_XLSX, **kwargs):
    import openpyxl
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb["Sheet1"]
    ws["B1"] = ws["A1"].value * 2
    wb.save(OUTPUT_XLSX)
```""",
        metadata={"domains": ["SpreadsheetBench"], "allowed_tools": ["openpyxl"]},
    )
    info = _write_spreadsheet_skill_library([skill], tmp_path, disclosure_mode="progressive")
    assert info["skills"][0]["manifest_path"] == "skills.md"
    assert info["skills"][0]["script_path"] == "skills/spreadsheet_script_double.py"
    manifest = tmp_path / "skills.md"
    assert manifest.exists()
    manifest_text = manifest.read_text()
    assert "## spreadsheet_script_double" in manifest_text
    assert "python skills/spreadsheet_script_double.py" in manifest_text
    script = tmp_path / "skills" / "spreadsheet_script_double.py"
    assert script.exists()

    env = os.environ.copy()
    env["INPUT_XLSX"] = str(input_xlsx)
    env["OUTPUT_XLSX"] = str(output_xlsx)
    result = subprocess.run(
        [sys.executable, str(script)],
        cwd=tmp_path,
        env=env,
        capture_output=True,
        text=True,
        timeout=10,
        check=False,
    )
    assert result.returncode == 0, result.stderr
    out = openpyxl.load_workbook(output_xlsx)
    assert out["Sheet1"]["B1"].value == 14


def test_spreadsheet_truncated_trace_body_is_not_callable() -> None:
    skill = SkillArtifact(
        name="spreadsheet_bad_trace_callable",
        kind="executable_tool",
        description="Bad truncated trace body.",
        body=(
            "Applicability: SpreadsheetBench trace transcript.\n\n"
            "Reusable openpyxl idiom from successful evidence:\n"
            "```python\n"
            "import openpyxl\n"
            "wb = openpyxl.load_workbook(INPUT_XLSX)\n"
        ),
    )

    assert _is_spreadsheet_callable_skill(skill) is False


def test_spreadsheet_pending_retrieval_requires_explicit_include(tmp_path: Path) -> None:
    active = SkillArtifact(
        name="spreadsheet_active_double",
        kind="workflow_guardrail_card",
        description="Active double guidance.",
        body="Double A1 into B1.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
        status="active",
    )
    pending = SkillArtifact(
        name="spreadsheet_pending_double",
        kind="workflow_guardrail_card",
        description="Pending double guidance.",
        body="Double A1 into B1 with pending candidate.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"], "is_pending_skill": True},
        status="pending",
    )
    store = ArtifactStore([active, pending])

    assert [item.name for item in store.retrieve("double A1 into B1", top_k=5)] == ["spreadsheet_active_double"]
    assert "spreadsheet_pending_double" in [
        item.name
        for item in store.retrieve(
            "double A1 into B1",
            top_k=5,
            include_pending=True,
            predicate=lambda artifact: artifact.status == "pending",
        )
    ]


async def test_spreadsheet_pending_fraction_mixes_active_and_pending(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_pending_mix", "Double A1 into B1.", source=4, answer=8)
    active_a = SkillArtifact(
        name="spreadsheet_active_double_a",
        kind="workflow_guardrail_card",
        description="Active double A.",
        body="Double A1 into B1.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
        status="active",
    )
    active_b = SkillArtifact(
        name="spreadsheet_active_double_b",
        kind="workflow_guardrail_card",
        description="Active double B.",
        body="Double A1 into B1.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
        status="active",
    )
    pending = SkillArtifact(
        name="spreadsheet_pending_double",
        kind="workflow_guardrail_card",
        description="Pending double candidate.",
        body="Double A1 into B1.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"], "is_pending_skill": True},
        status="pending",
    )

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        return TextLLMResponse(
            content="""```python
import openpyxl
wb = openpyxl.load_workbook(INPUT_XLSX)
ws = wb["Sheet1"]
ws["B1"] = ws["A1"].value * 2
wb.save(OUTPUT_XLSX)
```""",
            prompt_tokens=10,
            completion_tokens=10,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)
    result = await run_spreadsheet_task(
        task,
        llm_config="mock",
        model_name="mock-model",
        artifact_store=ArtifactStore([active_a, active_b, pending]),
        top_k_skills=3,
        pending_skill_fraction=1 / 3,
        work_dir=tmp_path / "work_pending_mix",
    )

    assert result.success is True
    assert "spreadsheet_pending_double" in result.metrics["retrieved_skills"]
    assert len([name for name in result.metrics["retrieved_skills"] if "active" in name]) == 2
    retrieval_events = [
        event for event in result.metrics["injector_events"]
        if event.get("type") == "spreadsheet_pending_mixed_retrieval"
    ]
    assert retrieval_events
    assert [item["name"] for item in retrieval_events[0]["pending_selected"]] == ["spreadsheet_pending_double"]


def test_spreadsheet_pending_retrieval_respects_min_score(tmp_path: Path) -> None:
    pending = SkillArtifact(
        name="spreadsheet_pending_unrelated",
        kind="executable_tool",
        description="Conditional sum formula helper for invoice totals.",
        body="Use SUMIF for invoice totals.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["invoice", "sumif"]},
        status="pending",
    )
    pending.metadata["is_pending_skill"] = True
    store = ArtifactStore([pending])

    retrieved, event = _retrieve_spreadsheet_skills(
        store,
        query="Sort the student names alphabetically in column A.",
        top_k=3,
        min_score=0.01,
        pending_skill_fraction=1 / 3,
    )

    assert retrieved == []
    assert event is not None
    assert event["pending_selected"] == []


def test_spreadsheet_notebook_session_accepts_relative_work_dir(tmp_path: Path, monkeypatch) -> None:
    from academic.benchmarks.spreadsheet.adapter import _NotebookPythonSession

    monkeypatch.chdir(tmp_path)
    session = _NotebookPythonSession(Path("relative_nb_work"))
    try:
        first = session.run_cell("x = 41\nprint('x', x)", timeout=5)
        second = session.run_cell("print('x_plus_one', x + 1)", timeout=5)
    finally:
        session.close()

    assert first["returncode"] == 0
    assert "x 41" in first["stdout"]
    assert second["returncode"] == 0
    assert "x_plus_one 42" in second["stdout"]


async def test_spreadsheet_micro_extracts_actionable_pending_skill_from_success(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_train_1", "Double the value in A1 and write it to B1.")
    detail = _detail(
        task,
        success=True,
        score=1.0,
        retrieved=[],
        code="import openpyxl\nwb=openpyxl.load_workbook(INPUT_XLSX)\nws=wb['Sheet1']\nws['B1']=ws['A1'].value*2\nwb.save(OUTPUT_XLSX)",
    )
    calls: List[str] = []

    async def fake_ask_json(**kwargs: Any) -> Dict[str, Any]:
        calls.append(kwargs["role"])
        return {
            "artifacts": [
                {
                    "name": "spreadsheet_double_a1_to_b1",
                    "kind": "executable_tool",
                    "description": "Double a numeric source cell and write the result with openpyxl.",
                    "body": "Applicability: tasks asking to double a numeric source cell. ```python\nws['B1'] = ws['A1'].value * 2\nwb.save(OUTPUT_XLSX)\n``` Non-applicability: do not hard-code values.",
                    "interface": {
                        "summary": "Double source cell into answer cell.",
                        "usage": "Use for formula_generation doubling tasks.",
                        "input_contract": {"benchmark": "SpreadsheetBench", "source_cell": "numeric"},
                        "output_contract": {"answer_cell": "source*2"},
                        "invocation_contract": {"injection_type": "informational"},
                        "compatibility_notes": "Inspect the workbook before applying.",
                    },
                    "metadata": {
                        "domains": ["SpreadsheetBench", "formula_generation"],
                        "allowed_tools": ["openpyxl"],
                        "intent_keywords": ["double", "numeric", "formula"],
                        "source_task_ids": ["sheet_train_1"],
                        "evidence_span": "successful trace wrote B1=A1*2",
                        "non_applicability": "not for unrelated aggregations",
                    },
                }
            ]
        }

    async def fake_bundle_tests(**kwargs: Any) -> SkillTestResult:
        return _passing_skill_test_result(kwargs["artifact"])

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._compat_ask_json", fake_ask_json)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter._execute_spreadsheet_bundle_tests", fake_bundle_tests)
    store = ArtifactStore()
    report = await SpreadsheetMaintenanceAdapter().run_micro_maintenance(
        detail=detail,
        credit_events=[],
        credit_bundle_cases=[],
        store=store,
        config=MaintenanceRunConfig(
            llm_config="mock",
            model_name="mock-model",
            extra={"spreadsheet_prestore_refine_max_rounds": 0},
        ),
        round_index=0,
        task_index=0,
    )

    assert "spreadsheet_extractor" in calls
    assert report["extraction_reports"][0]["skill_name"] == "spreadsheet_double_a1_to_b1"
    assert "spreadsheet_double_a1_to_b1" in report["maintenance_targets"]
    assert report["maintenance_test_results"]
    artifact = store.get("spreadsheet_double_a1_to_b1")
    assert artifact is not None
    assert artifact.status == "pending"
    assert artifact.injection_type() == "functional"
    assert artifact.interface.invocation_contract["injection_type"] == "functional"
    assert artifact.bundle.positive_cases
    assert "openpyxl" in artifact.body


async def test_spreadsheet_extractor_rewrites_oversized_code_before_store(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_extract_rewrite", "Double A1 into B1.", source=4, answer=8)
    detail = _detail(
        task,
        success=True,
        score=1.0,
        retrieved=[],
        code="import openpyxl\nwb=openpyxl.load_workbook(INPUT_XLSX)\nws=wb['Sheet1']\nws['B1']=ws['A1'].value*2\nwb.save(OUTPUT_XLSX)",
    )
    roles: List[str] = []
    oversized_code = "\n".join(f"print({idx})" for idx in range(45))

    async def fake_ask_json(**kwargs: Any) -> Dict[str, Any]:
        roles.append(kwargs["role"])
        if kwargs["role"] == "spreadsheet_extractor":
            return {
                "artifacts": [
                    {
                        "name": "spreadsheet_oversized_trace",
                        "kind": "executable_tool",
                        "description": "Oversized copied trace.",
                        "body": f"Applicability: copied trace.\n```python\n{oversized_code}\n```\nNon-applicability: none.",
                        "metadata": {"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
                    }
                ]
            }
        return {
            "artifacts": [
                {
                    "name": "spreadsheet_compact_double",
                    "kind": "executable_tool",
                    "description": "Double A1 into B1 with openpyxl.",
                    "body": "Applicability: double A1 into B1.\n```python\nws['B1'] = ws['A1'].value * 2\nwb.save(OUTPUT_XLSX)\n```\nNon-applicability: use only after verifying A1 and B1 match the task.",
                    "metadata": {"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
                }
            ]
        }

    async def fake_bundle_tests(**kwargs: Any) -> SkillTestResult:
        return _passing_skill_test_result(kwargs["artifact"])

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._compat_ask_json", fake_ask_json)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter._execute_spreadsheet_bundle_tests", fake_bundle_tests)
    store = ArtifactStore()
    report = await SpreadsheetMaintenanceAdapter().run_micro_maintenance(
        detail=detail,
        credit_events=[],
        credit_bundle_cases=[],
        store=store,
        config=MaintenanceRunConfig(
            llm_config="mock",
            model_name="mock-model",
            extra={"spreadsheet_prestore_refine_max_rounds": 0},
        ),
        round_index=0,
        task_index=0,
    )

    assert roles[:2] == ["spreadsheet_extractor", "spreadsheet_extractor_rubric_rewrite"]
    assert store.get("spreadsheet_oversized_trace") is None
    assert store.get("spreadsheet_compact_double") is not None
    assert report["extraction_reports"][0]["skill_name"] == "spreadsheet_compact_double"


async def test_spreadsheet_extractor_rejects_candidate_that_fails_prestore_bundle(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_extract_reject", "Double A1 into B1.", source=4, answer=8)
    detail = _detail(
        task,
        success=True,
        score=1.0,
        retrieved=[],
        code="import openpyxl\nwb=openpyxl.load_workbook(INPUT_XLSX)\nws=wb['Sheet1']\nws['B1']=ws['A1'].value*2\nwb.save(OUTPUT_XLSX)",
    )

    async def fake_ask_json(**kwargs: Any) -> Dict[str, Any]:
        assert kwargs["role"] == "spreadsheet_extractor"
        return {
            "artifacts": [
                {
                    "name": "spreadsheet_bad_double",
                    "kind": "executable_tool",
                    "description": "Bad helper for doubling A1 into B1.",
                    "body": "Applicability: double A1 into B1.\n```python\nws['B1'] = 0\nwb.save(OUTPUT_XLSX)\n```\nNon-applicability: none.",
                    "metadata": {"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
                }
            ]
        }

    async def fake_bundle_tests(**kwargs: Any) -> SkillTestResult:
        artifact = kwargs["artifact"]
        return SkillTestResult(
            result_id=f"{artifact.name}:mock_prestore_fail",
            skill_name=artifact.name,
            skill_version=artifact.version,
            bundle_id=artifact.bundle.bundle_id,
            bundle_version=artifact.bundle.bundle_version,
            run_label="spreadsheet_bundle_unit",
            aggregate={"passed": False, "pass_all_tests": False, "n_cases": len(artifact.bundle.all_cases())},
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._compat_ask_json", fake_ask_json)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter._execute_spreadsheet_bundle_tests", fake_bundle_tests)
    store = ArtifactStore()
    report = await SpreadsheetMaintenanceAdapter().run_micro_maintenance(
        detail=detail,
        credit_events=[],
        credit_bundle_cases=[],
        store=store,
        config=MaintenanceRunConfig(
            llm_config="mock",
            model_name="mock-model",
            extra={"spreadsheet_prestore_refine_max_rounds": 0},
        ),
        round_index=0,
        task_index=0,
    )

    assert store.get("spreadsheet_bad_double") is None
    assert report["maintenance_targets"] == []
    assert report["extraction_reports"][0]["status"] == "rejected"
    assert report["extraction_reports"][0]["rejection_reason"] == "prestore_bundle_test_failed"


async def test_spreadsheet_extractor_refines_candidate_before_prestore_reject(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_extract_refine", "Double A1 into B1.", source=4, answer=8)
    detail = _detail(
        task,
        success=True,
        score=1.0,
        retrieved=[],
        code="import openpyxl\nwb=openpyxl.load_workbook(INPUT_XLSX)\nws=wb['Sheet1']\nws['B1']=ws['A1'].value*2\nwb.save(OUTPUT_XLSX)",
    )

    async def fake_ask_json(**kwargs: Any) -> Dict[str, Any]:
        assert kwargs["role"] == "spreadsheet_extractor"
        return {
            "artifacts": [
                {
                    "name": "spreadsheet_repaired_double",
                    "kind": "executable_tool",
                    "description": "Helper for doubling A1 into B1.",
                    "body": "Applicability: double A1 into B1.\n```python\nws['B1'] = 0\nwb.save(OUTPUT_XLSX)\n```",
                    "metadata": {"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
                }
            ]
        }

    async def fake_bundle_tests(**kwargs: Any) -> SkillTestResult:
        artifact = kwargs["artifact"]
        passed = "* 2" in artifact.body or "*2" in artifact.body
        return SkillTestResult(
            result_id=f"{artifact.name}:mock_prestore_{'pass' if passed else 'fail'}",
            skill_name=artifact.name,
            skill_version=artifact.version,
            bundle_id=artifact.bundle.bundle_id,
            bundle_version=artifact.bundle.bundle_version,
            run_label="spreadsheet_bundle_unit",
            aggregate={"passed": passed, "pass_all_tests": passed, "n_cases": len(artifact.bundle.all_cases())},
        )

    async def fake_refine(*args: Any, **kwargs: Any) -> Dict[str, Any]:
        return {
            "decision": {"action": "refine_minor", "reason": "Repair wrong output assignment.", "version_kind": "minor"},
            "artifact": {
                "body": "Applicability: double A1 into B1.\n```python\nws['B1'] = ws['A1'].value * 2\nwb.save(OUTPUT_XLSX)\n```",
                "metadata": {"intent_keywords": ["double"]},
            },
        }

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._compat_ask_json", fake_ask_json)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter._execute_spreadsheet_bundle_tests", fake_bundle_tests)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.refine_skill_artifact_llm", fake_refine)
    store = ArtifactStore()
    report = await SpreadsheetMaintenanceAdapter().run_micro_maintenance(
        detail=detail,
        credit_events=[],
        credit_bundle_cases=[],
        store=store,
        config=MaintenanceRunConfig(
            llm_config="mock",
            model_name="mock-model",
            extra={"spreadsheet_prestore_refine_max_rounds": 1},
        ),
        round_index=0,
        task_index=0,
    )

    accepted = store.get("spreadsheet_repaired_double")
    assert accepted is not None
    assert accepted.status == "pending"
    assert "* 2" in accepted.body
    assert report["extraction_reports"][0]["status"] == "pending"
    assert len(report["extraction_reports"][0]["prestore_test_results"]) == 2
    assert report["extraction_reports"][0]["prestore_refine_decisions"][0]["action"] == "refine_minor"


def test_spreadsheet_coerce_executable_tool_overrides_informational_metadata(tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_coerce_functional", "Double A1 into B1.", source=3, answer=6)
    detail = _detail(
        task,
        success=True,
        score=1.0,
        retrieved=[],
        code="import openpyxl\nwb=openpyxl.load_workbook(INPUT_XLSX)\nws=wb['Sheet1']\nws['B1']=ws['A1'].value*2\nwb.save(OUTPUT_XLSX)",
    )

    artifact = _coerce_spreadsheet_artifact(
        {
            "name": "spreadsheet_double_from_llm",
            "kind": "executable_tool",
            "description": "Double A1 into B1 with openpyxl.",
            "body": """```python
import openpyxl
wb = openpyxl.load_workbook(INPUT_XLSX)
ws = wb["Sheet1"]
ws["B1"] = ws["A1"].value * 2
wb.save(OUTPUT_XLSX)
```""",
            "metadata": {"injection_type": "informational"},
            "interface": {"invocation_contract": {"injection_type": "informational"}},
        },
        detail=detail,
    )

    assert artifact is not None
    assert artifact.kind == "executable_tool"
    assert artifact.injection_type() == "functional"
    assert artifact.metadata["injection_type"] == "functional"
    assert artifact.interface.invocation_contract["injection_type"] == "functional"
    assert artifact.metadata["source_success"] is True
    assert artifact.bundle.positive_cases


async def test_spreadsheet_refiner_preserves_executable_injection_contract(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_refine_contract", "Double A1 into B1.", source=3, answer=6)
    skill = SkillArtifact(
        name="spreadsheet_refine_contract",
        kind="executable_tool",
        description="Double A1 into B1.",
        body="""```python
import openpyxl
wb = openpyxl.load_workbook(INPUT_XLSX)
ws = wb["Sheet1"]
ws["B1"] = ws["A1"].value * 2
wb.save(OUTPUT_XLSX)
```""",
        metadata={"domains": ["SpreadsheetBench"], "injection_type": "functional"},
        interface=SkillInterface(invocation_contract={"injection_type": "functional"}),
    )
    detail = _detail(
        task,
        success=False,
        score=0.0,
        retrieved=[skill.name],
        code="ws['B1']=0\nwb.save(OUTPUT_XLSX)",
    )

    async def fake_refine(artifact: SkillArtifact, **kwargs: Any) -> Dict[str, Any]:
        return {
            "decision": {"action": "refine_minor", "reason": "disable noisy helper", "version_kind": "minor"},
            "artifact": {"interface": {"summary": "Refined executable helper."}},
        }

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.refine_skill_artifact_llm", fake_refine)

    decision = await _refine_spreadsheet_skill_from_credit(
        artifact=skill,
        credit_context=[
            {
                "skill_name": skill.name,
                "judgment": "harmful",
                "effect_type": "retrieval_noise",
                "confidence": 0.95,
                "reason": "Injected helper was harmful.",
            }
        ],
        detail=detail,
        store=ArtifactStore([skill]),
        config=MaintenanceRunConfig(llm_config="mock"),
    )

    updated = decision["updated_artifact"]
    assert updated.metadata["injection_type"] == "functional"
    assert updated.interface.invocation_contract["injection_type"] == "functional"
    assert _is_spreadsheet_callable_skill(updated) is True


def test_spreadsheet_coerce_failure_repair_does_not_create_positive_case(tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_coerce_failure", "Double A1 into B1.", source=3, answer=6)
    detail = _detail(
        task,
        success=False,
        score=0.0,
        retrieved=[],
        code="import openpyxl\nwb=openpyxl.load_workbook(INPUT_XLSX)\nws=wb['Sheet1']\nws['B1']=0\nwb.save(OUTPUT_XLSX)",
    )

    artifact = _coerce_spreadsheet_artifact(
        {
            "name": "spreadsheet_repair_double_failure",
            "kind": "workflow_guardrail_card",
            "description": "Avoid writing zero for doubling tasks.",
            "body": "For doubling tasks, write source*2 instead of 0.",
            "metadata": {},
            "interface": {"invocation_contract": {"injection_type": "workflow"}},
        },
        detail=detail,
    )

    assert artifact is not None
    assert artifact.injection_type() == "workflow"
    assert artifact.metadata["source_success"] is False
    assert artifact.bundle.positive_cases == []
    assert artifact.bundle.negative_cases


async def test_spreadsheet_credit_creates_focused_case_and_micro_refines_before_bundle(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_train_2", "Double A1 into B1.", source=7, answer=14)
    skill = SkillArtifact(
        name="spreadsheet_bad_double",
        kind="executable_tool",
        description="Bad doubling pattern.",
        body="Wrongly write ws['B1'] = 0 for doubling tasks.",
        metadata={
            "domains": ["SpreadsheetBench", "formula_generation"],
            "allowed_tools": ["openpyxl"],
            "intent_keywords": ["double"],
            "source_task_ids": ["old"],
        },
        interface=SkillInterface(summary="Bad", usage="Use for doubling."),
    )
    store = ArtifactStore([skill])
    detail = _detail(task, success=False, score=0.0, retrieved=[skill.name], code="ws['B1']=0\nwb.save(OUTPUT_XLSX)")
    call_order: List[str] = []

    async def fake_ask_json(**kwargs: Any) -> Dict[str, Any]:
        call_order.append(kwargs["role"])
        return {
            "task_summary": {"task_id": "sheet_train_2", "score": 0.0, "success": False, "total_tokens": 120},
            "skill_judgments": [
                {
                    "skill_name": skill.name,
                    "judgment": "harmful",
                    "effect_type": "correctness_harm",
                    "confidence": 0.9,
                    "reason": "The retrieved skill writes the wrong answer cell value.",
                    "maintenance_actions": [{"action": "refine_workflow", "reason": "replace hard-coded zero", "target_scope": "doubling"}],
                    "refine_required": True,
                    "filter_candidate": False,
                    "evidence_strength": "strong",
                    "attribution_scope": "prompt_influence",
                    "bundle_case_suggestions": [
                        {
                            "polarity": "negative",
                            "reason": "Guard against writing zero instead of source*2.",
                            "source_task_id": "sheet_train_2",
                            "focus_turn_indices": [0],
                            "required_context_turn_indices": [],
                            "state_requirements": {},
                            "expected_contract": "B1 should match the official golden workbook.",
                            "task_fragment_policy": "reuse_official_fragment",
                        }
                    ],
                    "evidence": {"retrieved": True, "used": False, "trace_signals": ["B1 mismatch"]},
                }
            ],
        }

    async def fake_refine(*args: Any, **kwargs: Any) -> Dict[str, Any]:
        call_order.append("refiner")
        return {
            "decision": {"action": "refine_minor", "reason": "Use source*2, not zero.", "version_kind": "minor", "pinned_dependencies": []},
            "artifact": {
                "body": "Applicability: doubling tasks. ```python\nws['B1'] = ws['A1'].value * 2\nwb.save(OUTPUT_XLSX)\n``` Non-applicability: verify cells dynamically.",
                "metadata": {"last_refine_reason": "Use source*2."},
                "interface": {
                    "summary": "Double source cell.",
                    "usage": "Use for SpreadsheetBench doubling tasks.",
                    "input_contract": {"source_cell": "numeric"},
                    "output_contract": {"answer": "source*2"},
                    "invocation_contract": {"injection_type": "informational"},
                    "compatibility_notes": "Do not hard-code output values.",
                },
            },
            "bundle": {"maintenance_notes": "", "positive_cases": [], "negative_cases": [], "integration_cases": []},
        }

    async def fake_bundle_tests(**kwargs: Any):
        call_order.append("bundle_test")
        artifact = kwargs["artifact"]
        return SkillTestResult(
            result_id="mock_bundle_result",
            skill_name=artifact.name,
            skill_version=artifact.version,
            bundle_id=artifact.bundle.bundle_id,
            bundle_version=artifact.bundle.bundle_version,
            run_label="spreadsheet_bundle_unit",
            aggregate={"passed": True},
        )

    async def no_extract(*args: Any, **kwargs: Any) -> List[SkillArtifact]:
        return []

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._compat_ask_json", fake_ask_json)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._extract_spreadsheet_skills_from_detail", no_extract)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.refine_skill_artifact_llm", fake_refine)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter._execute_spreadsheet_bundle_tests", fake_bundle_tests)

    credit = await SpreadsheetMaintenanceAdapter().assign_credit(
        detail=detail,
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        task_index=0,
    )
    cases = await SpreadsheetMaintenanceAdapter().apply_credit_bundle_cases(
        detail=detail,
        credit_events=credit,
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        task_index=0,
    )
    report = await SpreadsheetMaintenanceAdapter().run_micro_maintenance(
        detail=detail,
        credit_events=credit,
        credit_bundle_cases=cases,
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        task_index=0,
    )

    assert cases and cases[0]["polarity"] == "negative"
    assert call_order.index("refiner") < call_order.index("bundle_test")
    assert report["refine_decisions"][0]["action"] == "refine_minor"
    assert "ws['A1'].value * 2" in store.get(skill.name).body


async def test_spreadsheet_credit_uses_compact_skill_projection_without_dropping_executable_code(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_compact_credit", "Double A1 into B1.", source=7, answer=14)
    long_guidance = "Applicability: double source cells.\n\n" + ("workflow note " * 900)
    workflow = SkillArtifact(
        name="spreadsheet_long_workflow",
        kind="workflow_guardrail_card",
        description="Long workflow guidance.",
        body=long_guidance,
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
    )
    code = """def spreadsheet_full_code(INPUT_XLSX, OUTPUT_XLSX, **kwargs):
    import openpyxl
    wb = openpyxl.load_workbook(INPUT_XLSX)
    ws = wb["Sheet1"]
    ws["B1"] = ws["A1"].value * 2
    wb.save(OUTPUT_XLSX)
    return OUTPUT_XLSX
"""
    executable = SkillArtifact(
        name="spreadsheet_full_code",
        kind="executable_tool",
        description="Full executable code.",
        body=f"Applicability: double A1.\n```python\n{code}```\nNon-applicability: other layouts.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
    )
    projection = {
        "retrieved_skills": [workflow.name, executable.name],
        "prompt_injected_skills": [workflow.name],
        "called_skill_functions": [executable.name],
    }
    workflow_view = _spreadsheet_skill_projection(workflow, projection=projection)
    executable_view = _spreadsheet_skill_projection(executable, projection=projection)

    assert workflow_view["body_projection"]["body_truncated"] is False
    assert "body" not in workflow_view
    assert workflow_view["body_projection"]["body"] == long_guidance
    assert workflow_view["body_projection"]["code_preview"] == ""
    assert executable_view["body_projection"]["projection_kind"] == "full_executable_body"
    assert code.strip() in executable_view["body_projection"]["code_preview"]


async def test_spreadsheet_refiner_prompt_preserves_long_body(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_compact_refiner", "Double A1 into B1.", source=7, answer=14)
    original_body = "Applicability: double cells.\n\n" + ("very long workflow body " * 700)
    skill = SkillArtifact(
        name="spreadsheet_refiner_compact",
        kind="workflow_guardrail_card",
        description="Long body skill.",
        body=original_body,
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
    )
    store = ArtifactStore([skill])
    detail = _detail(task, success=False, score=0.0, retrieved=[skill.name], prompt_injected=[skill.name], code="ws['B1']=0")
    captured: Dict[str, Any] = {}

    async def fake_refine(artifact: SkillArtifact, **kwargs: Any) -> Dict[str, Any]:
        captured["body"] = artifact.body
        captured["metadata"] = dict(artifact.metadata)
        return {
            "decision": {"action": "refine_minor", "reason": "append caveat", "version_kind": "minor"},
            "artifact": {"metadata": {"refined": True}},
            "bundle": {},
        }

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.refine_skill_artifact_llm", fake_refine)

    decision = await _refine_spreadsheet_skill_from_credit(
        artifact=skill,
        credit_context=[
            {
                "skill_name": skill.name,
                "judgment": "harmful",
                "confidence": 0.9,
                "reason": "wrong formula guidance",
            }
        ],
        detail=detail,
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
    )

    assert captured["body"] == original_body
    assert captured["metadata"]["spreadsheet_compact_projection"]["body_truncated"] is False
    assert captured["metadata"]["spreadsheet_compact_projection"]["prompt_body_preserved"] is True
    assert decision["updated_artifact"].body == original_body
    assert decision["updated_artifact"].metadata["refined"] is True


async def test_spreadsheet_micro_does_not_refine_helpful_positive_credit(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_helpful_micro", "Double A1 into B1.", source=7, answer=14)
    skill = SkillArtifact(
        name="spreadsheet_helpful_double",
        kind="workflow_guardrail_card",
        description="Helpful doubling guidance.",
        body="Use the source cell value instead of hard-coded constants.",
        metadata={"domains": ["SpreadsheetBench", "formula_generation"], "intent_keywords": ["double"]},
    )
    store = ArtifactStore([skill])
    detail = _detail(
        task,
        success=True,
        score=1.0,
        retrieved=[skill.name],
        prompt_injected=[skill.name],
        code="ws['B1']=ws['A1'].value*2\nwb.save(OUTPUT_XLSX)",
    )
    calls: List[str] = []

    async def fake_refine(*args: Any, **kwargs: Any) -> Dict[str, Any]:
        calls.append("refine")
        return {"decision": {"action": "keep"}, "artifact": {}, "bundle": {}}

    async def fake_bundle_tests(**kwargs: Any) -> SkillTestResult:
        calls.append("bundle_test")
        artifact = kwargs["artifact"]
        return SkillTestResult(
            result_id="unexpected_helpful_bundle",
            skill_name=artifact.name,
            skill_version=artifact.version,
            bundle_id=artifact.bundle.bundle_id,
            bundle_version=artifact.bundle.bundle_version,
            aggregate={"passed": True},
        )

    async def no_extract(*args: Any, **kwargs: Any) -> List[SkillArtifact]:
        return []

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._extract_spreadsheet_skills_from_detail", no_extract)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.refine_skill_artifact_llm", fake_refine)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter._execute_spreadsheet_bundle_tests", fake_bundle_tests)

    positive_case = {
        "skill_name": skill.name,
        "case_id": f"{skill.name}:positive:sheet_helpful_micro",
        "polarity": "positive",
        "created": True,
    }
    report = await SpreadsheetMaintenanceAdapter().run_micro_maintenance(
        detail=detail,
        credit_events=[
            {
                "skill_name": skill.name,
                "judgment": "helpful",
                "effect_type": "correctness_gain",
                "confidence": 0.95,
                "reason": "The injected skill aligned the formula.",
                "refine_required": False,
                "filter_candidate": False,
            }
        ],
        credit_bundle_cases=[positive_case],
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        task_index=0,
    )

    assert report["maintenance_targets"] == []
    assert report["micro_target_reasons"] == {}
    assert report["credit_bundle_cases"] == [positive_case]
    assert report["micro_maintenance_credit_events"] == []
    assert report["micro_maintenance_credit_bundle_cases"] == []
    assert calls == []


async def test_spreadsheet_micro_refines_strong_harmful_without_bundle_case(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_harmful_micro", "Double A1 into B1.", source=7, answer=14)
    skill = SkillArtifact(
        name="spreadsheet_harmful_double",
        kind="workflow_guardrail_card",
        description="Bad doubling guidance.",
        body="Write zero for doubling tasks.",
        metadata={"domains": ["SpreadsheetBench", "formula_generation"], "intent_keywords": ["double"]},
    )
    store = ArtifactStore([skill])
    detail = _detail(
        task,
        success=False,
        score=0.0,
        retrieved=[skill.name],
        prompt_injected=[skill.name],
        code="ws['B1']=0\nwb.save(OUTPUT_XLSX)",
    )
    calls: List[str] = []

    async def fake_refine(*args: Any, **kwargs: Any) -> Dict[str, Any]:
        calls.append("refine")
        return {"decision": {"action": "keep", "reason": "mock"}, "artifact": {}, "bundle": {}}

    async def no_extract(*args: Any, **kwargs: Any) -> List[SkillArtifact]:
        return []

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._extract_spreadsheet_skills_from_detail", no_extract)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.refine_skill_artifact_llm", fake_refine)

    report = await SpreadsheetMaintenanceAdapter().run_micro_maintenance(
        detail=detail,
        credit_events=[
            {
                "skill_name": skill.name,
                "judgment": "harmful",
                "effect_type": "correctness_harm",
                "confidence": 0.9,
                "reason": "The skill caused a wrong output.",
                "refine_required": False,
                "filter_candidate": False,
            }
        ],
        credit_bundle_cases=[],
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        task_index=0,
    )

    assert report["maintenance_targets"] == [skill.name]
    assert report["micro_target_reasons"] == {skill.name: ["strong_harmful_credit"]}
    assert report["maintenance_test_results"][0]["aggregate"]["reason"] == "no_bundle_cases"
    assert calls == ["refine"]


async def test_spreadsheet_micro_refines_relevant_retrieved_only_skill(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_retrieved_only_refine", "Double A1 into B1.", source=7, answer=14)
    skill = SkillArtifact(
        name="spreadsheet_retrieved_only_double",
        kind="workflow_guardrail_card",
        description="Ambiguous doubling guidance.",
        body="Use for arithmetic edits.",
        metadata={"domains": ["SpreadsheetBench", "formula_generation"], "intent_keywords": ["double"]},
    )
    store = ArtifactStore([skill])
    detail = _detail(
        task,
        success=True,
        score=1.0,
        retrieved=[skill.name],
        prompt_injected=[],
        called=[],
        code="ws['B1']=ws['A1'].value*2\nwb.save(OUTPUT_XLSX)",
    )
    calls: List[str] = []

    async def fake_refine(*args: Any, **kwargs: Any) -> Dict[str, Any]:
        calls.append("refine")
        return {
            "decision": {"action": "refine_minor", "reason": "Clarify when to apply to A1-to-B1 doubling.", "version_kind": "minor"},
            "artifact": {
                "body": "Use for SpreadsheetBench tasks that double a source cell into an answer cell; do not require exact old layout names.",
                "metadata": {"retrieval_scope_refined": True},
            },
            "bundle": {},
        }

    async def no_extract(*args: Any, **kwargs: Any) -> List[SkillArtifact]:
        return []

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._extract_spreadsheet_skills_from_detail", no_extract)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.refine_skill_artifact_llm", fake_refine)

    report = await SpreadsheetMaintenanceAdapter().run_micro_maintenance(
        detail=detail,
        credit_events=[
            {
                "skill_name": skill.name,
                "judgment": "neutral",
                "effect_type": "domain_match",
                "confidence": 0.8,
                "reason": "The skill is semantically relevant but its usage guidance was too vague to adopt.",
                "maintenance_actions": [{"action": "refine_workflow", "reason": "broaden and clarify applicability"}],
                "refine_required": True,
                "filter_candidate": False,
                "failure_mode": "skill_scope_too_broad",
                "evidence_strength": "medium",
                "attribution_scope": "retrieval_noise",
                "evidence": {"retrieved": True, "retrieved_only": True, "used": False},
                "projection": {
                    "retrieved_skills": [skill.name],
                    "retrieved_only_skills": [skill.name],
                    "prompt_injected_skills": [],
                    "called_skill_functions": [],
                },
            }
        ],
        credit_bundle_cases=[],
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        task_index=0,
    )

    assert report["maintenance_targets"] == [skill.name]
    assert report["micro_target_reasons"] == {skill.name: ["refine_required", "retrieved_only_scope_refine"]}
    assert report["refine_decisions"][0]["action"] == "refine_minor"
    assert store.get(skill.name).metadata["retrieval_scope_refined"] is True
    assert calls == ["refine"]


async def test_spreadsheet_micro_does_not_refine_irrelevant_retrieved_only_skill(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_irrelevant_retrieved_only", "Double A1 into B1.", source=7, answer=14)
    skill = SkillArtifact(
        name="spreadsheet_irrelevant_date_skill",
        kind="workflow_guardrail_card",
        description="Date extraction guidance.",
        body="Use for date parsing tasks.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["date"]},
    )
    store = ArtifactStore([skill])
    detail = _detail(task, success=True, score=1.0, retrieved=[skill.name], prompt_injected=[], code="ws['B1']=14")
    calls: List[str] = []

    async def fake_refine(*args: Any, **kwargs: Any) -> Dict[str, Any]:
        calls.append("refine")
        return {"decision": {"action": "keep"}, "artifact": {}, "bundle": {}}

    async def no_extract(*args: Any, **kwargs: Any) -> List[SkillArtifact]:
        return []

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._extract_spreadsheet_skills_from_detail", no_extract)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.refine_skill_artifact_llm", fake_refine)

    report = await SpreadsheetMaintenanceAdapter().run_micro_maintenance(
        detail=detail,
        credit_events=[
            {
                "skill_name": skill.name,
                "judgment": "neutral",
                "effect_type": "domain_mismatch",
                "confidence": 0.9,
                "reason": "The date skill is unrelated to doubling.",
                "refine_required": False,
                "filter_candidate": False,
                "failure_mode": "irrelevant_retrieval",
                "evidence_strength": "medium",
                "attribution_scope": "retrieval_noise",
                "evidence": {"retrieved": True, "retrieved_only": True, "used": False},
                "projection": {"retrieved_only_skills": [skill.name]},
            }
        ],
        credit_bundle_cases=[],
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        task_index=0,
    )

    assert report["maintenance_targets"] == []
    assert report["micro_target_reasons"] == {}
    assert calls == []


async def test_spreadsheet_refiner_failure_disables_high_confidence_harmful_skill(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_refiner_failure_harm", "Double A1 into B1.", source=7, answer=14)
    skill = SkillArtifact(
        name="spreadsheet_noisy_harmful",
        kind="workflow_guardrail_card",
        description="Noisy irrelevant guidance.",
        body="Use only for unrelated date extraction tasks.",
        metadata={"domains": ["SpreadsheetBench"], "allowed_tools": ["openpyxl"]},
        status="active",
    )
    detail = _detail(
        task,
        success=False,
        score=0.0,
        retrieved=[skill.name],
        prompt_injected=[skill.name],
        code="ws['B1'] = 0",
    )

    async def failing_refiner(*args: Any, **kwargs: Any) -> Dict[str, Any]:
        raise ValueError("bad json")

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.refine_skill_artifact_llm", failing_refiner)
    decision = await _refine_spreadsheet_skill_from_credit(
        artifact=skill,
        credit_context=[
            {
                "skill_name": skill.name,
                "judgment": "harmful",
                "effect_type": "workflow_pollution",
                "confidence": 0.9,
                "refine_required": True,
                "filter_candidate": False,
            }
        ],
        detail=detail,
        store=ArtifactStore([skill]),
        config=MaintenanceRunConfig(llm_config="mock"),
    )

    assert decision["action"] == "disable"
    assert decision["updated_artifact"].status == "disabled"
    assert decision["reason"] == "spreadsheet_refiner_unavailable_after_high_confidence_harmful_credit:ValueError"


async def test_spreadsheet_micro_tests_integration_case_even_with_neutral_credit(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_integration_micro", "Double A1 into B1.", source=7, answer=14)
    skill = SkillArtifact(
        name="spreadsheet_integration_guard",
        kind="workflow_guardrail_card",
        description="Integration guard.",
        body="Keep source workbook state intact while writing the answer.",
        metadata={"domains": ["SpreadsheetBench", "formula_generation"], "intent_keywords": ["double"]},
    )
    store = ArtifactStore([skill])
    detail = _detail(task, success=False, score=0.0, retrieved=[skill.name], prompt_injected=[skill.name], code="")
    calls: List[str] = []

    async def fake_refine(*args: Any, **kwargs: Any) -> Dict[str, Any]:
        calls.append("refine")
        return {"decision": {"action": "keep"}, "artifact": {}, "bundle": {}}

    async def fake_bundle_tests(**kwargs: Any) -> SkillTestResult:
        calls.append("bundle_test")
        artifact = kwargs["artifact"]
        return SkillTestResult(
            result_id="integration_bundle_result",
            skill_name=artifact.name,
            skill_version=artifact.version,
            bundle_id=artifact.bundle.bundle_id,
            bundle_version=artifact.bundle.bundle_version,
            aggregate={"passed": True},
        )

    async def no_extract(*args: Any, **kwargs: Any) -> List[SkillArtifact]:
        return []

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._extract_spreadsheet_skills_from_detail", no_extract)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.refine_skill_artifact_llm", fake_refine)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter._execute_spreadsheet_bundle_tests", fake_bundle_tests)

    skill.bundle.integration_cases.append(
        SkillBundleCase(
            case_id="integration_case",
            source="manual_integration",
            prompt="Double A1 into B1.",
            polarity="integration",
        )
    )
    report = await SpreadsheetMaintenanceAdapter().run_micro_maintenance(
        detail=detail,
        credit_events=[
            {
                "skill_name": skill.name,
                "judgment": "neutral",
                "confidence": 0.5,
                "reason": "No direct causal evidence.",
            }
        ],
        credit_bundle_cases=[
            {
                "skill_name": skill.name,
                "case_id": "integration_case",
                "polarity": "integration",
                "created": True,
            }
        ],
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        task_index=0,
    )

    assert report["maintenance_targets"] == [skill.name]
    assert report["micro_target_reasons"] == {skill.name: ["integration_bundle_case"]}
    assert report["refine_decisions"][0]["action"] == "keep"
    assert report["refine_decisions"][0]["reason"] == "no_strong_credit_signal"
    assert calls == ["bundle_test"]


async def test_spreadsheet_credit_includes_retrieved_only_for_relevance_judgment(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_credit_candidates", "Double A1 into B1.", source=7, answer=14)
    retrieved_only = SkillArtifact(
        name="spreadsheet_retrieved_only",
        kind="workflow_guardrail_card",
        description="Should receive relevance judgment even if not injected.",
        body="Retrieved-only guidance.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
    )
    injected = SkillArtifact(
        name="spreadsheet_injected_workflow",
        kind="workflow_guardrail_card",
        description="Injected workflow guidance.",
        body="Injected guidance.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
    )
    called = SkillArtifact(
        name="spreadsheet_called_function",
        kind="executable_tool",
        description="Called function skill.",
        body="```python\ndef spreadsheet_called_function(INPUT_XLSX, OUTPUT_XLSX, **kwargs):\n    return OUTPUT_XLSX\n```",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
    )
    store = ArtifactStore([retrieved_only, injected, called])
    detail = _detail(
        task,
        success=True,
        score=1.0,
        retrieved=[retrieved_only.name, injected.name, called.name],
        prompt_injected=[injected.name],
        called=[called.name],
        code="from skill_library import spreadsheet_called_function\nspreadsheet_called_function(INPUT_XLSX, OUTPUT_XLSX)",
    )
    captured: Dict[str, Any] = {}

    async def fake_ask_json(**kwargs: Any) -> Dict[str, Any]:
        captured["payload"] = kwargs["user"]
        payload = json.loads(kwargs["user"])
        candidate_names = [item["skill_name"] for item in payload["candidate_skills"]]
        assert candidate_names == [injected.name, called.name, retrieved_only.name]
        assert payload["retrieval_audit"]["retrieved_only_skills"] == [retrieved_only.name]
        return {
            "task_summary": {"task_id": task.task_id, "score": 1.0, "success": True},
            "skill_judgments": [
                {
                    "skill_name": retrieved_only.name,
                    "judgment": "neutral",
                    "effect_type": "domain_match",
                    "confidence": 0.7,
                    "reason": "relevant by wording but not adopted by the executor",
                    "maintenance_actions": [],
                    "refine_required": False,
                    "filter_candidate": False,
                    "failure_mode": "insufficient_evidence",
                    "evidence_strength": "weak",
                    "attribution_scope": "retrieval_noise",
                    "bundle_case_suggestions": [],
                    "evidence": {"retrieved": True, "retrieved_only": True, "used": False},
                },
                {
                    "skill_name": injected.name,
                    "judgment": "neutral",
                    "effect_type": "no_material_effect",
                    "confidence": 0.5,
                    "reason": "present but not causal",
                    "maintenance_actions": [],
                    "refine_required": False,
                    "filter_candidate": False,
                    "evidence_strength": "weak",
                    "attribution_scope": "prompt_influence",
                    "bundle_case_suggestions": [],
                    "evidence": {"retrieved": True, "injected": True, "used": False},
                },
                {
                    "skill_name": called.name,
                    "judgment": "helpful",
                    "effect_type": "correctness_gain",
                    "confidence": 0.8,
                    "reason": "called helper matched task",
                    "maintenance_actions": [],
                    "refine_required": False,
                    "filter_candidate": False,
                    "evidence_strength": "strong",
                    "attribution_scope": "direct_use",
                    "bundle_case_suggestions": [],
                    "evidence": {"retrieved": True, "injected": False, "used": True},
                },
            ],
        }

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter._ask_json", fake_ask_json)

    events = await SpreadsheetMaintenanceAdapter().assign_credit(
        detail=detail,
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        task_index=0,
    )

    assert [event["skill_name"] for event in events] == [injected.name, called.name, retrieved_only.name]
    assert "retrieved_only_skills" in captured["payload"]
    assert "prompt_injected_or_called_plus_retrieved_only_relevance_gate" in captured["payload"]
    assert store.get(retrieved_only.name).usage_count == 0
    assert store.get(injected.name).usage_count == 1
    assert store.get(called.name).success_count == 1


async def test_spreadsheet_credit_fallback_ignores_retrieved_only_skill(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_credit_fallback", "Double A1 into B1.", source=7, answer=14)
    retrieved_only = SkillArtifact(
        name="spreadsheet_fallback_retrieved_only",
        kind="workflow_guardrail_card",
        description="Should not receive fallback credit if not injected.",
        body="Retrieved-only guidance.",
        metadata={"domains": ["SpreadsheetBench", "formula_generation"], "intent_keywords": ["double"]},
    )
    injected = SkillArtifact(
        name="spreadsheet_fallback_injected",
        kind="workflow_guardrail_card",
        description="Injected guidance.",
        body="Injected guidance.",
        metadata={"domains": ["SpreadsheetBench", "formula_generation"], "intent_keywords": ["double"]},
    )
    store = ArtifactStore([retrieved_only, injected])
    detail = _detail(
        task,
        success=True,
        score=1.0,
        retrieved=[retrieved_only.name, injected.name],
        prompt_injected=[injected.name],
        code="ws['B1']=ws['A1'].value*2\nwb.save(OUTPUT_XLSX)",
    )

    async def failing_ask_json(**kwargs: Any) -> Dict[str, Any]:
        raise RuntimeError("mock credit failure")

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter._ask_json", failing_ask_json)

    events = await SpreadsheetMaintenanceAdapter().assign_credit(
        detail=detail,
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        task_index=0,
    )

    assert [event["skill_name"] for event in events] == [injected.name, retrieved_only.name]
    assert events[1]["judgment"] == "neutral"
    assert events[1]["evidence"]["retrieved_only"] is True
    assert events[1]["evidence"]["retrieved"] is True
    assert events[1]["evidence"]["injected"] is False
    assert events[0]["evidence"]["injected"] is True
    assert store.get(retrieved_only.name).usage_count == 0


async def test_spreadsheet_failed_formula_mismatch_extracts_repair_skill_when_llm_empty(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_formula_fail", "Fill C2:C3 with A+B formulas, leaving blank if either input is blank.")
    detail = _detail(
        task,
        success=False,
        score=0.0,
        retrieved=[],
        code='ws["C2"] = "=IF(AND(A2<>\\"\\", B2<>\\"\\"), A2+B2, \\"\\")"',
    )
    detail["runs"][0]["metrics"]["mismatched_cells"] = [
        {
            "cell": "C2",
            "predicted": '=IF(AND(A2<>"", B2<>""), A2+B2, "")',
            "expected": '=IF(OR(A2="",B2=""),"",A2+B2)',
        }
    ]

    async def fake_ask_json(**kwargs: Any) -> Dict[str, Any]:
        assert kwargs["role"] == "spreadsheet_extractor"
        return {"artifacts": []}

    async def fake_bundle_tests(**kwargs: Any) -> SkillTestResult:
        return _passing_skill_test_result(kwargs["artifact"])

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter._ask_json", fake_ask_json)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter._execute_spreadsheet_bundle_tests", fake_bundle_tests)
    store = ArtifactStore()
    report = await SpreadsheetMaintenanceAdapter().run_micro_maintenance(
        detail=detail,
        credit_events=[],
        credit_bundle_cases=[],
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        task_index=0,
    )

    assert report["extraction_reports"]
    artifact = store.get(report["extraction_reports"][0]["skill_name"])
    assert artifact is not None
    assert artifact.status == "pending"
    assert artifact.metadata["extracted_from_failure"] is True
    assert "expected" in artifact.body


async def test_spreadsheet_bundle_replay_uses_real_workbook_and_injected_skill(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_bundle_1", "Double A1 into B1.", source=9, answer=18)
    skill = SkillArtifact(
        name="spreadsheet_double_openpyxl",
        kind="executable_tool",
        description="Double A1 to B1.",
        body="Use openpyxl to set B1 to A1*2.",
        metadata={"domains": ["SpreadsheetBench"], "allowed_tools": ["openpyxl"], "intent_keywords": ["double"]},
    )
    detail = _detail(task, success=True, score=1.0, retrieved=[], code="")
    from academic.benchmarks.spreadsheet.adapter import _spreadsheet_case_from_task

    skill.bundle.positive_cases.append(
        _spreadsheet_case_from_task(
            detail=detail,
            skill_name=skill.name,
            polarity="positive",
            reason="replay doubling task",
            source="manual",
            confidence=1.0,
        )
    )

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        return TextLLMResponse(
            content="""```python
import openpyxl
wb = openpyxl.load_workbook(INPUT_XLSX)
ws = wb["Sheet1"]
ws["B1"] = ws["A1"].value * 2
wb.save(OUTPUT_XLSX)
```""",
            prompt_tokens=20,
            completion_tokens=10,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)
    result = await _execute_spreadsheet_bundle_tests(
        artifact=skill,
        config=MaintenanceRunConfig(llm_config="mock", model_name="mock-model"),
    )

    assert result.aggregate["passed"] is True
    assert result.aggregate["avg_accuracy"] == 1.0
    assert [run.variant for run in result.unit_case_runs] == ["with_skill", "without_skill"]
    assert result.unit_case_runs[0].tokens == 30
    assert result.aggregate["with_before_without"] is True
    assert result.counterfactual["with_skill_valid_by_task"] == {"sheet_bundle_1": True}
    assert result.counterfactual["without_skill_valid_by_task"] == {"sheet_bundle_1": True}


async def test_spreadsheet_bundle_strict_gate_rejects_missing_paths_before_llm(monkeypatch) -> None:
    skill = SkillArtifact(
        name="spreadsheet_strict_gate_skill",
        kind="workflow_guardrail_card",
        description="Gate test skill.",
        body="Use openpyxl carefully.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
    )
    skill.bundle.positive_cases.append(
        SkillBundleCase(
            case_id="strict_gate_missing_paths",
            source="manual",
            prompt="Double A1 into B1.",
            expected={"verifier": "spreadsheet_golden_range"},
            context={
                "task_fragment": {
                    "task_id": "strict_gate_missing_paths",
                    "question": "Double A1 into B1.",
                    "expected": {"answer_sheet": "Sheet1", "answer_position": "B1"},
                    "input_artifacts": {},
                    "metadata": {},
                }
            },
            polarity="positive",
        )
    )
    calls: List[str] = []

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        calls.append("llm")
        return TextLLMResponse(content="```python\npass\n```")

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)

    result = await _execute_spreadsheet_bundle_tests(
        artifact=skill,
        config=MaintenanceRunConfig(llm_config="mock", model_name="mock-model"),
    )

    assert calls == []
    assert result.aggregate["passed"] is False
    assert result.aggregate["strict_contract_gate"] is True
    assert result.aggregate["n_strict_failures"] == 1
    assert result.unit_case_runs[0].variant == "strict_gate"
    assert result.unit_case_runs[0].metadata["contract_failures"][0]["type"] == "missing_input_xlsx"


async def test_spreadsheet_bundle_skips_without_skill_when_with_skill_fails(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_bundle_with_fail", "Double A1 into B1.", source=9, answer=18)
    skill = SkillArtifact(
        name="spreadsheet_bundle_with_fail",
        kind="workflow_guardrail_card",
        description="Bad replay skill.",
        body="Use openpyxl.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
    )
    from academic.benchmarks.spreadsheet.adapter import _spreadsheet_case_from_task

    skill.bundle.positive_cases.append(
        _spreadsheet_case_from_task(
            detail=_detail(task, success=True, score=1.0, retrieved=[], code=""),
            skill_name=skill.name,
            polarity="positive",
            reason="with skill failure",
            source="manual",
            confidence=1.0,
        )
    )
    calls: List[str] = []

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        calls.append("llm")
        return TextLLMResponse(
            content="""```python
import openpyxl
wb = openpyxl.load_workbook(INPUT_XLSX)
ws = wb["Sheet1"]
ws["B1"] = 0
wb.save(OUTPUT_XLSX)
```""",
            prompt_tokens=5,
            completion_tokens=5,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)

    result = await _execute_spreadsheet_bundle_tests(
        artifact=skill,
        config=MaintenanceRunConfig(llm_config="mock", model_name="mock-model"),
    )

    assert calls == ["llm"]
    assert [run.variant for run in result.unit_case_runs] == ["with_skill"]
    assert result.unit_case_runs[0].passed is False
    assert result.aggregate["passed"] is False
    assert result.counterfactual["without_skill_valid_by_task"] == {}


async def test_spreadsheet_bundle_runs_without_skill_only_after_with_skill_passes(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_bundle_with_then_without", "Double A1 into B1.", source=9, answer=18)
    skill = SkillArtifact(
        name="spreadsheet_bundle_with_then_without",
        kind="workflow_guardrail_card",
        description="Replay skill.",
        body="Use openpyxl.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
    )
    from academic.benchmarks.spreadsheet.adapter import _spreadsheet_case_from_task

    skill.bundle.positive_cases.append(
        _spreadsheet_case_from_task(
            detail=_detail(task, success=True, score=1.0, retrieved=[], code=""),
            skill_name=skill.name,
            polarity="positive",
            reason="with skill pass",
            source="manual",
            confidence=1.0,
        )
    )
    injected_by_call: List[List[str]] = []

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        injected_by_call.append(
            [skill.name] if skill.name in kwargs["system"] else []
        )
        return TextLLMResponse(
            content="""```python
import openpyxl
wb = openpyxl.load_workbook(INPUT_XLSX)
ws = wb["Sheet1"]
ws["B1"] = ws["A1"].value * 2
wb.save(OUTPUT_XLSX)
```""",
            prompt_tokens=5,
            completion_tokens=5,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)

    result = await _execute_spreadsheet_bundle_tests(
        artifact=skill,
        config=MaintenanceRunConfig(llm_config="mock", model_name="mock-model"),
    )

    assert [run.variant for run in result.unit_case_runs] == ["with_skill", "without_skill"]
    assert injected_by_call == [[skill.name], []]
    assert result.aggregate["passed"] is True
    assert result.aggregate["n_comparable_cases"] == 1


async def test_spreadsheet_bundle_replay_respects_bash_react_execution_mode(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_bundle_bash_react", "Double A1 into B1.", source=9, answer=18)
    skill = SkillArtifact(
        name="spreadsheet_bundle_bash_react",
        kind="executable_tool",
        description="Replay skill.",
        body="Use openpyxl.",
        metadata={"domains": ["SpreadsheetBench"], "intent_keywords": ["double"]},
    )
    from academic.benchmarks.spreadsheet.adapter import _spreadsheet_case_from_task

    skill.bundle.positive_cases.append(
        _spreadsheet_case_from_task(
            detail=_detail(task, success=True, score=1.0, retrieved=[], code=""),
            skill_name=skill.name,
            polarity="positive",
            reason="bash_react replay",
            source="manual",
            confidence=1.0,
        )
    )
    calls: List[Dict[str, Any]] = []

    async def fake_run_bash(task: BenchmarkTask, **kwargs: Any) -> BenchmarkResult:
        calls.append(kwargs)
        return BenchmarkResult(
            benchmark="spreadsheet",
            task_id=task.task_id,
            success=True,
            score=1.0,
            metrics={
                "answer_sheet": "Sheet1",
                "answer_position": "B1",
                "checked_cells": 1,
                "mismatched_cells": [],
                "execution_ok": True,
                "total_tokens": 10,
                "execution_mode": "bash_react",
            },
            trace={},
        )

    async def fail_single(*args: Any, **kwargs: Any) -> BenchmarkResult:
        raise AssertionError("bundle replay should not use single-shot executor")

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.run_spreadsheet_task_bash_react", fake_run_bash)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.run_spreadsheet_task", fail_single)

    result = await _execute_spreadsheet_bundle_tests(
        artifact=skill,
        config=MaintenanceRunConfig(
            llm_config="mock",
            model_name="mock-model",
            max_task_seconds=99,
            extra={"spreadsheet_execution_mode": "bash_react", "spreadsheet_max_turns": 17, "llm_request_timeout_s": 88},
        ),
    )

    assert result.aggregate["passed"] is True
    assert [run.variant for run in result.unit_case_runs] == ["with_skill", "without_skill"]
    assert [call["top_k_skills"] for call in calls] == [1, 0]
    assert [len(call["artifact_store"].all()) for call in calls] == [1, 0]
    assert all(call["max_turns"] == 17 for call in calls)
    assert all(call["llm_request_timeout_s"] == 88 for call in calls)
    assert calls[0]["skill_injector_mode"] == "direct"
    assert calls[0]["min_skill_score"] == 0.0
    assert calls[1]["skill_injector_mode"] is None


def _package_double_skill() -> SkillArtifact:
    return SkillArtifact(
        name="spreadsheet_package_double",
        kind="skill_package",
        description="Folder skill that doubles A1 into B1.",
        body="Read SKILL.md, then run scripts/apply_double.py when the task asks for A1 doubled into B1.",
        metadata={
            "domains": ["SpreadsheetBench", "formula_generation"],
            "intent_keywords": ["double", "A1", "B1"],
            "package_format": "skills_md",
            "package_files": {
                "SKILL.md": "---\nname: spreadsheet_package_double\ndescription: Double A1 into B1.\n---\n\n# Double A1\nUse when B1 should equal A1 * 2.",
                "scripts/apply_double.py": (
                    "from __future__ import annotations\n"
                    "import sys\n"
                    "import openpyxl\n\n"
                    "def apply(input_xlsx, output_xlsx):\n"
                    "    wb = openpyxl.load_workbook(input_xlsx)\n"
                    "    ws = wb['Sheet1']\n"
                    "    ws['B1'] = ws['A1'].value * 2\n"
                    "    wb.save(output_xlsx)\n\n"
                    "if __name__ == '__main__':\n"
                    "    apply(sys.argv[1], sys.argv[2])\n"
                ),
                "references/notes.md": "Only for Sheet1 A1 -> B1 doubling tasks.",
            },
            "bundle_files": {
                "run_tests.py": (
                    "from pathlib import Path\n"
                    "import sys\n"
                    "import openpyxl\n\n"
                    "ROOT = Path(__file__).resolve().parents[2]\n"
                    "sys.path.insert(0, str(ROOT / 'skills' / 'spreadsheet_package_double' / 'scripts'))\n"
                    "from apply_double import apply\n\n"
                    "inp = Path('unit_input.xlsx')\n"
                    "out = Path('unit_output.xlsx')\n"
                    "wb = openpyxl.Workbook()\n"
                    "ws = wb.active\n"
                    "ws.title = 'Sheet1'\n"
                    "ws['A1'] = 11\n"
                    "wb.save(inp)\n"
                    "apply(inp, out)\n"
                    "wb2 = openpyxl.load_workbook(out)\n"
                    "assert wb2['Sheet1']['B1'].value == 22\n"
                ),
            },
        },
        interface=SkillInterface(invocation_contract={"injection_type": "functional"}),
    )


def test_spreadsheet_package_skill_materializes_folder_files(tmp_path: Path) -> None:
    skill = _package_double_skill()
    result = _write_spreadsheet_skill_packages([skill], tmp_path)

    assert _is_spreadsheet_package_skill(skill) is True
    assert result["skills"][0]["skill_dir"] == "skills/spreadsheet_package_double"
    assert (tmp_path / "skills" / "spreadsheet_package_double" / "SKILL.md").exists()
    assert (tmp_path / "skills" / "spreadsheet_package_double" / "scripts" / "apply_double.py").exists()
    assert (tmp_path / "bundles" / "spreadsheet_package_double" / "run_tests.py").exists()
    assert "Folder-style Spreadsheet skills" in result["prompt"]
    assert "skills/spreadsheet_package_double/SKILL.md" in result["prompt"]


async def test_spreadsheet_bash_react_injects_folder_skill_package(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_package_bash", "Double A1 into B1.", source=8, answer=16)
    skill = _package_double_skill()
    skill.status = "pending"

    class StepStore(ArtifactStore):
        def __init__(self) -> None:
            super().__init__([skill])

        def retrieve(self, query: str, top_k: int = 5, **kwargs: Any) -> List[SkillArtifact]:
            if "NameError" in query:
                return [skill]
            return []

    prompts: List[str] = []

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        prompts.append(str(kwargs.get("prompt") or ""))
        if len(prompts) == 1:
            return TextLLMResponse(content="```bash\npython - <<'PY'\nraise NameError('need package')\nPY\n```")
        assert "New folder-style Spreadsheet skill packages:" in kwargs["prompt"]
        assert "skills/spreadsheet_package_double/SKILL.md" in kwargs["prompt"]
        assert "skills/spreadsheet_package_double/scripts/apply_double.py" in kwargs["prompt"]
        assert "Entry manifest:" in kwargs["prompt"]
        return TextLLMResponse(
            content=f"""```bash
python skills/spreadsheet_package_double/scripts/apply_double.py "$INPUT_XLSX" "$OUTPUT_XLSX"
```
{SPREADSHEET_DONE_PATTERN}""",
            prompt_tokens=10,
            completion_tokens=5,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.ask_text_llm", fake_ask_text_llm)
    result = await run_spreadsheet_task_bash_react(
        task,
        llm_config="mock",
        model_name="mock-model",
        artifact_store=StepStore(),
        top_k_skills=1,
        pending_skill_fraction=1.0,
        max_turns=3,
        work_dir=tmp_path / "package_bash_work",
    )

    assert result.success is True
    assert result.metrics["package_skills"][0]["skill_name"] == "spreadsheet_package_double"
    assert (tmp_path / "package_bash_work" / "skills" / "spreadsheet_package_double" / "scripts" / "apply_double.py").exists()


async def test_spreadsheet_package_bundle_unit_test_runs_without_replay(tmp_path: Path) -> None:
    skill = _package_double_skill()
    result = await _execute_spreadsheet_bundle_tests(
        artifact=skill,
        config=MaintenanceRunConfig(llm_config="mock", model_name="mock-model", max_task_seconds=20),
    )

    assert result.aggregate["passed"] is True
    assert result.aggregate["n_cases"] == 1
    assert [run.variant for run in result.unit_case_runs] == ["package_unit"]
    assert result.unit_case_runs[0].metadata["package_unit"] is True


async def test_spreadsheet_package_bundle_unit_supports_scripts_package_import(tmp_path: Path) -> None:
    skill = _package_double_skill()
    skill.metadata["bundle_files"] = {
        "run_tests.py": (
            "from pathlib import Path\n"
            "import openpyxl\n"
            "from scripts.apply_double import apply\n\n"
            "inp = Path('unit_input.xlsx')\n"
            "out = Path('unit_output.xlsx')\n"
            "wb = openpyxl.Workbook()\n"
            "ws = wb.active\n"
            "ws.title = 'Sheet1'\n"
            "ws['A1'] = 13\n"
            "wb.save(inp)\n"
            "apply(inp, out)\n"
            "wb2 = openpyxl.load_workbook(out)\n"
            "assert wb2['Sheet1']['B1'].value == 26\n"
        )
    }

    result = await _execute_spreadsheet_bundle_tests(
        artifact=skill,
        config=MaintenanceRunConfig(llm_config="mock", model_name="mock-model", max_task_seconds=20),
    )

    assert result.aggregate["passed"] is True
    assert result.unit_case_runs[0].variant == "package_unit"


def test_spreadsheet_extractor_accepts_skill_package_raw_artifact(tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_package_extract", "Double A1 into B1.", source=4, answer=8)
    detail = _detail(task, success=True, score=1.0, retrieved=[], code="")
    raw = {
        "name": "spreadsheet_package_extract_double",
        "kind": "skill_package",
        "description": "Folder skill for doubling A1 into B1.",
        "body": "Use only for Sheet1 A1 to B1 doubling.",
        "interface": {"invocation_contract": {"injection_type": "functional"}},
        "metadata": {
            "domains": ["SpreadsheetBench"],
            "intent_keywords": ["double"],
            "package_format": "skills_md",
            "package_files": {
                "SKILL.md": "---\nname: spreadsheet_package_extract_double\ndescription: Double A1 into B1.\n---\n\n# Double A1\nUse only when B1 should be A1*2.",
                "scripts/apply.py": "import openpyxl\n",
            },
            "bundle_files": {"run_tests.py": "print('ok')\n"},
        },
        "dependencies": [],
    }

    assert _spreadsheet_extraction_rubric_failures([raw], limits=_spreadsheet_extract_limits()) == []
    artifact = _coerce_spreadsheet_artifact(raw, detail=detail)
    assert artifact is not None
    assert artifact.kind == "skill_package"
    assert artifact.injection_type() == "functional"
    assert artifact.metadata["package_files"]["SKILL.md"].startswith("---")


async def test_spreadsheet_skill_format_function_filters_out_package(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_format_function", "Double A1 into B1.", source=4, answer=8)
    detail = _detail(task, success=True, score=1.0, retrieved=[], code="")

    async def fake_ask_json(**kwargs: Any) -> Dict[str, Any]:
        assert kwargs["role"] == "spreadsheet_extractor"
        assert '"skill_format": "function"' in kwargs["user"]
        return {
            "artifacts": [
                {
                    "name": "spreadsheet_format_package",
                    "kind": "skill_package",
                    "description": "Folder skill.",
                    "body": "Use package.",
                    "metadata": {
                        "domains": ["SpreadsheetBench"],
                        "package_files": {"SKILL.md": "# Skill", "scripts/apply.py": "print('x')"},
                    },
                }
            ]
        }

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._compat_ask_json", fake_ask_json)
    artifacts = await _extract_spreadsheet_skills_from_detail(
        detail,
        store=ArtifactStore(),
        config=MaintenanceRunConfig(llm_config="mock", extra={"spreadsheet_skill_format": "function"}),
        task_index=0,
    )

    assert artifacts == []


async def test_spreadsheet_skill_format_folder_heuristic_creates_package(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_format_folder", "Double A1 into B1.", source=4, answer=8)
    detail = _detail(
        task,
        success=True,
        score=1.0,
        retrieved=[],
        code="import openpyxl\nwb = openpyxl.load_workbook(OUTPUT_XLSX)\nws = wb['Sheet1']\nws['B1'] = ws['A1'].value * 2\nwb.save(OUTPUT_XLSX)",
    )

    async def fail_ask_json(**kwargs: Any) -> Dict[str, Any]:
        raise RuntimeError("force heuristic")

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._compat_ask_json", fail_ask_json)
    artifacts = await _extract_spreadsheet_skills_from_detail(
        detail,
        store=ArtifactStore(),
        config=MaintenanceRunConfig(llm_config="mock", extra={"spreadsheet_skill_format": "folder"}),
        task_index=0,
    )

    assert artifacts
    assert artifacts[0].kind == "skill_package"
    assert artifacts[0].metadata["package_format"] == "skills_md"
    assert "scripts/apply.py" in artifacts[0].metadata["package_files"]


async def test_spreadsheet_folder_extractor_writes_package_files(monkeypatch, tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_folder_terminal", "Double A1 into B1.", source=4, answer=8)
    detail = _detail(
        task,
        success=True,
        score=1.0,
        retrieved=[],
        code="",
    )
    detail["runs"][0]["trace"]["notebook_turns"] = [
        {
            "turn_index": 0,
            "code": """python3 << 'PY'
import openpyxl, os
wb = openpyxl.load_workbook(os.environ["INPUT_XLSX"])
ws = wb.active
ws["B1"] = ws["A1"].value * 2
wb.save(os.environ["OUTPUT_XLSX"])
PY""",
            "stdout": "saved",
            "stderr": "",
            "returncode": 0,
        }
    ]

    async def fail_ask_json(**kwargs: Any) -> Dict[str, Any]:
        raise AssertionError("folder extractor should not use JSON artifact output")

    async def fake_ask_text_llm(**kwargs: Any) -> TextLLMResponse:
        assert "write skill package files" in kwargs["system"].lower()
        return TextLLMResponse(
            content="""```bash
mkdir -p skills/double_a1_to_b1/scripts bundles/double_a1_to_b1 reports
cat > skills/double_a1_to_b1/SKILL.md <<'EOF'
---
name: double_a1_to_b1
description: Double numeric A1 into B1.
---

# Double A1 To B1
Use when Sheet1 has a numeric value in A1 and the task asks to write twice that value into B1.
Non-applicability: unrelated sheets or target ranges.
EOF
cat > skills/double_a1_to_b1/scripts/apply.py <<'EOF'
from __future__ import annotations
import sys
import openpyxl

def apply(input_xlsx, output_xlsx):
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.active
    ws["B1"] = ws["A1"].value * 2
    wb.save(output_xlsx)

if __name__ == "__main__":
    apply(sys.argv[1], sys.argv[2])
EOF
cat > bundles/double_a1_to_b1/run_tests.py <<'EOF'
from pathlib import Path
import sys
ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "skills" / "double_a1_to_b1" / "scripts"))
import apply
assert hasattr(apply, "apply")
EOF
cat > reports/double_a1_to_b1.txt <<'EOF'
# NAME
double_a1_to_b1
# DESCRIPTION
Double numeric A1 into B1.
# INTENT_KEYWORDS
double,A1,B1
# SOURCE_TASK_IDS
sheet_folder_terminal
# NON_APPLICABILITY
Do not use for unrelated sheets or target ranges.
EOF
```""",
            prompt_tokens=100,
            completion_tokens=100,
            model_name="mock-model",
            api_style="mock",
        )

    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter._compat_ask_json", fail_ask_json)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.maintenance.adapter.ask_text_llm", fake_ask_text_llm)
    artifacts = await _extract_spreadsheet_skills_from_detail(
        detail,
        store=ArtifactStore(),
        config=MaintenanceRunConfig(
            llm_config="mock",
            model_name="mock-model",
            extra={"spreadsheet_skill_format": "folder"},
        ),
        task_index=0,
    )

    assert len(artifacts) == 1
    artifact = artifacts[0]
    assert artifact.name == "double_a1_to_b1"
    assert artifact.kind == "skill_package"
    assert artifact.injection_type() == "functional"
    assert "SKILL.md" in artifact.metadata["package_files"]
    assert "scripts/apply.py" in artifact.metadata["package_files"]
    assert "run_tests.py" in artifact.metadata["bundle_files"]


async def test_spreadsheet_macro_promotes_dedupes_and_filters(tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_train_3", "Double A1 into B1.")
    detail = _detail(task, success=True, score=1.0, retrieved=[], code="ws['B1']=ws['A1'].value*2")
    from academic.benchmarks.spreadsheet.adapter import _spreadsheet_case_from_task

    pending = SkillArtifact(
        name="spreadsheet_double_pending",
        kind="executable_tool",
        description="Double source cell.",
        body="Double source cell with openpyxl.",
        metadata={
            "domains": ["SpreadsheetBench", "formula_generation"],
            "allowed_tools": ["openpyxl"],
            "intent_keywords": ["double"],
            "source_task_ids": ["sheet_train_3"],
            "instruction_type": "formula_generation",
        },
        status="pending",
    )
    pending.bundle.positive_cases.append(
        _spreadsheet_case_from_task(
            detail=detail,
            skill_name=pending.name,
            polarity="positive",
            reason="source success",
            source="distilled_success",
            confidence=0.8,
        )
    )
    duplicate = copy.deepcopy(pending)
    duplicate.name = "spreadsheet_double_duplicate"
    duplicate.status = "active"
    duplicate.success_count = 0
    harmful = SkillArtifact(
        name="spreadsheet_harmful",
        kind="workflow_guardrail_card",
        description="Bad generic sheet rule.",
        body="Always write zero.",
        metadata={"domains": ["SpreadsheetBench"], "allowed_tools": ["openpyxl"], "intent_keywords": ["zero"]},
        status="active",
    )
    store = ArtifactStore([pending, duplicate, harmful])
    pending.evidence.helpful_cases.append(
        {
            "task_id": "sheet_train_3_reuse",
            "judgment": "helpful",
            "effect_type": "correctness_gain",
            "confidence": 0.9,
        }
    )
    store.add_test_result(
        SkillTestResult(
            result_id="spreadsheet_double_pending:passing_bundle",
            skill_name="spreadsheet_double_pending",
            skill_version=pending.version,
            bundle_id=pending.bundle.bundle_id,
            bundle_version=pending.bundle.bundle_version,
            run_label="spreadsheet_bundle_unit",
            aggregate={"passed": True, "pass_all_tests": True},
        )
    )
    credit = [
        {"skill_name": "spreadsheet_harmful", "judgment": "harmful", "confidence": 0.9},
        {"skill_name": "spreadsheet_harmful", "judgment": "harmful", "confidence": 0.95},
    ]

    report = await SpreadsheetMaintenanceAdapter().run_macro_maintenance(
        window_details=[detail],
        all_train_details=[detail],
        credit_events=credit,
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        window_index=0,
        final_window=True,
    )

    assert store.get("spreadsheet_double_pending").status == "active"
    assert store.get("spreadsheet_harmful").status == "disabled"
    assert report["overlap_refactor"]["promoted_pending_skills"] == ["spreadsheet_double_pending"]
    assert "spreadsheet_harmful" in report["overlap_refactor"]["filtered_skills"]


async def test_spreadsheet_macro_requires_posterior_helpful_credit_for_promotion(tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_train_no_credit", "Double A1 into B1.")
    detail = _detail(task, success=True, score=1.0, retrieved=[], code="ws['B1']=ws['A1'].value*2")
    from academic.benchmarks.spreadsheet.adapter import _spreadsheet_case_from_task

    pending = SkillArtifact(
        name="spreadsheet_pending_without_helpful_credit",
        kind="workflow_guardrail_card",
        description="Pending skill without posterior helpful credit.",
        body="Double A1 into B1.",
        metadata={
            "domains": ["SpreadsheetBench"],
            "allowed_tools": ["openpyxl"],
            "source_task_ids": ["sheet_train_no_credit"],
            "source_success": True,
        },
        status="pending",
    )
    pending.bundle.positive_cases.append(
        _spreadsheet_case_from_task(
            detail=detail,
            skill_name=pending.name,
            polarity="positive",
            reason="source success",
            source="distilled_success",
            confidence=0.8,
        )
    )
    store = ArtifactStore([pending])
    store.add_test_result(
        SkillTestResult(
            result_id="spreadsheet_pending_without_helpful_credit:passing_bundle",
            skill_name=pending.name,
            skill_version=pending.version,
            bundle_id=pending.bundle.bundle_id,
            bundle_version=pending.bundle.bundle_version,
            run_label="spreadsheet_bundle_unit",
            aggregate={"passed": True, "pass_all_tests": True},
        )
    )

    report = await SpreadsheetMaintenanceAdapter().run_macro_maintenance(
        window_details=[detail],
        all_train_details=[detail],
        credit_events=[],
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        window_index=0,
        final_window=True,
    )

    artifact = store.get(pending.name)
    assert artifact.status == "pending"
    assert artifact.metadata["promotion_blocked_reason"] == "requires_successful_source_posterior_helpful_credit_and_passing_bundle_test"
    assert artifact.metadata["promotion_helpful_credit_count"] == 0
    assert report["overlap_refactor"]["promoted_pending_skills"] == []


async def test_spreadsheet_macro_does_not_promote_failure_derived_pending(tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_failed_source", "Double A1 into B1.")
    detail = _detail(task, success=False, score=0.0, retrieved=[], code="ws['B1']=0")
    from academic.benchmarks.spreadsheet.adapter import _spreadsheet_case_from_task

    pending = SkillArtifact(
        name="spreadsheet_failed_pending",
        kind="workflow_guardrail_card",
        description="Failure-derived repair hint.",
        body="Avoid writing zero.",
        metadata={
            "domains": ["SpreadsheetBench"],
            "allowed_tools": ["openpyxl"],
            "source_task_ids": ["sheet_failed_source"],
            "source_success": False,
        },
        status="pending",
    )
    pending.bundle.positive_cases.append(
        _spreadsheet_case_from_task(
            detail=detail,
            skill_name=pending.name,
            polarity="positive",
            reason="incorrect legacy positive case on failed source",
            source="distilled_failure_legacy",
            confidence=0.8,
        )
    )
    store = ArtifactStore([pending])

    report = await SpreadsheetMaintenanceAdapter().run_macro_maintenance(
        window_details=[detail],
        all_train_details=[detail],
        credit_events=[],
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        window_index=0,
        final_window=True,
    )

    assert store.get("spreadsheet_failed_pending").status == "pending"
    assert store.get("spreadsheet_failed_pending").metadata["promotion_blocked_reason"] == "requires_successful_source_or_passing_bundle_test"
    assert report["overlap_refactor"]["promoted_pending_skills"] == []


async def test_spreadsheet_macro_filters_single_filter_candidate(tmp_path: Path) -> None:
    task = _task(tmp_path, "sheet_filter_candidate", "Double A1 into B1.")
    detail = _detail(task, success=False, score=0.0, retrieved=["spreadsheet_bad"], code="ws['B1']=0")
    bad = SkillArtifact(
        name="spreadsheet_bad",
        kind="workflow_guardrail_card",
        description="Bad rule.",
        body="Always write zero.",
        metadata={"domains": ["SpreadsheetBench"], "allowed_tools": ["openpyxl"]},
        status="active",
    )
    store = ArtifactStore([bad])
    credit = [
        {
            "skill_name": "spreadsheet_bad",
            "judgment": "harmful",
            "confidence": 0.9,
            "filter_candidate": True,
        }
    ]

    report = await SpreadsheetMaintenanceAdapter().run_macro_maintenance(
        window_details=[detail],
        all_train_details=[detail],
        credit_events=credit,
        store=store,
        config=MaintenanceRunConfig(llm_config="mock"),
        round_index=0,
        window_index=0,
        final_window=True,
    )

    assert store.get("spreadsheet_bad").status == "disabled"
    assert store.get("spreadsheet_bad").metadata["disabled_reason"] == "spreadsheet_macro_filter_candidate_credit"
    assert "spreadsheet_bad" in report["overlap_refactor"]["filtered_skills"]


async def test_spreadsheet_generic_evolve_produces_real_skills_and_retrieves_on_heldout(monkeypatch, tmp_path: Path) -> None:
    train = [
        _task(tmp_path, "train_double", "Double the value in A1 and write it to B1.", source=5, answer=10),
        _task(tmp_path, "train_double_reuse", "Double the value in A1 and write it to B1.", source=6, answer=12),
    ]
    test = [_task(tmp_path, "test_double", "Double the value in A1 and write it to B1.", source=8, answer=16)]
    seen_store_sizes: List[int] = []

    async def fake_run_task(task: BenchmarkTask, **kwargs: Any) -> BenchmarkResult:
        store = kwargs["artifact_store"]
        include_pending = task.task_id == "train_double_reuse"
        retrieved = store.retrieve(
            task.question,
            top_k=kwargs.get("top_k_skills", 5),
            include_pending=include_pending,
        ) if store else []
        seen_store_sizes.append(len(store.all()) if store else 0)
        return BenchmarkResult(
            benchmark="spreadsheet",
            task_id=task.task_id,
            success=True,
            score=1.0,
            metrics={
                "answer_sheet": "Sheet1",
                "answer_position": "B1",
                "checked_cells": 1,
                "mismatched_cells": [],
                "execution_ok": True,
                "returncode": 0,
                "total_tokens": 100,
                "retrieved_skills": [item.name for item in retrieved],
                "prompt_injected_skills": [item.name for item in retrieved],
                "called_skill_functions": [],
            },
            trace={
                "retrieved_skills": [item.name for item in retrieved],
                "prompt_injected_skills": [item.name for item in retrieved],
                "called_skill_functions": [],
                "code": "ws['B1']=ws['A1'].value*2\nwb.save(OUTPUT_XLSX)",
                "stderr": "",
                "stdout": "",
            },
        )

    async def fake_ask_json(**kwargs: Any) -> Dict[str, Any]:
        if kwargs["role"] == "spreadsheet_extractor":
            return {
                "artifacts": [
                    {
                        "name": "spreadsheet_double_a1_to_b1",
                        "kind": "executable_tool",
                        "description": "Double a numeric source cell and write the result with openpyxl.",
                        "body": "Applicability: double A1 into B1. ```python\nws['B1']=ws['A1'].value*2\nwb.save(OUTPUT_XLSX)\n``` Non-applicability: no hard-coded source values.",
                        "metadata": {
                            "domains": ["SpreadsheetBench", "formula_generation"],
                            "allowed_tools": ["openpyxl"],
                            "intent_keywords": ["double", "value"],
                            "source_task_ids": ["train_double"],
                        },
                    }
                ]
            }
        payload = json.loads(kwargs["user"])
        task_id = ((payload.get("task") or {}).get("task_id")) or ""
        candidate_names = [row.get("skill_name") for row in payload.get("candidate_skills") or []]
        if kwargs["role"] == "spreadsheet_credit_assigner" and task_id == "train_double_reuse" and "spreadsheet_double_a1_to_b1" in candidate_names:
            return {
                "task_summary": {"task_id": task_id, "score": 1.0, "success": True},
                "skill_judgments": [
                    {
                        "skill_name": "spreadsheet_double_a1_to_b1",
                        "judgment": "helpful",
                        "effect_type": "correctness_gain",
                        "confidence": 0.9,
                        "reason": "Pending double skill matched and helped the second training case.",
                        "maintenance_actions": [{"action": "keep", "reason": "posterior positive reuse", "target_scope": "double A1"}],
                        "refine_required": False,
                        "filter_candidate": False,
                        "failure_mode": "insufficient_evidence",
                        "evidence_strength": "strong",
                        "attribution_scope": "prompt_influence",
                        "bundle_case_suggestions": [
                            {
                                "polarity": "positive",
                                "reason": "second successful reuse",
                                "source_task_id": task_id,
                                "task_fragment_policy": "reuse_official_fragment",
                            }
                        ],
                        "evidence": {"retrieved": True, "used": False, "relevant_turn_indices": [0], "related_tool_names": ["openpyxl"], "trace_signals": []},
                    }
                ],
            }
        return {"task_summary": {}, "skill_judgments": []}

    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter.run_spreadsheet_task", fake_run_task)
    monkeypatch.setattr("academic.benchmarks.spreadsheet.adapter._ask_json", fake_ask_json)
    runner = OnlineSkillEvolutionRunner(
        adapter=SpreadsheetMaintenanceAdapter(),
        config=MaintenanceRunConfig(
            llm_config="mock",
            model_name="mock-model",
            micro_maintenance_step=1,
            macro_maintenance_step=1,
            test_concurrency=1,
            top_k_skills=3,
            extra={"spreadsheet_pending_skill_fraction": 1.0},
        ),
    )

    summary = await runner.run(train_tasks=train, test_tasks=test)

    assert summary["store_snapshot"]["n_active"] >= 1
    assert summary["skills"]
    assert summary["micro_maintenance_reports"][0]["extraction_reports"]
    assert summary["test_details"][0]["runs"][0]["metrics"]["retrieved_skills"] == ["spreadsheet_double_a1_to_b1"]
    assert max(seen_store_sizes) >= 1
